# -*- coding: utf-8 -*-
"""API unit test — MI excel decimals + sheets (no browser)."""
import io
import os
import sys
from pathlib import Path

if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

_ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(_ROOT / 'app'))
os.chdir(str(_ROOT / 'app'))

import sqlite3
import openpyxl
from config import Config
from app import app

results = []


def ok(n, c, d=''):
    results.append((n, c, d))
    print(f'  [{"PASS" if c else "FAIL"}] {n}' + (f' — {d}' if d else ''))


pw = sqlite3.connect(Config.MOCK_DB_PATH).execute(
    "SELECT Sifre FROM sistem_kullanici WHERE KullaniciAdi='admin'"
).fetchone()[0]
c = app.test_client()
c.post('/giris', data={'kullanici': 'admin', 'sifre': pw})

payload = {
    'siparis_no': 'PZM-2026-0012',
    'toplu': [
        {'stok_kod': 'A', 'stok_ad': 'Pigment', 'gerekli_kg': 0.035, 'kullanilabilir_kg': 0.034,
         'net_eksik_kg': 0.001, 'birim': 'KG', 'yeterli': False},
        {'stok_kod': 'B', 'stok_ad': 'Kimyasal', 'gerekli_kg': 0.066, 'kullanilabilir_kg': 1.0,
         'net_eksik_kg': 0, 'birim': 'KG', 'yeterli': True},
        {'stok_kod': 'C', 'stok_ad': 'Katkı', 'gerekli_kg': 0.015, 'kullanilabilir_kg': 0.02,
         'net_eksik_kg': 0, 'birim': 'KG', 'yeterli': True},
    ],
    'detay': [
        {'plan_kodu': 'P1', 'formul_kod': 'F1', 'rv_ad': 'Lacivert', 'stok_kod': 'A',
         'stok_ad': 'Pigment', 'gerekli_kg': 0.035, 'bir_formulde_kg': 0.035, 'birim': 'KG', 'boyut': 'L'},
        {'plan_kodu': 'P2', 'formul_kod': 'F2', 'stok_kod': 'B',
         'stok_ad': 'Kimyasal', 'gerekli_kg': 0.066, 'bir_formulde_kg': 83.45, 'birim': 'KG'},
        {'plan_kodu': 'P3', 'formul_kod': 'F3', 'stok_kod': 'B',
         'stok_ad': 'Kimyasal', 'gerekli_kg': 0.066, 'bir_formulde_kg': 75.20, 'birim': 'KG'},
        {'plan_kodu': 'P4', 'formul_kod': 'F4', 'stok_kod': 'B',
         'stok_ad': 'Kimyasal', 'gerekli_kg': 0.01, 'bir_formulde_kg': 83.45, 'birim': 'KG'},
    ],
    'alt_emir': [
        {'plan_kodu': 'P1', 'boyut': 'L', 'tam_formul_kg': 12.5, 'formul_adedi': 2,
         'uretilecek_kg': 25, 'fazla_kg': 0.5, 'durum': 'Hazır'},
    ],
}

r = c.post('/nexgen/api/pazarlama/mi-excel', json=payload)
ok('status_200', r.status_code == 200, str(r.status_code))
ok('xlsx_mime', 'spreadsheetml' in (r.content_type or ''), r.content_type)
cd = r.headers.get('Content-Disposition') or ''
ok('filename', 'PZM-2026-0012_Malzeme_Ihtiyac_' in cd and '.xlsx' in cd, cd)

wb = openpyxl.load_workbook(io.BytesIO(r.data), data_only=True)
ok('sheets_3', len(wb.sheetnames) >= 3, str(wb.sheetnames))
ws1 = wb[wb.sheetnames[0]]
hdr = [ws1.cell(1, c).value for c in range(1, 9)]
ok('hdr_siparis_ihtiyac', 'Sipariş İhtiyacı' in hdr and 'Toplam İhtiyaç' not in hdr, str(hdr))
ok('hdr_formul', 'Formülde Kullanılan' in hdr, str(hdr))
# Kolonlar: 1 kod, 2 ad, 3 mevcut stok, 4 formül, 5 sipariş ihtiyacı, 6 eksik, 7 birim, 8 durum
ihtiyac = {ws1.cell(row=i, column=1).value: ws1.cell(row=i, column=5).value for i in range(2, ws1.max_row + 1)}
ok('dec_035', abs(float(ihtiyac['A']) - 0.035) < 1e-12, str(ihtiyac.get('A')))
ok('dec_066', abs(float(ihtiyac['B']) - 0.066) < 1e-12, str(ihtiyac.get('B')))
ok('dec_015', abs(float(ihtiyac['C']) - 0.015) < 1e-12, str(ihtiyac.get('C')))
ok('stok_034', abs(float(ws1.cell(2, 3).value) - 0.034) < 1e-12, str(ws1.cell(2, 3).value))
formul_b = str(ws1.cell(3, 4).value or '')
ok('formul_ab', '75.2' in formul_b and '83.45' in formul_b and formul_b.count('+') == 1, formul_b)
ok('detay_rows', wb[wb.sheetnames[1]].max_row >= 2)
ok('alt_rows', wb[wb.sheetnames[2]].max_row >= 2)

# empty reject
r2 = c.post('/nexgen/api/pazarlama/mi-excel', json={'siparis_no': 'X'})
ok('empty_400', r2.status_code == 400, str(r2.status_code))

# save sample
out = _ROOT / 'backup' / 'screenshots' / 'pzm_mi_excel_api_sample.xlsx'
out.parent.mkdir(parents=True, exist_ok=True)
out.write_bytes(r.data)
print('SAMPLE', out)

passed = sum(1 for _, c, _ in results if c)
print(f'=== API {passed}/{len(results)} PASS ===')
sys.exit(0 if passed == len(results) else 1)
