# -*- coding: utf-8 -*-
"""
Parça 8a: Çin Ofis Excel parse + validation + template üretimi.
"""
import os
from datetime import date, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


TEMPLATE_VERSION = 'v1.0'
REVISION_DATE = '2026-04-23'

# Desteklenen para birimleri
CURRENCIES = ['USD', 'CNY', 'EUR', 'TRY']

# Label sözlüğü — EN ve CN
# Çince: Çin ofisinin gerçek iş dilinde kullandığı sade karşılıklar.
LABELS = {
    # Header & meta
    'title':                    ('CPS Çin Ofis Sipariş Excel',       'CPS 中国办公室订单模板'),
    'language':                 ('Language',                          '语言'),
    'template_version':         ('Template Version',                  '模板版本'),
    'revision_date':            ('Revision Date',                     '修订日期'),

    # INFO sheet
    'info_title':               ('ORDER INFORMATION',                 '订单信息'),
    'order_code':               ('Order Code',                        '订单号'),
    'supplier_name':            ('Supplier Name',                     '供应商名称'),
    'supplier_contact':         ('Supplier Contact',                  '供应商联系人'),
    'total_container_count':    ('Total Container Count',             '集装箱总数'),
    'currency':                 ('Currency',                          '币种'),
    'currency_rate':            ('Currency Rate (optional)',          '汇率（可选）'),
    'shipment_type':            ('Shipment Type',                     '运输方式'),
    'expected_etd':             ('Expected ETD',                      '预计发货日期'),
    'expected_eta':             ('Expected ETA',                      '预计到港日期'),
    'loading_port':             ('Loading Port',                      '装货港'),
    'discharge_port':           ('Discharge Port',                    '卸货港'),
    'notes':                    ('Notes',                             '备注'),
    'required_mark':            ('* required',                        '* 必填'),

    # ITEMS sheet
    'items_title':              ('LINE ITEMS (Container Based)',      '订单明细（按集装箱）'),
    'line_no':                  ('Line #',                            '行号'),
    'container_group':          ('Container Group',                   '集装箱分组'),
    'container_type':           ('Container Type',                    '集装箱类型'),
    'container_qty':            ('Container Qty',                     '集装箱数量'),
    'quality':                  ('Quality',                           '品质等级'),
    'product_name':             ('Product Name',                      '产品名称'),
    'description':              ('Description',                       '备注说明'),
    'qty':                      ('Qty',                               '数量'),
    'unit':                     ('Unit',                              '单位'),
    'unit_price':               ('Unit Price',                        '单价'),
    'weight_kg':                ('Weight (kg)',                       '重量（公斤）'),
    'line_total':               ('Line Total',                        '小计'),
    'total_items':              ('Total Items',                       '项目总数'),
    'total_containers':         ('Total Containers',                  '集装箱总数'),
    'grand_total':              ('Grand Total',                       '总金额'),

    # PAYMENT sheet
    'payment_title':            ('PAYMENT PLAN',                      '付款计划'),
    'installment':              ('#',                                 '序号'),
    'payment_type':             ('Payment Type',                      '付款方式'),
    'ratio':                    ('Ratio (%)',                         '比例 (%)'),
    'amount':                   ('Amount',                            '金额'),
    'due_date':                 ('Due Date',                          '到期日'),
    'trigger':                  ('Trigger Event',                     '触发条件'),
    'total_ratio':              ('Total Ratio (must be 100%)',        '总比例（必须 100%）'),
    'total_amount':             ('Total Amount',                      '总金额'),

    # FILES sheet
    'files_title':              ('FILE REFERENCES',                   '文件清单'),
    'file_name':                ('File Name',                         '文件名'),
    'file_type':                ('File Type',                         '文件类型'),
    'required':                 ('Required (Yes/No)',                 '是否必需 (是/否)'),
    'file_description':         ('Description',                       '说明'),

    # UI actions (Excel'de kullanılmıyor ama label seti tutarlı olsun)
    'preview':                  ('Preview',                           '预览'),
    'create_order':             ('Create Order',                      '创建订单'),
    'continue_btn':             ('Continue',                          '继续'),
    'success':                  ('Success',                           '成功'),
    'cancel':                   ('Cancel',                            '取消'),
}


# Dropdown listeleri
SHIPMENT_TYPES = ['SEA', 'AIR', 'LAND', 'DHL']
CONTAINER_TYPES = ['20GP', '40GP', '40HQ', 'LCL', 'Break Bulk']
QUALITY_TYPES = ['Main', 'Sample', 'Other']
UNITS = ['kg', 'pair', 'pcs', 'box', 'ton']
PAYMENT_TYPES = ['Advance TT', 'Pre-shipment TT', 'Post-shipment TT', 'L/C', 'D/A', 'D/P', 'Cash', 'Check', 'Other']
FILE_TYPES = ['Proforma', 'Invoice', 'Image', 'Test Report', 'Certificate', 'Contract', 'Other']
YES_NO = ['Yes', 'No']


# =========================================================
# PARSE
# =========================================================
def parse_excel(filepath):
    """
    Excel'i parse eder, structured dict döndürür.
    Hatalar raise etmez, bir kısmı çıkarıp hata listesi döner.
    """
    data = {
        '_filepath': filepath,
        '_lang': 'CN',
        '_template_version': 'v1.0',
        '_hatalar': [],
        'INFO': {},
        'ITEMS': [],
        'PAYMENT': [],
        'FILES': [],
    }

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        data['_hatalar'].append(f'Excel dosyası okunamadı: {e}')
        return data

    # Template versiyonu
    if 'INFO' in wb.sheetnames:
        info = wb['INFO']
        # Versiyon ve dil (sabit hücrelerden)
        try:
            data['_lang'] = (info['B2'].value or 'CN').strip().upper()
            if data['_lang'] not in ('EN', 'CN'):
                data['_lang'] = 'CN'
        except Exception:
            pass
        try:
            data['_template_version'] = (info['B3'].value or 'v1.0').strip()
        except Exception:
            pass

        # INFO alanları — B sütunundan sırayla oku (6-17 arası)
        info_fields = [
            ('Order Code',            6),
            ('Supplier Name',         7),
            ('Supplier Contact',      8),
            ('Total Container Count', 9),
            ('Currency',              10),
            ('Currency Rate',         11),
            ('Shipment Type',         12),
            ('Expected ETD',          13),
            ('Expected ETA',          14),
            ('Loading Port',          15),
            ('Discharge Port',        16),
            ('Notes',                 17),
        ]
        for key, row in info_fields:
            val = info[f'B{row}'].value
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d')
            elif isinstance(val, date):
                val = val.strftime('%Y-%m-%d')
            data['INFO'][key] = val
    else:
        data['_hatalar'].append('INFO sheet bulunamadı.')
        return data

    # ITEMS
    if 'ITEMS' in wb.sheetnames:
        ws = wb['ITEMS']
        for row in ws.iter_rows(min_row=3, max_row=50, values_only=True):
            if not row or not row[0]:
                continue
            if not any(row[5:]):  # Product Name ve sonraki boş
                continue
            item = {
                'Line #':           row[0],
                'Container Group':  row[1],
                'Container Type':   row[2],
                'Container Qty':    row[3],
                'Quality':          row[4],
                'Product Name':     row[5],
                'Description':      row[6],
                'Qty':              row[7],
                'Unit':             row[8],
                'Unit Price':       row[9],
                'Weight (kg)':      row[10],
            }
            data['ITEMS'].append(item)

    # PAYMENT
    if 'PAYMENT' in wb.sheetnames:
        ws = wb['PAYMENT']
        for row in ws.iter_rows(min_row=3, max_row=15, values_only=True):
            if not row or not row[0]:
                continue
            if not row[1]:  # Type yoksa atla
                continue
            due = row[4]
            if isinstance(due, (date, datetime)):
                due = due.strftime('%Y-%m-%d')
            data['PAYMENT'].append({
                '#':        row[0],
                'Type':     row[1],
                'Ratio':    row[2],
                'Amount':   row[3],
                'Due Date': due,
                'Trigger':  row[5],
                'Notes':    row[6],
            })

    # FILES
    if 'FILES' in wb.sheetnames:
        ws = wb['FILES']
        for row in ws.iter_rows(min_row=3, max_row=20, values_only=True):
            if not row or not row[0]:
                continue
            if not row[1]:
                continue
            data['FILES'].append({
                '#':            row[0],
                'File Name':    row[1],
                'Type':         row[2],
                'Required':     row[3],
                'Description':  row[4],
            })

    return data


# =========================================================
# VALIDATION
# =========================================================
def validate(data):
    """
    Parse edilmiş data'yı doğrular. (hatalar, uyarilar) tuple döndürür.
    """
    hatalar = list(data.get('_hatalar', []))
    uyarilar = []

    # Template versiyon kontrolü
    if data.get('_template_version') != TEMPLATE_VERSION:
        hatalar.append(f'Template versiyonu uyumsuz: "{data.get("_template_version")}" '
                      f'(beklenen: {TEMPLATE_VERSION}). Güncel template indirin.')

    info = data['INFO']

    # INFO zorunlu alanlar
    for key, ad in [('Order Code', 'Sipariş Kodu'), ('Supplier Name', 'Tedarikçi'),
                    ('Total Container Count', 'Konteyner adedi'),
                    ('Currency', 'Para birimi'), ('Shipment Type', 'Nakliye tipi')]:
        if not info.get(key):
            hatalar.append(f'INFO zorunlu alan eksik: {key} ({ad})')

    # Currency geçerliliği
    if info.get('Currency') and info['Currency'] not in CURRENCIES:
        hatalar.append(f'Currency geçersiz: {info["Currency"]} (desteklenen: {", ".join(CURRENCIES)})')

    # Currency Rate negatif olamaz
    try:
        rate = info.get('Currency Rate')
        if rate is not None and rate != '' and float(rate) < 0:
            hatalar.append(f'Currency Rate negatif olamaz: {rate}')
    except (TypeError, ValueError):
        if info.get('Currency Rate'):
            hatalar.append(f'Currency Rate sayısal olmalı: {info["Currency Rate"]}')

    # ITEMS
    items = data['ITEMS']
    if not items:
        hatalar.append('ITEMS sheet boş. En az 1 kalem zorunlu.')
    else:
        toplam_cnt = 0
        for i, it in enumerate(items, 1):
            if not it.get('Product Name'):
                hatalar.append(f'ITEMS #{i}: Product Name zorunlu.')
            if not it.get('Qty') or float(it.get('Qty') or 0) <= 0:
                hatalar.append(f'ITEMS #{i}: Qty sıfırdan büyük olmalı.')
            if not it.get('Unit Price') or float(it.get('Unit Price') or 0) <= 0:
                hatalar.append(f'ITEMS #{i}: Unit Price sıfırdan büyük olmalı.')
            if it.get('Container Type') and it['Container Type'] not in CONTAINER_TYPES:
                uyarilar.append(f'ITEMS #{i}: Container Type bilinmiyor: {it["Container Type"]}')
            if it.get('Quality') and it['Quality'] not in QUALITY_TYPES:
                uyarilar.append(f'ITEMS #{i}: Quality bilinmiyor: {it["Quality"]}')
            try:
                toplam_cnt += int(it.get('Container Qty') or 0)
            except (TypeError, ValueError):
                pass

        # Toplam konteyner kontrolü
        if info.get('Total Container Count'):
            try:
                info_cnt = int(info['Total Container Count'])
                if toplam_cnt > 0 and toplam_cnt != info_cnt:
                    uyarilar.append(f'Konteyner sayısı uyumsuz: ITEMS toplamı {toplam_cnt}, '
                                   f'INFO {info_cnt}')
            except (TypeError, ValueError):
                pass

    # PAYMENT — boş olabilir, ama doluysa toplam %100
    payments = data['PAYMENT']
    if not payments:
        uyarilar.append('PAYMENT sheet boş. Ödeme planı import sonrası manuel eklenmelidir.')
    else:
        try:
            toplam_oran = sum(float(p.get('Ratio', 0) or 0) for p in payments)
            if abs(toplam_oran - 100) > 0.5:
                hatalar.append(f'PAYMENT toplam ratio %{toplam_oran:.1f} (100 olmalı)')
        except (TypeError, ValueError):
            hatalar.append('PAYMENT ratio değerleri sayısal olmalı.')

    # FILES — Required bilgisi
    for f in data['FILES']:
        req = (f.get('Required') or '').strip().lower()
        if req and req not in ('yes', 'no', 'y', 'n', '是', '否'):
            uyarilar.append(f'FILES: "{f.get("File Name")}" Required değeri bilinmiyor: {req}')

    return (hatalar, uyarilar)


# =========================================================
# KUR ÇÖZÜMLEME
# =========================================================
def kur_coz(data):
    """
    Parse edilmiş data'dan kur çözümleme sonucunu döndürür.
    Sistem kuru yoksa ValueError.
    """
    from modules.yonetim.queries import get_kur_by_date

    info = data['INFO']
    pb = info.get('Currency') or 'USD'
    islem_tarih = info.get('Expected ETD') or date.today().strftime('%Y-%m-%d')
    # Excel rate
    excel_rate = info.get('Currency Rate')
    try:
        excel_rate = float(excel_rate) if excel_rate not in (None, '', 0) else None
    except (TypeError, ValueError):
        excel_rate = None

    if pb == 'TRY':
        return {
            'pb': 'TRY', 'kur': 1.0, 'kaynak': 'TRY_NATIVE',
            'islem_tarih': islem_tarih, 'sistem_kur': 1.0,
            'excel_rate': excel_rate, 'fark_yuzde': 0.0,
            'override_mumkun': False, 'buyuk_fark': False,
            'kur_tarih': islem_tarih, 'bulma_tipi': 'TRY_NATIVE',
        }

    sistem_kur, bulma_tipi, kur_tarih = get_kur_by_date(pb, islem_tarih)
    if not sistem_kur:
        raise ValueError(
            f'{pb} için {islem_tarih} tarihinde sistemde kur bulunamadı. '
            f'Yönetim /yonetim/kur sayfasından kur girişi yapmalıdır.'
        )

    if not excel_rate or excel_rate <= 0:
        return {
            'pb': pb, 'kur': sistem_kur, 'kaynak': 'SISTEM_OTOMATIK',
            'islem_tarih': islem_tarih, 'sistem_kur': sistem_kur,
            'excel_rate': None, 'fark_yuzde': 0.0,
            'override_mumkun': False, 'buyuk_fark': False,
            'kur_tarih': kur_tarih, 'bulma_tipi': bulma_tipi,
        }

    fark_yuzde = abs(excel_rate - sistem_kur) / sistem_kur * 100
    return {
        'pb': pb, 'kur': sistem_kur, 'kaynak': 'SISTEM_OTOMATIK',  # default
        'islem_tarih': islem_tarih, 'sistem_kur': sistem_kur,
        'excel_rate': excel_rate, 'fark_yuzde': round(fark_yuzde, 2),
        'override_mumkun': True, 'buyuk_fark': fark_yuzde > 10,
        'kur_tarih': kur_tarih, 'bulma_tipi': bulma_tipi,
    }


# =========================================================
# TEMPLATE ÜRETİMİ
# =========================================================
def _label(key, lang):
    """LABELS sözlüğünden dile göre çeker."""
    e, c = LABELS.get(key, ('?', '?'))
    return e if lang == 'EN' else c


def _style_header(cell):
    cell.font = Font(bold=True, size=11, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='F97316')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='CC6B1A')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_required(cell):
    cell.fill = PatternFill('solid', fgColor='FEF3C7')
    thin = Side(border_style='thin', color='999999')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_label(cell):
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill('solid', fgColor='F3F4F6')
    thin = Side(border_style='thin', color='D1D5DB')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _formula_label(key):
    """
    _LABELS sheet'inden VLOOKUP ile label çeken formül.
    Key'i row bazında bulur.
    """
    return f'=VLOOKUP("{key}",_LABELS!$A:$C,IF(LANG="EN",2,3),FALSE)'


def template_olustur(hedef_path, doldurulmus=False):
    """
    Boş veya EVA senaryosu doldurulmuş template üretir.
    """
    wb = Workbook()

    # _LABELS sheet (gizli)
    ws_labels = wb.create_sheet('_LABELS')
    ws_labels.append(['Key', 'EN', 'CN'])
    for key, (en, cn) in LABELS.items():
        ws_labels.append([key, en, cn])
    ws_labels.sheet_state = 'hidden'

    # Ana aktif sheet'i sil
    del wb['Sheet']

    # INFO sheet
    ws = wb.create_sheet('INFO', 0)
    # A1: başlık, B1 boş (gelecek için)
    ws['A1'] = _label('title', 'CN')
    ws['A1'].font = Font(bold=True, size=14, color='F97316')
    ws.merge_cells('A1:C1')

    ws['A2'] = _label('language', 'CN')
    _style_label(ws['A2'])
    ws['B2'] = 'CN'  # default
    ws['B2'].fill = PatternFill('solid', fgColor='FEF3C7')
    ws['B2'].font = Font(bold=True, size=11, color='F97316')

    ws['A3'] = _label('template_version', 'CN')
    _style_label(ws['A3'])
    ws['B3'] = TEMPLATE_VERSION

    ws['A4'] = _label('revision_date', 'CN')
    _style_label(ws['A4'])
    ws['B4'] = REVISION_DATE

    # Named range LANG
    defname = DefinedName('LANG', attr_text=f"INFO!$B$2")
    wb.defined_names['LANG'] = defname

    # ORDER INFORMATION başlığı
    ws['A5'] = _label('info_title', 'CN')
    ws['A5'].font = Font(bold=True, size=12)
    ws.merge_cells('A5:C5')

    # Alanlar — sabit sıra (parse_excel ile eşleşir)
    field_defs = [
        ('order_code',             6,  True,  'CN-OF-2026-001' if doldurulmus else ''),
        ('supplier_name',          7,  True,  'Dongguan Light Industry' if doldurulmus else ''),
        ('supplier_contact',       8,  False, 'Mr. Wang +86 139-5555-5555' if doldurulmus else ''),
        ('total_container_count',  9,  True,  3 if doldurulmus else ''),
        ('currency',               10, True,  'USD' if doldurulmus else 'USD'),
        ('currency_rate',          11, False, 32.10 if doldurulmus else ''),
        ('shipment_type',          12, True,  'SEA' if doldurulmus else 'SEA'),
        ('expected_etd',           13, False, '2026-05-15' if doldurulmus else ''),
        ('expected_eta',           14, False, '2026-06-20' if doldurulmus else ''),
        ('loading_port',           15, False, 'Shenzhen' if doldurulmus else ''),
        ('discharge_port',         16, False, 'Istanbul' if doldurulmus else ''),
        ('notes',                  17, False, ('2 ana + 1 numune konteyner. Numune için ayrı kalite raporu gönderildi. '
                                              'Navlun $1850 konteyner başı. Vergi %18 tahmini.')
                                              if doldurulmus else ''),
    ]
    for key, row, zorunlu, deger in field_defs:
        # Label — formüllü dil desteği
        ws[f'A{row}'] = _formula_label(key)
        _style_label(ws[f'A{row}'])
        # Değer
        ws[f'B{row}'] = deger
        if zorunlu:
            _style_required(ws[f'B{row}'])
        if zorunlu:
            ws[f'C{row}'] = _label('required_mark', 'CN')
            ws[f'C{row}'].font = Font(italic=True, color='DC2626', size=9)

    # Data Validation — Currency
    dv_cur = DataValidation(type='list', formula1=f'"{",".join(CURRENCIES)}"', allow_blank=False)
    ws.add_data_validation(dv_cur)
    dv_cur.add('B10')

    # Data Validation — Shipment
    dv_ship = DataValidation(type='list', formula1=f'"{",".join(SHIPMENT_TYPES)}"', allow_blank=False)
    ws.add_data_validation(dv_ship)
    dv_ship.add('B12')

    # Language DV
    dv_lang = DataValidation(type='list', formula1='"CN,EN"', allow_blank=False)
    ws.add_data_validation(dv_lang)
    dv_lang.add('B2')

    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 14

    # =========================================================
    # ITEMS sheet
    # =========================================================
    ws_items = wb.create_sheet('ITEMS')
    ws_items['A1'] = _label('items_title', 'CN')
    ws_items['A1'].font = Font(bold=True, size=14, color='F97316')
    ws_items.merge_cells('A1:K1')

    headers = ['line_no', 'container_group', 'container_type', 'container_qty',
               'quality', 'product_name', 'description', 'qty', 'unit',
               'unit_price', 'weight_kg']
    for col_idx, key in enumerate(headers, 1):
        cell = ws_items.cell(row=2, column=col_idx)
        cell.value = _formula_label(key)
        _style_header(cell)

    # Sütun genişlikleri
    widths = [8, 16, 16, 14, 12, 26, 26, 12, 10, 12, 14]
    for i, w in enumerate(widths, 1):
        ws_items.column_dimensions[get_column_letter(i)].width = w

    # DV: Container Type, Quality, Unit
    dv_ct = DataValidation(type='list', formula1=f'"{",".join(CONTAINER_TYPES)}"', allow_blank=True)
    ws_items.add_data_validation(dv_ct)
    dv_ct.add('C3:C50')
    dv_q = DataValidation(type='list', formula1=f'"{",".join(QUALITY_TYPES)}"', allow_blank=True)
    ws_items.add_data_validation(dv_q)
    dv_q.add('E3:E50')
    dv_u = DataValidation(type='list', formula1=f'"{",".join(UNITS)}"', allow_blank=True)
    ws_items.add_data_validation(dv_u)
    dv_u.add('I3:I50')

    # Örnek veri
    if doldurulmus:
        ws_items['A3'] = 1; ws_items['B3'] = 'Main-A'; ws_items['C3'] = '40HQ'
        ws_items['D3'] = 2; ws_items['E3'] = 'Main'
        ws_items['F3'] = 'EVA Grade A Compound'
        ws_items['G3'] = 'White, shore 55'
        ws_items['H3'] = 40000; ws_items['I3'] = 'kg'
        ws_items['J3'] = 2.10; ws_items['K3'] = 40000

        ws_items['A4'] = 2; ws_items['B4'] = 'Sample-B'; ws_items['C4'] = '20GP'
        ws_items['D4'] = 1; ws_items['E4'] = 'Sample'
        ws_items['F4'] = 'EVA Grade B Compound'
        ws_items['G4'] = 'Trial run, black'
        ws_items['H4'] = 18000; ws_items['I4'] = 'kg'
        ws_items['J4'] = 2.35; ws_items['K4'] = 18000

    # =========================================================
    # PAYMENT sheet
    # =========================================================
    ws_pay = wb.create_sheet('PAYMENT')
    ws_pay['A1'] = _label('payment_title', 'CN')
    ws_pay['A1'].font = Font(bold=True, size=14, color='F97316')
    ws_pay.merge_cells('A1:G1')

    pay_headers = ['installment', 'payment_type', 'ratio', 'amount',
                   'due_date', 'trigger', 'notes']
    for col_idx, key in enumerate(pay_headers, 1):
        cell = ws_pay.cell(row=2, column=col_idx)
        cell.value = _formula_label(key)
        _style_header(cell)

    widths = [8, 22, 12, 14, 14, 28, 24]
    for i, w in enumerate(widths, 1):
        ws_pay.column_dimensions[get_column_letter(i)].width = w

    dv_pt = DataValidation(type='list', formula1=f'"{",".join(PAYMENT_TYPES)}"', allow_blank=True)
    ws_pay.add_data_validation(dv_pt)
    dv_pt.add('B3:B15')

    if doldurulmus:
        ws_pay['A3'] = 1; ws_pay['B3'] = 'Advance TT'; ws_pay['C3'] = 30
        ws_pay['D3'] = 37890; ws_pay['E3'] = '2026-04-25'
        ws_pay['F3'] = 'Order Confirmation'; ws_pay['G3'] = 'PO sonrası 5 gün'

        ws_pay['A4'] = 2; ws_pay['B4'] = 'Pre-shipment TT'; ws_pay['C4'] = 40
        ws_pay['D4'] = 50520; ws_pay['E4'] = '2026-05-12'
        ws_pay['F4'] = 'Before Loading'; ws_pay['G4'] = 'B/L kopyası öncesi'

        ws_pay['A5'] = 3; ws_pay['B5'] = 'Post-shipment'; ws_pay['C5'] = 30
        ws_pay['D5'] = 37890; ws_pay['E5'] = '2026-06-15'
        ws_pay['F5'] = 'After B/L Copy'; ws_pay['G5'] = 'Sevk sonrası'

    # =========================================================
    # FILES sheet
    # =========================================================
    ws_f = wb.create_sheet('FILES')
    ws_f['A1'] = _label('files_title', 'CN')
    ws_f['A1'].font = Font(bold=True, size=14, color='F97316')
    ws_f.merge_cells('A1:E1')

    file_headers = ['line_no', 'file_name', 'file_type', 'required', 'file_description']
    for col_idx, key in enumerate(file_headers, 1):
        cell = ws_f.cell(row=2, column=col_idx)
        cell.value = _formula_label(key)
        _style_header(cell)

    widths = [8, 32, 16, 12, 32]
    for i, w in enumerate(widths, 1):
        ws_f.column_dimensions[get_column_letter(i)].width = w

    dv_ft = DataValidation(type='list', formula1=f'"{",".join(FILE_TYPES)}"', allow_blank=True)
    ws_f.add_data_validation(dv_ft)
    dv_ft.add('C3:C25')
    dv_yn = DataValidation(type='list', formula1=f'"{",".join(YES_NO)}"', allow_blank=True)
    ws_f.add_data_validation(dv_yn)
    dv_yn.add('D3:D25')

    if doldurulmus:
        ws_f['A3'] = 1; ws_f['B3'] = 'proforma_2026_001.pdf'
        ws_f['C3'] = 'Proforma'; ws_f['D3'] = 'Yes'
        ws_f['E3'] = 'Ana proforma fatura'

        ws_f['A4'] = 2; ws_f['B4'] = 'eva_grade_a_photo.jpg'
        ws_f['C4'] = 'Image'; ws_f['D4'] = 'No'
        ws_f['E4'] = 'Ana ürün görseli'

        ws_f['A5'] = 3; ws_f['B5'] = 'sample_lab_report.pdf'
        ws_f['C5'] = 'Test Report'; ws_f['D5'] = 'Yes'
        ws_f['E5'] = 'Sample lab raporu'

    wb.save(hedef_path)
    return hedef_path
