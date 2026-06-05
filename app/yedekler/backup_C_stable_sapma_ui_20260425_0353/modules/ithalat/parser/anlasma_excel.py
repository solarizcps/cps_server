# -*- coding: utf-8 -*-
"""
CPS DEV - Proforma / Anlasma Excel Parser
==========================================
BLOK 4.6-a

Test dosyasi: Copy of pi-PI20260402-2 lighting.xls (Chengyu Global Trading)
Format desteklemesi:
  - .xls  (eski format, xlrd ile)
  - .xlsx (yeni format, openpyxl ile)

Cikardigi alanlar:
  Zorunlu:
    - PI No         (kaynak_ref, idempotency)
    - Tedarikci ad
    - FOB toplam (USD)
  Opsiyonel:
    - Komisyon     (ayri kalem olarak)
    - Yukleme tarihi
    - Para birimi
    - Urun kalemleri (icerigi not'a yazilir)
    - Teslim sekli (EXW / FOB / CIF)
    - Kur (USD/RMB)
    - Odeme sarti

Karar notlari (user onayli):
  1. Komisyon ayri kalem (KOMISYON tipi)
  2. Teslim sekli kalem not'una yazilir ("Teslim: EXW")
  3. Yukleme tarihi parti_bilgi icinde doner (parti.YuklemeTarih bos ise uygulanir)
  4. Odeme sarti sadece uyarida, otomatik plan olusturulmaz
"""
import logging
import os
import re
from datetime import datetime, date
from modules.ithalat.parser._base import ParseSonuc

log = logging.getLogger("cps.ithalat.parser.anlasma_excel")


# =====================================================================
# SABITLER (genel)
# =====================================================================
# Tip eslemesi - serbest metinden Tip cikart
TIP_MAP = {
    'fob':       'FOB',
    'freight':   'NAVLUN',
    'navlun':    'NAVLUN',
    'gumruk':    'GUMRUK',
    'customs':   'GUMRUK',
    'sigorta':   'SIGORTA',
    'insurance': 'SIGORTA',
    'depo':      'DEPOLAMA',
    'storage':   'DEPOLAMA',
    'komisyon':  'KOMISYON',
    'commission': 'KOMISYON',
    'liman':     'LIMAN',
    'port':      'LIMAN',
}

# Ay adlari - shipment date parse icin
AY_MAP = {
    'january': 1, 'february': 2, 'march': 3, 'april': 4,
    'may': 5, 'june': 6, 'july': 7, 'august': 8,
    'september': 9, 'october': 10, 'november': 11, 'december': 12,
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'sept': 9,
    'oct': 10, 'nov': 11, 'dec': 12,
}

# Baslik satirini bulmak icin anahtar kelimeler
TABLO_BASLIK_ANAHTARLARI = {
    'unit price', 'unitprice', 'quantity', 'qty',
    'amount', 'item', 'goods description', 'price',
}


# =====================================================================
# HUCRE OKUMA ARAYUZU (xls/xlsx ortak)
# =====================================================================
class _SheetProxy:
    """
    xlrd ve openpyxl arasinda tek arayuz.
    Hucre okuma, satir sayisi, sutun sayisi.
    """
    def __init__(self, sheet, motor):
        self.sheet = sheet
        self.motor = motor  # 'xlrd' veya 'openpyxl'
        if motor == 'xlrd':
            self.nrows = sheet.nrows
            self.ncols = sheet.ncols
        else:
            # openpyxl
            self.nrows = sheet.max_row or 0
            self.ncols = sheet.max_column or 0

    def hucre(self, r, c):
        """0-indexli hucre degeri. Bos/hata ise None."""
        try:
            if self.motor == 'xlrd':
                if r >= self.nrows or c >= self.ncols:
                    return None
                v = self.sheet.cell_value(r, c)
                return v if v != '' else None
            else:
                # openpyxl 1-indexli
                if r >= self.nrows or c >= self.ncols:
                    return None
                v = self.sheet.cell(row=r + 1, column=c + 1).value
                return v if v != '' else None
        except Exception:
            return None


def _aktif_sheet_ac(dosya_yol):
    """
    Dosyayi uygun kutuphaneyle ac, en cok dolu olan ilk sheet'i dondur.
    Donus: (_SheetProxy, sheet_ad, uyarilar_listesi) veya (None, None, hatalar)
    """
    ext = os.path.splitext(dosya_yol)[1].lower()

    if ext in ('.xls',):
        try:
            import xlrd
        except ImportError:
            return None, None, ["xlrd kurulu degil. 'pip install xlrd' ile kurun."]
        try:
            wb = xlrd.open_workbook(dosya_yol)
            # En cok satira sahip sheet
            en_iyi_idx = 0
            en_iyi_nrows = 0
            for i in range(wb.nsheets):
                sh = wb.sheet_by_index(i)
                if sh.nrows > en_iyi_nrows:
                    en_iyi_nrows = sh.nrows
                    en_iyi_idx = i
            sheet = wb.sheet_by_index(en_iyi_idx)
            return _SheetProxy(sheet, 'xlrd'), sheet.name, []
        except Exception as e:
            return None, None, [f"xlrd ile acilamadi: {str(e)[:100]}"]

    elif ext in ('.xlsx', '.xlsm'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return None, None, ["openpyxl kurulu degil."]
        try:
            wb = load_workbook(dosya_yol, data_only=True, read_only=True)
            # En cok satira sahip sheet
            en_iyi_ad = None
            en_iyi_nrows = 0
            for ad in wb.sheetnames:
                sh = wb[ad]
                nr = sh.max_row or 0
                if nr > en_iyi_nrows:
                    en_iyi_nrows = nr
                    en_iyi_ad = ad
            if not en_iyi_ad:
                en_iyi_ad = wb.sheetnames[0]
            sheet = wb[en_iyi_ad]
            return _SheetProxy(sheet, 'openpyxl'), en_iyi_ad, []
        except Exception as e:
            return None, None, [f"openpyxl ile acilamadi: {str(e)[:100]}"]

    else:
        return None, None, [f"'{ext}' formati desteklenmiyor"]


# =====================================================================
# YARDIMCILAR
# =====================================================================
def _norm(v):
    """Hucreyi normalize et: kucuk harf, strip, coklu bosluk tekle."""
    if v is None:
        return ''
    return ' '.join(str(v).strip().lower().split())


def _floata_cevir(v):
    """Sayi gibi duran degeri float'a. Basarisizsa None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    # Para sembolleri ve harfler
    for ch in ['$', '€', '₺', '¥', 'USD', 'EUR', 'TRY', 'CNY', 'RMB',
               'usd', 'eur', 'try', 'cny', 'rmb']:
        s = s.replace(ch, '')
    s = s.strip()
    # Turkce format: 1.234,56 -> 1234.56
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


# =====================================================================
# ALAN CIKARICILAR
# =====================================================================
def _pi_no_bul(sp):
    """
    'PI No' / 'P/I No' / 'Proforma No' / 'Invoice No' / 'Commercial Invoice No' /
    'Contract No' / 'Order No' / 'Fatura No' etiketini bul, yanindaki/altindaki
    alfasayisal referansi al.

    Ayrica TEK HUCREDE "Invoice NO.:PI20251218" gibi bitisik formati da yakalar.

    Donus: (ref, (row, col)) veya (None, None)
    """
    ref_kaliplari = [
        r'^pi\s*no',
        r'^p/i\s*no',
        r'^proforma\s*no',
        r'^invoice\s*no',
        r'^commercial\s*invoice\s*no',
        r'^contract\s*no',
        r'^order\s*no',
        r'^fatura\s*no',
    ]
    # Tek hucrede "Label: Value" formati (icinde iki nokta var)
    bitisik_pattern = re.compile(
        r'^(?:pi\s*no|p/i\s*no|proforma\s*no|invoice\s*no\.?|'
        r'commercial\s*invoice\s*no|contract\s*no|order\s*no|fatura\s*no)\s*[:：]\s*([A-Z0-9][A-Z0-9\-_/]{2,})',
        re.I,
    )

    for r in range(min(30, sp.nrows)):
        for c in range(sp.ncols):
            v = _norm(sp.hucre(r, c))
            if not v:
                continue

            # Once bitisik format kontrol ("Invoice NO.:PI20251218")
            m = bitisik_pattern.match(v)
            if m:
                return m.group(1), (r, c)

            # Sonra ayri hucre formati
            if any(re.match(p, v) for p in ref_kaliplari):
                for dr, dc in [(0, 1), (0, 2), (1, 0), (1, 1)]:
                    ref = sp.hucre(r + dr, c + dc)
                    if ref is None:
                        continue
                    s = str(ref).strip()
                    if s and re.match(r'^[A-Z0-9][A-Z0-9\-_/]{2,}$', s, re.I):
                        return s, (r, c)
    return None, None


def _tedarikci_bul(sp):
    """Genelde A1, bazen A1-A3 arasinda ilk 'firma adi' gibi duran hucre."""
    for r in range(min(5, sp.nrows)):
        v = sp.hucre(r, 0)
        if not v:
            continue
        s = str(v).strip()
        n = _norm(s)
        # Eleme kriterleri
        if any(k in n for k in ['tel', 'proforma', 'invoice', 'the seller',
                                 'the buyer', 'pi no', 'date', 'bank']):
            continue
        # Sirket adi tipik karakterler (Co., Ltd, Limited, Inc, ..)
        if len(s) < 5:
            continue
        return s
    return None


def _baslik_satiri_bul(sp):
    """
    Tablo baslik satiri - ekseriyetle 'Amount', 'Quantity', 'Unit Price'
    gibi kelimeler iceren satir. Iki satir birlesik baslik olabilir.
    Donus: baslik_satir_no (0-indexli) veya None.
    """
    en_iyi = None
    en_iyi_puan = 0
    for r in range(min(25, sp.nrows)):
        puan = 0
        satir_metni = []
        for c in range(sp.ncols):
            v = _norm(sp.hucre(r, c))
            satir_metni.append(v)
            for anahtar in TABLO_BASLIK_ANAHTARLARI:
                if anahtar in v:
                    puan += 1
                    break
        if puan > en_iyi_puan:
            en_iyi_puan = puan
            en_iyi = r
    if en_iyi_puan < 2:
        return None
    return en_iyi


def _kolon_harita(sp, baslik_r):
    """
    Baslik + alt baslik satirlarini birlestirerek kolonlari haritala.
    Donus: {sutun_no: alan_adi}
    """
    if baslik_r is None:
        return {}
    harita = {}
    for c in range(sp.ncols):
        ust = _norm(sp.hucre(baslik_r, c))
        alt = _norm(sp.hucre(baslik_r + 1, c)) if baslik_r + 1 < sp.nrows else ''
        birlesik_bosluksuz = (ust + alt).replace(' ', '').replace('\n', '')
        birlesik = (ust + ' ' + alt).strip()

        # USD tutar sutunu
        if 'amount(usd)' in birlesik_bosluksuz or (
                'amount' in ust and 'usd' in ust):
            harita[c] = 'amount_usd'
            continue
        # RMB tutar sutunu
        if 'amount(rmb)' in birlesik_bosluksuz or (
                'amount' in ust and 'rmb' in ust):
            harita[c] = 'amount_rmb'
            continue
        # EUR tutar sutunu
        if 'amount(eur)' in birlesik_bosluksuz or (
                'amount' in ust and 'eur' in ust):
            harita[c] = 'amount_eur'
            continue
        # TRY tutar sutunu
        if 'amount(try)' in birlesik_bosluksuz or (
                'amount' in ust and ('try' in ust or 'tl' in ust)):
            harita[c] = 'amount_try'
            continue
        # Unit price
        if 'unitprice' in birlesik_bosluksuz or 'unit price' in birlesik:
            harita[c] = 'unit_price'
            continue
        # Quantity / pairs
        if 'quantity' in ust or 'qty' in ust:
            harita[c] = 'qty'
            continue
        # Item / kod
        if 'item' in ust or 'item' in alt or 'goods' in ust or 'description' in ust:
            harita[c] = 'item'
            continue
        # Spec
        if 'spec' in ust or 'spec' in alt:
            harita[c] = 'spec'
            continue
        # Unit of measure (pair/pc)
        if 'by pair' in ust or 'pair' in alt or 'pc' in ust:
            harita[c] = 'unit_of_measure'
            continue
    return harita


def _amount_kolon_sec(harita):
    """
    Oncelik: USD > EUR > TRY > RMB
    Donus: (sutun_no, para_birimi) veya (None, None)
    """
    oncelikler = [
        ('amount_usd', 'USD'),
        ('amount_eur', 'EUR'),
        ('amount_try', 'TRY'),
        ('amount_rmb', 'RMB'),
    ]
    for k, para in oncelikler:
        for c, alan in harita.items():
            if alan == k:
                return c, para
    return None, None


def _urun_satirlari_oku(sp, baslik_r, harita, tutar_kol, para_birimi):
    """
    Baslik + 2 satir sonrasindan baslayarak urun ve toplam/komisyon satirlarini
    oku. Total: ve commission satirlarini ayristir.

    Donus: {
        'urunler': [dict, ...],
        'toplam': float | None,
        'komisyon': float | None,
    }
    """
    basla = baslik_r + 2
    urunler = []
    toplam = None
    komisyon = None

    item_kol = next((c for c, k in harita.items() if k == 'item'), None)

    for r in range(basla, sp.nrows):
        satir_metinleri = [_norm(sp.hucre(r, c)) for c in range(sp.ncols)]
        satir_birlesik = ' '.join(satir_metinleri).strip()
        if not satir_birlesik:
            continue

        # Total satiri
        if 'total' in satir_birlesik and tutar_kol is not None:
            v = _floata_cevir(sp.hucre(r, tutar_kol))
            if v and v > 0:
                toplam = v
            continue
        # Commission satiri
        if 'commission' in satir_birlesik and tutar_kol is not None:
            v = _floata_cevir(sp.hucre(r, tutar_kol))
            if v and v > 0:
                komisyon = v
            continue
        # "Amount" label-only row (bos toplam etiketi, ornek: H16='Amount')
        if ('amount' in satir_birlesik and
            'ex-t' not in satir_birlesik and
            not any(h in satir_birlesik for h in ['usd', 'rmb', 'eur', 'try'])):
            # Bu muhtemelen 'Amount' baslik veya etiket satiri - atla
            # Ama tutar kolonunda deger varsa grand total olabilir
            v = _floata_cevir(sp.hucre(r, tutar_kol)) if tutar_kol else None
            if v and v > 0 and toplam is not None and v > toplam:
                # Grand total (urun + komisyon) - loglamak yeterli
                pass
            continue

        # Urun satiri tespit - item dolu + tutar pozitif
        item_val = sp.hucre(r, item_kol) if item_kol is not None else None
        tutar_val = _floata_cevir(sp.hucre(r, tutar_kol)) if tutar_kol is not None else None

        if item_val and tutar_val and tutar_val > 0:
            # Curly quote'lari da temizle
            item_temiz = str(item_val).strip(" '‘’\"“”\t\n")
            urun = {
                'item':  item_temiz,
                'tutar': tutar_val,
                'para_birimi': para_birimi,
            }
            for c, k in harita.items():
                if k in ('spec', 'qty', 'unit_price', 'amount_rmb'):
                    val = sp.hucre(r, c)
                    if val is not None:
                        urun[k] = val
            # D sutunu extra desc (genelde renk/model)
            d_val = sp.hucre(r, 3)
            if d_val and 'extra_desc' not in harita.values():
                urun['extra_desc'] = str(d_val).replace('\n', ' ').strip()
            urunler.append(urun)

    return {'urunler': urunler, 'toplam': toplam, 'komisyon': komisyon}


def _yukleme_tarihi_bul(sp, pi_tarihi_yil=None):
    """
    Shipment date: April. 30  tarzi serbest metin yakala.
    Yil belirtilmemisse PI tarihi yili veya mevcut yil.
    Donus: 'YYYY-MM-DD' veya None.
    """
    for r in range(sp.nrows):
        for c in range(sp.ncols):
            v = sp.hucre(r, c)
            if not v:
                continue
            s = str(v).lower()
            if 'shipment' not in s and 'delivery' not in s and 'loading' not in s:
                continue
            if 'date' not in s:
                continue
            # Ay adi + gun
            for ay_ad, ay_no in AY_MAP.items():
                m = re.search(
                    r'\b' + re.escape(ay_ad) + r'\.?\s+(\d{1,2})(?:st|nd|rd|th)?',
                    s, re.I,
                )
                if m:
                    try:
                        gun = int(m.group(1))
                        if not (1 <= gun <= 31):
                            continue
                        yil = pi_tarihi_yil or datetime.now().year
                        return f"{yil:04d}-{ay_no:02d}-{gun:02d}"
                    except Exception:
                        pass
            # ISO format: 2026-04-30
            m = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', s)
            if m:
                try:
                    y, mth, d = map(int, m.groups())
                    datetime(y, mth, d)  # validate
                    return f"{y:04d}-{mth:02d}-{d:02d}"
                except Exception:
                    pass
    return None


def _pi_tarihi_yil_bul(sp):
    """A34 tarzi 'DATE: 2026/4/9' satirindan yil cikart."""
    for r in range(sp.nrows):
        for c in range(sp.ncols):
            v = sp.hucre(r, c)
            if not v:
                continue
            s = str(v).lower()
            if 'date' not in s:
                continue
            # 2026/4/9 veya 2026-04-09 vs
            m = re.search(r'(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})', s)
            if m:
                try:
                    y = int(m.group(1))
                    if 2000 <= y <= 2100:
                        return y
                except Exception:
                    pass
    return None


def _teslim_sekli_bul(sp):
    """EXW / FOB / CIF / CFR arar."""
    # EXW varyantlari
    exw_patterns = ['ex-t factory', 'exit factory', 'ex factory',
                     'ex works', 'exworks', ' exw ']
    for r in range(sp.nrows):
        for c in range(sp.ncols):
            v = _norm(sp.hucre(r, c))
            if not v:
                continue
            # EXW
            if any(p in v for p in exw_patterns):
                return 'EXW'
            # 'FOB' - standalone (included/not included ayrimi zor)
            if re.search(r'\bfob\b', v):
                # 'not include fob' dedik ediyse EXW olabilir
                if 'not include' in v and 'fob' in v:
                    return 'EXW'
                return 'FOB'
            if re.search(r'\bcif\b', v):
                return 'CIF'
            if re.search(r'\bcfr\b', v) or re.search(r'\bc&f\b', v):
                return 'CFR'
    return None


def _kur_bul(sp):
    """
    USD:RMB 6.80 benzeri kur bilgisini yakala.

    Onemli: "4) usd:rmb based on 6.80" gibi satirlarda '4' madde
    numarasi da sayi olarak yakalanir. Bu yuzden once anlamli
    anahtar kelimelerden (based on, rate, usd:rmb) sonraki sayiyi
    tercih eder.
    """
    # Sirasiyla daha guvenli kaliplar
    kaliplar = [
        r'based\s+on\s+(\d+\.?\d*)',
        r'exchange\s+rate\s*[:=]?\s*(\d+\.?\d*)',
        r'rate\s*[:=]\s*(\d+\.?\d*)',
        r'(?:usd\s*[:/]\s*rmb|rmb\s*[:/]\s*usd)[^\d]{0,20}(\d+\.?\d*)',
    ]
    for r in range(sp.nrows):
        for c in range(sp.ncols):
            v = sp.hucre(r, c)
            if not v:
                continue
            s = str(v).lower()
            # Kurdan bahsediyor mu?
            if not any(k in s for k in ['exchange rate', 'usd:rmb', 'rmb:usd',
                                         'usd/rmb', 'rmb/usd', 'based on']):
                continue
            for kalip in kaliplar:
                m = re.search(kalip, s, re.I)
                if m:
                    try:
                        sayi = float(m.group(1))
                        if 3 <= sayi <= 50:
                            return sayi
                    except Exception:
                        pass
    return None


def _odeme_sarti_bul(sp):
    """'30% deposit, balance before shipment' benzeri metin."""
    aday = None
    for r in range(sp.nrows):
        for c in range(sp.ncols):
            v = sp.hucre(r, c)
            if not v:
                continue
            s = str(v)
            if '%' in s and any(k in s.lower() for k in
                                 ['deposit', 'deposite', 'balance',
                                  't/t', 'tt', 'payment']):
                adayi = s.strip()
                if aday is None or len(adayi) > len(aday):
                    aday = adayi
    return aday


# =====================================================================
# ANA PARSE FONKSIYONU
# =====================================================================
def _guven_skoru_hesapla(
    kalemler, kaynak_ref, tedarikci_ad, para_birimi,
    baslik_r, harita, tutar_kol, urunler, toplam,
    parti_toplam_cift=None,
):
    """
    Guven skoru: 0-100.

    Kurallar:
      - Baslangic 100 puan
      - Her problem puan dusurur
      - Matematik kontrolu FAIL -> puan dusur + otomatik uygulanmayi engelle
      - Mantiksiz deger -> direkt max tavan koyar

    Donus: {
        'skor': int,
        'seviye': 'yuksek' | 'orta' | 'dusuk' | 'red',
        'detay': [{madde, durum, puan}],
        'matematik_ok': bool,
        'mantiksizlik_uyarilari': [str],
    }
    """
    skor = 100
    detay = []
    matematik_ok = True
    mantiksizlik = []
    tavan = 100  # Guvenlik kurali bazli max skor (mantiksizlik olursa dusuyor)

    # --- 1) Header satiri ---
    if baslik_r is None:
        skor -= 20
        detay.append({'madde': 'Header satiri tespit edilemedi', 'durum': 'hata', 'puan': -20})
    else:
        detay.append({'madde': f'Header satiri bulundu (L{baslik_r})',
                      'durum': 'ok', 'puan': 0})

    # --- 2) PI / Invoice No ---
    if not kaynak_ref:
        skor -= 25
        detay.append({'madde': 'PI/Invoice No bulunamadi',
                      'durum': 'hata', 'puan': -25})
    else:
        detay.append({'madde': f'Kaynak referans: {kaynak_ref}',
                      'durum': 'ok', 'puan': 0})

    # --- 3) Tedarikci ---
    if not tedarikci_ad:
        skor -= 10
        detay.append({'madde': 'Tedarikci adi okunamadi',
                      'durum': 'uyari', 'puan': -10})
    else:
        detay.append({'madde': f'Tedarikci: {tedarikci_ad[:50]}',
                      'durum': 'ok', 'puan': 0})

    # --- 4) Para birimi ---
    if not para_birimi:
        skor -= 10
        detay.append({'madde': 'Para birimi belirsiz',
                      'durum': 'uyari', 'puan': -10})
    else:
        detay.append({'madde': f'Para birimi: {para_birimi}',
                      'durum': 'ok', 'puan': 0})

    # --- 5) Kolonlar numerik mi ---
    if urunler:
        miktar_bozuk = sum(1 for u in urunler
                           if u.get('qty') is not None
                           and not isinstance(u['qty'], (int, float)))
        if miktar_bozuk > 0:
            skor -= 20
            detay.append({'madde': f'{miktar_bozuk} satirda miktar numerik degil',
                          'durum': 'hata', 'puan': -20})

        fiyat_bozuk = sum(1 for u in urunler
                          if u.get('unit_price') is not None
                          and not isinstance(u['unit_price'], (int, float)))
        if fiyat_bozuk > 0:
            skor -= 15
            detay.append({'madde': f'{fiyat_bozuk} satirda birim fiyat numerik degil',
                          'durum': 'hata', 'puan': -15})

        tutar_bozuk = sum(1 for u in urunler
                          if u.get('tutar') is not None
                          and not isinstance(u['tutar'], (int, float)))
        if tutar_bozuk > 0:
            skor -= 20
            detay.append({'madde': f'{tutar_bozuk} satirda tutar numerik degil',
                          'durum': 'hata', 'puan': -20})

    # --- 6) MATEMATIK KONTROLU (KRITIK - kuralina gore: fail ise otomatik uygulama YOK) ---
    math_fail_var = False
    if urunler:
        for u in urunler:
            miktar = u.get('qty')
            fiyat = u.get('unit_price')
            utu = u.get('tutar')
            if (miktar and fiyat and utu
                    and isinstance(miktar, (int, float))
                    and isinstance(fiyat, (int, float))
                    and isinstance(utu, (int, float))
                    and miktar > 0 and fiyat > 0 and utu > 0):
                hesap = miktar * fiyat
                fark = abs(hesap - utu)
                tolerans = max(utu * 0.01, 0.5)  # %1 veya 0.5 birim
                if fark > tolerans:
                    math_fail_var = True
                    detay.append({
                        'madde': (f'Matematik FAIL: {miktar} × {fiyat} = {hesap:.2f} '
                                  f'≠ {utu:.2f} (fark: {fark:.2f})'),
                        'durum': 'hata',
                        'puan': -30,
                    })
                    break  # Tek fail yeter

    if math_fail_var:
        skor -= 30
        matematik_ok = False
    elif urunler:
        # Urun satirinda miktar/fiyat yoksa matematik kontrol edilemiyor
        # bu durumda "atlandi" de, fail sayma
        test_edilebilir = any(
            u.get('qty') and u.get('unit_price') and u.get('tutar')
            for u in urunler
        )
        if test_edilebilir:
            detay.append({'madde': 'Matematik kontrolu gecti (miktar × fiyat = tutar)',
                          'durum': 'ok', 'puan': 0})
        else:
            detay.append({'madde': 'Matematik kontrolu yapilamadi (miktar/fiyat eksik)',
                          'durum': 'uyari', 'puan': 0})
            # Test yapilamadiysa matematik_ok True kalsin ama skora ekstra -5
            skor -= 5

    # --- 7) MANTIKSIZ DEGERLER (guvenlik filtreleri) ---
    # Toplam cok buyuk
    if toplam and toplam > 10_000_000:
        tavan = 0
        mantiksizlik.append(f'Toplam tutar cok yuksek: {toplam:.2f} {para_birimi or ""}')

    # Birim fiyat makul mu
    if urunler:
        for u in urunler:
            bf = u.get('unit_price')
            if bf is not None and isinstance(bf, (int, float)):
                if bf < 0:
                    tavan = min(tavan, 0)
                    mantiksizlik.append(f'Birim fiyat negatif: {bf}')
                elif bf > 1000:
                    tavan = min(tavan, 30)
                    mantiksizlik.append(
                        f'Birim fiyat cok yuksek: {bf} {para_birimi or ""} '
                        f'(terlik/ayakkabi icin $1-$20 bekleniyor)')

    # Miktar makul mu
    if urunler:
        for u in urunler:
            mk = u.get('qty')
            if mk is not None and isinstance(mk, (int, float)):
                if mk < 0:
                    tavan = min(tavan, 0)
                    mantiksizlik.append(f'Miktar negatif: {mk}')
                elif mk > 10_000_000:
                    tavan = min(tavan, 30)
                    mantiksizlik.append(f'Miktar cok yuksek: {mk}')

    # Cift maliyet kontrolu (parti.ToplamCift biliniyorsa)
    if parti_toplam_cift and parti_toplam_cift > 0 and toplam and toplam > 0:
        cift_mal = toplam / parti_toplam_cift
        if cift_mal > 100:
            tavan = min(tavan, 55)
            mantiksizlik.append(
                f'Cift maliyet cok yuksek: {cift_mal:.2f} {para_birimi or ""}/cift '
                f'(${0.50}-${50} aralıgi bekleniyor)')
        elif cift_mal < 0.10:
            tavan = min(tavan, 55)
            mantiksizlik.append(
                f'Cift maliyet cok dusuk: {cift_mal:.4f} {para_birimi or ""}/cift')

    # Tavani uygula
    if tavan < skor:
        detay.append({'madde': f'Guvenlik filtreleri tavani dusurdu: max {tavan}',
                      'durum': 'uyari', 'puan': -(skor - tavan)})
        skor = tavan

    # Aralik kontrolu
    if skor < 0: skor = 0
    if skor > 100: skor = 100

    # --- 8) Seviye belirle ---
    # KRITIK KURAL: Matematik FAIL ise skor ne olursa olsun >=90 olamaz
    if not matematik_ok and skor >= 90:
        detay.append({
            'madde': 'Matematik FAIL oldugu icin otomatik uygulama iptal (max 85)',
            'durum': 'uyari',
            'puan': -(skor - 85),
        })
        skor = 85

    if skor >= 90:
        seviye = 'yuksek'
    elif skor >= 60:
        seviye = 'orta'
    elif skor > 0:
        seviye = 'dusuk'
    else:
        seviye = 'red'

    return {
        'skor': int(skor),
        'seviye': seviye,
        'detay': detay,
        'matematik_ok': matematik_ok,
        'mantiksizlik_uyarilari': mantiksizlik,
    }


def parse(belge_row, dosya_yol):
    """
    Proforma / Anlasma Excel parse et.

    Donus: ParseSonuc
    """
    try:
        # Dosya var mi?
        if not os.path.isfile(dosya_yol):
            return ParseSonuc.hata(
                f"Dosya bulunamadi: {os.path.basename(dosya_yol)}",
                durum=ParseSonuc.DURUM_DOSYA_YOK,
            )

        # Dosyayi ac
        sp, sheet_ad, acma_uyarilari = _aktif_sheet_ac(dosya_yol)
        if sp is None:
            mesaj = '; '.join(acma_uyarilari) or 'Dosya acilamadi'
            return ParseSonuc.hata(
                f"Excel acilamadi: {mesaj}",
                durum=ParseSonuc.DURUM_FORMAT_HATA,
            )

        log.info(
            "Parse basliyor: dosya=%s sheet=%s (%dx%d)",
            os.path.basename(dosya_yol), sheet_ad, sp.nrows, sp.ncols,
        )

        uyarilar = []

        # --- 1) PI No (idempotency anahtari) ---
        kaynak_ref, pi_konum = _pi_no_bul(sp)

        # --- 2) Tedarikci ---
        tedarikci_ad = _tedarikci_bul(sp)

        # --- 3) PI tarihi yili ---
        pi_yil = _pi_tarihi_yil_bul(sp)

        # --- 4) Yukleme tarihi ---
        yukleme_tarih = _yukleme_tarihi_bul(sp, pi_tarihi_yil=pi_yil)

        # --- 5) Tablo baslik + kolon haritasi ---
        baslik_r = _baslik_satiri_bul(sp)
        if baslik_r is None:
            return ParseSonuc.hata(
                "Proforma tablo basligi bulunamadi (Amount/Quantity/Item "
                "gibi sutunlar gorunmuyor).",
                durum=ParseSonuc.DURUM_FORMAT_HATA,
            )
        harita = _kolon_harita(sp, baslik_r)
        tutar_kol, para_birimi = _amount_kolon_sec(harita)
        if tutar_kol is None:
            return ParseSonuc.hata(
                "Tutar sutunu (Amount USD/EUR/TRY/RMB) bulunamadi.",
                durum=ParseSonuc.DURUM_FORMAT_HATA,
            )

        # --- 6) Urun satirlari, total, komisyon ---
        tablo_veri = _urun_satirlari_oku(
            sp, baslik_r, harita, tutar_kol, para_birimi,
        )
        urunler = tablo_veri['urunler']
        toplam = tablo_veri['toplam']
        komisyon = tablo_veri['komisyon']

        # Eger total yoksa urun toplamini al
        if toplam is None and urunler:
            toplam = sum(u.get('tutar', 0) for u in urunler)

        if toplam is None or toplam <= 0:
            return ParseSonuc.hata(
                f"Proforma okundu ama tutar cikarilamadi "
                f"({len(urunler)} urun satiri bulundu).",
                durum=ParseSonuc.DURUM_FORMAT_HATA,
            )

        # --- 7) Teslim sekli / kur / odeme sarti ---
        teslim_sekli = _teslim_sekli_bul(sp)
        kur = _kur_bul(sp)
        odeme_sarti = _odeme_sarti_bul(sp)

        # --- 8) Kalemleri insa et (FOB + opsiyonel KOMISYON) ---
        kalemler = []

        # FOB kalem
        fob_aciklama_parcalari = []
        if urunler:
            adet_str = ' + '.join(
                f"{u['item']}"
                + (f" ({int(u['qty'])} {u.get('unit_of_measure', 'adet')})"
                   if 'qty' in u and u['qty'] else '')
                for u in urunler[:3]  # ilk 3
            )
            if len(urunler) > 3:
                adet_str += f" ...+{len(urunler) - 3}"
            fob_aciklama_parcalari.append(f"Urun: {adet_str}")

        if teslim_sekli:
            fob_aciklama_parcalari.append(f"Teslim: {teslim_sekli}")

        fob_aciklama = ' | '.join(fob_aciklama_parcalari) if fob_aciklama_parcalari else None

        fob_not_parcalari = []
        if teslim_sekli:
            fob_not_parcalari.append(
                f"Teslim: {teslim_sekli}"
                + (' (EX-T FACTORY)' if teslim_sekli == 'EXW' else '')
            )
        if kur:
            fob_not_parcalari.append(f"Kur: USD/RMB {kur}")
        if urunler and any('qty' in u for u in urunler):
            toplam_adet = sum(int(u.get('qty') or 0) for u in urunler)
            if toplam_adet:
                fob_not_parcalari.append(f"Toplam miktar: {toplam_adet}")

        fob_not = ' | '.join(fob_not_parcalari) if fob_not_parcalari else None

        kalemler.append({
            'tip': 'FOB',
            'kaynak': 'TAHMINI',
            'tutar': round(float(toplam), 4),
            'para_birimi': para_birimi,
            'aciklama': fob_aciklama,
            'cari_ad': tedarikci_ad,
            'fatura_no': kaynak_ref,  # PI No'yu fatura no olarak da kaydet
            'not_metni': fob_not,
        })

        # Komisyon kalem (varsa)
        if komisyon and komisyon > 0:
            kalemler.append({
                'tip': 'KOMISYON',
                'kaynak': 'TAHMINI',
                'tutar': round(float(komisyon), 4),
                'para_birimi': para_birimi,
                'aciklama': 'Proforma komisyon',
                'cari_ad': tedarikci_ad,
                'fatura_no': kaynak_ref,
                'not_metni': f"Proforma satiri: 'commission' (toplam uzerinden)",
            })

        # --- 9) Parti bilgisi (parti dolduramaya yardimci) ---
        parti_bilgi = {}
        if tedarikci_ad:
            parti_bilgi['tedarikci_ad'] = tedarikci_ad
        if yukleme_tarih:
            parti_bilgi['yukleme_tarih'] = yukleme_tarih
        if para_birimi:
            parti_bilgi['para_birimi'] = para_birimi
        if urunler:
            toplam_cift = sum(
                int(u.get('qty') or 0) for u in urunler
                if u.get('unit_of_measure', '').lower() in ('pair', 'pairs', 'cift')
                or (not u.get('unit_of_measure') and u.get('qty'))
            )
            if toplam_cift > 0:
                parti_bilgi['toplam_cift'] = toplam_cift

        # YENI: Ozet metrikleri (preview ekranında karar için)
        if toplam is not None and toplam > 0:
            parti_bilgi['toplam_tutar'] = round(float(toplam), 2)
        if komisyon and komisyon > 0:
            parti_bilgi['komisyon_tutar'] = round(float(komisyon), 2)

        # Birim maliyet (cift basi)
        _cift = parti_bilgi.get('toplam_cift')
        if _cift and _cift > 0 and toplam and toplam > 0:
            parti_bilgi['birim_maliyet'] = round(float(toplam) / _cift, 4)

        # --- 10) Uyarilar ---
        if not kaynak_ref:
            uyarilar.append("PI No bulunamadi - idempotency kontrolu yapilamaz")
        if not tedarikci_ad:
            uyarilar.append("Tedarikci adi otomatik okunamadi")
        if not yukleme_tarih:
            uyarilar.append("Yukleme tarihi serbest metin icinde bulunamadi")
        if teslim_sekli:
            uyarilar.append(f"Teslim sekli: {teslim_sekli}")
        if kur:
            uyarilar.append(f"Proforma'da belirtilen kur: USD/RMB {kur}")
        if odeme_sarti:
            uyarilar.append(f"Odeme sarti (otomatik plan olusturulmadi): {odeme_sarti}")
        if urunler:
            uyarilar.append(f"{len(urunler)} urun satiri okundu")

        # --- 11) GUVEN SKORU HESAPLA ---
        # parti.ToplamCift bilgisi varsa al (belge_row icinde parti_id var)
        parti_toplam_cift = None
        try:
            parti_id = belge_row.get('PartiId') or belge_row.get('parti_id')
            if parti_id:
                # parti getir - lazy import
                from modules.ithalat import queries as _ith_q
                parti = _ith_q.parti_getir(parti_id)
                if parti:
                    pc = parti.get('ToplamCift')
                    if pc and pc > 0:
                        parti_toplam_cift = int(pc)
        except Exception as _e:
            pass  # parti bilgisi yoksa mantiksizlik kontrolu atlanir

        tespit_edilen_kolonlar = {}
        if harita:
            for k, idx in harita.items():
                if idx is not None:
                    # Kolon harfini excel indexi'nden cikar (0=A, 1=B, ...)
                    try:
                        harf = chr(ord('A') + int(idx))
                    except Exception:
                        harf = str(idx)
                    tespit_edilen_kolonlar[k] = f"Kolon {harf}"

        guven = _guven_skoru_hesapla(
            kalemler=kalemler,
            kaynak_ref=kaynak_ref,
            tedarikci_ad=tedarikci_ad,
            para_birimi=para_birimi,
            baslik_r=baslik_r,
            harita=harita,
            tutar_kol=tutar_kol,
            urunler=urunler,
            toplam=toplam,
            parti_toplam_cift=parti_toplam_cift,
        )

        # --- 12) KARAR ---
        # KRITIK 3 KURAL:
        # 1) Matematik FAIL ise asla otomatik uygulama (ParseSonuc icinde 85'e cekiliyor)
        # 2) Skor >= 90 -> otomatik uygula
        # 3) Skor 60-89 -> BEKLIYOR_ONAY (preview)
        # 4) Skor < 60 -> REDDEDILDI (maliyete yazmaz)

        base_mesaj = (f"{len(kalemler)} kalem (FOB: {toplam:.2f} {para_birimi}"
                      + (f", Komisyon: {komisyon:.2f}" if komisyon else '') + ")"
                      + (f" - PI: {kaynak_ref}" if kaynak_ref else ''))

        if guven['skor'] < 60:
            # REDDEDILDI - maliyete yazmaz
            mesaj = f"Güven skoru düşük ({guven['skor']}/100). "
            if guven['mantiksizlik_uyarilari']:
                mesaj += "Mantıksız değer tespit edildi. "
            mesaj += "Kalem otomatik eklenmedi — manuel kontrol gerekli."
            return ParseSonuc.reddedildi(
                guven_skoru=guven['skor'],
                mesaj=mesaj,
                guven_detay=guven['detay'],
                mantiksizlik_uyarilari=guven['mantiksizlik_uyarilari'],
                kalemler=kalemler,
            )

        # === ILK TEST FAZI: Skor >=90 bile otomatik yazmiyoruz ===
        # Her durumda BEKLIYOR_ONAY - kullanici preview'de karar versin
        # Ileride otomatik acilacak (guven_skoru >=90 AND matematik_ok)
        if guven['skor'] >= 90 and guven['matematik_ok']:
            onay_mesaji = "Yüksek güven (test fazı: manuel onay)"
        else:
            onay_mesaji = "Önizleme onayı bekleniyor"

        mesaj = (f"Güven skoru {guven['skor']}/100 — {onay_mesaji}. "
                 f"{base_mesaj}")

        # PI No yoksa uyari ek (ama hala BEKLIYOR_ONAY akisinda)
        if not kaynak_ref:
            uyarilar = list(uyarilar or [])
            uyarilar.append("DIKKAT: Kaynak referans (PI/Invoice No) yok - idempotency kontrolu yapilamayacak")

        return ParseSonuc.bekliyor_onay(
            kalemler=kalemler,
            guven_skoru=guven['skor'],
            mesaj=mesaj,
            kaynak_ref=kaynak_ref,
            parti_bilgi=parti_bilgi,
            uyarilar=uyarilar,
            guven_detay=guven['detay'],
            tespit_edilen_kolonlar=tespit_edilen_kolonlar,
            matematik_ok=guven['matematik_ok'],
            mantiksizlik_uyarilari=guven['mantiksizlik_uyarilari'],
        )

    except Exception as e:
        log.exception("anlasma_excel.parse fail-safe: %s", e)
        return ParseSonuc.hata(
            f"Parse sirasinda beklenmedik hata: {str(e)[:200]}",
            durum=ParseSonuc.DURUM_PARSER_HATA,
        )
