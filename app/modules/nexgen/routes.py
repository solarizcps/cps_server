# -*- coding: utf-8 -*-
"""
SOLARIZ CPS — NexGen Hammadde Yönetimi
=======================================
FAZ-1A : Modül iskeleti, menü, yetki
FAZ-1B : Stok Kart Master + Hareket Motoru altyapısı
FAZ-2  : Satın Alma Merkezi — Tedarikçi CRUD + Sipariş akışı
FAZ-2.6: Haftalık Fiyat Geçmişi — Excel import, preview/onay akışı
FAZ-3A : Depo Mal Kabul — GIRIS hareketi Depo yetkisiyle

Yetki kodları:
  nexgen.view                — modül geneli görüntüleme
  nexgen.stok.view           — stok kartları görüntüleme
  nexgen.stok.manage         — stok kartı ekleme/düzenleme
  nexgen.satinalma.view      — satın alma sipariş görüntüleme
  nexgen.satinalma.manage    — yeni sipariş oluşturma/düzenleme
  nexgen.satinalma.approve   — sipariş onaylama (sadece Yönetim)
  nexgen.satinalma.fiyat     — fiyat/kur/maliyet görüntüleme
  nexgen.tedarikci.view      — tedarikçi listesi görüntüleme
  nexgen.tedarikci.manage    — tedarikçi ekleme/düzenleme
  nexgen.fiyat.view          — fiyat geçmişini görüntüleme
  nexgen.fiyat.manage        — fiyat girişi (manuel + Excel)
  nexgen.fiyat.approve       — batch onaylama
  nexgen.fiyat.admin         — fiyat pasife alma (sadece Yönetim)
  nexgen.depo.view           — depo ekranı görüntüleme
  nexgen.depo.giris          — mal kabul + GIRIS hareketi oluşturma

KURAL: nexgen_stok_hareket INSERT SADECE Depo mal kabulünde yapılır (FAZ-3A+).
       Satın Alma, Fiyat, Yönetim ekranları stok hareketi oluşturmaz.
"""

import sqlite3
import os
import io
from datetime import datetime, date

from flask import (
    Blueprint, render_template, abort,
    request, jsonify, session, g,
    Response, redirect, url_for, flash
)
from modules.auth import yetki_gerekli, yetki_var, login_gerekli

nexgen_bp = Blueprint('nexgen', __name__, url_prefix='/nexgen')

DB_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'mock_data.db')


# ─────────────────────────────────────────────────────────────
# Migration 074 — AR-GE kimlik alanları (startup, idempotent)
# ─────────────────────────────────────────────────────────────
def _mig074_arge_identity():
    """nexgen_arge_test tablosuna FAZ-2A-P1 alanlarını ekler."""
    try:
        _con = sqlite3.connect(DB_PATH)
        cols = [c[1] for c in _con.execute(
            "PRAGMA table_info(nexgen_arge_test)"
        ).fetchall()]
        eklemeler = [
            ("cari_id",         "ALTER TABLE nexgen_arge_test ADD COLUMN cari_id INTEGER"),
            ("shore_hedef",     "ALTER TABLE nexgen_arge_test ADD COLUMN shore_hedef REAL"),
            ("lot_no",          "ALTER TABLE nexgen_arge_test ADD COLUMN lot_no TEXT"),
            ("talep_referansi", "ALTER TABLE nexgen_arge_test ADD COLUMN talep_referansi TEXT"),
        ]
        degisti = False
        for kolon, sql in eklemeler:
            if kolon not in cols:
                _con.execute(sql)
                degisti = True
        if degisti:
            _con.commit()
            print("[nexgen] Migration 074: AR-GE kimlik alanları eklendi")
        _con.close()
    except Exception as _e:
        print(f"[nexgen] Migration 074 uyarı: {_e}")

_mig074_arge_identity()


# ─────────────────────────────────────────────────────────────
# Migration 075 — AR-GE onay notu (startup, idempotent)
# ─────────────────────────────────────────────────────────────
def _mig075_arge_onay_notu():
    """nexgen_arge_test tablosuna onay_notu alanını ekler."""
    try:
        _con = sqlite3.connect(DB_PATH)
        cols = [c[1] for c in _con.execute(
            "PRAGMA table_info(nexgen_arge_test)"
        ).fetchall()]
        if 'onay_notu' not in cols:
            _con.execute("ALTER TABLE nexgen_arge_test ADD COLUMN onay_notu TEXT")
            _con.commit()
            print("[nexgen] Migration 075: onay_notu eklendi")
        _con.close()
    except Exception as _e:
        print(f"[nexgen] Migration 075 uyarı: {_e}")

_mig075_arge_onay_notu()


# ─────────────────────────────────────────────────────────────
# Migration 076 — AR-GE rf_renk_id FK (startup, idempotent)
# ─────────────────────────────────────────────────────────────
def _mig076_rf_renk_id():
    """nexgen_arge_test tablosuna rf_renk_id alanini ekler."""
    try:
        _con = sqlite3.connect(DB_PATH)
        cols = [c[1] for c in _con.execute(
            "PRAGMA table_info(nexgen_arge_test)"
        ).fetchall()]
        if 'rf_renk_id' not in cols:
            _con.execute("ALTER TABLE nexgen_arge_test ADD COLUMN rf_renk_id INTEGER")
            _con.commit()
            print("[nexgen] Migration 076: rf_renk_id eklendi")
        _con.close()
    except Exception as _e:
        print(f"[nexgen] Migration 076 uyari: {_e}")

_mig076_rf_renk_id()


# ─────────────────────────────────────────────────────────────
# Yardımcı: DB bağlantısı
# ─────────────────────────────────────────────────────────────
def _db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con


def _kullanici_id():
    u = session.get('kullanici')
    return u.get('Id') if u else None


def _kullanici_ad():
    u = session.get('kullanici')
    return u.get('KullaniciAdi', 'sistem') if u else 'sistem'


def _taban_kalemler(kalemler):
    """Ana formül görünümü: BOYA satırları hariç (RF/renk panelinde yönetilir)."""
    return [k for k in kalemler if (k.get('kategori') or '').upper() != 'BOYA']


def _boya_kalemler_filtre(kalemler):
    return [k for k in kalemler if (k.get('kategori') or '').upper() == 'BOYA']


def _kg_to_gr_etiket(kg):
    """Test kalem kg → tablet boya özeti (gram)."""
    try:
        g = float(kg) * 1000
    except (TypeError, ValueError):
        return '—'
    if g >= 10:
        s = f'{g:.1f}'.rstrip('0').rstrip('.')
    else:
        s = f'{g:.2f}'.rstrip('0').rstrip('.')
    return f'{s} gr'


def _stok_boya_dropdown_etiket(kod, ad):
    """Tablet boya select: KOD — AD."""
    k = (kod or '').strip()
    a = (ad or '').strip()
    if k and a:
        return f'{k} — {a}'
    return a or k or 'Boya'


def _stok_boya_ozet_ad(kod, ad):
    """Sol panel boya özeti — operatör adı (N550, P.Red 48:3)."""
    return (ad or kod or 'Boya').strip()


def _arge_test_boya_ozet(con, test_id, limit=3):
    """Tablet sol panel: BOYA kalem özeti satırları."""
    rows = con.execute("""
        SELECT sk.ad, sk.kod, tk.test_miktar_kg
        FROM nexgen_arge_test_kalem tk
        JOIN nexgen_stok_kart sk ON sk.id = tk.stok_kart_id
        WHERE tk.test_id = ? AND UPPER(COALESCE(sk.kategori, '')) = 'BOYA'
        ORDER BY tk.sira, tk.id
    """, (test_id,)).fetchall()
    satirlar = []
    for r in rows[:limit]:
        ad = _stok_boya_ozet_ad(r['kod'], r['ad'])
        satirlar.append({'ad': ad, 'gram': _kg_to_gr_etiket(r['test_miktar_kg'])})
    return satirlar, max(0, len(rows) - limit)


# ─────────────────────────────────────────────────────────────
# Stok miktarı hesapla (hareket toplamı)
# ─────────────────────────────────────────────────────────────
def _mevcut_stok(con, kart_id):
    row = con.execute(
        "SELECT COALESCE(SUM(miktar_kg), 0) AS toplam "
        "FROM nexgen_stok_hareket WHERE stok_kart_id=?",
        (kart_id,)
    ).fetchone()
    return round(row["toplam"], 3) if row else 0.0


# ─────────────────────────────────────────────────────────────
# Ana Sayfa
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/')
@yetki_gerekli('nexgen.view', 'can_view')
def index():
    can_yonetim = yetki_var('nexgen.yonetim.manage', 'can_view')
    can_depo    = yetki_var('nexgen.depo.view', 'can_view')
    can_recete  = yetki_var('nexgen.recete.view', 'can_view')
    can_tablet  = yetki_var('nexgen.tablet.view', 'can_view')
    return render_template('nexgen/index.html', active='nexgen',
                           can_yonetim=can_yonetim,
                           can_depo=can_depo,
                           can_recete=can_recete,
                           can_tablet=can_tablet)


@nexgen_bp.route('/formuller')
@nexgen_bp.route('/formuller/')
@yetki_gerekli('nexgen.recete.view', 'can_view')
def formuller_redirect():
    """Eski /formuller linklerini Reçete Merkezine yönlendir."""
    return redirect('/nexgen/recete/')


# ─────────────────────────────────────────────────────────────
# Stok Kartları — Liste
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/stok-kartlari')
@yetki_gerekli('nexgen.stok.view', 'can_view')
def stok_kartlari():
    con = _db()
    try:
        kartlar_raw = con.execute("""
            SELECT k.id, k.kod, k.ad, k.kategori, k.birim,
                   k.minimum_stok, k.kritik_stok, k.aktif,
                   COALESCE(SUM(h.miktar_kg), 0) AS mevcut_stok
            FROM nexgen_stok_kart k
            LEFT JOIN nexgen_stok_hareket h ON h.stok_kart_id = k.id
            GROUP BY k.id
            ORDER BY k.kategori, k.ad
        """).fetchall()
    finally:
        con.close()

    kartlar = []
    for r in kartlar_raw:
        ms = r["mevcut_stok"]
        mn = r["minimum_stok"]
        kr = r["kritik_stok"]
        if ms <= kr:
            durum = "kritik"
        elif ms <= mn:
            durum = "uyari"
        else:
            durum = "normal"
        kartlar.append({
            "id": r["id"], "kod": r["kod"], "ad": r["ad"],
            "kategori": r["kategori"], "birim": r["birim"],
            "minimum_stok": mn, "kritik_stok": kr,
            "mevcut_stok": ms, "durum": durum,
            "aktif": r["aktif"],
        })

    can_manage = yetki_var('nexgen.stok.manage', 'can_manage') or yetki_var('nexgen.stok.manage', 'can_create')
    return render_template(
        'nexgen/stok_kartlari.html',
        active='nexgen',
        kartlar=kartlar,
        can_manage=can_manage,
    )


# ─────────────────────────────────────────────────────────────
# Stok Kart Detay
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/stok-kartlari/<int:kart_id>')
@yetki_gerekli('nexgen.stok.view', 'can_view')
def stok_detay(kart_id):
    con = _db()
    try:
        kart = con.execute(
            "SELECT * FROM nexgen_stok_kart WHERE id=?", (kart_id,)
        ).fetchone()
        if not kart:
            abort(404)

        mevcut = _mevcut_stok(con, kart_id)

        hareketler_raw = con.execute("""
            SELECT h.id, h.hareket_tipi, h.miktar_kg,
                   h.onceki_stok, h.sonraki_stok,
                   h.aciklama, h.referans_tip, h.referans_id,
                   h.olusturma_tarihi,
                   sk.KullaniciAdi AS olusturan_ad
            FROM nexgen_stok_hareket h
            LEFT JOIN sistem_kullanici sk ON sk.Id = h.olusturan_id
            WHERE h.stok_kart_id = ?
            ORDER BY h.id DESC
            LIMIT 200
        """, (kart_id,)).fetchall()

        # FAZ-2.7: Fiyat geçmişi — sadece nexgen.fiyat.admin yetkisiyle
        # Yönetim / Adem / Alpay / Altan görür; Satın Alma / Depo / Üretim görmez
        can_fiyat_admin = yetki_var('nexgen.fiyat.admin', 'can_view') or \
                          yetki_var('nexgen.fiyat.admin', 'can_manage')
        fiyat_gecmisi_raw = []
        if can_fiyat_admin:
            # nexgen_hammadde_fiyat tablosu FAZ-2.6'da oluşturuldu
            try:
                fiyat_gecmisi_raw = con.execute("""
                    SELECT hf.id, hf.fiyat, hf.para_birimi, hf.kur, hf.fiyat_try,
                           hf.vade_gun, hf.fiyat_tarihi, hf.kaynak, hf.aktif,
                           hf.notlar, hf.olusturma_tarihi,
                           t.ad  AS tedarikci_ad,  t.kod AS tedarikci_kod,
                           sk.KullaniciAdi AS olusturan_ad
                    FROM nexgen_hammadde_fiyat hf
                    LEFT JOIN nexgen_tedarikci  t  ON t.id  = hf.tedarikci_id
                    LEFT JOIN sistem_kullanici  sk ON sk.Id = hf.olusturan_id
                    WHERE hf.stok_kart_id = ?
                    ORDER BY hf.fiyat_tarihi DESC, hf.id DESC
                    LIMIT 100
                """, (kart_id,)).fetchall()
            except Exception:
                fiyat_gecmisi_raw = []

    finally:
        con.close()

    ms = mevcut
    mn = kart["minimum_stok"]
    kr = kart["kritik_stok"]
    if ms <= kr:
        durum = "kritik"
    elif ms <= mn:
        durum = "uyari"
    else:
        durum = "normal"

    hareketler = [dict(h) for h in hareketler_raw]
    can_manage  = yetki_var('nexgen.stok.manage', 'can_manage') or \
                  yetki_var('nexgen.stok.manage', 'can_create')

    # Hareket listelerini sekmelere göre ayır
    GIRIS_TIPLERI = {'ACILIS_DEVIR', 'GIRIS', 'URETIM_CIKTI'}
    CIKIS_TIPLERI = {'CIKIS', 'URETIM_TUKETIM', 'SEVK'}

    girişler  = []
    cikislar  = []
    for h in hareketler:
        tip = h.get('hareket_tipi', '')
        kg  = h.get('miktar_kg', 0) or 0
        if tip in GIRIS_TIPLERI or (tip == 'SAYIM_DUZELTME' and kg >= 0):
            girişler.append(h)
        elif tip in CIKIS_TIPLERI or (tip == 'SAYIM_DUZELTME' and kg < 0):
            cikislar.append(h)
        else:
            girişler.append(h)  # bilinmeyen tipler giriş tarafında göster

    # Fiyat geçmişi — önceki AKTİF fiyata göre fark/yüzde hesapla
    # Pasif (aktif=0) kayıtlar fark hesabından atlanır; geçmişte görünür ama
    # referans alınmaz ve kendileri için de fark gösterilmez.
    fiyat_gecmisi = [dict(f) for f in fiyat_gecmisi_raw]
    for i, f in enumerate(fiyat_gecmisi):
        # Pasif kayıt — fark gösterme
        if not f.get('aktif', 1):
            f['fark'] = f['yuzde'] = None
            continue
        # Sonraki kayıtlar arasında ilk aktif olanı bul
        onceki_fiyat = None
        for j in range(i + 1, len(fiyat_gecmisi)):
            kandidat = fiyat_gecmisi[j]
            if kandidat.get('aktif', 1):
                onceki_fiyat = kandidat.get('fiyat')
                break
        if onceki_fiyat:
            fark   = round(f['fiyat'] - onceki_fiyat, 4)
            yuzde  = round((fark / onceki_fiyat) * 100, 2)
            f['fark']  = fark
            f['yuzde'] = yuzde
        else:
            f['fark'] = f['yuzde'] = None

    return render_template(
        'nexgen/stok_detay.html',
        active='nexgen',
        kart=dict(kart),
        mevcut_stok=ms,
        durum=durum,
        hareketler=hareketler,
        girişler=girişler,
        cikislar=cikislar,
        can_manage=can_manage,
        can_fiyat_admin=can_fiyat_admin,
        fiyat_gecmisi=fiyat_gecmisi,
    )


# ─────────────────────────────────────────────────────────────
# API — Hızlı Panel Özet (FAZ-1D)
# GET /nexgen/api/stok-kart/<id>/ozet
# Yetki: nexgen.stok.view can_view
# Fiyat/maliyet bilgisi dönmez.
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/stok-kart/<int:kart_id>/ozet')
@yetki_gerekli('nexgen.stok.view', 'can_view')
def api_stok_kart_ozet(kart_id):
    con = _db()
    try:
        # SELECT * — Migration 048 çalıştırılmamış olsa bile çalışır
        kart_row = con.execute(
            "SELECT * FROM nexgen_stok_kart WHERE id=?",
            (kart_id,)
        ).fetchone()
        if not kart_row:
            return jsonify({"ok": False, "hata": "Kart bulunamadı."}), 404

        kart = dict(kart_row)
        ms = _mevcut_stok(con, kart_id)

        hareketler_raw = con.execute("""
            SELECT h.hareket_tipi, h.miktar_kg, h.sonraki_stok,
                   h.aciklama, h.olusturma_tarihi,
                   sk.KullaniciAdi AS yapan
            FROM nexgen_stok_hareket h
            LEFT JOIN sistem_kullanici sk ON sk.Id = h.olusturan_id
            WHERE h.stok_kart_id = ?
            ORDER BY h.id DESC LIMIT 5
        """, (kart_id,)).fetchall()

    except Exception as e:
        return jsonify({"ok": False, "hata": f"Veritabanı hatası: {str(e)}"}), 500
    finally:
        con.close()

    mn = kart.get("minimum_stok", 0) or 0
    kr = kart.get("kritik_stok", 0) or 0
    if ms <= kr:
        durum = "kritik"
    elif ms <= mn:
        durum = "uyari"
    else:
        durum = "normal"

    hareketler = []
    for h in hareketler_raw:
        hareketler.append({
            "tip":      h["hareket_tipi"],
            "miktar":   h["miktar_kg"],
            "stok":     h["sonraki_stok"],
            "aciklama": h["aciklama"] or "",
            "tarih":    (h["olusturma_tarihi"] or "")[:16],
            "yapan":    h["yapan"] or "—",
        })

    return jsonify({
        "ok": True,
        "kart": {
            "id":            kart.get("id"),
            "kod":           kart.get("kod", ""),
            "ad":            kart.get("ad", ""),
            "kategori":      kart.get("kategori", ""),
            "birim":         kart.get("birim", "KG"),
            "minimum_stok":  mn,
            "kritik_stok":   kr,
            "aktif":         kart.get("aktif", 1),
            # FAZ-1C opsiyonel alanlar — Migration 048 yoksa None
            "renk":          kart.get("renk"),
            "alt_kategori":  kart.get("alt_kategori"),
            "kalite_sinifi": kart.get("kalite_sinifi"),
            "shore_degeri":  kart.get("shore_degeri"),
            "notlar":        kart.get("notlar"),
        },
        "mevcut_stok": ms,
        "durum":       durum,
        "hareketler":  hareketler,
    })


# ─────────────────────────────────────────────────────────────
# API — Yeni Stok Kartı Ekle
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/stok-kart-ekle', methods=['POST'])
@yetki_gerekli('nexgen.stok.manage', 'can_create')
def api_stok_kart_ekle():
    data = request.get_json(silent=True) or {}
    kod      = (data.get('kod') or '').strip().upper()
    ad       = (data.get('ad') or '').strip()
    kategori = (data.get('kategori') or 'HAMMADDE').strip().upper()
    birim    = (data.get('birim') or 'KG').strip().upper()
    # FAZ-1C opsiyonel alanlar — boşsa None kaydedilir
    renk          = (data.get('renk') or '').strip() or None
    alt_kategori  = (data.get('alt_kategori') or '').strip() or None
    kalite_sinifi = (data.get('kalite_sinifi') or '').strip() or None
    shore_degeri  = (data.get('shore_degeri') or '').strip() or None
    notlar        = (data.get('notlar') or '').strip() or None

    try:
        min_stok  = float(data.get('minimum_stok') or 0)
        krit_stok = float(data.get('kritik_stok') or 0)
    except (ValueError, TypeError):
        return jsonify({"ok": False, "hata": "Geçersiz sayısal değer."}), 400

    GECERLI_KATEGORILER = {
        'HAMMADDE', 'MAMUL_COMPOUND', 'RECYCLE', 'KATKI', 'BOYA', 'DIGER'
    }

    if not kod:
        return jsonify({"ok": False, "hata": "Kod zorunludur."}), 400
    if not ad:
        return jsonify({"ok": False, "hata": "Ad zorunludur."}), 400
    if kategori not in GECERLI_KATEGORILER:
        return jsonify({"ok": False, "hata": f"Geçersiz kategori: {kategori}"}), 400

    con = _db()
    try:
        mevcut = con.execute(
            "SELECT id FROM nexgen_stok_kart WHERE kod=?", (kod,)
        ).fetchone()
        if mevcut:
            return jsonify({"ok": False, "hata": f"'{kod}' kodu zaten mevcut."}), 409

        con.execute("""
            INSERT INTO nexgen_stok_kart
              (kod, ad, kategori, birim, minimum_stok, kritik_stok,
               renk, alt_kategori, kalite_sinifi, shore_degeri, notlar,
               aktif, olusturan_id, olusturma_tarihi)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, datetime('now'))
        """, (kod, ad, kategori, birim, min_stok, krit_stok,
              renk, alt_kategori, kalite_sinifi, shore_degeri, notlar,
              _kullanici_id()))
        yeni_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({"ok": True, "id": yeni_id, "kod": kod, "ad": ad})


# ─────────────────────────────────────────────────────────────
# API — Stok Hareketi Ekle (manuel giriş/çıkış/sayım)
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/stok-hareket-ekle', methods=['POST'])
@yetki_gerekli('nexgen.stok.manage', 'can_create')
def api_stok_hareket_ekle():
    data = request.get_json(silent=True) or {}
    kart_id       = data.get('stok_kart_id')
    hareket_tipi  = (data.get('hareket_tipi') or '').strip().upper()
    aciklama      = (data.get('aciklama') or '').strip()
    try:
        miktar = float(data.get('miktar_kg') or 0)
    except (ValueError, TypeError):
        return jsonify({"ok": False, "hata": "Geçersiz miktar."}), 400

    GECERLI_TIPLER = {
        'GIRIS', 'CIKIS', 'URETIM_TUKETIM', 'URETIM_CIKTI',
        'SAYIM_DUZELTME', 'SEVK'
    }
    CIKIS_TIPLERI = {'CIKIS', 'URETIM_TUKETIM', 'SEVK'}

    if not kart_id:
        return jsonify({"ok": False, "hata": "stok_kart_id zorunludur."}), 400
    if hareket_tipi not in GECERLI_TIPLER:
        return jsonify({"ok": False, "hata": f"Geçersiz hareket tipi: {hareket_tipi}"}), 400
    if miktar <= 0:
        return jsonify({"ok": False, "hata": "Miktar sıfırdan büyük olmalıdır."}), 400

    # Çıkış hareketlerinde miktarı negatife çevir
    if hareket_tipi in CIKIS_TIPLERI:
        miktar_kayit = -abs(miktar)
    else:
        miktar_kayit = abs(miktar)

    con = _db()
    try:
        kart = con.execute(
            "SELECT id FROM nexgen_stok_kart WHERE id=? AND aktif=1", (kart_id,)
        ).fetchone()
        if not kart:
            return jsonify({"ok": False, "hata": "Stok kartı bulunamadı veya pasif."}), 404

        onceki = _mevcut_stok(con, kart_id)
        sonraki = round(onceki + miktar_kayit, 3)

        con.execute("""
            INSERT INTO nexgen_stok_hareket
              (stok_kart_id, hareket_tipi, miktar_kg,
               onceki_stok, sonraki_stok,
               aciklama, olusturan_id, olusturma_tarihi)
            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'))
        """, (
            kart_id, hareket_tipi, miktar_kayit,
            onceki, sonraki,
            aciklama, _kullanici_id()
        ))
        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({
        "ok": True,
        "onceki_stok": onceki,
        "sonraki_stok": sonraki,
        "hareket_tipi": hareket_tipi,
    })


# ─────────────────────────────────────────────────────────────
# API — Stok minimal detay (çıkış formu canlı stok gösterimi)
# GET /nexgen/api/stok-detay-minimal?id=<kart_id>
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/stok-detay-minimal', methods=['GET'])
@yetki_gerekli('nexgen.depo.view', 'can_view')
def api_stok_detay_minimal():
    kart_id = request.args.get('id', type=int)
    if not kart_id:
        return jsonify({"ok": False, "hata": "id zorunlu"}), 400
    con = _db()
    try:
        kart = con.execute(
            "SELECT id, kod, ad FROM nexgen_stok_kart WHERE id=? AND aktif=1", (kart_id,)
        ).fetchone()
        if not kart:
            return jsonify({"ok": False, "hata": "Bulunamadı"}), 404
        mevcut = _mevcut_stok(con, kart_id)
    finally:
        con.close()
    return jsonify({"ok": True, "mevcut_stok": mevcut, "kod": kart["kod"], "ad": kart["ad"]})


# ─────────────────────────────────────────────────────────────
# API — Kart pasif/aktif yap (silme yok)
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/stok-kart-durum', methods=['POST'])
@yetki_gerekli('nexgen.stok.manage', 'can_update')
def api_stok_kart_durum():
    data = request.get_json(silent=True) or {}
    kart_id = data.get('id')
    aktif   = 1 if data.get('aktif') else 0

    if not kart_id:
        return jsonify({"ok": False, "hata": "id zorunludur."}), 400

    con = _db()
    try:
        kart = con.execute(
            "SELECT id FROM nexgen_stok_kart WHERE id=?", (kart_id,)
        ).fetchone()
        if not kart:
            return jsonify({"ok": False, "hata": "Kart bulunamadı."}), 404

        con.execute("""
            UPDATE nexgen_stok_kart
            SET aktif=?, guncelleyen_id=?, guncelleme_tarihi=datetime('now')
            WHERE id=?
        """, (aktif, _kullanici_id(), kart_id))
        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({"ok": True, "aktif": aktif})


# ─────────────────────────────────────────────────────────────
# Placeholder Alt Sayfalar (satinalma ve stok-kartlari hariç)
# ─────────────────────────────────────────────────────────────
_SAYFALAR = {
    'formuller': ('Formüller',       'Ürün başına hammadde formül tanımları.'),
    'uretim':    ('Üretim Tablet',   'Üretim sürecinde kullanılan hammadde giriş/çıkış ekranı.'),
    'recycle':   ('Recycle',         'Geri dönüşüm ve fire takip ekranı.'),
    'deneme':    ('Deneme',          'Deneme/prototip üretim kayıt ekranı.'),
    'sevk':      ('Sevk & Barkod',   'Ürün sevkiyat ve barkod yönetim ekranı.'),
    'maliyet':   ('Maliyet & Rapor', 'Hammadde maliyet analizi ve raporlama ekranı.'),
    'raporlar':  ('Raporlar',        'NexGen genel raporlar merkezi.'),
}


@nexgen_bp.route('/<string:sayfa>')
@yetki_gerekli('nexgen.view', 'can_view')
def alt_sayfa(sayfa):
    if sayfa not in _SAYFALAR:
        abort(404)
    baslik, aciklama = _SAYFALAR[sayfa]
    return render_template(
        'nexgen/placeholder.html',
        active='nexgen',
        sayfa_baslik=baslik,
        sayfa_aciklama=aciklama,
        sayfa_slug=sayfa,
    )


# ═════════════════════════════════════════════════════════════
# FAZ-4B: REÇETE / FORMÜL MERKEZİ
# ─────────────────────────────────────────────────────────────
# KURAL: Bu bölümde nexgen_stok_hareket INSERT YAPILMAZ.
#        nexgen_hammadde_fiyat dokunulmaz.
#        Sadece okuma + reçete master veri.
# ─────────────────────────────────────────────────────────────

@nexgen_bp.route('/recete/')
@yetki_gerekli('nexgen.recete.view', 'can_view')
def recete_liste():
    con = _db()
    try:
        formuller_raw = con.execute("""
            SELECT
                f.id, f.kod, f.ad, f.durum, f.onay_durumu,
                f.olusturma_tarihi, f.notlar,
                ku.KullaniciAdi AS olusturan_ad,
                COUNT(DISTINCT rv.id)  AS renk_say,
                COUNT(DISTINCT uv.id)  AS uretim_say,
                COUNT(DISTINCT rk.id)  AS kalem_say
            FROM nexgen_formul f
            LEFT JOIN sistem_kullanici ku ON ku.Id = f.olusturan_id
            LEFT JOIN nexgen_renk_varyant rv ON rv.formul_id = f.id AND rv.aktif = 1
            LEFT JOIN nexgen_uretim_varyant uv ON uv.renk_varyant_id = rv.id AND uv.aktif = 1
            LEFT JOIN nexgen_recete_kalem rk ON rk.uretim_varyant_id = uv.id AND rk.aktif = 1
            GROUP BY f.id
            ORDER BY f.id DESC
        """).fetchall()

        # Her formül için toplam batch KG ve KG maliyet (ilk üretim varyantı bazında özet)
        formuller = []
        for f in formuller_raw:
            f_dict = dict(f)
            # Tüm aktif kalemler ve fiyatları
            kalemler = con.execute("""
                SELECT rk.miktar_kg, rk.stok_kart_id
                FROM nexgen_recete_kalem rk
                JOIN nexgen_uretim_varyant uv ON uv.id = rk.uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                WHERE rv.formul_id = ? AND rv.aktif = 1
                  AND uv.aktif = 1 AND rk.aktif = 1
            """, (f_dict['id'],)).fetchall()

            if kalemler:
                toplam_kg = round(sum(k['miktar_kg'] for k in kalemler), 3)
                toplam_maliyet = 0.0
                for k in kalemler:
                    fiyat_row = con.execute("""
                        SELECT fiyat, para_birimi, kur, fiyat_try
                        FROM nexgen_hammadde_fiyat
                        WHERE stok_kart_id = ? AND aktif = 1
                        ORDER BY fiyat_tarihi DESC, id DESC LIMIT 1
                    """, (k['stok_kart_id'],)).fetchone()
                    if fiyat_row:
                        if fiyat_row['fiyat_try'] and fiyat_row['fiyat_try'] > 0:
                            birim_fiyat = float(fiyat_row['fiyat_try'])
                        elif fiyat_row['para_birimi'] == 'TRY':
                            birim_fiyat = float(fiyat_row['fiyat'] or 0)
                        elif fiyat_row['kur'] and fiyat_row['kur'] > 0:
                            birim_fiyat = float(fiyat_row['fiyat'] or 0) * float(fiyat_row['kur'])
                        else:
                            birim_fiyat = float(fiyat_row['fiyat'] or 0)
                        toplam_maliyet += float(k['miktar_kg']) * birim_fiyat
                f_dict['liste_toplam_kg']  = toplam_kg
                f_dict['liste_kg_maliyet'] = round(toplam_maliyet / toplam_kg, 2) if toplam_kg > 0 else 0.0
            else:
                f_dict['liste_toplam_kg']  = 0.0
                f_dict['liste_kg_maliyet'] = 0.0

            # recete_durum özeti: üretime açık varyant var mı?
            durum_rows = con.execute("""
                SELECT uv.recete_durum
                FROM nexgen_uretim_varyant uv
                JOIN nexgen_renk_varyant rv ON rv.id = uv.renk_varyant_id
                WHERE rv.formul_id = ? AND rv.aktif = 1 AND uv.aktif = 1
            """, (f_dict['id'],)).fetchall()
            durumlar = [r['recete_durum'] or 'TASLAK' for r in durum_rows]
            if 'URETIME_ACIK' in durumlar:
                f_dict['recete_durum_ozet'] = 'URETIME_ACIK'
            elif 'ONAYLI' in durumlar:
                f_dict['recete_durum_ozet'] = 'ONAYLI'
            elif 'DENEME' in durumlar:
                f_dict['recete_durum_ozet'] = 'DENEME'
            elif durumlar:
                f_dict['recete_durum_ozet'] = 'TASLAK'
            else:
                f_dict['recete_durum_ozet'] = None

            formuller.append(f_dict)

        # Her formül için renk/varyant özet listesi (liste kartı için)
        for f_dict in formuller:
            rv_ozet = con.execute("""
                SELECT rv.id AS rv_id, rv.ad AS rv_ad, rv.renk,
                       uv.id AS uv_id, uv.boyut, uv.recete_durum,
                       COUNT(rk.id) AS kalem_say,
                       SUM(rk.miktar_kg) AS toplam_kg
                FROM nexgen_renk_varyant rv
                JOIN nexgen_uretim_varyant uv ON uv.renk_varyant_id = rv.id AND uv.aktif = 1
                LEFT JOIN nexgen_recete_kalem rk ON rk.uretim_varyant_id = uv.id AND rk.aktif = 1
                WHERE rv.formul_id = ? AND rv.aktif = 1
                GROUP BY uv.id
                ORDER BY rv.id, uv.id
            """, (f_dict['id'],)).fetchall()

            # Renk bazında grupla
            rv_gruplar = {}
            for row in rv_ozet:
                rv_id = row['rv_id']
                if rv_id not in rv_gruplar:
                    rv_gruplar[rv_id] = {'rv_id': rv_id, 'ad': row['rv_ad'], 'renk': row['renk'], 'varyantlar': []}
                rv_gruplar[rv_id]['varyantlar'].append({
                    'uv_id':        row['uv_id'],
                    'boyut':        row['boyut'],
                    'recete_durum': row['recete_durum'] or 'TASLAK',
                    'kalem_say':    row['kalem_say'] or 0,
                    'toplam_kg':    round(float(row['toplam_kg'] or 0), 3),
                })
            f_dict['rv_ozet'] = list(rv_gruplar.values())

            # Her varyant için KG maliyet + kalem listesi ekle
            for rv_g in f_dict['rv_ozet']:
                for uv in rv_g['varyantlar']:
                    if uv['toplam_kg'] > 0:
                        uv_kalemler = con.execute("""
                            SELECT rk.miktar_kg, rk.stok_kart_id, rk.sira,
                                   sk.kod AS stok_kod, sk.ad AS stok_ad, sk.kategori
                            FROM nexgen_recete_kalem rk
                            JOIN nexgen_stok_kart sk ON sk.id = rk.stok_kart_id
                            WHERE rk.uretim_varyant_id = ? AND rk.aktif = 1
                            ORDER BY rk.sira, rk.id
                        """, (uv['uv_id'],)).fetchall()
                        mal = 0.0
                        kalemler_liste = []
                        for k in uv_kalemler:
                            fr = con.execute("""
                                SELECT fiyat, para_birimi, kur, fiyat_try
                                FROM nexgen_hammadde_fiyat
                                WHERE stok_kart_id = ? AND aktif = 1
                                ORDER BY fiyat_tarihi DESC, id DESC LIMIT 1
                            """, (k['stok_kart_id'],)).fetchone()
                            if fr:
                                if fr['fiyat_try'] and fr['fiyat_try'] > 0:
                                    bp = float(fr['fiyat_try'])
                                elif fr['para_birimi'] == 'TRY':
                                    bp = float(fr['fiyat'] or 0)
                                elif fr['kur'] and fr['kur'] > 0:
                                    bp = float(fr['fiyat'] or 0) * float(fr['kur'])
                                else:
                                    bp = 0.0
                                mal += float(k['miktar_kg']) * bp
                            kalemler_liste.append({
                                'stok_kod': k['stok_kod'],
                                'stok_ad':  k['stok_ad'],
                                'kategori': k['kategori'] or '',
                                'miktar_kg': float(k['miktar_kg']),
                            })
                        uv['kg_maliyet'] = round(mal / uv['toplam_kg'], 2) if uv['toplam_kg'] > 0 else 0.0
                        uv['kalemler']   = kalemler_liste
                    else:
                        uv['kg_maliyet'] = 0.0
                        uv['kalemler']   = []

    finally:
        con.close()

    can_create  = yetki_var('nexgen.recete.create',  'can_create')
    can_approve = yetki_var('nexgen.recete.approve', 'can_approve')
    can_manage  = yetki_var('nexgen.recete.manage',  'can_manage')

    return render_template(
        'nexgen/recete_liste.html',
        active='nexgen',
        formuller=formuller,
        can_create=can_create,
        can_approve=can_approve,
        can_manage=can_manage,
    )


@nexgen_bp.route('/recete/<int:formul_id>')
@yetki_gerekli('nexgen.recete.view', 'can_view')
def recete_detay(formul_id):
    con = _db()
    try:
        formul = con.execute("""
            SELECT f.*,
                   ku.KullaniciAdi AS olusturan_ad,
                   on_ku.KullaniciAdi AS onaylayan_ad
            FROM nexgen_formul f
            LEFT JOIN sistem_kullanici ku    ON ku.Id    = f.olusturan_id
            LEFT JOIN sistem_kullanici on_ku ON on_ku.Id = f.onaylayan_id
            WHERE f.id = ?
        """, (formul_id,)).fetchone()
        if not formul:
            abort(404)

        # Tüm renk varyantları
        renk_raw = con.execute("""
            SELECT rv.*
            FROM nexgen_renk_varyant rv
            WHERE rv.formul_id = ? AND rv.aktif = 1
            ORDER BY rv.id
        """, (formul_id,)).fetchall()

        # Stok kartı bazında son aktif TRY fiyatı (fiyat yoksa 0)
        # Öncelik: fiyat_try varsa → kullan; yoksa fiyat*kur; yoksa 0
        def _son_fiyat_try(stok_kart_id):
            row = con.execute("""
                SELECT fiyat, para_birimi, kur, fiyat_try
                FROM nexgen_hammadde_fiyat
                WHERE stok_kart_id = ? AND aktif = 1
                ORDER BY fiyat_tarihi DESC, id DESC
                LIMIT 1
            """, (stok_kart_id,)).fetchone()
            if not row:
                return 0.0
            if row['fiyat_try'] and row['fiyat_try'] > 0:
                return float(row['fiyat_try'])
            if row['para_birimi'] == 'TRY':
                return float(row['fiyat'] or 0)
            if row['kur'] and row['kur'] > 0:
                return round(float(row['fiyat'] or 0) * float(row['kur']), 4)
            return float(row['fiyat'] or 0)

        # Her renk için üretim varyantları + kalemleri + maliyet
        agac = []
        for rv in renk_raw:
            uretim_raw = con.execute("""
                SELECT uv.*,
                       ku.KullaniciAdi AS onaylayan_ad
                FROM nexgen_uretim_varyant uv
                LEFT JOIN sistem_kullanici ku ON ku.Id = uv.onaylayan_id
                WHERE uv.renk_varyant_id = ? AND uv.aktif = 1
                ORDER BY uv.boyut
            """, (rv['id'],)).fetchall()

            uretim_listesi = []
            for uv in uretim_raw:
                kalemler_raw = con.execute("""
                    SELECT rk.*, sk.kod AS stok_kod, sk.ad AS stok_ad,
                           sk.kategori, sk.birim
                    FROM nexgen_recete_kalem rk
                    JOIN nexgen_stok_kart sk ON sk.id = rk.stok_kart_id
                    WHERE rk.uretim_varyant_id = ? AND rk.aktif = 1
                    ORDER BY rk.sira, rk.id
                """, (uv['id'],)).fetchall()
                kalemler = []
                kalemler_boya = []
                toplam_kg    = 0.0
                toplam_maliyet = 0.0
                for k in kalemler_raw:
                    k_dict = dict(k)
                    miktar = float(k_dict.get('miktar_kg') or 0)
                    birim_fiyat = _son_fiyat_try(k_dict['stok_kart_id'])
                    satir_maliyet = round(miktar * birim_fiyat, 4)
                    k_dict['birim_fiyat_try'] = birim_fiyat
                    k_dict['satir_maliyet']   = satir_maliyet
                    if (k_dict.get('kategori') or '').upper() == 'BOYA':
                        kalemler_boya.append(k_dict)
                    else:
                        toplam_kg      += miktar
                        toplam_maliyet += satir_maliyet
                        kalemler.append(k_dict)

                toplam_kg      = round(toplam_kg, 3)
                toplam_maliyet = round(toplam_maliyet, 4)
                kg_maliyet = round(toplam_maliyet / toplam_kg, 4) if toplam_kg > 0 else 0.0

                uretim_listesi.append({
                    **dict(uv),
                    'kalemler':       kalemler,
                    'boya_kalem_say': len(kalemler_boya),
                    'toplam_kg':      toplam_kg,
                    'toplam_maliyet': toplam_maliyet,
                    'kg_maliyet':     kg_maliyet,
                })

            agac.append({
                **dict(rv),
                'uretim_listesi': uretim_listesi,
            })

        # FAZ-4F: her varyant için recycle izinleri (con kapanmadan önce)
        for rv in agac:
            for uv in rv['uretim_listesi']:
                uv['recycle_izinleri'] = _uretim_varyant_recycle_izinleri(con, uv['id'])

        # FAZ-4F: modal için hammadde listesi (tüm aktif stok kartları — RECYCLE dışı)
        tum_hammaddeler = con.execute("""
            SELECT id, ad, kod, kategori FROM nexgen_stok_kart
            WHERE aktif=1 AND kategori != 'RECYCLE'
            ORDER BY ad
        """).fetchall()
        tum_hammaddeler = [dict(h) for h in tum_hammaddeler]

    finally:
        con.close()

    can_create       = yetki_var('nexgen.recete.create',   'can_create')
    can_approve      = yetki_var('nexgen.recete.approve',  'can_approve')
    can_manage       = yetki_var('nexgen.recete.manage',   'can_manage')
    can_recycle_mgr  = yetki_var('nexgen.recycle.manage',  'can_recycle_mgr')

    return render_template(
        'nexgen/recete_detay.html',
        active='nexgen',
        formul=dict(formul),
        agac=agac,
        can_create=can_create,
        can_approve=can_approve,
        can_manage=can_manage,
        can_recycle_mgr=can_recycle_mgr,
        tum_hammaddeler=tum_hammaddeler,
    )

# ─────────────────────────────────────────────────────────────
# FAZ-4C-2: REÇETE KLONLAMA + DURUM
# KURAL: nexgen_stok_hareket / nexgen_hammadde_fiyat dokunulmaz.
# ─────────────────────────────────────────────────────────────

# Geçerli recete_durum değerleri
RECETE_DURUM_GECERLI = {'TASLAK', 'DENEME', 'ONAYLI', 'URETIME_ACIK', 'PASIF'}


def _uretime_acik_receteler(con):
    """FAZ-5 hazırlık: üretime açık üretim varyantlarını döner.
    recete_durum = 'URETIME_ACIK' olan aktif varyantlar.
    """
    return con.execute("""
        SELECT uv.id, uv.boyut, uv.ad, uv.recete_durum,
               rv.ad AS renk_ad, rv.renk,
               f.id AS formul_id, f.kod AS formul_kod, f.ad AS formul_ad,
               COUNT(rk.id) AS kalem_say
        FROM nexgen_uretim_varyant uv
        JOIN nexgen_renk_varyant rv ON rv.id = uv.renk_varyant_id
        JOIN nexgen_formul f        ON f.id  = rv.formul_id
        LEFT JOIN nexgen_recete_kalem rk
               ON rk.uretim_varyant_id = uv.id AND rk.aktif = 1
        WHERE uv.aktif = 1 AND rv.aktif = 1 AND f.aktif = 1
          AND uv.recete_durum = 'URETIME_ACIK'
        GROUP BY uv.id
        ORDER BY f.kod, rv.ad, uv.boyut
    """).fetchall()


@nexgen_bp.route('/api/recete/klon', methods=['POST'])
@yetki_gerekli('nexgen.recete.create', 'can_create')
def api_recete_klon():
    """Bir üretim varyantının reçete kalemlerini başka bir üretim varyantına kopyalar.

    POST JSON:
        kaynak_uv_id : int   — kopyalanacak kaynak üretim varyantı
        hedef_uv_id  : int   — hedef üretim varyantı
        notlar       : str   — opsiyonel
    """
    data = request.get_json(silent=True) or {}
    kaynak_uv_id = data.get('kaynak_uv_id')
    hedef_uv_id  = data.get('hedef_uv_id')
    notlar       = (data.get('notlar') or '').strip() or None

    if not kaynak_uv_id or not hedef_uv_id:
        return jsonify({"ok": False, "hata": "kaynak_uv_id ve hedef_uv_id zorunludur."}), 400
    if kaynak_uv_id == hedef_uv_id:
        return jsonify({"ok": False, "hata": "Kaynak ve hedef aynı olamaz."}), 400

    kullanici_id = _kullanici_id()
    con = _db()
    try:
        # Kaynak varyant var mı?
        kaynak = con.execute(
            "SELECT id, ad FROM nexgen_uretim_varyant WHERE id=? AND aktif=1",
            (kaynak_uv_id,)
        ).fetchone()
        if not kaynak:
            return jsonify({"ok": False, "hata": "Kaynak üretim varyantı bulunamadı."}), 404

        # Hedef varyant var mı?
        hedef = con.execute(
            "SELECT id, ad, renk_varyant_id FROM nexgen_uretim_varyant WHERE id=? AND aktif=1",
            (hedef_uv_id,)
        ).fetchone()
        if not hedef:
            return jsonify({"ok": False, "hata": "Hedef üretim varyantı bulunamadı."}), 404

        # Güvenlik: hedefte zaten aktif kalem var mı?
        mevcut_kalem = con.execute(
            "SELECT COUNT(*) AS say FROM nexgen_recete_kalem WHERE uretim_varyant_id=? AND aktif=1",
            (hedef_uv_id,)
        ).fetchone()['say']
        if mevcut_kalem > 0:
            return jsonify({
                "ok": False,
                "hata": f"Hedef varyantta zaten {mevcut_kalem} aktif reçete kalemi var. "
                        "Önce temizleyin veya farklı hedef seçin."
            }), 409

        # Kaynak kalemleri al
        kalemler = con.execute("""
            SELECT stok_kart_id, sira, miktar_kg, aciklama
            FROM nexgen_recete_kalem
            WHERE uretim_varyant_id = ? AND aktif = 1
            ORDER BY sira, id
        """, (kaynak_uv_id,)).fetchall()

        if not kalemler:
            return jsonify({"ok": False, "hata": "Kaynak varyantta kopyalanacak reçete kalemi yok."}), 400

        # Kalemleri hedefe kopyala — miktar_kg REAL olarak aynen
        for k in kalemler:
            con.execute("""
                INSERT INTO nexgen_recete_kalem
                    (uretim_varyant_id, stok_kart_id, sira, miktar_kg, aciklama,
                     aktif, olusturma_tarihi)
                VALUES (?, ?, ?, ?, ?, 1, datetime('now'))
            """, (hedef_uv_id, k['stok_kart_id'], k['sira'],
                  float(k['miktar_kg']),  # REAL korunuyor
                  k['aciklama']))

        # Hedef varyantı TASLAK olarak işaretle (yeni klon)
        con.execute(
            "UPDATE nexgen_uretim_varyant SET recete_durum='TASLAK' WHERE id=?",
            (hedef_uv_id,)
        )

        # Audit log
        con.execute("""
            INSERT INTO nexgen_recete_klon_log
                (kaynak_uv_id, hedef_uv_id, kalem_sayisi, yapan_id, notlar)
            VALUES (?, ?, ?, ?, ?)
        """, (kaynak_uv_id, hedef_uv_id, len(kalemler), kullanici_id, notlar))

        con.commit()

        # Hedef formül id'sini bul (redirect için)
        hedef_formul = con.execute("""
            SELECT rv.formul_id
            FROM nexgen_uretim_varyant uv
            JOIN nexgen_renk_varyant rv ON rv.id = uv.renk_varyant_id
            WHERE uv.id = ?
        """, (hedef_uv_id,)).fetchone()
        hedef_formul_id = hedef_formul['formul_id'] if hedef_formul else None

    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({
        "ok": True,
        "kalem_sayisi": len(kalemler),
        "hedef_uv_id":  hedef_uv_id,
        "hedef_formul_id": hedef_formul_id,
        "mesaj": f"{len(kalemler)} reçete kalemi '{kaynak['ad']}' → '{hedef['ad']}' kopyalandı.",
    })


@nexgen_bp.route('/api/recete/durum-guncelle', methods=['POST'])
@yetki_gerekli('nexgen.recete.manage', 'can_manage')
def api_recete_durum_guncelle():
    """Üretim varyantının recete_durum alanını günceller.
    Sadece nexgen.recete.manage yetkisi gerekir.

    POST JSON:
        uv_id        : int
        yeni_durum   : str  (TASLAK / DENEME / ONAYLI / URETIME_ACIK / PASIF)
        notlar       : str  opsiyonel
    """
    data = request.get_json(silent=True) or {}
    uv_id      = data.get('uv_id')
    yeni_durum = (data.get('yeni_durum') or '').strip().upper()
    notlar     = (data.get('notlar') or '').strip() or None

    if not uv_id:
        return jsonify({"ok": False, "hata": "uv_id zorunludur."}), 400
    if yeni_durum not in RECETE_DURUM_GECERLI:
        return jsonify({
            "ok": False,
            "hata": f"Geçersiz durum. Geçerliler: {', '.join(sorted(RECETE_DURUM_GECERLI))}"
        }), 400

    con = _db()
    try:
        uv = con.execute(
            "SELECT id, ad, recete_durum FROM nexgen_uretim_varyant WHERE id=? AND aktif=1",
            (uv_id,)
        ).fetchone()
        if not uv:
            return jsonify({"ok": False, "hata": "Üretim varyantı bulunamadı."}), 404

        eski_durum = uv['recete_durum'] or 'TASLAK'
        con.execute(
            "UPDATE nexgen_uretim_varyant SET recete_durum=? WHERE id=?",
            (yeni_durum, uv_id)
        )

        # Not: notlar parametresi ileride audit_log tablosuna yazılabilir
        # Şimdilik uv.notlar alanına ekle (mevcut notların üstüne yazmadan)
        if notlar:
            mevcut_not = uv['ad']  # sadece isim referans, notlar alanı güncellenmez bu versiyonda

        con.commit()

    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({
        "ok": True,
        "uv_id":      uv_id,
        "eski_durum": eski_durum,
        "yeni_durum": yeni_durum,
        "mesaj":      f"'{uv['ad']}' durumu {eski_durum} → {yeni_durum} olarak güncellendi.",
    })


@nexgen_bp.route('/api/recete/hedef-varyantlar', methods=['GET'])
@yetki_gerekli('nexgen.recete.view', 'can_view')
def api_recete_hedef_varyantlar():
    """Klonlama modalı için: tüm aktif üretim varyantlarını listeler.
    Opsiyonel: formul_id filtresi.
    """
    formul_id = request.args.get('formul_id', type=int)
    con = _db()
    try:
        where = "WHERE uv.aktif=1 AND rv.aktif=1 AND f.aktif=1"
        params = []
        if formul_id:
            where += " AND f.id=?"
            params.append(formul_id)

        rows = con.execute(f"""
            SELECT uv.id, uv.boyut, uv.ad, uv.recete_durum,
                   rv.id AS renk_id, rv.ad AS renk_ad,
                   f.id AS formul_id, f.kod AS formul_kod, f.ad AS formul_ad,
                   COUNT(rk.id) AS kalem_say
            FROM nexgen_uretim_varyant uv
            JOIN nexgen_renk_varyant rv ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f        ON f.id  = rv.formul_id
            LEFT JOIN nexgen_recete_kalem rk
                   ON rk.uretim_varyant_id = uv.id AND rk.aktif = 1
            {where}
            GROUP BY uv.id
            ORDER BY f.kod, rv.ad, uv.boyut
        """, params).fetchall()
    finally:
        con.close()

    return jsonify([dict(r) for r in rows])

# ═════════════════════════════════════════════════════════════
# FAZ-4C-3: ARGE RENK TEST MERKEZİ
# ─────────────────────────────────────────────────────────────
# KURAL: nexgen_stok_hareket / nexgen_hammadde_fiyat dokunulmaz.
#        Ana reçete kalemleri (nexgen_recete_kalem) değişmez.
#        Sadece nexgen_arge_test + nexgen_arge_test_kalem yazılır.
# ─────────────────────────────────────────────────────────────

def _arge_test_no_uret(con):
    """AT-YYYY-NNNN formatında benzersiz test numarası üretir."""
    from datetime import datetime
    yil = datetime.now().strftime('%Y')
    row = con.execute(
        "SELECT MAX(CAST(SUBSTR(test_no, -4) AS INTEGER)) AS son "
        "FROM nexgen_arge_test WHERE test_no LIKE ?",
        (f'AT-{yil}-%',)
    ).fetchone()
    son = row['son'] if row and row['son'] else 0
    return f'AT-{yil}-{son + 1:04d}'


def _arge_renk_calisma_kaydet(con, kaynak_uv_id, test_batch_kg, yeni_renk_adi, cari_id,
                              shore_hedef, talep_referansi, notlar, makina, kullanici_id):
    """RENK_TEST calismasi olusturur (test + kalem clone). Commit disarida yapilir."""
    uv = con.execute(
        "SELECT id FROM nexgen_uretim_varyant WHERE id=? AND aktif=1",
        (kaynak_uv_id,)
    ).fetchone()
    if not uv:
        raise ValueError("Kaynak uretim varyanti bulunamadi.")

    kalemler = con.execute("""
        SELECT rk.stok_kart_id, rk.sira, rk.miktar_kg, rk.aciklama,
               sk.kod AS stok_kod, sk.ad AS stok_ad
        FROM nexgen_recete_kalem rk
        JOIN nexgen_stok_kart sk ON sk.id = rk.stok_kart_id
        WHERE rk.uretim_varyant_id = ? AND rk.aktif = 1
        ORDER BY rk.sira, rk.id
    """, (kaynak_uv_id,)).fetchall()
    if not kalemler:
        raise ValueError("Kaynak varyantta recete kalemi yok.")

    kaynak_batch_kg = round(sum(float(k['miktar_kg']) for k in kalemler), 3)
    if kaynak_batch_kg <= 0:
        raise ValueError("Kaynak batch KG sifir — recete gecersiz.")

    carpan = test_batch_kg / kaynak_batch_kg
    test_no = _arge_test_no_uret(con)

    from datetime import datetime
    yil = datetime.now().year
    son = con.execute(
        "SELECT COUNT(*) FROM nexgen_arge_test WHERE lot_no LIKE ?",
        (f'ARGE-LOT-{yil}-%',)
    ).fetchone()[0]
    lot_no = f'ARGE-LOT-{yil}-{(son + 1):05d}'

    con.execute("""
        INSERT INTO nexgen_arge_test
            (kaynak_uretim_varyant_id, test_no, test_tipi,
             makina, test_batch_kg, kaynak_batch_kg,
             yeni_renk_adi, notlar, durum,
             cari_id, shore_hedef, lot_no, talep_referansi,
             olusturan_id, olusturma_tarihi)
        VALUES (?, ?, 'RENK_TEST', ?, ?, ?, ?, ?, 'TASLAK', ?, ?, ?, ?, ?, datetime('now'))
    """, (kaynak_uv_id, test_no, makina, test_batch_kg, kaynak_batch_kg,
          yeni_renk_adi, notlar, cari_id, shore_hedef, lot_no, talep_referansi,
          kullanici_id))
    test_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]

    for k in kalemler:
        orj_kg = float(k['miktar_kg'])
        test_kg = round(orj_kg * carpan, 4)
        con.execute("""
            INSERT INTO nexgen_arge_test_kalem
                (test_id, stok_kart_id, sira, orjinal_miktar_kg, test_miktar_kg, aciklama)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (test_id, k['stok_kart_id'], k['sira'], orj_kg, test_kg, k['aciklama']))

    return {
        'test_id': test_id,
        'test_no': test_no,
        'lot_no': lot_no,
        'kaynak_batch_kg': kaynak_batch_kg,
        'test_batch_kg': test_batch_kg,
        'kalem_sayisi': len(kalemler),
    }


def _rf_kod_uret(con):
    """RF-NNN formatinda benzersiz RF renk kodu uretir."""
    row = con.execute(
        "SELECT MAX(CAST(SUBSTR(rf_kod, 4) AS INTEGER)) AS son "
        "FROM nexgen_rf_renk WHERE rf_kod LIKE 'RF-%'"
    ).fetchone()
    son = row['son'] if row and row['son'] else 0
    return f'RF-{son + 1:03d}'


@nexgen_bp.route('/api/arge/test-olustur', methods=['POST'])
@login_gerekli
def api_arge_test_olustur():
    """Kaynak üretim varyantından AR-GE test kaydı oluşturur.
    Ana reçete (nexgen_recete_kalem) YAZILMAZ — sadece ARGE tabloları.

    POST JSON:
        kaynak_uv_id    : int
        test_tipi       : RENK_TEST | FORMUL_TEST
        makina          : str   (default: "7.5 LT")
        test_batch_kg   : float
        yeni_renk_adi   : str   opsiyonel
        notlar          : str   opsiyonel
        cari_id         : int   opsiyonel — hangi cari/firma için (nexgen_cari.id)
        shore_hedef     : float opsiyonel — hedef shore değeri
        lot_no          : str   opsiyonel — otomatik üretilir (ARGE-LOT-YYYY-NNNNN)
        talep_referansi : str   opsiyonel — müşteri talebi / referans notu

    Yetki: nexgen.recete.create VEYA nexgen.tablet.view (tablet hızlı başlatma).
    """
    if not (yetki_var('nexgen.recete.create', 'can_create')
            or yetki_var('nexgen.tablet.view', 'can_view')):
        abort(403)

    data = request.get_json(silent=True) or {}
    kaynak_uv_id    = data.get('kaynak_uv_id')
    test_tipi       = (data.get('test_tipi') or 'RENK_TEST').strip().upper()
    makina          = (data.get('makina') or '7.5 LT').strip()
    yeni_renk_adi   = (data.get('yeni_renk_adi') or '').strip() or None
    notlar          = (data.get('notlar') or data.get('aciklama') or '').strip() or None
    cari_id         = data.get('cari_id') or None
    talep_referansi = (data.get('talep_referansi') or '').strip() or None
    lot_no_giris    = (data.get('lot_no') or '').strip() or None

    try:
        shore_hedef = float(data['shore_hedef']) if data.get('shore_hedef') not in (None, '') else None
    except (ValueError, TypeError):
        shore_hedef = None

    try:
        test_batch_kg = float(data.get('test_batch_kg') or data.get('test_kg') or 0)
    except (ValueError, TypeError):
        return jsonify({"ok": False, "hata": "test_batch_kg geçersiz."}), 400

    if not kaynak_uv_id:
        return jsonify({"ok": False, "hata": "kaynak_uv_id zorunludur."}), 400
    if test_tipi not in {'RENK_TEST', 'FORMUL_TEST'}:
        return jsonify({"ok": False, "hata": "test_tipi: RENK_TEST veya FORMUL_TEST"}), 400
    if test_batch_kg <= 0:
        return jsonify({"ok": False, "hata": "test_batch_kg sıfırdan büyük olmalı."}), 400

    kullanici_id = _kullanici_id()
    con = _db()
    try:
        # Kaynak varyant var mı?
        uv = con.execute(
            "SELECT id, ad FROM nexgen_uretim_varyant WHERE id=? AND aktif=1",
            (kaynak_uv_id,)
        ).fetchone()
        if not uv:
            return jsonify({"ok": False, "hata": "Kaynak üretim varyantı bulunamadı."}), 404

        # Kaynak reçete kalemleri
        kalemler = con.execute("""
            SELECT rk.stok_kart_id, rk.sira, rk.miktar_kg, rk.aciklama,
                   sk.kod AS stok_kod, sk.ad AS stok_ad
            FROM nexgen_recete_kalem rk
            JOIN nexgen_stok_kart sk ON sk.id = rk.stok_kart_id
            WHERE rk.uretim_varyant_id = ? AND rk.aktif = 1
            ORDER BY rk.sira, rk.id
        """, (kaynak_uv_id,)).fetchall()

        if not kalemler:
            return jsonify({"ok": False, "hata": "Kaynak varyantta reçete kalemi yok."}), 400

        # Kaynak batch KG = kalem toplamı
        kaynak_batch_kg = round(sum(float(k['miktar_kg']) for k in kalemler), 3)
        if kaynak_batch_kg <= 0:
            return jsonify({"ok": False, "hata": "Kaynak batch KG sıfır — reçete geçersiz."}), 400

        # Ölçekleme çarpanı
        carpan = test_batch_kg / kaynak_batch_kg

        # Test no üret
        test_no = _arge_test_no_uret(con)

        # Lot no: giriş yoksa otomatik üret (ARGE-LOT-YYYY-NNNNN)
        if lot_no_giris:
            lot_no = lot_no_giris
        else:
            _yil = __import__('datetime').datetime.now().year
            _son = con.execute(
                "SELECT COUNT(*) FROM nexgen_arge_test WHERE lot_no LIKE ?",
                (f'ARGE-LOT-{_yil}-%',)
            ).fetchone()[0]
            lot_no = f'ARGE-LOT-{_yil}-{(_son + 1):05d}'

        # Ana test kaydı — yeni kimlik alanları dahil
        con.execute("""
            INSERT INTO nexgen_arge_test
                (kaynak_uretim_varyant_id, test_no, test_tipi,
                 makina, test_batch_kg, kaynak_batch_kg,
                 yeni_renk_adi, notlar, durum,
                 cari_id, shore_hedef, lot_no, talep_referansi,
                 olusturan_id, olusturma_tarihi)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'TASLAK', ?, ?, ?, ?, ?, datetime('now'))
        """, (kaynak_uv_id, test_no, test_tipi,
              makina, test_batch_kg, kaynak_batch_kg,
              yeni_renk_adi, notlar,
              cari_id, shore_hedef, lot_no, talep_referansi,
              kullanici_id))
        test_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Test kalemleri — REAL hassasiyet, round(x, 4) yeterli
        test_kalemler = []
        for k in kalemler:
            orj_kg  = float(k['miktar_kg'])
            test_kg = round(orj_kg * carpan, 4)
            con.execute("""
                INSERT INTO nexgen_arge_test_kalem
                    (test_id, stok_kart_id, sira,
                     orjinal_miktar_kg, test_miktar_kg, aciklama)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (test_id, k['stok_kart_id'], k['sira'],
                  orj_kg, test_kg, k['aciklama']))
            test_kalemler.append({
                'stok_kod':        k['stok_kod'],
                'stok_ad':         k['stok_ad'],
                'orjinal_kg':      orj_kg,
                'test_kg':         test_kg,
            })

        con.commit()

    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({
        "ok":              True,
        "test_id":         test_id,
        "test_no":         test_no,
        "lot_no":          lot_no,
        "kaynak_batch_kg": kaynak_batch_kg,
        "test_batch_kg":   test_batch_kg,
        "carpan":          round(carpan, 6),
        "kalem_sayisi":    len(test_kalemler),
        "kalemler":        test_kalemler,
        "mesaj":           f"{test_no} / {lot_no} oluşturuldu. {len(test_kalemler)} kalem ölçeklendi.",
    })


@nexgen_bp.route('/arge/renk-yeni')
@yetki_gerekli('nexgen.recete.create', 'can_create')
def arge_renk_yeni():
    """Masaustu RENK gelistirme calismasi olusturma formu."""
    con = _db()
    try:
        cariler = con.execute(
            "SELECT id, cari_kod, unvan FROM nexgen_cari WHERE aktif=1 ORDER BY cari_kod"
        ).fetchall()
        kaynak_uvler = con.execute("""
            SELECT uv.id, uv.boyut,
                   rv.ad AS renk_ad,
                   f.id AS formul_id, f.kod AS formul_kod, f.ad AS formul_ad,
                   ROUND(SUM(rk.miktar_kg), 3) AS kaynak_batch_kg,
                   COUNT(rk.id) AS kalem_sayisi
            FROM nexgen_uretim_varyant uv
            JOIN nexgen_renk_varyant rv ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f ON f.id = rv.formul_id
            JOIN nexgen_recete_kalem rk ON rk.uretim_varyant_id = uv.id AND rk.aktif = 1
            WHERE uv.aktif = 1 AND rv.aktif = 1 AND f.aktif = 1
            GROUP BY uv.id
            HAVING kalem_sayisi > 0
            ORDER BY f.ad, rv.ad, uv.boyut
        """).fetchall()
    finally:
        con.close()

    return render_template(
        'nexgen/arge_renk_olustur.html',
        active='nexgen',
        cariler=[dict(c) for c in cariler],
        kaynak_uvler=[dict(u) for u in kaynak_uvler],
    )


@nexgen_bp.route('/api/arge/renk-calisma-olustur', methods=['POST'])
@yetki_gerekli('nexgen.recete.create', 'can_create')
def api_arge_renk_calisma_olustur():
    """RENK gelistirme calismasi olusturur (TASLAK + recete clone).

    POST JSON:
        cari_id         : int   zorunlu
        kaynak_uv_id    : int   zorunlu
        yeni_renk_adi   : str   zorunlu
        test_batch_kg   : float zorunlu (>0)
        shore_hedef     : float opsiyonel
        talep_referansi : str   opsiyonel
        notlar          : str   opsiyonel
        makina          : str   opsiyonel (default 7.5 LT)
    """
    data = request.get_json(silent=True) or {}
    cari_id = data.get('cari_id')
    kaynak_uv_id = data.get('kaynak_uv_id')
    yeni_renk_adi = (data.get('yeni_renk_adi') or '').strip()
    notlar = (data.get('notlar') or '').strip() or None
    talep_referansi = (data.get('talep_referansi') or '').strip() or None
    makina = (data.get('makina') or '7.5 LT').strip() or '7.5 LT'

    try:
        shore_hedef = float(data['shore_hedef']) if data.get('shore_hedef') not in (None, '') else None
    except (ValueError, TypeError):
        shore_hedef = None

    try:
        test_batch_kg = float(data.get('test_batch_kg') or 0)
    except (ValueError, TypeError):
        return jsonify({"ok": False, "hata": "test_batch_kg gecersiz."}), 400

    if not cari_id:
        return jsonify({"ok": False, "hata": "cari_id zorunludur."}), 400
    if not kaynak_uv_id:
        return jsonify({"ok": False, "hata": "kaynak_uv_id zorunludur."}), 400
    if not yeni_renk_adi:
        return jsonify({"ok": False, "hata": "yeni_renk_adi zorunludur."}), 400
    if test_batch_kg <= 0:
        return jsonify({"ok": False, "hata": "test_batch_kg sifirdan buyuk olmali."}), 400

    kullanici_id = _kullanici_id()
    con = _db()
    try:
        cari = con.execute(
            "SELECT id FROM nexgen_cari WHERE id=? AND aktif=1", (cari_id,)
        ).fetchone()
        if not cari:
            return jsonify({"ok": False, "hata": "Cari kaydi bulunamadi."}), 400

        sonuc = _arge_renk_calisma_kaydet(
            con, kaynak_uv_id, test_batch_kg, yeni_renk_adi, cari_id,
            shore_hedef, talep_referansi, notlar, makina, kullanici_id,
        )
        con.commit()
    except ValueError as e:
        con.rollback()
        msg = str(e)
        code = 404 if 'bulunamadi' in msg.lower() else 400
        return jsonify({"ok": False, "hata": msg}), code
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({
        "ok": True,
        "test_id": sonuc['test_id'],
        "test_no": sonuc['test_no'],
        "lot_no": sonuc['lot_no'],
        "test_tipi": "RENK_TEST",
        "durum": "TASLAK",
        "kaynak_batch_kg": sonuc['kaynak_batch_kg'],
        "test_batch_kg": sonuc['test_batch_kg'],
        "kalem_sayisi": sonuc['kalem_sayisi'],
        "redirect_url": f"/nexgen/arge/test/{sonuc['test_id']}",
    })


@nexgen_bp.route('/arge')
@yetki_gerekli('nexgen.recete.view', 'can_view')
def arge_liste():
    """AR-GE Merkezi — tüm test kayıtları listesi."""
    con = _db()
    try:
        rows = con.execute("""
            SELECT t.id, t.test_no, t.lot_no, t.test_tipi, t.makina,
                   t.durum, t.shore_hedef, t.shore_degeri,
                   t.yeni_renk_adi, t.talep_referansi,
                   t.olusturma_tarihi, t.onay_tarihi,
                   uv.boyut,
                   rv.ad AS renk_ad,
                   f.id AS formul_id, f.kod AS formul_kod, f.ad AS formul_ad,
                   c.unvan AS cari_unvan, c.cari_kod,
                   ku.KullaniciAdi AS olusturan_ad
            FROM nexgen_arge_test t
            JOIN nexgen_uretim_varyant uv ON uv.id = t.kaynak_uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            LEFT JOIN nexgen_cari c       ON c.id  = t.cari_id
            LEFT JOIN sistem_kullanici ku ON ku.Id = t.olusturan_id
            WHERE t.aktif = 1
            ORDER BY t.id DESC
        """).fetchall()
        testler = [dict(r) for r in rows]
    finally:
        con.close()

    can_create = yetki_var('nexgen.recete.create', 'can_create')

    return render_template(
        'nexgen/arge_liste.html',
        active='nexgen',
        testler=testler,
        can_create=can_create,
    )


@nexgen_bp.route('/arge/test/<int:test_id>')
@yetki_gerekli('nexgen.recete.view', 'can_view')
def arge_test_detay(test_id):
    """AR-GE test detay sayfası."""
    con = _db()
    try:
        test = con.execute("""
            SELECT t.*,
                   uv.ad AS uv_ad, uv.boyut,
                   rv.ad AS renk_ad,
                   f.id  AS formul_id, f.kod AS formul_kod, f.ad AS formul_ad,
                   c.unvan AS cari_unvan, c.cari_kod,
                   ku.KullaniciAdi AS olusturan_ad,
                   on_ku.KullaniciAdi AS onaylayan_ad
            FROM nexgen_arge_test t
            JOIN nexgen_uretim_varyant uv ON uv.id = t.kaynak_uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            LEFT JOIN nexgen_cari c       ON c.id  = t.cari_id
            LEFT JOIN sistem_kullanici ku    ON ku.Id    = t.olusturan_id
            LEFT JOIN sistem_kullanici on_ku ON on_ku.Id = t.onaylayan_id
            WHERE t.id = ? AND t.aktif = 1
        """, (test_id,)).fetchone()
        if not test:
            abort(404)
        test_row = dict(test)

        kalemler = con.execute("""
            SELECT tk.id, tk.stok_kart_id, tk.sira, tk.orjinal_miktar_kg, tk.test_miktar_kg, tk.aciklama,
                   sk.kod AS stok_kod, sk.ad AS stok_ad, sk.kategori
            FROM nexgen_arge_test_kalem tk
            JOIN nexgen_stok_kart sk ON sk.id = tk.stok_kart_id
            WHERE tk.test_id = ?
            ORDER BY tk.sira
        """, (test_id,)).fetchall()
        kalem_list = [dict(k) for k in kalemler]
        boya_kalemler = _boya_kalemler_filtre(kalem_list)
        if test_row.get('test_tipi') == 'RENK_TEST':
            taban_kalemler = _taban_kalemler(kalem_list)
        else:
            taban_kalemler = kalem_list
        taban_orj_top = round(sum(k['orjinal_miktar_kg'] for k in taban_kalemler), 3)
        taban_test_top = round(sum(k['test_miktar_kg'] for k in taban_kalemler), 4)
        boya_stok_secenekleri = []
        if (test_row.get('durum') == 'TASLAK'
                and test_row.get('test_tipi') == 'RENK_TEST'):
            boya_stok_secenekleri = [
                dict(r) for r in con.execute("""
                    SELECT id, kod, ad, kategori
                    FROM nexgen_stok_kart
                    WHERE aktif = 1 AND kategori = 'BOYA'
                    ORDER BY kod
                """).fetchall()
            ]

        # Bağlantılı üretim varyantı bilgisi (FAZ-4C-4)
        olusan_uv = None
        rf_renk = None
        test_d = test_row
        if test_d.get('olusan_uretim_varyant_id'):
            row = con.execute("""
                SELECT uv.id, uv.ad, uv.boyut, uv.recete_durum,
                       rv.ad AS renk_ad, f.kod AS formul_kod, f.id AS formul_id
                FROM nexgen_uretim_varyant uv
                JOIN nexgen_renk_varyant rv ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f        ON f.id  = rv.formul_id
                WHERE uv.id = ?
            """, (test_d['olusan_uretim_varyant_id'],)).fetchone()
            if row:
                olusan_uv = dict(row)

        rf_renk = _arge_rf_mevcut_bul(con, test_id, test_row)

    finally:
        con.close()

    can_manage  = yetki_var('nexgen.recete.manage', 'can_manage')
    can_create  = yetki_var('nexgen.recete.create', 'can_create')
    can_approve = yetki_var('nexgen.recete.approve', 'can_approve')

    return render_template(
        'nexgen/arge_test_detay.html',
        active='nexgen',
        test=test_d,
        kalemler=kalem_list,
        taban_kalemler=taban_kalemler,
        taban_orj_top=taban_orj_top,
        taban_test_top=taban_test_top,
        boya_kalemler=boya_kalemler,
        boya_stok_secenekleri=boya_stok_secenekleri,
        can_manage=can_manage,
        can_create=can_create,
        can_approve=can_approve,
        olusan_uv=olusan_uv,
        rf_renk=rf_renk,
        boya_editable=(
            can_create
            and test_d.get('durum') == 'TASLAK'
            and test_d.get('test_tipi') == 'RENK_TEST'
        ),
    )


@nexgen_bp.route('/api/arge/renk-stok-secenekleri', methods=['GET'])
@login_gerekli
def api_arge_renk_stok_secenekleri():
    """RENK_TEST boya kalem edit icin aktif BOYA stok kartlari."""
    if not (yetki_var('nexgen.recete.view', 'can_view')
            or yetki_var('nexgen.tablet.view', 'can_view')):
        abort(403)
    con = _db()
    try:
        rows = con.execute("""
            SELECT id, kod, ad, kategori
            FROM nexgen_stok_kart
            WHERE aktif = 1 AND kategori = 'BOYA'
            ORDER BY kod
        """).fetchall()
    finally:
        con.close()
    return jsonify([dict(r) for r in rows])


@nexgen_bp.route('/api/arge/test/<int:test_id>/boya-kalemler', methods=['PUT'])
@login_gerekli
def api_arge_test_boya_kalemler(test_id):
    """TASLAK RENK_TEST icin yalnizca BOYA kalemlerini gunceller.

    PUT JSON:
        kalemler: [{stok_kart_id: int, test_miktar_kg: float}, ...]

    Yetki: nexgen.recete.create VEYA nexgen.tablet.view (tablet numune girisi).
    """
    if not (yetki_var('nexgen.recete.create', 'can_create')
            or yetki_var('nexgen.tablet.view', 'can_view')):
        abort(403)

    data = request.get_json(silent=True) or {}
    kalemler_in = data.get('kalemler')
    if kalemler_in is None or not isinstance(kalemler_in, list):
        return jsonify({"ok": False, "hata": "kalemler listesi zorunludur."}), 400
    if not kalemler_in:
        return jsonify({"ok": False, "hata": "En az bir BOYA kalemi gerekli."}), 400

    con = _db()
    try:
        test = con.execute(
            "SELECT id, durum, test_tipi, test_batch_kg, kaynak_batch_kg "
            "FROM nexgen_arge_test WHERE id=? AND aktif=1",
            (test_id,)
        ).fetchone()
        if not test:
            return jsonify({"ok": False, "hata": "Test kaydi bulunamadi."}), 404
        if test['durum'] != 'TASLAK':
            return jsonify({
                "ok": False,
                "hata": f"Sadece TASLAK testlerde boya duzenlenebilir. Mevcut: {test['durum']}"
            }), 400
        if test['test_tipi'] != 'RENK_TEST':
            return jsonify({
                "ok": False,
                "hata": f"Sadece RENK_TEST tipinde boya duzenlenebilir. Mevcut: {test['test_tipi']}"
            }), 400

        test_batch_kg = float(test['test_batch_kg'])
        kaynak_batch_kg = float(test['kaynak_batch_kg'])
        if test_batch_kg <= 0 or kaynak_batch_kg <= 0:
            return jsonify({"ok": False, "hata": "test_batch_kg veya kaynak_batch_kg gecersiz."}), 400

        non_boya_before = con.execute("""
            SELECT COUNT(*) AS c FROM nexgen_arge_test_kalem tk
            JOIN nexgen_stok_kart sk ON sk.id = tk.stok_kart_id
            WHERE tk.test_id = ? AND sk.kategori != 'BOYA'
        """, (test_id,)).fetchone()['c']

        seen_stok = set()
        parsed = []
        for i, k in enumerate(kalemler_in):
            try:
                stok_id = int(k.get('stok_kart_id'))
                test_kg = float(k.get('test_miktar_kg'))
            except (TypeError, ValueError):
                return jsonify({"ok": False, "hata": f"Kalem {i + 1}: gecersiz stok veya miktar."}), 400
            if stok_id in seen_stok:
                return jsonify({"ok": False, "hata": "Ayni stok_kart_id birden fazla kez gonderilemez."}), 400
            if test_kg <= 0:
                return jsonify({"ok": False, "hata": f"Kalem {i + 1}: test_miktar_kg sifirdan buyuk olmali."}), 400
            seen_stok.add(stok_id)
            parsed.append((stok_id, round(test_kg, 4)))

        for stok_id, _ in parsed:
            sk = con.execute(
                "SELECT id, kategori, aktif FROM nexgen_stok_kart WHERE id=?",
                (stok_id,)
            ).fetchone()
            if not sk or not sk['aktif']:
                return jsonify({"ok": False, "hata": f"Stok karti bulunamadi veya pasif (id={stok_id})."}), 400
            if sk['kategori'] != 'BOYA':
                return jsonify({
                    "ok": False,
                    "hata": f"Yalnizca BOYA stoklari kabul edilir (id={stok_id}, kategori={sk['kategori']})."
                }), 400

        max_sira_row = con.execute("""
            SELECT COALESCE(MAX(tk.sira), 0) AS maks
            FROM nexgen_arge_test_kalem tk
            JOIN nexgen_stok_kart sk ON sk.id = tk.stok_kart_id
            WHERE tk.test_id = ? AND sk.kategori != 'BOYA'
        """, (test_id,)).fetchone()
        base_sira = int(max_sira_row['maks'] or 0)

        con.execute("""
            DELETE FROM nexgen_arge_test_kalem
            WHERE test_id = ?
              AND stok_kart_id IN (
                  SELECT id FROM nexgen_stok_kart WHERE kategori = 'BOYA'
              )
        """, (test_id,))

        carpan = kaynak_batch_kg / test_batch_kg
        for idx, (stok_id, test_kg) in enumerate(parsed):
            orj_kg = round(test_kg * carpan, 4)
            con.execute("""
                INSERT INTO nexgen_arge_test_kalem
                    (test_id, stok_kart_id, sira, orjinal_miktar_kg, test_miktar_kg, aciklama)
                VALUES (?, ?, ?, ?, ?, NULL)
            """, (test_id, stok_id, base_sira + idx + 1, orj_kg, test_kg))

        non_boya_after = con.execute("""
            SELECT COUNT(*) AS c FROM nexgen_arge_test_kalem tk
            JOIN nexgen_stok_kart sk ON sk.id = tk.stok_kart_id
            WHERE tk.test_id = ? AND sk.kategori != 'BOYA'
        """, (test_id,)).fetchone()['c']

        if non_boya_after != non_boya_before:
            con.rollback()
            return jsonify({"ok": False, "hata": "Taban reçete kalemleri korunamadi."}), 500

        con.commit()
        boya_cnt = len(parsed)
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({
        "ok": True,
        "test_id": test_id,
        "boya_kalem_sayisi": boya_cnt,
    })

@nexgen_bp.route('/api/arge/sonuc-kaydet', methods=['POST'])
@login_gerekli
def api_arge_sonuc_kaydet():
    """Test sonucunu ve durum değişikliğini kaydeder (AR-GE).
    ONAYLANDI / REDDEDILDI bu endpoint'ten kabul edilmez.

    POST JSON:
        test_id           : int
        durum             : TEST_EDILDI | ONAYA_GONDERILDI
        renk_tuttu        : 1/0  opsiyonel
        shore_degeri      : float opsiyonel
        kopurme_notu      : str opsiyonel
        cekme_problemi    : 1/0 opsiyonel
        genel_aciklama    : str opsiyonel
        sonuc_notu        : str opsiyonel

    Tablet alias: urun_calisti, abras_var, pisme_saniye, enjeksiyon_saniye, not

    Yetki: nexgen.recete.create VEYA nexgen.tablet.view
    """
    if not (yetki_var('nexgen.recete.create', 'can_create')
            or yetki_var('nexgen.tablet.view', 'can_view')):
        abort(403)

    data = request.get_json(silent=True) or {}
    test_id = data.get('test_id')
    durum   = (data.get('durum') or '').strip().upper()

    GECERLI_DURUM = {'TEST_EDILDI', 'ONAYA_GONDERILDI'}
    if not test_id:
        return jsonify({"ok": False, "hata": "test_id zorunludur."}), 400
    if durum not in GECERLI_DURUM:
        return jsonify({"ok": False,
                        "hata": f"Geçerli durumlar: {', '.join(sorted(GECERLI_DURUM))}"}), 400

    renk_tuttu       = data.get('renk_tuttu')
    if renk_tuttu is None and data.get('urun_calisti') is not None:
        renk_tuttu = data.get('urun_calisti')
    shore_degeri     = data.get('shore_degeri')
    kopurme_notu     = (data.get('kopurme_notu') or '').strip() or None
    cekme_problemi   = data.get('cekme_problemi')
    if cekme_problemi is None and data.get('abras_var') is not None:
        cekme_problemi = data.get('abras_var')
    genel_aciklama   = (data.get('genel_aciklama') or '').strip() or None
    sonuc_notu       = (data.get('sonuc_notu') or data.get('not') or '').strip() or None

    pisme = data.get('pisme_saniye')
    if pisme in (None, ''):
        pisme = data.get('enjeksiyon_saniye')
    if pisme not in (None, ''):
        pisme_metin = f'Pişme/enjeksiyon: {pisme} sn'
        genel_aciklama = pisme_metin if not genel_aciklama else f'{pisme_metin}\n{genel_aciklama}'

    try:
        if shore_degeri is not None:
            shore_degeri = float(shore_degeri)
    except (ValueError, TypeError):
        shore_degeri = None

    con = _db()
    try:
        test = con.execute(
            "SELECT id, durum FROM nexgen_arge_test WHERE id=? AND aktif=1",
            (test_id,)
        ).fetchone()
        if not test:
            return jsonify({"ok": False, "hata": "Test kaydı bulunamadı."}), 404

        mevcut = test['durum']

        # Onay sürecindeki / kilitli kayıtlar — sadece REDDEDILDI → tekrar onaya gönder
        if mevcut in ('ONAYA_GONDERILDI', 'ONAYLANDI'):
            return jsonify({
                "ok": False,
                "hata": f"Bu durumda sonuç güncellenemez ({mevcut})."
            }), 400
        if mevcut == 'REDDEDILDI' and durum != 'ONAYA_GONDERILDI':
            return jsonify({
                "ok": False,
                "hata": "Reddedilmiş test yalnızca tekrar onaya gönderilebilir."
            }), 400

        # Durum geçiş kuralları
        if durum == 'TEST_EDILDI':
            if mevcut not in ('TASLAK', 'TEST_EDILDI'):
                return jsonify({
                    "ok": False,
                    "hata": f"TASLAK → TEST_EDILDI geçişi gerekli. Mevcut: {mevcut}"
                }), 400
        elif durum == 'ONAYA_GONDERILDI':
            if mevcut not in ('TEST_EDILDI', 'REDDEDILDI'):
                return jsonify({
                    "ok": False,
                    "hata": "Sadece TEST_EDILDI veya REDDEDILDI testler onaya gönderilebilir."
                }), 400

        if durum == 'ONAYA_GONDERILDI' and mevcut == 'REDDEDILDI':
            # Tekrar onaya gönder — sonuç alanları kilitli, sadece durum + eski onay temizle
            con.execute("""
                UPDATE nexgen_arge_test
                SET durum='ONAYA_GONDERILDI',
                    onaylayan_id=NULL, onay_tarihi=NULL, onay_notu=NULL
                WHERE id=?
            """, (test_id,))
        else:
            con.execute("""
                UPDATE nexgen_arge_test
                SET durum=?, renk_tuttu=?, shore_degeri=?,
                    kopurme_notu=?, cekme_problemi=?,
                    genel_aciklama=?, sonuc_notu=?
                WHERE id=?
            """, (durum,
                  int(renk_tuttu) if renk_tuttu is not None else None,
                  shore_degeri, kopurme_notu,
                  int(cekme_problemi) if cekme_problemi is not None else None,
                  genel_aciklama, sonuc_notu,
                  test_id))
        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({"ok": True, "test_id": test_id, "yeni_durum": durum})


@nexgen_bp.route('/api/arge/onay-islem', methods=['POST'])
@yetki_gerekli('nexgen.recete.approve', 'can_approve')
def api_arge_onay_islem():
    """Yönetim onayı — sadece ONAYA_GONDERILDI testler.

    POST JSON:
        test_id    : int
        eylem      : ONAYLA | REDDET
        onay_notu  : str opsiyonel
    """
    data = request.get_json(silent=True) or {}
    test_id   = data.get('test_id')
    eylem     = (data.get('eylem') or '').strip().upper()
    onay_notu = (data.get('onay_notu') or '').strip() or None

    if not test_id:
        return jsonify({"ok": False, "hata": "test_id zorunludur."}), 400
    if eylem not in ('ONAYLA', 'REDDET'):
        return jsonify({"ok": False, "hata": "eylem: ONAYLA veya REDDET"}), 400

    kullanici_id = _kullanici_id()
    yeni_durum = 'ONAYLANDI' if eylem == 'ONAYLA' else 'REDDEDILDI'

    con = _db()
    try:
        test = con.execute(
            "SELECT id, durum FROM nexgen_arge_test WHERE id=? AND aktif=1",
            (test_id,)
        ).fetchone()
        if not test:
            return jsonify({"ok": False, "hata": "Test kaydı bulunamadı."}), 404
        if test['durum'] != 'ONAYA_GONDERILDI':
            return jsonify({
                "ok": False,
                "hata": f"Sadece ONAYA_GONDERILDI testler işlenebilir. Mevcut: {test['durum']}"
            }), 400

        con.execute("""
            UPDATE nexgen_arge_test
            SET durum=?, onaylayan_id=?, onay_tarihi=datetime('now'), onay_notu=?
            WHERE id=?
        """, (yeni_durum, kullanici_id, onay_notu, test_id))
        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({"ok": True, "test_id": test_id, "yeni_durum": yeni_durum})


def _arge_rf_mevcut_bul(con, test_id, test_row=None):
    """Test için mevcut RF kaydını döner; test.rf_renk_id ile senkron eksikse tamamlar."""
    if test_row is None:
        test_row = con.execute(
            "SELECT id, rf_renk_id FROM nexgen_arge_test WHERE id=? AND aktif=1",
            (test_id,)
        ).fetchone()
    if not test_row:
        return None
    rf_id = test_row['rf_renk_id']
    if rf_id:
        rf = con.execute(
            "SELECT id, rf_kod, ad, durum, cari_id, ilk_talep_cari_id "
            "FROM nexgen_rf_renk WHERE id=? AND aktif=1",
            (rf_id,)
        ).fetchone()
        if rf:
            return dict(rf)
    rf = con.execute(
        "SELECT id, rf_kod, ad, durum, cari_id, ilk_talep_cari_id "
        "FROM nexgen_rf_renk WHERE kaynak_arge_test_id=? AND aktif=1",
        (test_id,)
    ).fetchone()
    if rf and not test_row['rf_renk_id']:
        con.execute(
            "UPDATE nexgen_arge_test SET rf_renk_id=? WHERE id=?",
            (rf['id'], test_id)
        )
    return dict(rf) if rf else None


def _arge_rf_olustur_core(con, test_id, cari_id_override=None, siparis_id_override=None):
    """ONAYLANDI testten RF oluşturur veya mevcut RF döner. commit yapmaz."""
    test = con.execute("""
        SELECT t.*,
               rv.ad AS kaynak_renk_ad,
               f.id AS formul_id, f.kod AS formul_kod, f.ad AS formul_ad
        FROM nexgen_arge_test t
        JOIN nexgen_uretim_varyant uv ON uv.id = t.kaynak_uretim_varyant_id
        JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
        JOIN nexgen_formul f          ON f.id  = rv.formul_id
        WHERE t.id = ? AND t.aktif = 1
    """, (test_id,)).fetchone()
    if not test:
        return {'ok': False, 'hata': 'Test kaydi bulunamadi.', 'status': 404}
    if test['durum'] != 'ONAYLANDI':
        return {
            'ok': False,
            'hata': f"Sadece ONAYLANDI testlerden RF olusturulabilir. Mevcut: {test['durum']}",
            'status': 400,
        }

    mevcut = _arge_rf_mevcut_bul(con, test_id, test)
    if mevcut:
        return {
            'ok': True, 'mevcut': True, 'test_id': test_id,
            'rf_renk_id': mevcut['id'], 'rf_kod': mevcut['rf_kod'],
            'rf_ad': mevcut['ad'], 'formul_id': test['formul_id'],
            'formul_ad': test['formul_ad'],
            'cari_id': mevcut.get('cari_id'), 'siparis_id': None,
            'boya_kalem_sayisi': con.execute(
                "SELECT COUNT(*) FROM nexgen_rf_kalem WHERE rf_renk_id=? AND aktif=1",
                (mevcut['id'],)
            ).fetchone()[0],
        }

    cari_id = test['cari_id']
    if cari_id_override is not None and cari_id_override != '':
        try:
            cari_id = int(cari_id_override)
        except (TypeError, ValueError):
            return {'ok': False, 'hata': 'cari_id gecersiz.', 'status': 400}
        if not con.execute(
            "SELECT id FROM nexgen_cari WHERE id=? AND aktif=1", (cari_id,)
        ).fetchone():
            return {'ok': False, 'hata': f'Cari bulunamadi (id={cari_id}).', 'status': 400}

    siparis_id = None
    if siparis_id_override is not None and siparis_id_override != '':
        try:
            siparis_id = int(siparis_id_override)
        except (TypeError, ValueError):
            return {'ok': False, 'hata': 'siparis_id gecersiz.', 'status': 400}
        if not con.execute(
            "SELECT id FROM nexgen_uretim_plan WHERE id=?", (siparis_id,)
        ).fetchone():
            return {
                'ok': False,
                'hata': f'Uretim plani bulunamadi (siparis_id={siparis_id}).',
                'status': 400,
            }

    test_batch_kg = float(test['test_batch_kg'])
    kaynak_batch_kg = float(test['kaynak_batch_kg'])
    if test_batch_kg <= 0 or kaynak_batch_kg <= 0:
        return {
            'ok': False,
            'hata': 'test_batch_kg veya kaynak_batch_kg sifir — RF olusturulamaz.',
            'status': 400,
        }

    boya_kalemler = con.execute("""
        SELECT tk.stok_kart_id, tk.sira, tk.test_miktar_kg, tk.aciklama
        FROM nexgen_arge_test_kalem tk
        JOIN nexgen_stok_kart sk ON sk.id = tk.stok_kart_id
        WHERE tk.test_id = ?
          AND sk.kategori = 'BOYA'
          AND sk.aktif = 1
          AND tk.test_miktar_kg > 0
        ORDER BY tk.sira
    """, (test_id,)).fetchall()
    if not boya_kalemler:
        return {
            'ok': False,
            'hata': 'BOYA kategorisinde kalem bulunamadi — RF olusturulamaz.',
            'status': 400,
        }

    rf_ad = (test['yeni_renk_adi'] or test['kaynak_renk_ad'] or '').strip()
    if not rf_ad:
        return {'ok': False, 'hata': 'RF renk adi belirlenemedi.', 'status': 400}

    rf_kod = _rf_kod_uret(con)
    carpan = kaynak_batch_kg / test_batch_kg
    aciklama = (test['genel_aciklama'] or test['sonuc_notu'] or '').strip() or None
    ilk_cari = test['cari_id']

    con.execute("""
        INSERT INTO nexgen_rf_renk
            (rf_kod, ad, durum, kaynak_arge_test_id, ilk_talep_cari_id,
             cari_id, siparis_id, aciklama, olusturan_id, onaylayan_id,
             onay_tarihi, aktif)
        VALUES (?, ?, 'ONAYLI', ?, ?, ?, ?, ?, ?, ?, ?, 1)
    """, (rf_kod, rf_ad, test_id, ilk_cari, cari_id, siparis_id,
          aciklama, test['olusturan_id'], test['onaylayan_id'], test['onay_tarihi']))
    rf_renk_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]

    for k in boya_kalemler:
        miktar_kg = round(float(k['test_miktar_kg']) * carpan, 4)
        if miktar_kg <= 0:
            continue
        con.execute("""
            INSERT INTO nexgen_rf_kalem
                (rf_renk_id, stok_kart_id, miktar_kg, sira, aciklama, aktif)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (rf_renk_id, k['stok_kart_id'], miktar_kg, k['sira'], k['aciklama']))

    con.execute("""
        INSERT INTO nexgen_rf_formul_uygunluk
            (rf_renk_id, formul_id, kaynak_arge_test_id, durum,
             ilk_talep_cari_id, shore_hedef, shore_sonuc,
             renk_sonucu, numune_sonucu, onay_tarihi, aktif)
        VALUES (?, ?, ?, 'ONAYLI', ?, ?, ?, ?, ?, ?, 1)
    """, (rf_renk_id, test['formul_id'], test_id, ilk_cari,
          test['shore_hedef'], test['shore_degeri'],
          test['renk_tuttu'], test['cekme_problemi'], test['onay_tarihi']))

    con.execute(
        "UPDATE nexgen_arge_test SET rf_renk_id=? WHERE id=?",
        (rf_renk_id, test_id)
    )

    return {
        'ok': True, 'mevcut': False, 'test_id': test_id,
        'rf_renk_id': rf_renk_id, 'rf_kod': rf_kod, 'rf_ad': rf_ad,
        'formul_id': test['formul_id'], 'formul_ad': test['formul_ad'],
        'cari_id': cari_id, 'siparis_id': siparis_id,
        'boya_kalem_sayisi': len(boya_kalemler),
    }


@nexgen_bp.route('/api/arge/rf-olustur', methods=['POST'])
@yetki_gerekli('nexgen.recete.manage', 'can_manage')
def api_arge_rf_olustur():
    """Onayli AR-GE testinden global RF renk + formul uygunluk olusturur.
    nexgen_recete_kalem / uretim_varyant DOKUNULMAZ.
    Mevcut RF varsa ayni kaydi doner (idempotent).

    POST JSON:
        test_id   : int
        cari_id   : int (opsiyonel; yoksa test.cari_id)
        siparis_id: int (opsiyonel; nexgen_uretim_plan.id)
    """
    data = request.get_json(silent=True) or {}
    test_id = data.get('test_id')
    if not test_id:
        return jsonify({"ok": False, "hata": "test_id zorunludur."}), 400

    con = _db()
    try:
        sonuc = _arge_rf_olustur_core(
            con, test_id,
            cari_id_override=data.get('cari_id'),
            siparis_id_override=data.get('siparis_id'),
        )
        if not sonuc.get('ok'):
            con.rollback()
            return jsonify({"ok": False, "hata": sonuc.get('hata')}), sonuc.get('status', 400)
        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify(sonuc)


@nexgen_bp.route('/api/rf/kullanim-log', methods=['POST'])
@yetki_gerekli('nexgen.recete.manage', 'can_manage')
def api_rf_kullanim_log():
    """RF renk kullanim kaydi olusturur (planlama / siparis izi).

    POST JSON:
        rf_renk_id : int (zorunlu)
        formul_id  : int (opsiyonel)
        cari_id    : int (opsiyonel)
        siparis_id : int (opsiyonel; nexgen_uretim_plan.id)
        aciklama   : str (opsiyonel)
    """
    data = request.get_json(silent=True) or {}
    rf_renk_id = data.get('rf_renk_id')
    if not rf_renk_id:
        return jsonify({"ok": False, "hata": "rf_renk_id zorunludur."}), 400

    try:
        rf_renk_id = int(rf_renk_id)
    except (TypeError, ValueError):
        return jsonify({"ok": False, "hata": "rf_renk_id gecersiz."}), 400

    formul_id = data.get('formul_id')
    cari_id = data.get('cari_id')
    siparis_id = data.get('siparis_id')
    aciklama = (data.get('aciklama') or '').strip() or None

    con = _db()
    try:
        rf = con.execute(
            "SELECT id, rf_kod, durum FROM nexgen_rf_renk WHERE id=? AND aktif=1",
            (rf_renk_id,)
        ).fetchone()
        if not rf:
            return jsonify({"ok": False, "hata": "RF renk kaydi bulunamadi."}), 404
        if rf['durum'] not in ('ONAYLI',):
            return jsonify({
                "ok": False,
                "hata": f"RF durumu kullanim icin uygun degil: {rf['durum']}"
            }), 400

        if formul_id is not None and formul_id != '':
            try:
                formul_id = int(formul_id)
            except (TypeError, ValueError):
                return jsonify({"ok": False, "hata": "formul_id gecersiz."}), 400
            if not con.execute(
                "SELECT id FROM nexgen_formul WHERE id=? AND aktif=1", (formul_id,)
            ).fetchone():
                return jsonify({"ok": False, "hata": f"Formul bulunamadi (id={formul_id})."}), 400
        else:
            formul_id = None

        if cari_id is not None and cari_id != '':
            try:
                cari_id = int(cari_id)
            except (TypeError, ValueError):
                return jsonify({"ok": False, "hata": "cari_id gecersiz."}), 400
            if not con.execute(
                "SELECT id FROM nexgen_cari WHERE id=? AND aktif=1", (cari_id,)
            ).fetchone():
                return jsonify({"ok": False, "hata": f"Cari bulunamadi (id={cari_id})."}), 400
        else:
            cari_id = None

        if siparis_id is not None and siparis_id != '':
            try:
                siparis_id = int(siparis_id)
            except (TypeError, ValueError):
                return jsonify({"ok": False, "hata": "siparis_id gecersiz."}), 400
            if not con.execute(
                "SELECT id FROM nexgen_uretim_plan WHERE id=?", (siparis_id,)
            ).fetchone():
                return jsonify({
                    "ok": False,
                    "hata": f"Uretim plani bulunamadi (siparis_id={siparis_id})."
                }), 400
        else:
            siparis_id = None

        kullanici_id = _kullanici_id()
        cols = _rf_kullanim_kolonlari(con)
        miktar_kg = 0.0
        if 'miktar_kg' in cols:
            try:
                miktar_kg = round(float(data.get('miktar_kg') or 0), 3)
            except (TypeError, ValueError):
                miktar_kg = 0.0

        if 'durum' in cols:
            con.execute("""
                INSERT INTO nexgen_rf_kullanim
                    (rf_renk_id, formul_id, cari_id, siparis_id, aciklama,
                     olusturan_id, aktif, durum, miktar_kg)
                VALUES (?, ?, ?, ?, ?, ?, 1, 'MANUEL', ?)
            """, (rf_renk_id, formul_id, cari_id, siparis_id, aciklama,
                  kullanici_id, miktar_kg))
        else:
            con.execute("""
                INSERT INTO nexgen_rf_kullanim
                    (rf_renk_id, formul_id, cari_id, siparis_id, aciklama, olusturan_id, aktif)
                VALUES (?, ?, ?, ?, ?, ?, 1)
            """, (rf_renk_id, formul_id, cari_id, siparis_id, aciklama, kullanici_id))
        kullanim_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({
        "ok":          True,
        "kullanim_id": kullanim_id,
        "rf_renk_id":  rf_renk_id,
        "rf_kod":      rf['rf_kod'],
        "formul_id":   formul_id,
        "cari_id":     cari_id,
        "siparis_id":  siparis_id,
    })


def _analitik_view_var(con):
    return con.execute(
        "SELECT name FROM sqlite_master WHERE type='view' AND name='v_nexgen_rf_kullanim_ozet'"
    ).fetchone() is not None


def _siparis_kontrol_view_var(con):
    return con.execute(
        "SELECT name FROM sqlite_master WHERE type='view' "
        "AND name='v_nexgen_siparis_uretim_kontrol'"
    ).fetchone() is not None


def _rf_kullanim_tablosu_var(con):
    return con.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='nexgen_rf_kullanim'"
    ).fetchone() is not None


def _rf_kullanim_uretilen_kg(con, siparis_id=None, batch_kodu=None):
    """Tek kaynak: nexgen_rf_kullanim SUM(miktar_kg)."""
    if not _rf_kullanim_tablosu_var(con):
        return 0.0
    if batch_kodu:
        row = con.execute("""
            SELECT ROUND(COALESCE(SUM(miktar_kg), 0), 3) AS kg
            FROM nexgen_rf_kullanim
            WHERE aktif = 1 AND tablet_session_id = ?
        """, (batch_kodu,)).fetchone()
        kg = float(row['kg'] or 0) if row else 0.0
        if kg > 0:
            return kg
    if siparis_id:
        row = con.execute("""
            SELECT ROUND(COALESCE(SUM(miktar_kg), 0), 3) AS kg
            FROM nexgen_rf_kullanim
            WHERE aktif = 1 AND siparis_id = ?
        """, (siparis_id,)).fetchone()
        return float(row['kg'] or 0) if row else 0.0
    return 0.0


def _uretim_kontrol_durumu(siparis_toplam_kg, uretilen_kg):
    """Validation: asim / devam / tam."""
    st = float(siparis_toplam_kg or 0)
    ur = float(uretilen_kg or 0)
    if st <= 0:
        return {'kontrol_durum': 'BELIRSIZ', 'uyari': None}
    if ur > st + 0.001:
        return {
            'kontrol_durum': 'ASIM',
            'uyari': f'Uretilen ({ur:g} KG) siparis toplamini ({st:g} KG) asiyor.',
        }
    if ur < st - 0.001:
        return {
            'kontrol_durum': 'DEVAM',
            'uyari': None,
            'mesaj': 'Devam eden uretim',
        }
    return {'kontrol_durum': 'TAM', 'uyari': None}


def _siparis_kontrol_get(con, siparis_id):
    """Siparis KG kontrol ozeti (VIEW veya fallback)."""
    if _siparis_kontrol_view_var(con):
        row = con.execute(
            "SELECT * FROM v_nexgen_siparis_uretim_kontrol WHERE siparis_id=?",
            (siparis_id,)
        ).fetchone()
        if row:
            d = dict(row)
            d.update(_uretim_kontrol_durumu(d['siparis_toplam_kg'], d['uretilen_kg']))
            return d
    plan = con.execute(
        "SELECT id, planlanan_kg FROM nexgen_uretim_plan WHERE id=?",
        (siparis_id,)
    ).fetchone()
    if not plan:
        return None
    dagitim = 0.0
    if _parca_tablosu_var(con):
        r = con.execute("""
            SELECT ROUND(COALESCE(SUM(p.hedef_kg), 0), 3) AS kg
            FROM nexgen_uretim_parca p
            JOIN nexgen_uretim_batch b ON b.batch_kodu = p.batch_kodu
            WHERE b.plan_id = ?
        """, (siparis_id,)).fetchone()
        dagitim = float(r['kg'] or 0) if r else 0.0
    uretilen = _rf_kullanim_uretilen_kg(con, siparis_id=siparis_id)
    siparis_toplam = float(plan['planlanan_kg'])
    d = {
        'siparis_id': siparis_id,
        'siparis_toplam_kg': round(siparis_toplam, 3),
        'planlanan_kg': round(dagitim, 3),
        'uretilen_kg': round(uretilen, 3),
        'fark_kg': round(siparis_toplam - uretilen, 3),
    }
    d.update(_uretim_kontrol_durumu(siparis_toplam, uretilen))
    return d


def _batch_uretim_kontrol(con, batch_kodu):
    """Batch/tablet ekrani icin siparis master + rf uretilen."""
    batch = con.execute("""
        SELECT nb.batch_kodu, nb.planlanan_kg, nb.plan_id
        FROM nexgen_uretim_batch nb
        WHERE nb.batch_kodu = ?
    """, (batch_kodu,)).fetchone()
    if not batch:
        return None
    plan_id = batch['plan_id']
    siparis_toplam = float(batch['planlanan_kg'])
    if plan_id:
        sk = _siparis_kontrol_get(con, plan_id)
        if sk:
            return {
                **sk,
                'batch_kodu': batch_kodu,
                'batch_planlanan_kg': round(float(batch['planlanan_kg']), 3),
                'kalan_kg': max(0.0, float(sk['fark_kg'])),
            }
    uretilen = _rf_kullanim_uretilen_kg(con, batch_kodu=batch_kodu)
    d = {
        'siparis_id': plan_id,
        'batch_kodu': batch_kodu,
        'siparis_toplam_kg': round(siparis_toplam, 3),
        'batch_planlanan_kg': round(float(batch['planlanan_kg']), 3),
        'planlanan_kg': 0.0,
        'uretilen_kg': round(uretilen, 3),
        'fark_kg': round(siparis_toplam - uretilen, 3),
        'kalan_kg': max(0.0, siparis_toplam - uretilen),
    }
    d.update(_uretim_kontrol_durumu(siparis_toplam, uretilen))
    return d


def _batch_kg_rf_enrich(con, b_dict, liste_modu=False):
    """KG gosterim alanlarini rf_kullanim tek kaynagindan doldurur.

    liste_modu=True: batch kartlari (L/S) — uretilen = batch rf SUM, hedef = batch planlanan_kg
    liste_modu=False: siparis ozeti — plan master + siparis rf SUM
    """
    k = _batch_uretim_kontrol(con, b_dict.get('batch_kodu'))
    if not k:
        return b_dict
    b_dict['siparis_toplam_kg'] = k['siparis_toplam_kg']
    b_dict['kontrol_durum'] = k.get('kontrol_durum')
    b_dict['kontrol_uyari'] = k.get('uyari')
    b_dict['kontrol_mesaj'] = k.get('mesaj')
    if liste_modu:
        batch_ur = 0.0
        if _rf_kullanim_tablosu_var(con) and b_dict.get('batch_kodu'):
            row = con.execute("""
                SELECT ROUND(COALESCE(SUM(miktar_kg), 0), 3) AS kg
                FROM nexgen_rf_kullanim
                WHERE aktif = 1 AND tablet_session_id = ?
            """, (b_dict['batch_kodu'],)).fetchone()
            batch_ur = float(row['kg'] or 0) if row else 0.0
        bpl = float(b_dict.get('planlanan_kg') or 0)
        b_dict['uretilen_kg'] = round(batch_ur, 3)
        b_dict['kalan_kg'] = round(max(0.0, bpl - batch_ur), 3)
    else:
        b_dict['uretilen_kg'] = k['uretilen_kg']
        b_dict['kalan_kg'] = k['kalan_kg']
    return b_dict


def _durum_dagilimi_satir(row):
    """VIEW veya aggregation satirindan durum dagilimi dict."""
    if not row:
        return {'MANUEL': 0, 'URETIM': 0, 'TAMAMLANDI': 0}
    keys = row.keys() if hasattr(row, 'keys') else []
    if 'adet_manuel' in keys:
        return {
            'MANUEL': int(row['adet_manuel'] or 0),
            'URETIM': int(row['adet_uretim'] or 0),
            'TAMAMLANDI': int(row['adet_tamamlandi'] or 0),
        }
    return {'MANUEL': 0, 'URETIM': 0, 'TAMAMLANDI': 0}


@nexgen_bp.route('/api/nexgen/siparis-kontrol/<int:siparis_id>', methods=['GET'])
@yetki_gerekli('nexgen.recete.view', 'can_view')
def api_nexgen_siparis_kontrol(siparis_id):
    """Siparis KG kontrol ozeti (VIEW v_nexgen_siparis_uretim_kontrol)."""
    con = _db()
    try:
        row = _siparis_kontrol_get(con, siparis_id)
        if not row:
            return jsonify({'ok': False, 'hata': 'Siparis bulunamadi.'}), 404
        return jsonify({'ok': True, **row})
    finally:
        con.close()


@nexgen_bp.route('/api/nexgen/uretim-ozet', methods=['GET'])
@yetki_gerekli('nexgen.recete.view', 'can_view')
def api_nexgen_uretim_ozet():
    """RF kullanim + tablet uretim ozet analitigi (JSON)."""
    con = _db()
    try:
        if not con.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='nexgen_rf_kullanim'"
        ).fetchone():
            return jsonify({
                'ok': True,
                'toplam_kg': 0.0,
                'aktif_uretim': 0,
                'tamamlanan_batch': 0,
                'en_cok_kullanilan_rf': None,
                'view_hazir': False,
            })

        ozet = con.execute("""
            SELECT ROUND(COALESCE(SUM(miktar_kg), 0), 3) AS toplam_kg,
                   SUM(CASE WHEN durum = 'URETIM' THEN 1 ELSE 0 END) AS aktif_uretim
            FROM nexgen_rf_kullanim WHERE aktif = 1
        """).fetchone()

        tamamlanan = con.execute("""
            SELECT COUNT(DISTINCT tablet_session_id) AS c
            FROM nexgen_rf_kullanim
            WHERE aktif = 1 AND durum = 'TAMAMLANDI'
              AND tablet_session_id IS NOT NULL AND tablet_session_id != ''
        """).fetchone()
        tamamlanan_batch = int(tamamlanan['c'] or 0) if tamamlanan else 0

        if tamamlanan_batch == 0:
            batch_bit = con.execute(
                "SELECT COUNT(*) AS c FROM nexgen_uretim_batch WHERE durum = 'BITTI'"
            ).fetchone()
            tamamlanan_batch = int(batch_bit['c'] or 0) if batch_bit else 0

        en_cok = con.execute("""
            SELECT k.rf_renk_id,
                   COUNT(*) AS kullanim_sayisi,
                   ROUND(SUM(COALESCE(k.miktar_kg, 0)), 3) AS toplam_kg,
                   rf.rf_kod, rf.ad AS rf_ad
            FROM nexgen_rf_kullanim k
            LEFT JOIN nexgen_rf_renk rf ON rf.id = k.rf_renk_id
            WHERE k.aktif = 1
            GROUP BY k.rf_renk_id
            ORDER BY kullanim_sayisi DESC, toplam_kg DESC
            LIMIT 1
        """).fetchone()

        en_cok_rf = None
        if en_cok and en_cok['rf_renk_id']:
            en_cok_rf = {
                'rf_renk_id': en_cok['rf_renk_id'],
                'rf_kod': en_cok['rf_kod'],
                'rf_ad': en_cok['rf_ad'],
                'kullanim_sayisi': int(en_cok['kullanim_sayisi'] or 0),
                'toplam_kg': float(en_cok['toplam_kg'] or 0),
            }

        return jsonify({
            'ok': True,
            'toplam_kg': float(ozet['toplam_kg'] or 0) if ozet else 0.0,
            'aktif_uretim': int(ozet['aktif_uretim'] or 0) if ozet else 0,
            'tamamlanan_batch': tamamlanan_batch,
            'en_cok_kullanilan_rf': en_cok_rf,
            'view_hazir': _analitik_view_var(con),
        })
    finally:
        con.close()


@nexgen_bp.route('/api/nexgen/cari-rf-analiz/<int:cari_id>', methods=['GET'])
@yetki_gerekli('nexgen.recete.view', 'can_view')
def api_nexgen_cari_rf_analiz(cari_id):
    """Cari bazli RF kullanim analizi."""
    con = _db()
    try:
        cari = con.execute(
            "SELECT id, cari_kod, unvan FROM nexgen_cari WHERE id=? AND aktif=1",
            (cari_id,)
        ).fetchone()
        if not cari:
            return jsonify({'ok': False, 'hata': 'Cari bulunamadi.'}), 404

        if _analitik_view_var(con):
            rows = con.execute("""
                SELECT v.*, rf.rf_kod, rf.ad AS rf_ad
                FROM v_nexgen_rf_kullanim_ozet v
                LEFT JOIN nexgen_rf_renk rf ON rf.id = v.rf_renk_id
                WHERE v.cari_id = ?
                ORDER BY v.toplam_uretim_kg DESC
            """, (cari_id,)).fetchall()
        else:
            rows = con.execute("""
                SELECT k.rf_renk_id, k.cari_id, k.siparis_id, k.formul_id,
                       ROUND(SUM(COALESCE(k.miktar_kg, 0)), 3) AS toplam_uretim_kg,
                       COUNT(*) AS kayit_sayisi,
                       SUM(CASE WHEN COALESCE(k.durum,'MANUEL')='MANUEL' THEN 1 ELSE 0 END) AS adet_manuel,
                       SUM(CASE WHEN k.durum='URETIM' THEN 1 ELSE 0 END) AS adet_uretim,
                       SUM(CASE WHEN k.durum='TAMAMLANDI' THEN 1 ELSE 0 END) AS adet_tamamlandi,
                       rf.rf_kod, rf.ad AS rf_ad
                FROM nexgen_rf_kullanim k
                LEFT JOIN nexgen_rf_renk rf ON rf.id = k.rf_renk_id
                WHERE k.aktif = 1 AND k.cari_id = ?
                GROUP BY k.rf_renk_id, k.cari_id, k.siparis_id, k.formul_id
                ORDER BY toplam_uretim_kg DESC
            """, (cari_id,)).fetchall()

        rf_map = {}
        siparis_set = set()
        toplam_kg = 0.0
        for r in rows:
            rd = dict(r)
            siparis_set.add(rd.get('siparis_id'))
            kg = float(rd.get('toplam_uretim_kg') or 0)
            toplam_kg += kg
            rid = rd['rf_renk_id']
            if rid not in rf_map:
                rf_map[rid] = {
                    'rf_renk_id': rid,
                    'rf_kod': rd.get('rf_kod'),
                    'rf_ad': rd.get('rf_ad'),
                    'toplam_kg': 0.0,
                    'kayit_sayisi': 0,
                    'durum_dagilimi': {'MANUEL': 0, 'URETIM': 0, 'TAMAMLANDI': 0},
                }
            rf_map[rid]['toplam_kg'] = round(rf_map[rid]['toplam_kg'] + kg, 3)
            rf_map[rid]['kayit_sayisi'] += int(rd.get('kayit_sayisi') or 1)
            dd = _durum_dagilimi_satir(rd)
            for k, v in dd.items():
                rf_map[rid]['durum_dagilimi'][k] += v

        rf_listesi = sorted(rf_map.values(), key=lambda x: x['toplam_kg'], reverse=True)

        return jsonify({
            'ok': True,
            'cari_id': cari_id,
            'cari_kod': cari['cari_kod'],
            'cari_unvan': cari['unvan'],
            'toplam_kg': round(toplam_kg, 3),
            'siparis_sayisi': len([s for s in siparis_set if s is not None]),
            'rf_listesi': rf_listesi,
        })
    finally:
        con.close()


@nexgen_bp.route('/api/nexgen/formul-rf-analiz/<int:formul_id>', methods=['GET'])
@yetki_gerekli('nexgen.recete.view', 'can_view')
def api_nexgen_formul_rf_analiz(formul_id):
    """Formul bazli RF kullanim analizi."""
    con = _db()
    try:
        formul = con.execute(
            "SELECT id, kod, ad FROM nexgen_formul WHERE id=? AND aktif=1",
            (formul_id,)
        ).fetchone()
        if not formul:
            return jsonify({'ok': False, 'hata': 'Formul bulunamadi.'}), 404

        if _analitik_view_var(con):
            rows = con.execute("""
                SELECT v.*, rf.rf_kod, rf.ad AS rf_ad
                FROM v_nexgen_rf_kullanim_ozet v
                LEFT JOIN nexgen_rf_renk rf ON rf.id = v.rf_renk_id
                WHERE v.formul_id = ?
                ORDER BY v.toplam_uretim_kg DESC
            """, (formul_id,)).fetchall()
        else:
            rows = con.execute("""
                SELECT k.rf_renk_id, k.cari_id, k.siparis_id, k.formul_id,
                       ROUND(SUM(COALESCE(k.miktar_kg, 0)), 3) AS toplam_uretim_kg,
                       COUNT(*) AS kayit_sayisi,
                       SUM(CASE WHEN COALESCE(k.durum,'MANUEL')='MANUEL' THEN 1 ELSE 0 END) AS adet_manuel,
                       SUM(CASE WHEN k.durum='URETIM' THEN 1 ELSE 0 END) AS adet_uretim,
                       SUM(CASE WHEN k.durum='TAMAMLANDI' THEN 1 ELSE 0 END) AS adet_tamamlandi,
                       rf.rf_kod, rf.ad AS rf_ad
                FROM nexgen_rf_kullanim k
                LEFT JOIN nexgen_rf_renk rf ON rf.id = k.rf_renk_id
                WHERE k.aktif = 1 AND k.formul_id = ?
                GROUP BY k.rf_renk_id, k.cari_id, k.siparis_id, k.formul_id
                ORDER BY toplam_uretim_kg DESC
            """, (formul_id,)).fetchall()

        rf_map = {}
        cari_set = set()
        toplam_kg = 0.0
        for r in rows:
            rd = dict(r)
            cari_set.add(rd.get('cari_id'))
            kg = float(rd.get('toplam_uretim_kg') or 0)
            toplam_kg += kg
            rid = rd['rf_renk_id']
            if rid not in rf_map:
                rf_map[rid] = {
                    'rf_renk_id': rid,
                    'rf_kod': rd.get('rf_kod'),
                    'rf_ad': rd.get('rf_ad'),
                    'toplam_kg': 0.0,
                    'kayit_sayisi': 0,
                    'cari_sayisi': set(),
                    'durum_dagilimi': {'MANUEL': 0, 'URETIM': 0, 'TAMAMLANDI': 0},
                }
            rf_map[rid]['toplam_kg'] = round(rf_map[rid]['toplam_kg'] + kg, 3)
            rf_map[rid]['kayit_sayisi'] += int(rd.get('kayit_sayisi') or 1)
            if rd.get('cari_id'):
                rf_map[rid]['cari_sayisi'].add(rd['cari_id'])
            dd = _durum_dagilimi_satir(rd)
            for k, v in dd.items():
                rf_map[rid]['durum_dagilimi'][k] += v

        rf_listesi = []
        for item in sorted(rf_map.values(), key=lambda x: x['toplam_kg'], reverse=True):
            item['cari_sayisi'] = len(item.pop('cari_sayisi'))
            rf_listesi.append(item)

        return jsonify({
            'ok': True,
            'formul_id': formul_id,
            'formul_kod': formul['kod'],
            'formul_ad': formul['ad'],
            'toplam_kg': round(toplam_kg, 3),
            'cari_sayisi': len([c for c in cari_set if c is not None]),
            'rf_listesi': rf_listesi,
        })
    finally:
        con.close()


@nexgen_bp.route('/api/arge/gecmis/<int:uv_id>', methods=['GET'])
@yetki_gerekli('nexgen.recete.view', 'can_view')
def api_arge_gecmis(uv_id):
    """Belirli bir üretim varyantının ARGE test geçmişini döner."""
    con = _db()
    try:
        rows = con.execute("""
            SELECT t.id, t.test_no, t.test_tipi, t.makina,
                   t.test_batch_kg, t.kaynak_batch_kg,
                   t.yeni_renk_adi, t.durum,
                   t.olusturma_tarihi, t.onay_tarihi,
                   ku.KullaniciAdi AS olusturan_ad
            FROM nexgen_arge_test t
            LEFT JOIN sistem_kullanici ku ON ku.Id = t.olusturan_id
            WHERE t.kaynak_uretim_varyant_id = ? AND t.aktif = 1
            ORDER BY t.id DESC
        """, (uv_id,)).fetchall()
    finally:
        con.close()
    return jsonify([dict(r) for r in rows])

# ═════════════════════════════════════════════════════════════
# FAZ-4C-4: ARGE TESTTEN ÜRETİM REÇETESİNE AKTARIM
# ─────────────────────────────────────────────────────────────
# KURAL: Sadece ONAYLANDI durumdaki testler aktarılabilir.
#        nexgen_stok_hareket YAZILMAZ.
#        Kaynak renk_varyant / uretim_varyant kalemleri değişmez.
#        Geri ölçekleme: uretim_kg = test_kg / test_batch * kaynak_batch
# ─────────────────────────────────────────────────────────────

@nexgen_bp.route('/api/arge/hedef-formuller', methods=['GET'])
@yetki_gerekli('nexgen.recete.view', 'can_view')
def api_arge_hedef_formuller():
    """Modal için: aktif formüller + her formülün mevcut renk varyantları."""
    con = _db()
    try:
        formuller = con.execute(
            "SELECT id, kod, ad FROM nexgen_formul WHERE aktif=1 ORDER BY kod"
        ).fetchall()
        renkler = con.execute("""
            SELECT rv.id, rv.formul_id, rv.ad, rv.kod AS rv_kod,
                   GROUP_CONCAT(uv.boyut) AS boyutlar
            FROM nexgen_renk_varyant rv
            LEFT JOIN nexgen_uretim_varyant uv ON uv.renk_varyant_id=rv.id AND uv.aktif=1
            WHERE rv.aktif=1
            GROUP BY rv.id
            ORDER BY rv.formul_id, rv.id
        """).fetchall()
    finally:
        con.close()

    return jsonify({
        "formuller": [dict(f) for f in formuller],
        "renkler":   [dict(r) for r in renkler],
    })


@nexgen_bp.route('/api/arge/uretim-recetesi-olustur', methods=['POST'])
@yetki_gerekli('nexgen.recete.manage', 'can_manage')
def api_arge_uretim_recetesi_olustur():
    """Onaylı ARGE testinden üretim reçetesi oluşturur.

    Adımlar:
    1) Test ONAYLANDI mı kontrol et.
    2) Test kalemlerini geri ölçekle (test_kg / test_batch * kaynak_batch).
    3) Renk varyantı: varsa mevcut seç, yoksa yeni oluştur.
    4) Üretim varyantı oluştur (recete_durum=ONAYLI veya URETIME_ACIK).
    5) Reçete kalemlerini nexgen_recete_kalem'e yaz.
    6) nexgen_arge_test.olusan_uretim_varyant_id güncelle.

    KURAL: nexgen_stok_hareket YAZILMAZ.
           Kaynak reçete kalemleri DEĞİŞMEZ.

    POST JSON:
        test_id         : int
        hedef_formul_id : int
        yeni_renk_adi   : str   (yeni renk adı veya mevcut renk adı)
        mevcut_renk_id  : int   opsiyonel — varsa bu renke ekle
        boyut           : SMALL | LARGE | STANDART
        recete_durum    : ONAYLI | URETIME_ACIK  (default: ONAYLI)
        notlar          : str opsiyonel
    """
    data = request.get_json(silent=True) or {}

    test_id         = data.get('test_id')
    hedef_formul_id = data.get('hedef_formul_id')
    yeni_renk_adi   = (data.get('yeni_renk_adi') or '').strip()
    mevcut_renk_id  = data.get('mevcut_renk_id')   # opsiyonel
    boyut           = (data.get('boyut') or 'LARGE').strip().upper()
    recete_durum    = (data.get('recete_durum') or 'ONAYLI').strip().upper()
    notlar          = (data.get('notlar') or '').strip() or None

    GECERLI_BOYUT  = {'SMALL', 'LARGE', 'STANDART'}
    GECERLI_DURUM  = {'ONAYLI', 'URETIME_ACIK'}

    if not test_id:
        return jsonify({"ok": False, "hata": "test_id zorunludur."}), 400
    if not hedef_formul_id:
        return jsonify({"ok": False, "hata": "hedef_formul_id zorunludur."}), 400
    if not yeni_renk_adi and not mevcut_renk_id:
        return jsonify({"ok": False, "hata": "yeni_renk_adi veya mevcut_renk_id gerekli."}), 400
    if boyut not in GECERLI_BOYUT:
        return jsonify({"ok": False, "hata": f"boyut: {', '.join(GECERLI_BOYUT)}"}), 400
    if recete_durum not in GECERLI_DURUM:
        return jsonify({"ok": False, "hata": f"recete_durum: {', '.join(GECERLI_DURUM)}"}), 400

    kullanici_id = _kullanici_id()
    con = _db()
    try:
        # ── 1) Test kaydı ve ONAYLANDI kontrolü ──────────────
        test = con.execute(
            "SELECT * FROM nexgen_arge_test WHERE id=? AND aktif=1",
            (test_id,)
        ).fetchone()
        if not test:
            return jsonify({"ok": False, "hata": "Test kaydı bulunamadı."}), 404
        if test['durum'] != 'ONAYLANDI':
            return jsonify({
                "ok":   False,
                "hata": f"Sadece ONAYLANDI testler aktarılabilir. Mevcut durum: {test['durum']}"
            }), 400
        if test['olusan_uretim_varyant_id']:
            return jsonify({
                "ok":   False,
                "hata": f"Bu test zaten aktarıldı. Oluşan varyant id={test['olusan_uretim_varyant_id']}"
            }), 409

        rf_info = _arge_rf_mevcut_bul(con, test_id, test)
        if not rf_info:
            return jsonify({
                "ok": False,
                "hata": "Önce RF oluşturulmalı."
            }), 400

        # ── 2) Hedef formül var mı? ───────────────────────────
        formul = con.execute(
            "SELECT id, kod, ad FROM nexgen_formul WHERE id=? AND aktif=1",
            (hedef_formul_id,)
        ).fetchone()
        if not formul:
            return jsonify({"ok": False, "hata": "Hedef formül bulunamadı."}), 404

        # ── 3) Test kalemleri (BOYA hariç — RF renk formülünde) ──
        test_kalemler = con.execute("""
            SELECT tk.stok_kart_id, tk.sira, tk.orjinal_miktar_kg, tk.test_miktar_kg,
                   tk.aciklama, sk.kategori
            FROM nexgen_arge_test_kalem tk
            JOIN nexgen_stok_kart sk ON sk.id = tk.stok_kart_id
            WHERE tk.test_id = ?
            ORDER BY tk.sira
        """, (test_id,)).fetchall()
        if not test_kalemler:
            return jsonify({"ok": False, "hata": "Test kalemlerinde veri yok."}), 400

        test_batch_kg   = float(test['test_batch_kg'])
        kaynak_batch_kg = float(test['kaynak_batch_kg'])
        if test_batch_kg <= 0 or kaynak_batch_kg <= 0:
            return jsonify({"ok": False, "hata": "test_batch_kg veya kaynak_batch_kg sıfır."}), 400

        geri_carpan = kaynak_batch_kg / test_batch_kg

        uretim_kalemleri = []
        atlanan_sifir = 0
        for k in test_kalemler:
            if (k['kategori'] or '').upper() == 'BOYA':
                continue
            uretim_kg = round(float(k['test_miktar_kg']) * geri_carpan, 3)
            if uretim_kg <= 0:
                atlanan_sifir += 1
                continue
            uretim_kalemleri.append({
                'stok_kart_id': k['stok_kart_id'],
                'sira':         k['sira'],
                'uretim_kg':    uretim_kg,
                'aciklama':     k['aciklama'],
            })

        if not uretim_kalemleri:
            return jsonify({
                "ok": False,
                "hata": "Taban kalem bulunamadi (BOYA haric) veya tum miktarlar sifir."
            }), 400

        toplam_uretim_kg = round(sum(x['uretim_kg'] for x in uretim_kalemleri), 3)

        # ── 4) Renk varyantı: mevcut veya yeni ───────────────
        renk_varyant_id = None
        renk_yeni_mi    = False

        if mevcut_renk_id:
            rv = con.execute(
                "SELECT id FROM nexgen_renk_varyant WHERE id=? AND formul_id=? AND aktif=1",
                (mevcut_renk_id, hedef_formul_id)
            ).fetchone()
            if not rv:
                return jsonify({"ok": False, "hata": "Seçilen renk varyantı bu formüle ait değil."}), 400
            renk_varyant_id = rv['id']
        else:
            # Aynı ad + formül kombinasyonu var mı?
            mevcut = con.execute(
                "SELECT id FROM nexgen_renk_varyant WHERE formul_id=? AND ad=? AND aktif=1",
                (hedef_formul_id, yeni_renk_adi)
            ).fetchone()
            if mevcut:
                renk_varyant_id = mevcut['id']
            else:
                # Yeni renk varyantı oluştur
                rv_kod = f"RV-{hedef_formul_id}-{yeni_renk_adi[:8].upper().replace(' ','')}"
                con.execute("""
                    INSERT INTO nexgen_renk_varyant
                        (formul_id, kod, ad, renk, notlar, aktif, olusturma_tarihi)
                    VALUES (?, ?, ?, ?, ?, 1, datetime('now'))
                """, (hedef_formul_id, rv_kod, yeni_renk_adi,
                      yeni_renk_adi, notlar))
                renk_varyant_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
                renk_yeni_mi = True

        # ── 5) Üretim varyantı: aynı renk+boyut var mı? ──────
        mevcut_uv = con.execute(
            "SELECT id FROM nexgen_uretim_varyant WHERE renk_varyant_id=? AND boyut=? AND aktif=1",
            (renk_varyant_id, boyut)
        ).fetchone()
        if mevcut_uv:
            # Zaten var — bu varyantın kalemlerine dokunma, hata dön
            return jsonify({
                "ok":   False,
                "hata": f"Bu renk / boyut kombinasyonu zaten mevcut (uv_id={mevcut_uv['id']}). "
                         "Mevcut reçeteyi klonlayın veya farklı boyut seçin."
            }), 409

        uv_ad = f"{formul['ad']} {yeni_renk_adi or 'Yeni Renk'}"
        con.execute("""
            INSERT INTO nexgen_uretim_varyant
                (renk_varyant_id, boyut, ad, onay_durumu,
                 kaynak_varyant_id, notlar, aktif,
                 olusturma_tarihi, recete_durum)
            VALUES (?, ?, ?, 'TASLAK', ?, ?, 1, datetime('now'), ?)
        """, (renk_varyant_id, boyut, uv_ad,
              test['kaynak_uretim_varyant_id'],
              f"ARGE test {test['test_no']}'den aktarıldı. {notlar or ''}".strip(),
              recete_durum))
        yeni_uv_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]

        # ── 6) Reçete kalemleri yaz ───────────────────────────
        for k in uretim_kalemleri:
            con.execute("""
                INSERT INTO nexgen_recete_kalem
                    (uretim_varyant_id, stok_kart_id, sira,
                     miktar_kg, aciklama, aktif, olusturma_tarihi)
                VALUES (?, ?, ?, ?, ?, 1, datetime('now'))
            """, (yeni_uv_id, k['stok_kart_id'], k['sira'],
                  k['uretim_kg'], k['aciklama']))

        # ── 7) ARGE test bağlantı güncelle ───────────────────
        con.execute("""
            UPDATE nexgen_arge_test
            SET olusan_uretim_varyant_id=?,
                olusan_renk_varyant_id=?,
                onaylayan_id=?,
                onay_tarihi=datetime('now')
            WHERE id=?
        """, (yeni_uv_id, renk_varyant_id, kullanici_id, test_id))

        con.commit()

    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({
        "ok":                True,
        "yeni_uv_id":        yeni_uv_id,
        "yeni_renk_varyant_id": renk_varyant_id,
        "renk_yeni_mi":      renk_yeni_mi,
        "boyut":             boyut,
        "recete_durum":      recete_durum,
        "toplam_uretim_kg":  toplam_uretim_kg,
        "kalem_sayisi":      len(uretim_kalemleri),
        "geri_carpan":       round(geri_carpan, 6),
        "rf_renk_id":        rf_info['id'],
        "rf_kod":            rf_info['rf_kod'],
        "rf_ad":             rf_info['ad'],
        "atlanan_sifir_kalem": atlanan_sifir,
        "mesaj": (
            f"{'Yeni renk oluşturuldu: ' + yeni_renk_adi + '. ' if renk_yeni_mi else ''}"
            f"Üretim reçetesi oluşturuldu (BOYA → {rf_info['rf_kod']}). "
            f"Toplam: {toplam_uretim_kg} KG, {len(uretim_kalemleri)} taban kalem."
        ),
    })

# ─────────────────────────────────────────────────────────────
# KURAL: Bu bölümde nexgen_stok_hareket INSERT YAPILMAZ.
# ═════════════════════════════════════════════════════════════

def _siparis_no_uret(con):
    """SP-YYYY-NNNN formatında benzersiz sipariş numarası üretir."""
    from datetime import datetime
    yil = datetime.now().strftime('%Y')
    row = con.execute("""
        SELECT MAX(CAST(SUBSTR(siparis_no, -4) AS INTEGER)) AS son
        FROM nexgen_satin_siparis
        WHERE siparis_no LIKE ?
    """, (f'SP-{yil}-%',)).fetchone()
    son = row['son'] if row and row['son'] else 0
    return f'SP-{yil}-{son + 1:04d}'


def _tedarikci_veya_404(con, tedarikci_id):
    t = con.execute(
        "SELECT * FROM nexgen_tedarikci WHERE id=? AND aktif=1", (tedarikci_id,)
    ).fetchone()
    if not t:
        abort(404)
    return t


# ─────────────────────────────────────────────────────────────
# Satın Alma Ana Sayfa
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/satinalma')
@yetki_gerekli('nexgen.satinalma.view', 'can_view')
def satinalma_index():
    con = _db()
    try:
        durum_filtre = request.args.get('durum', '')
        tedarikci_filtre = request.args.get('tedarikci_id', '')

        sorgu = """
            SELECT s.id, s.siparis_no, s.siparis_tarihi, s.beklenen_teslim,
                   s.siparis_miktari_kg, s.para_birimi, s.durum, s.onay_durumu,
                   s.birim_fiyat, s.birim_fiyat_try, s.toplam_tutar_try,
                   s.vade_tarihi, s.aciklama,
                   t.ad AS tedarikci_ad, t.id AS tedarikci_id,
                   k.kod AS stok_kod, k.ad AS stok_ad,
                   COALESCE(
                       (SELECT SUM(mk.miktar_kg)
                        FROM nexgen_mal_kabul mk
                        WHERE mk.satin_siparis_id = s.id), 0
                   ) AS gelen_kg
            FROM nexgen_satin_siparis s
            JOIN nexgen_tedarikci t ON t.id = s.tedarikci_id
            JOIN nexgen_stok_kart k ON k.id = s.stok_kart_id
            WHERE 1=1
        """
        params = []
        if durum_filtre:
            sorgu += " AND s.durum = ?"
            params.append(durum_filtre)
        if tedarikci_filtre:
            sorgu += " AND s.tedarikci_id = ?"
            params.append(tedarikci_filtre)
        sorgu += " ORDER BY s.id DESC"

        siparisler_raw = con.execute(sorgu, params).fetchall()
        tedarikciler   = con.execute(
            "SELECT id, ad FROM nexgen_tedarikci WHERE aktif=1 ORDER BY ad"
        ).fetchall()

    finally:
        con.close()

    can_manage     = yetki_var('nexgen.satinalma.manage', 'can_create')
    can_approve    = yetki_var('nexgen.satinalma.approve', 'can_approve')
    can_fiyat      = yetki_var('nexgen.satinalma.fiyat', 'can_view')
    # Tedarikçi yönetimi: sadece Yönetim rolü — Satın Alma sadece görüntüler
    can_ted_manage = yetki_var('nexgen.tedarikci.manage', 'can_create')
    can_fiyat_view = yetki_var('nexgen.fiyat.view', 'can_view')

    siparisler = []
    for s in siparisler_raw:
        row = dict(s)
        siparis_kg = row.get('siparis_miktari_kg') or 0
        gelen_kg   = row.get('gelen_kg') or 0
        row['gelen_kg']  = round(gelen_kg, 3)
        row['kalan_kg']  = round(siparis_kg - gelen_kg, 3)
        # Teslim durumu DB'deki durum alanından okunur (depo mal kabul sonrası güncellendi)
        # Görsel etiket için ayrıca hesapla
        if gelen_kg <= 0:
            row['teslim_durum'] = 'BEKLIYOR'
        elif gelen_kg >= siparis_kg:
            row['teslim_durum'] = 'TAMAMLANDI'
        else:
            row['teslim_durum'] = 'KISMI_TESLIM'
        siparisler.append(row)

    return render_template(
        'nexgen/satinalma_index.html',
        active='nexgen',
        siparisler=siparisler,
        tedarikciler=[dict(t) for t in tedarikciler],
        durum_filtre=durum_filtre,
        tedarikci_filtre=tedarikci_filtre,
        can_manage=can_manage,
        can_approve=can_approve,
        can_fiyat=can_fiyat,
        can_ted_manage=can_ted_manage,
        can_fiyat_view=can_fiyat_view,
    )


# ─────────────────────────────────────────────────────────────
# Sipariş Detay
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/satinalma/siparis/<int:siparis_id>')
@yetki_gerekli('nexgen.satinalma.view', 'can_view')
def satinalma_siparis_detay(siparis_id):
    con = _db()
    try:
        siparis = con.execute("""
            SELECT s.*,
                   t.ad AS tedarikci_ad, t.kod AS tedarikci_kod,
                   t.ulke, t.iletisim_email,
                   k.kod AS stok_kod, k.ad AS stok_ad, k.kategori AS stok_kategori,
                   oc.KullaniciAdi AS olusturan_ad,
                   ap.KullaniciAdi AS onaylayan_ad
            FROM nexgen_satin_siparis s
            JOIN nexgen_tedarikci t ON t.id = s.tedarikci_id
            JOIN nexgen_stok_kart k ON k.id = s.stok_kart_id
            LEFT JOIN sistem_kullanici oc ON oc.Id = s.olusturan_id
            LEFT JOIN sistem_kullanici ap ON ap.Id = s.onaylayan_id
            WHERE s.id = ?
        """, (siparis_id,)).fetchone()
        if not siparis:
            abort(404)
    finally:
        con.close()

    can_approve = yetki_var('nexgen.satinalma.approve', 'can_approve')
    can_manage  = yetki_var('nexgen.satinalma.manage', 'can_update')
    can_fiyat   = yetki_var('nexgen.satinalma.fiyat', 'can_view')

    return render_template(
        'nexgen/satinalma_siparis_detay.html',
        active='nexgen',
        siparis=dict(siparis),
        can_approve=can_approve,
        can_manage=can_manage,
        can_fiyat=can_fiyat,
    )


# ─────────────────────────────────────────────────────────────
# Tedarikçi Listesi
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/satinalma/tedarikci')
@yetki_gerekli('nexgen.tedarikci.view', 'can_view')
def satinalma_tedarikci():
    con = _db()
    try:
        tedarikciler = con.execute("""
            SELECT t.*,
                   COUNT(s.id) AS siparis_sayisi
            FROM nexgen_tedarikci t
            LEFT JOIN nexgen_satin_siparis s ON s.tedarikci_id = t.id
            GROUP BY t.id
            ORDER BY t.aktif DESC, t.ad
        """).fetchall()
    finally:
        con.close()

    can_manage = yetki_var('nexgen.tedarikci.manage', 'can_create')

    return render_template(
        'nexgen/satinalma_tedarikci.html',
        active='nexgen',
        tedarikciler=[dict(t) for t in tedarikciler],
        can_manage=can_manage,
    )


# ─────────────────────────────────────────────────────────────
# API — Yeni Sipariş Oluştur
# POST /nexgen/api/satinalma/siparis-ekle
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/satinalma/siparis-ekle', methods=['POST'])
@yetki_gerekli('nexgen.satinalma.manage', 'can_create')
def api_satinalma_siparis_ekle():
    data = request.get_json(silent=True) or {}

    tedarikci_id  = data.get('tedarikci_id')
    stok_kart_id  = data.get('stok_kart_id')
    siparis_tarihi = (data.get('siparis_tarihi') or '').strip()
    beklenen_teslim = (data.get('beklenen_teslim') or '').strip() or None
    aciklama = (data.get('aciklama') or '').strip() or None
    para_birimi = (data.get('para_birimi') or 'TRY').strip().upper()
    onay_durumu = 'TASLAK' if data.get('taslak') else 'ONAY_BEKLIYOR'

    try:
        miktar = float(data.get('siparis_miktari_kg') or 0)
        if miktar <= 0:
            return jsonify({"ok": False, "hata": "Sipariş miktarı sıfırdan büyük olmalı."}), 400
    except (ValueError, TypeError):
        return jsonify({"ok": False, "hata": "Geçersiz sipariş miktarı."}), 400

    birim_fiyat = None
    kur = None
    birim_fiyat_try = None
    toplam_tutar_try = None
    vade_gun = None
    vade_tarihi = None

    try:
        if data.get('birim_fiyat') not in (None, ''):
            birim_fiyat = float(data['birim_fiyat'])
            if birim_fiyat < 0:
                return jsonify({"ok": False, "hata": "Birim fiyat negatif olamaz."}), 400
    except (ValueError, TypeError):
        return jsonify({"ok": False, "hata": "Geçersiz birim fiyat."}), 400

    try:
        if data.get('kur') not in (None, ''):
            kur = float(data['kur'])
            if kur <= 0:
                return jsonify({"ok": False, "hata": "Kur sıfırdan büyük olmalı."}), 400
    except (ValueError, TypeError):
        return jsonify({"ok": False, "hata": "Geçersiz kur değeri."}), 400

    if birim_fiyat is not None:
        k = kur if (kur and para_birimi != 'TRY') else 1.0
        birim_fiyat_try = round(birim_fiyat * k, 4)
        toplam_tutar_try = round(birim_fiyat_try * miktar, 2)

    try:
        if data.get('vade_gun') not in (None, ''):
            vade_gun = int(data['vade_gun'])
            if siparis_tarihi and vade_gun >= 0:
                from datetime import date, timedelta
                base = date.fromisoformat(siparis_tarihi)
                vade_tarihi = (base + timedelta(days=vade_gun)).isoformat()
    except (ValueError, TypeError):
        return jsonify({"ok": False, "hata": "Geçersiz vade değeri."}), 400

    GECERLI_PARA = {'TRY', 'USD', 'EUR', 'GBP', 'CNY'}
    if para_birimi not in GECERLI_PARA:
        return jsonify({"ok": False, "hata": f"Geçersiz para birimi: {para_birimi}"}), 400

    if not tedarikci_id:
        return jsonify({"ok": False, "hata": "tedarikci_id zorunludur."}), 400
    if not stok_kart_id:
        return jsonify({"ok": False, "hata": "stok_kart_id zorunludur."}), 400
    if not siparis_tarihi:
        return jsonify({"ok": False, "hata": "siparis_tarihi zorunludur."}), 400

    con = _db()
    try:
        ted = con.execute(
            "SELECT id FROM nexgen_tedarikci WHERE id=? AND aktif=1", (tedarikci_id,)
        ).fetchone()
        if not ted:
            return jsonify({"ok": False, "hata": "Tedarikçi bulunamadı veya pasif."}), 404

        kart = con.execute(
            "SELECT id FROM nexgen_stok_kart WHERE id=? AND aktif=1", (stok_kart_id,)
        ).fetchone()
        if not kart:
            return jsonify({"ok": False, "hata": "Stok kartı bulunamadı veya pasif."}), 404

        siparis_no = _siparis_no_uret(con)

        con.execute("""
            INSERT INTO nexgen_satin_siparis
              (siparis_no, tedarikci_id, stok_kart_id,
               siparis_tarihi, beklenen_teslim,
               siparis_miktari_kg, birim_fiyat, para_birimi, kur,
               birim_fiyat_try, toplam_tutar_try,
               vade_gun, vade_tarihi,
               durum, onay_durumu, aciklama,
               olusturan_id, olusturma_tarihi)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'BEKLIYOR', ?, ?, ?, datetime('now'))
        """, (
            siparis_no, tedarikci_id, stok_kart_id,
            siparis_tarihi, beklenen_teslim,
            miktar, birim_fiyat, para_birimi, kur,
            birim_fiyat_try, toplam_tutar_try,
            vade_gun, vade_tarihi,
            onay_durumu, aciklama,
            _kullanici_id(),
        ))
        yeni_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({"ok": True, "id": yeni_id, "siparis_no": siparis_no})


# ─────────────────────────────────────────────────────────────
# API — Sipariş Onay/Red/İptal
# POST /nexgen/api/satinalma/siparis-durum
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/satinalma/siparis-durum', methods=['POST'])
@yetki_gerekli('nexgen.satinalma.view', 'can_view')
def api_satinalma_siparis_durum():
    data = request.get_json(silent=True) or {}
    siparis_id = data.get('id')
    eylem = (data.get('eylem') or '').strip().upper()

    if not siparis_id:
        return jsonify({"ok": False, "hata": "id zorunludur."}), 400

    EYLEMLER = {'ONAYLA', 'REDDET', 'IPTAL', 'ONAYA_GONDER'}
    if eylem not in EYLEMLER:
        return jsonify({"ok": False, "hata": f"Geçersiz eylem: {eylem}"}), 400

    # Onay eylemi sadece approve yetkisiyle
    if eylem in ('ONAYLA', 'REDDET'):
        if not yetki_var('nexgen.satinalma.approve', 'can_approve'):
            return jsonify({"ok": False, "hata": "Onay/red yetkisi yok."}), 403

    # İptal: manage yetkisi yeterli
    if eylem == 'IPTAL':
        if not yetki_var('nexgen.satinalma.manage', 'can_update'):
            return jsonify({"ok": False, "hata": "İptal yetkisi yok."}), 403

    con = _db()
    try:
        siparis = con.execute(
            "SELECT id, onay_durumu, durum FROM nexgen_satin_siparis WHERE id=?",
            (siparis_id,)
        ).fetchone()
        if not siparis:
            return jsonify({"ok": False, "hata": "Sipariş bulunamadı."}), 404

        mevcut_onay = siparis['onay_durumu']
        mevcut_durum = siparis['durum']

        # Durum geçiş kuralları
        if eylem == 'ONAYA_GONDER':
            if mevcut_onay not in ('TASLAK', 'REDDEDILDI'):
                return jsonify({"ok": False, "hata": "Sadece Taslak veya Reddedilmiş sipariş onaya gönderilebilir."}), 400
            con.execute("""
                UPDATE nexgen_satin_siparis
                SET onay_durumu='ONAY_BEKLIYOR',
                    guncelleyen_id=?, guncelleme_tarihi=datetime('now')
                WHERE id=?
            """, (_kullanici_id(), siparis_id))

        elif eylem == 'ONAYLA':
            if mevcut_onay != 'ONAY_BEKLIYOR':
                return jsonify({"ok": False, "hata": "Sadece Onay Bekleyen sipariş onaylanabilir."}), 400
            con.execute("""
                UPDATE nexgen_satin_siparis
                SET onay_durumu='ONAYLANDI',
                    onaylayan_id=?, onay_tarihi=datetime('now'),
                    guncelleyen_id=?, guncelleme_tarihi=datetime('now')
                WHERE id=?
            """, (_kullanici_id(), _kullanici_id(), siparis_id))

        elif eylem == 'REDDET':
            if mevcut_onay != 'ONAY_BEKLIYOR':
                return jsonify({"ok": False, "hata": "Sadece Onay Bekleyen sipariş reddedilebilir."}), 400
            con.execute("""
                UPDATE nexgen_satin_siparis
                SET onay_durumu='REDDEDILDI',
                    onaylayan_id=?, onay_tarihi=datetime('now'),
                    guncelleyen_id=?, guncelleme_tarihi=datetime('now')
                WHERE id=?
            """, (_kullanici_id(), _kullanici_id(), siparis_id))

        elif eylem == 'IPTAL':
            if mevcut_durum == 'TAMAMLANDI':
                return jsonify({"ok": False, "hata": "Tamamlanmış sipariş iptal edilemez."}), 400
            con.execute("""
                UPDATE nexgen_satin_siparis
                SET durum='IPTAL',
                    guncelleyen_id=?, guncelleme_tarihi=datetime('now')
                WHERE id=?
            """, (_kullanici_id(), siparis_id))

        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({"ok": True, "eylem": eylem})


# ─────────────────────────────────────────────────────────────
# API — Tedarikçi Ekle
# POST /nexgen/api/satinalma/tedarikci-ekle
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/satinalma/tedarikci-ekle', methods=['POST'])
@yetki_gerekli('nexgen.tedarikci.manage', 'can_create')
def api_tedarikci_ekle():
    data = request.get_json(silent=True) or {}
    kod   = (data.get('kod') or '').strip().upper()
    ad    = (data.get('ad') or '').strip()
    ulke  = (data.get('ulke') or 'TR').strip().upper()
    pb    = (data.get('para_birimi') or 'TRY').strip().upper()
    notlar = (data.get('notlar') or '').strip() or None
    iletisim_ad    = (data.get('iletisim_ad') or '').strip() or None
    iletisim_tel   = (data.get('iletisim_tel') or '').strip() or None
    iletisim_email = (data.get('iletisim_email') or '').strip() or None

    try:
        vade = int(data.get('varsayilan_vade') or 30)
    except (ValueError, TypeError):
        return jsonify({"ok": False, "hata": "Geçersiz vade değeri."}), 400

    if not kod:
        return jsonify({"ok": False, "hata": "Kod zorunludur."}), 400
    if not ad:
        return jsonify({"ok": False, "hata": "Ad zorunludur."}), 400

    GECERLI_PARA = {'TRY', 'USD', 'EUR', 'GBP', 'CNY'}
    if pb not in GECERLI_PARA:
        return jsonify({"ok": False, "hata": f"Geçersiz para birimi: {pb}"}), 400

    con = _db()
    try:
        mevcut = con.execute(
            "SELECT id FROM nexgen_tedarikci WHERE kod=?", (kod,)
        ).fetchone()
        if mevcut:
            return jsonify({"ok": False, "hata": f"'{kod}' kodu zaten mevcut."}), 409

        con.execute("""
            INSERT INTO nexgen_tedarikci
              (kod, ad, ulke, para_birimi, varsayilan_vade,
               iletisim_ad, iletisim_tel, iletisim_email, notlar,
               aktif, olusturan_id, olusturma_tarihi)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, datetime('now'))
        """, (kod, ad, ulke, pb, vade,
              iletisim_ad, iletisim_tel, iletisim_email, notlar,
              _kullanici_id()))
        yeni_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({"ok": True, "id": yeni_id, "kod": kod, "ad": ad})


# ─────────────────────────────────────────────────────────────
# API — Tedarikçi Düzenle
# POST /nexgen/api/satinalma/tedarikci-guncelle
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/satinalma/tedarikci-guncelle', methods=['POST'])
@yetki_gerekli('nexgen.tedarikci.manage', 'can_update')
def api_tedarikci_guncelle():
    data = request.get_json(silent=True) or {}
    tid  = data.get('id')
    if not tid:
        return jsonify({"ok": False, "hata": "id zorunludur."}), 400

    ad    = (data.get('ad') or '').strip()
    ulke  = (data.get('ulke') or 'TR').strip().upper()
    pb    = (data.get('para_birimi') or 'TRY').strip().upper()
    notlar = (data.get('notlar') or '').strip() or None
    iletisim_ad    = (data.get('iletisim_ad') or '').strip() or None
    iletisim_tel   = (data.get('iletisim_tel') or '').strip() or None
    iletisim_email = (data.get('iletisim_email') or '').strip() or None

    try:
        vade = int(data.get('varsayilan_vade') or 30)
    except (ValueError, TypeError):
        return jsonify({"ok": False, "hata": "Geçersiz vade değeri."}), 400

    if not ad:
        return jsonify({"ok": False, "hata": "Ad zorunludur."}), 400

    con = _db()
    try:
        mevcut = con.execute(
            "SELECT id FROM nexgen_tedarikci WHERE id=?", (tid,)
        ).fetchone()
        if not mevcut:
            return jsonify({"ok": False, "hata": "Tedarikçi bulunamadı."}), 404

        con.execute("""
            UPDATE nexgen_tedarikci
            SET ad=?, ulke=?, para_birimi=?, varsayilan_vade=?,
                iletisim_ad=?, iletisim_tel=?, iletisim_email=?, notlar=?,
                guncelleyen_id=?, guncelleme_tarihi=datetime('now')
            WHERE id=?
        """, (ad, ulke, pb, vade,
              iletisim_ad, iletisim_tel, iletisim_email, notlar,
              _kullanici_id(), tid))
        con.commit()
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# API — Stok Kartları (satın alma formu için dropdown)
# GET /nexgen/api/satinalma/stok-listesi
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/satinalma/stok-listesi')
@yetki_gerekli('nexgen.satinalma.manage', 'can_create')
def api_satinalma_stok_listesi():
    con = _db()
    try:
        kartlar = con.execute("""
            SELECT id, kod, ad, kategori, birim
            FROM nexgen_stok_kart
            WHERE aktif=1
            ORDER BY kategori, ad
        """).fetchall()
    finally:
        con.close()
    return jsonify({"ok": True, "kartlar": [dict(k) for k in kartlar]})


# ─────────────────────────────────────────────────────────────
# API — Tedarikçi Listesi (satın alma formu için dropdown)
# GET /nexgen/api/satinalma/tedarikci-listesi
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/satinalma/tedarikci-listesi')
@yetki_gerekli('nexgen.satinalma.manage', 'can_create')
def api_satinalma_tedarikci_listesi():
    con = _db()
    try:
        tedarikciler = con.execute("""
            SELECT id, kod, ad, para_birimi, varsayilan_vade
            FROM nexgen_tedarikci
            WHERE aktif=1
            ORDER BY ad
        """).fetchall()
    finally:
        con.close()
    return jsonify({"ok": True, "tedarikciler": [dict(t) for t in tedarikciler]})


# ═════════════════════════════════════════════════════════════
# FAZ-2.6 — HAFTALlK FİYAT YÖNETİMİ
# ═════════════════════════════════════════════════════════════
# KURAL: Bu bölümde nexgen_stok_hareket INSERT YAPILMAZ.
# Fiyat geçmişi ≠ stok hareketi.
# ─────────────────────────────────────────────────────────────


def _isoweek():
    """Geçerli ISO hafta kodu: '2026-W26'"""
    today = date.today()
    iso = today.isocalendar()
    return f"{iso[0]}-W{iso[1]:02d}"


def _son_fiyat(con, tedarikci_id, stok_kart_id):
    """
    Son aktif fiyat kaydını döndür.
    Kural: aktif=1 AND ORDER BY fiyat_tarihi DESC, id DESC LIMIT 1
    """
    return con.execute("""
        SELECT nhf.*, t.ad AS tedarikci_ad, sk.ad AS stok_ad,
               sk.kod AS stok_kod
        FROM nexgen_hammadde_fiyat nhf
        JOIN nexgen_tedarikci t ON t.id = nhf.tedarikci_id
        JOIN nexgen_stok_kart sk ON sk.id = nhf.stok_kart_id
        WHERE nhf.tedarikci_id = ?
          AND nhf.stok_kart_id = ?
          AND nhf.aktif = 1
        ORDER BY nhf.fiyat_tarihi DESC, nhf.id DESC
        LIMIT 1
    """, (tedarikci_id, stok_kart_id)).fetchone()


def _fiyat_farki(eski_fiyat, eski_pb, yeni_fiyat, yeni_pb):
    """Aynı para birimi varsa fark ve yüzde hesapla."""
    if eski_fiyat is None or yeni_fiyat is None:
        return None, None
    if eski_pb != yeni_pb:
        return None, None
    fark = yeni_fiyat - eski_fiyat
    yuzde = (fark / eski_fiyat * 100) if eski_fiyat != 0 else None
    return round(fark, 4), round(yuzde, 2) if yuzde is not None else None


# ─────────────────────────────────────────────────────────────
# Fiyat Listesi — GET /nexgen/satinalma/fiyat
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/satinalma/fiyat')
@yetki_gerekli('nexgen.fiyat.view', 'can_view')
def fiyat_listesi():
    con = _db()
    try:
        tedarikci_id = request.args.get('tedarikci_id', type=int)
        stok_kart_id = request.args.get('stok_kart_id', type=int)

        q = """
            SELECT nhf.id, nhf.fiyat, nhf.para_birimi, nhf.kur, nhf.fiyat_try,
                   nhf.vade_gun, nhf.fiyat_tarihi, nhf.kaynak, nhf.notlar,
                   nhf.aktif, nhf.batch_id,
                   t.kod AS tedarikci_kodu, t.ad AS tedarikci_ad,
                   sk.kod AS stok_kodu, sk.ad AS stok_ad
            FROM nexgen_hammadde_fiyat nhf
            JOIN nexgen_tedarikci t ON t.id = nhf.tedarikci_id
            JOIN nexgen_stok_kart sk ON sk.id = nhf.stok_kart_id
            WHERE 1=1
        """
        params = []
        if tedarikci_id:
            q += " AND nhf.tedarikci_id = ?"
            params.append(tedarikci_id)
        if stok_kart_id:
            q += " AND nhf.stok_kart_id = ?"
            params.append(stok_kart_id)
        q += " ORDER BY nhf.fiyat_tarihi DESC, nhf.id DESC LIMIT 200"

        fiyatlar = con.execute(q, params).fetchall()
        tedarikciler = con.execute(
            "SELECT id, kod, ad FROM nexgen_tedarikci WHERE aktif=1 ORDER BY ad"
        ).fetchall()
        stok_kartlari = con.execute(
            "SELECT id, kod, ad FROM nexgen_stok_kart WHERE aktif=1 ORDER BY kod"
        ).fetchall()
        son_batch = con.execute(
            "SELECT * FROM nexgen_fiyat_batch ORDER BY id DESC LIMIT 5"
        ).fetchall()

        can_manage = yetki_var('nexgen.fiyat.manage', 'can_create')
        can_admin  = yetki_var('nexgen.fiyat.admin', 'can_manage')
    finally:
        con.close()

    return render_template(
        'nexgen/fiyat_listesi.html',
        fiyatlar=fiyatlar,
        tedarikciler=tedarikciler,
        stok_kartlari=stok_kartlari,
        son_batch=son_batch,
        filtre_tedarikci=tedarikci_id,
        filtre_stok=stok_kart_id,
        can_manage=can_manage,
        can_admin=can_admin,
        active='nexgen'
    )


# ─────────────────────────────────────────────────────────────
# Excel Şablon İndir — GET /nexgen/satinalma/fiyat-sablonu-indir
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/satinalma/fiyat-sablonu-indir')
@yetki_gerekli('nexgen.fiyat.manage', 'can_create')
def fiyat_sablonu_indir():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({"ok": False, "hata": "openpyxl yüklü değil"}), 500

    con = _db()
    try:
        eslesmeler = con.execute("""
            SELECT ts.id, t.kod AS t_kod, t.ad AS t_ad, t.varsayilan_vade,
                   sk.kod AS s_kod, sk.ad AS s_ad,
                   t.id AS tedarikci_id, sk.id AS stok_kart_id
            FROM nexgen_tedarikci_stok ts
            JOIN nexgen_tedarikci t ON t.id = ts.tedarikci_id
            JOIN nexgen_stok_kart sk ON sk.id = ts.stok_kart_id
            WHERE ts.aktif = 1 AND t.aktif = 1 AND sk.aktif = 1
            ORDER BY t.kod, sk.kod
        """).fetchall()

        son_fiyatlar = {}
        for e in eslesmeler:
            sf = _son_fiyat(con, e['tedarikci_id'], e['stok_kart_id'])
            son_fiyatlar[(e['tedarikci_id'], e['stok_kart_id'])] = sf
    finally:
        con.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fiyat Girişi"

    bugun = date.today().isoformat()
    hafta = _isoweek()

    baslik_font = Font(bold=True, color="FFFFFF")
    baslik_dolu = PatternFill("solid", fgColor="2C3E50")
    kilitli_dolu = PatternFill("solid", fgColor="ECF0F1")
    orta = Alignment(horizontal="center")

    basliklar = [
        "tedarikci_kodu", "tedarikci_adi", "stok_kodu", "stok_adi",
        "fiyat", "para_birimi", "kur", "vade_gun",
        "fiyat_tarihi", "gecerlilik_bas", "gecerlilik_bitis", "not",
        "son_fiyat_bilgi"
    ]
    ws.append(basliklar)
    for col, hdr in enumerate(basliklar, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = baslik_font
        cell.fill = baslik_dolu
        cell.alignment = orta

    for e in eslesmeler:
        sf = son_fiyatlar.get((e['tedarikci_id'], e['stok_kart_id']))
        son_fiyat_bilgi = ""
        if sf:
            son_fiyat_bilgi = f"{sf['fiyat']} {sf['para_birimi']} ({sf['fiyat_tarihi']})"

        row = [
            e['t_kod'], e['t_ad'], e['s_kod'], e['s_ad'],
            None, "USD", None, e['varsayilan_vade'],
            bugun, bugun, None, None,
            son_fiyat_bilgi
        ]
        ws.append(row)

        r_idx = ws.max_row
        for col in [1, 2, 3, 4, 13]:
            ws.cell(row=r_idx, column=col).fill = kilitli_dolu

    for col_idx, width in zip(range(1, 14), [14, 22, 14, 28, 10, 12, 10, 10, 13, 13, 14, 20, 25]):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "E2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    dosya_adi = f"NexGen_Fiyat_Sablonu_{hafta}.xlsx"
    return Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={dosya_adi}"}
    )


# ─────────────────────────────────────────────────────────────
# Excel Yükle + Preview — POST /nexgen/satinalma/fiyat-yukle
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/satinalma/fiyat-yukle', methods=['POST'])
@yetki_gerekli('nexgen.fiyat.manage', 'can_create')
def fiyat_yukle():
    try:
        import openpyxl
    except ImportError:
        return jsonify({"ok": False, "hata": "openpyxl yüklü değil"}), 500

    f = request.files.get('dosya')
    if not f or not f.filename.endswith('.xlsx'):
        return jsonify({"ok": False, "hata": "Geçerli bir .xlsx dosyası seçin"}), 400

    kullanici_id = _kullanici_id()
    hafta = _isoweek()

    con = _db()
    try:
        # Batch kaydı oluştur
        con.execute("""
            INSERT INTO nexgen_fiyat_batch
              (hafta_kodu, dosya_adi, durum, yukleyen_id)
            VALUES (?, ?, 'ONAY_BEKLIYOR', ?)
        """, (hafta, f.filename, kullanici_id))
        con.commit()
        batch_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Excel parse
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active

        tedarikci_map = {r[0]: r[1] for r in con.execute(
            "SELECT kod, id FROM nexgen_tedarikci"
        ).fetchall()}
        stok_map = {r[0]: r[1] for r in con.execute(
            "SELECT kod, id FROM nexgen_stok_kart"
        ).fetchall()}

        toplam = gecerli = hatali = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            t_kod = str(row[0]).strip() if row[0] else None
            s_kod = str(row[2]).strip() if row[2] else None
            fiyat_raw = row[4]
            pb = str(row[5]).strip().upper() if row[5] else 'USD'
            kur = row[6]
            vade = row[7]
            f_tarih = str(row[8]).strip() if row[8] else date.today().isoformat()
            g_bas = str(row[9]).strip() if row[9] else None
            g_bit = str(row[10]).strip() if row[10] else None
            notlar = str(row[11]).strip() if row[11] else None

            toplam += 1

            if fiyat_raw is None or str(fiyat_raw).strip() == '':
                continue

            hata = None
            t_id = s_id = None
            try:
                fiyat = float(fiyat_raw)
                if fiyat <= 0:
                    hata = "Fiyat sıfır veya negatif olamaz"
            except (ValueError, TypeError):
                hata = f"Geçersiz fiyat: {fiyat_raw}"
                fiyat = None

            if t_kod not in tedarikci_map:
                hata = f"Tedarikçi kodu bulunamadı: {t_kod}"
            else:
                t_id = tedarikci_map[t_kod]

            if s_kod not in stok_map:
                hata = (hata + " | " if hata else "") + f"Stok kodu bulunamadı: {s_kod}"
            else:
                s_id = stok_map[s_kod]

            onceki = None
            fark = yuzde = None
            if t_id and s_id:
                onceki = _son_fiyat(con, t_id, s_id)
            if onceki and fiyat and not hata:
                fark, yuzde = _fiyat_farki(
                    onceki['fiyat'], onceki['para_birimi'], fiyat, pb
                )

            fiyat_try = None
            if fiyat and kur:
                try:
                    fiyat_try = round(float(fiyat) * float(kur), 4)
                except Exception:
                    pass

            gecerli_mi = 0 if hata else 1
            if gecerli_mi:
                gecerli += 1
            else:
                hatali += 1

            con.execute("""
                INSERT INTO nexgen_fiyat_batch_detay
                  (batch_id, tedarikci_kodu, stok_kodu, tedarikci_id, stok_kart_id,
                   fiyat, para_birimi, kur, fiyat_try, vade_gun,
                   fiyat_tarihi, gecerlilik_bas, gecerlilik_bitis, notlar,
                   onceki_fiyat, onceki_pb, onceki_tarih, fark, yuzde_degisim,
                   gecerli_mi, hata_sebebi)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                batch_id, t_kod, s_kod, t_id, s_id,
                fiyat, pb, kur, fiyat_try, vade,
                f_tarih, g_bas, g_bit, notlar,
                onceki['fiyat'] if onceki else None,
                onceki['para_birimi'] if onceki else None,
                onceki['fiyat_tarihi'] if onceki else None,
                fark, yuzde,
                gecerli_mi, hata
            ))

        con.execute("""
            UPDATE nexgen_fiyat_batch
            SET toplam_satir=?, gecerli_satir=?, hatali_satir=?
            WHERE id=?
        """, (toplam, gecerli, hatali, batch_id))
        con.commit()
    finally:
        con.close()

    return redirect(url_for('nexgen.fiyat_preview', batch_id=batch_id))


# ─────────────────────────────────────────────────────────────
# Preview — GET /nexgen/satinalma/fiyat-preview/<batch_id>
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/satinalma/fiyat-preview/<int:batch_id>')
@yetki_gerekli('nexgen.fiyat.manage', 'can_create')
def fiyat_preview(batch_id):
    con = _db()
    try:
        batch = con.execute(
            "SELECT * FROM nexgen_fiyat_batch WHERE id=?", (batch_id,)
        ).fetchone()
        if not batch:
            abort(404)

        if batch['durum'] != 'ONAY_BEKLIYOR':
            flash(f"Bu batch artık işlenemez: durum={batch['durum']}", "warning")
            return redirect(url_for('nexgen.fiyat_listesi'))

        detaylar = con.execute("""
            SELECT * FROM nexgen_fiyat_batch_detay
            WHERE batch_id=?
            ORDER BY tedarikci_kodu, stok_kodu
        """, (batch_id,)).fetchall()

        can_approve = yetki_var('nexgen.fiyat.approve', 'can_approve')
    finally:
        con.close()

    return render_template(
        'nexgen/fiyat_preview.html',
        batch=batch,
        detaylar=detaylar,
        can_approve=can_approve,
        active='nexgen'
    )


# ─────────────────────────────────────────────────────────────
# Onayla — POST /nexgen/satinalma/fiyat-onayla/<batch_id>
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/satinalma/fiyat-onayla/<int:batch_id>', methods=['POST'])
@yetki_gerekli('nexgen.fiyat.approve', 'can_approve')
def fiyat_onayla(batch_id):
    kullanici_id = _kullanici_id()
    con = _db()
    try:
        batch = con.execute(
            "SELECT * FROM nexgen_fiyat_batch WHERE id=?", (batch_id,)
        ).fetchone()
        if not batch or batch['durum'] != 'ONAY_BEKLIYOR':
            return jsonify({"ok": False, "hata": "Geçersiz veya işlenmiş batch"}), 400

        detaylar = con.execute("""
            SELECT * FROM nexgen_fiyat_batch_detay
            WHERE batch_id=? AND gecerli_mi=1
              AND tedarikci_id IS NOT NULL AND stok_kart_id IS NOT NULL
              AND fiyat IS NOT NULL
        """, (batch_id,)).fetchall()

        eklendi = 0
        for d in detaylar:
            con.execute("""
                INSERT INTO nexgen_hammadde_fiyat
                  (tedarikci_id, stok_kart_id, fiyat, para_birimi, kur, fiyat_try,
                   vade_gun, fiyat_tarihi, gecerlilik_bas, gecerlilik_bitis,
                   kaynak, batch_id, notlar, aktif, olusturan_id)
                VALUES (?,?,?,?,?,?,?,?,?,?,'EXCEL_IMPORT',?,?,1,?)
            """, (
                d['tedarikci_id'], d['stok_kart_id'],
                d['fiyat'], d['para_birimi'], d['kur'], d['fiyat_try'],
                d['vade_gun'], d['fiyat_tarihi'], d['gecerlilik_bas'], d['gecerlilik_bitis'],
                batch_id, d['notlar'], kullanici_id
            ))
            eklendi += 1

        con.execute("""
            UPDATE nexgen_fiyat_batch
            SET durum='ONAYLANDI', onaylayan_id=?, onay_tarihi=datetime('now')
            WHERE id=?
        """, (kullanici_id, batch_id))
        con.commit()
    finally:
        con.close()

    return jsonify({"ok": True, "eklendi": eklendi})


# ─────────────────────────────────────────────────────────────
# İptal — POST /nexgen/satinalma/fiyat-iptal/<batch_id>
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/satinalma/fiyat-iptal/<int:batch_id>', methods=['POST'])
@yetki_gerekli('nexgen.fiyat.manage', 'can_create')
def fiyat_iptal(batch_id):
    con = _db()
    try:
        batch = con.execute(
            "SELECT * FROM nexgen_fiyat_batch WHERE id=?", (batch_id,)
        ).fetchone()
        if not batch or batch['durum'] != 'ONAY_BEKLIYOR':
            return jsonify({"ok": False, "hata": "Geçersiz veya işlenmiş batch"}), 400

        con.execute("DELETE FROM nexgen_fiyat_batch_detay WHERE batch_id=?", (batch_id,))
        con.execute(
            "UPDATE nexgen_fiyat_batch SET durum='IPTAL' WHERE id=?", (batch_id,)
        )
        con.commit()
    finally:
        con.close()

    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# API — Manuel Tekil Fiyat Girişi
# POST /nexgen/api/satinalma/fiyat-manuel
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/satinalma/fiyat-manuel', methods=['POST'])
@yetki_gerekli('nexgen.fiyat.manage', 'can_create')
def api_fiyat_manuel():
    if not request.is_json:
        return jsonify({"ok": False, "hata": "JSON bekleniyor"}), 400
    d = request.json
    tedarikci_id = d.get('tedarikci_id')
    stok_kart_id = d.get('stok_kart_id')
    fiyat        = d.get('fiyat')
    para_birimi  = d.get('para_birimi', 'USD')
    kur          = d.get('kur')
    vade_gun     = d.get('vade_gun')
    fiyat_tarihi = d.get('fiyat_tarihi')
    notlar       = d.get('notlar')

    if not all([tedarikci_id, stok_kart_id, fiyat, fiyat_tarihi]):
        return jsonify({"ok": False, "hata": "Zorunlu alanlar eksik"}), 400
    if para_birimi not in ('USD', 'EUR', 'TRY', 'GBP', 'CNY'):
        return jsonify({"ok": False, "hata": f"Geçersiz para birimi: {para_birimi}"}), 400

    fiyat_try = None
    if kur:
        try:
            fiyat_try = round(float(fiyat) * float(kur), 4)
        except Exception:
            pass

    kullanici_id = _kullanici_id()
    con = _db()
    try:
        ted  = con.execute("SELECT id FROM nexgen_tedarikci WHERE id=?", (tedarikci_id,)).fetchone()
        stok = con.execute("SELECT id FROM nexgen_stok_kart WHERE id=?", (stok_kart_id,)).fetchone()
        if not ted:
            return jsonify({"ok": False, "hata": "Tedarikçi bulunamadı"}), 400
        if not stok:
            return jsonify({"ok": False, "hata": "Stok kartı bulunamadı"}), 400

        con.execute("""
            INSERT INTO nexgen_hammadde_fiyat
              (tedarikci_id, stok_kart_id, fiyat, para_birimi, kur, fiyat_try,
               vade_gun, fiyat_tarihi, kaynak, aktif, olusturan_id)
            VALUES (?,?,?,?,?,?,?,?,'MANUEL',1,?)
        """, (tedarikci_id, stok_kart_id, fiyat, para_birimi, kur, fiyat_try,
              vade_gun, fiyat_tarihi, kullanici_id))
        con.commit()
        fiyat_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
    finally:
        con.close()

    return jsonify({"ok": True, "id": fiyat_id})


# ─────────────────────────────────────────────────────────────
# API — Son Fiyat Öner (sipariş formu pre-fill)
# GET /nexgen/api/satinalma/son-fiyat?tedarikci_id=X&stok_kart_id=Y
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/satinalma/son-fiyat')
@yetki_gerekli('nexgen.satinalma.manage', 'can_create')
def api_son_fiyat():
    tedarikci_id = request.args.get('tedarikci_id', type=int)
    stok_kart_id = request.args.get('stok_kart_id', type=int)
    if not tedarikci_id or not stok_kart_id:
        return jsonify({"ok": False, "hata": "tedarikci_id ve stok_kart_id gerekli"}), 400

    con = _db()
    try:
        sf = _son_fiyat(con, tedarikci_id, stok_kart_id)
    finally:
        con.close()

    if not sf:
        return jsonify({"ok": True, "fiyat": None})

    return jsonify({
        "ok": True,
        "fiyat": {
            "id": sf['id'],
            "fiyat": sf['fiyat'],
            "para_birimi": sf['para_birimi'],
            "kur": sf['kur'],
            "vade_gun": sf['vade_gun'],
            "fiyat_tarihi": sf['fiyat_tarihi'],
        }
    })


# ─────────────────────────────────────────────────────────────
# API — Fiyat Pasife Al (sadece Yönetim)
# POST /nexgen/api/satinalma/fiyat-pasife-al/<fiyat_id>
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/satinalma/fiyat-pasife-al/<int:fiyat_id>', methods=['POST'])
@yetki_gerekli('nexgen.fiyat.admin', 'can_manage')
def api_fiyat_pasife_al(fiyat_id):
    sebep = request.json.get('sebep', '') if request.is_json else ''
    kullanici_id = _kullanici_id()
    con = _db()
    try:
        kayit = con.execute(
            "SELECT id, aktif FROM nexgen_hammadde_fiyat WHERE id=?", (fiyat_id,)
        ).fetchone()
        if not kayit:
            return jsonify({"ok": False, "hata": "Kayıt bulunamadı"}), 404
        if not kayit['aktif']:
            return jsonify({"ok": False, "hata": "Zaten pasif"}), 400

        con.execute("""
            UPDATE nexgen_hammadde_fiyat
            SET aktif=0, iptal_sebebi=?, iptal_eden_id=?, iptal_tarihi=datetime('now')
            WHERE id=?
        """, (sebep, kullanici_id, fiyat_id))
        con.commit()
    finally:
        con.close()

    return jsonify({"ok": True})


# =============================================================
# FAZ-2.8 — NexGen Yönetim Merkezi
# =============================================================

# ─────────────────────────────────────────────────────────────
# Yönetim Merkezi Ana Sayfa
# GET /nexgen/yonetim/
# Yetki: nexgen.yonetim.manage can_view
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/yonetim/')
@yetki_gerekli('nexgen.yonetim.manage', 'can_view')
def yonetim_merkezi():
    con = _db()
    try:
        # Stok kartları — aile adıyla birlikte
        kartlar_raw = con.execute("""
            SELECT k.id, k.kod, k.ad, k.kategori, k.birim,
                   k.minimum_stok, k.kritik_stok, k.aktif,
                   k.renk, k.alt_kategori, k.kalite_sinifi,
                   k.shore_degeri, k.notlar, k.aile_id,
                   a.ad AS aile_ad, a.aa_kodu AS aile_aa,
                   COALESCE(SUM(h.miktar_kg), 0) AS mevcut_stok
            FROM nexgen_stok_kart k
            LEFT JOIN nexgen_stok_aile a ON a.id = k.aile_id
            LEFT JOIN nexgen_stok_hareket h ON h.stok_kart_id = k.id
            GROUP BY k.id
            ORDER BY a.sira NULLS LAST, k.kod
        """).fetchall()

        # Tedarikçiler — bağlı hammadde sayısıyla
        tedarikciler_raw = con.execute("""
            SELECT t.id, t.kod, t.ad, t.ulke, t.para_birimi,
                   t.varsayilan_vade, t.iletisim_ad, t.iletisim_tel,
                   t.iletisim_email, t.notlar, t.aktif,
                   COUNT(ts.id) AS esleme_sayisi
            FROM nexgen_tedarikci t
            LEFT JOIN nexgen_tedarikci_stok ts ON ts.tedarikci_id = t.id AND ts.aktif=1
            GROUP BY t.id
            ORDER BY t.aktif DESC, t.ad
        """).fetchall()

        # Eşleştirmeler — tümü
        eslesme_raw = con.execute("""
            SELECT ts.id, ts.tedarikci_id, ts.stok_kart_id,
                   ts.tercih_sirasi, ts.aktif, ts.notlar,
                   t.kod  AS ted_kod,  t.ad  AS ted_ad,
                   sk.kod AS stok_kod, sk.ad AS stok_ad
            FROM nexgen_tedarikci_stok ts
            JOIN nexgen_tedarikci  t  ON t.id  = ts.tedarikci_id
            JOIN nexgen_stok_kart  sk ON sk.id = ts.stok_kart_id
            ORDER BY t.ad, ts.tercih_sirasi, sk.kod
        """).fetchall()

        # Stok aile listesi (yeni kart modalında dropdown için)
        aileler_raw = con.execute(
            "SELECT id, aa_kodu, ad FROM nexgen_stok_aile WHERE aktif=1 ORDER BY sira"
        ).fetchall()

        # Cari listesi
        try:
            cariler_raw = con.execute(
                "SELECT id, cari_kod, unvan, aktif FROM nexgen_cari ORDER BY aktif DESC, cari_kod"
            ).fetchall()
            cariler = [dict(c) for c in cariler_raw]
        except Exception:
            cariler = []

    finally:
        con.close()

    return render_template(
        'nexgen/yonetim.html',
        active='nexgen',
        kartlar=[dict(k) for k in kartlar_raw],
        tedarikciler=[dict(t) for t in tedarikciler_raw],
        eslesme_listesi=[dict(e) for e in eslesme_raw],
        aileler=[dict(a) for a in aileler_raw],
        cariler=cariler,
    )


# ─────────────────────────────────────────────────────────────
# API — Stok Kartı Güncelle (Yönetim)
# POST /nexgen/api/yonetim/stok-kart-guncelle
# Yetki: nexgen.yonetim.manage can_update
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/yonetim/stok-kart-guncelle', methods=['POST'])
@yetki_gerekli('nexgen.yonetim.manage', 'can_update')
def api_yonetim_stok_kart_guncelle():
    d = request.get_json(silent=True) or {}
    kart_id = d.get('id')
    if not kart_id:
        return jsonify({"ok": False, "hata": "id gerekli"}), 400

    guncellenecek = {}
    for alan in ('ad', 'kategori', 'birim', 'minimum_stok', 'kritik_stok',
                 'renk', 'alt_kategori', 'kalite_sinifi', 'shore_degeri',
                 'notlar', 'aile_id'):
        if alan in d:
            guncellenecek[alan] = d[alan] if d[alan] != '' else None

    if not guncellenecek:
        return jsonify({"ok": False, "hata": "Güncellenecek alan yok"}), 400

    kullanici_id = _kullanici_id()
    set_clause = ', '.join(f"{k}=?" for k in guncellenecek)
    vals = list(guncellenecek.values()) + [kullanici_id, kart_id]

    con = _db()
    try:
        mev = con.execute("SELECT id FROM nexgen_stok_kart WHERE id=?", (kart_id,)).fetchone()
        if not mev:
            return jsonify({"ok": False, "hata": "Stok kartı bulunamadı"}), 404
        con.execute(
            f"UPDATE nexgen_stok_kart SET {set_clause}, "
            f"guncelleyen_id=?, guncelleme_tarihi=datetime('now') WHERE id=?",
            vals
        )
        con.commit()
    finally:
        con.close()

    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# API — Stok Kartı Aktif/Pasif (Yönetim)
# POST /nexgen/api/yonetim/stok-kart-durum
# Yetki: nexgen.yonetim.manage can_update
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/yonetim/stok-kart-durum', methods=['POST'])
@yetki_gerekli('nexgen.yonetim.manage', 'can_update')
def api_yonetim_stok_kart_durum():
    d = request.get_json(silent=True) or {}
    kart_id = d.get('id')
    yeni_aktif = d.get('aktif')
    if kart_id is None or yeni_aktif is None:
        return jsonify({"ok": False, "hata": "id ve aktif gerekli"}), 400

    con = _db()
    try:
        con.execute(
            "UPDATE nexgen_stok_kart SET aktif=?, guncelleyen_id=?, "
            "guncelleme_tarihi=datetime('now') WHERE id=?",
            (1 if yeni_aktif else 0, _kullanici_id(), kart_id)
        )
        con.commit()
    finally:
        con.close()
    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# API — Tedarikçi Güncelle (Yönetim) — mevcut endpoint wrapper
# POST /nexgen/api/yonetim/tedarikci-guncelle
# Yetki: nexgen.yonetim.manage can_update
# Not: Mevcut /api/satinalma/tedarikci-guncelle zaten nexgen.tedarikci.manage
#      gerektiriyor; Yönetim SuperAdmin olduğundan geçer.
#      Yönetim ekranından doğrudan mevcut endpoint çağrılacak (JS'te URL değişir).
# ─────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────
# API — Eşleştirme Ekle (Yönetim)
# POST /nexgen/api/yonetim/eslestirme-ekle
# Yetki: nexgen.yonetim.manage can_create
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/yonetim/eslestirme-ekle', methods=['POST'])
@yetki_gerekli('nexgen.yonetim.manage', 'can_create')
def api_yonetim_eslestirme_ekle():
    d = request.get_json(silent=True) or {}
    tedarikci_id = d.get('tedarikci_id')
    stok_kart_id = d.get('stok_kart_id')
    tercih_sirasi = d.get('tercih_sirasi', 1)
    notlar = d.get('notlar', '')

    if not tedarikci_id or not stok_kart_id:
        return jsonify({"ok": False, "hata": "tedarikci_id ve stok_kart_id gerekli"}), 400

    con = _db()
    try:
        # Çift kayıt kontrolü
        mev = con.execute(
            "SELECT id FROM nexgen_tedarikci_stok WHERE tedarikci_id=? AND stok_kart_id=?",
            (tedarikci_id, stok_kart_id)
        ).fetchone()
        if mev:
            # Pasifse reaktive et
            con.execute(
                "UPDATE nexgen_tedarikci_stok SET aktif=1, tercih_sirasi=?, notlar=? WHERE id=?",
                (tercih_sirasi, notlar, mev['id'])
            )
            con.commit()
            return jsonify({"ok": True, "yeni": False, "id": mev['id']})

        con.execute("""
            INSERT INTO nexgen_tedarikci_stok
              (tedarikci_id, stok_kart_id, tercih_sirasi, aktif, notlar)
            VALUES (?, ?, ?, 1, ?)
        """, (tedarikci_id, stok_kart_id, tercih_sirasi, notlar))
        yeni_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
        con.commit()
    finally:
        con.close()

    return jsonify({"ok": True, "yeni": True, "id": yeni_id})


# ─────────────────────────────────────────────────────────────
# API — Eşleştirme Kaldır (Yönetim)
# POST /nexgen/api/yonetim/eslestirme-kaldir
# Yetki: nexgen.yonetim.manage can_delete
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/yonetim/eslestirme-kaldir', methods=['POST'])
@yetki_gerekli('nexgen.yonetim.manage', 'can_delete')
def api_yonetim_eslestirme_kaldir():
    d = request.get_json(silent=True) or {}
    eslesme_id = d.get('id')
    if not eslesme_id:
        return jsonify({"ok": False, "hata": "id gerekli"}), 400

    con = _db()
    try:
        kayit = con.execute(
            "SELECT id FROM nexgen_tedarikci_stok WHERE id=?", (eslesme_id,)
        ).fetchone()
        if not kayit:
            return jsonify({"ok": False, "hata": "Kayıt bulunamadı"}), 404
        # Silmek yerine pasife al — geçmişi koru
        con.execute(
            "UPDATE nexgen_tedarikci_stok SET aktif=0 WHERE id=?", (eslesme_id,)
        )
        con.commit()
    finally:
        con.close()

    return jsonify({"ok": True})


# ─────────────────────────────────────────────────────────────
# CARİ MASTER API — FAZ-5E-3
# ─────────────────────────────────────────────────────────────

@nexgen_bp.route('/api/yonetim/cari-sonraki-kod', methods=['GET'])
@yetki_gerekli('nexgen.yonetim.manage', 'can_create')
def api_cari_sonraki_kod():
    """Standart formattaki (120.NX.XXX) mevcut en büyük numarayı bulur, +1 verir."""
    import re
    con = _db()
    try:
        rows = con.execute("SELECT cari_kod FROM nexgen_cari").fetchall()
        maks = 0
        pat = re.compile(r'^120\.NX\.(\d+)$')
        for row in rows:
            m = pat.match((row[0] or '').strip())
            if m:
                n = int(m.group(1))
                if n > maks:
                    maks = n
        yeni_no = maks + 1
        yeni_kod = f"120.NX.{yeni_no:03d}"
        return jsonify({'ok': True, 'kod': yeni_kod})
    except Exception as e:
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/yonetim/cari-ekle', methods=['POST'])
@yetki_gerekli('nexgen.yonetim.manage', 'can_create')
def api_cari_ekle():
    """Yeni cari ekle."""
    d = request.get_json(silent=True) or {}
    kod   = (d.get('cari_kod') or '').strip()
    unvan = (d.get('unvan') or '').strip()
    if not kod or not unvan:
        return jsonify({'ok': False, 'hata': 'cari_kod ve unvan zorunlu'}), 400
    con = _db()
    try:
        mev = con.execute("SELECT id FROM nexgen_cari WHERE cari_kod=?", (kod,)).fetchone()
        if mev:
            return jsonify({'ok': False, 'hata': f"'{kod}' kodu zaten mevcut"}), 400
        con.execute(
            "INSERT INTO nexgen_cari(cari_kod, unvan, aktif) VALUES(?,?,1)",
            (kod, unvan)
        )
        con.commit()
        yeni_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
        return jsonify({'ok': True, 'id': yeni_id})
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/yonetim/cari-guncelle', methods=['POST'])
@yetki_gerekli('nexgen.yonetim.manage', 'can_update')
def api_cari_guncelle():
    """Cari unvan güncelle."""
    d = request.get_json(silent=True) or {}
    cari_id = d.get('id')
    unvan   = (d.get('unvan') or '').strip()
    if not cari_id or not unvan:
        return jsonify({'ok': False, 'hata': 'id ve unvan zorunlu'}), 400
    con = _db()
    try:
        kayit = con.execute("SELECT id FROM nexgen_cari WHERE id=?", (cari_id,)).fetchone()
        if not kayit:
            return jsonify({'ok': False, 'hata': 'Cari bulunamadı'}), 404
        con.execute(
            "UPDATE nexgen_cari SET unvan=?, updated_at=datetime('now','localtime') WHERE id=?",
            (unvan, cari_id)
        )
        con.commit()
        return jsonify({'ok': True})
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/yonetim/cari-durum', methods=['POST'])
@yetki_gerekli('nexgen.yonetim.manage', 'can_update')
def api_cari_durum():
    """Cari aktif/pasif toggle."""
    d = request.get_json(silent=True) or {}
    cari_id = d.get('id')
    if not cari_id:
        return jsonify({'ok': False, 'hata': 'id zorunlu'}), 400
    con = _db()
    try:
        kayit = con.execute("SELECT id, aktif FROM nexgen_cari WHERE id=?", (cari_id,)).fetchone()
        if not kayit:
            return jsonify({'ok': False, 'hata': 'Cari bulunamadı'}), 404
        yeni = 0 if kayit['aktif'] else 1
        con.execute(
            "UPDATE nexgen_cari SET aktif=?, updated_at=datetime('now','localtime') WHERE id=?",
            (yeni, cari_id)
        )
        con.commit()
        return jsonify({'ok': True, 'aktif': yeni})
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


# =============================================================
# FAZ-3A — NexGen Depo Mal Kabul
# =============================================================

@nexgen_bp.route('/depo')
def depo_slash_yonlendir():
    """Depo ana sayfa slash uyumu — /nexgen/depo → /nexgen/depo/"""
    from flask import redirect
    return redirect('/nexgen/depo/', code=302)


# ─────────────────────────────────────────────────────────────
# Depo Ana Sayfa
# GET /nexgen/depo/
# Yetki: nexgen.depo.view can_view
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/depo/')
@yetki_gerekli('nexgen.depo.view', 'can_view')
def depo():
    can_giris = yetki_var('nexgen.depo.giris', 'can_create')
    con = _db()
    try:
        # Bekleyen / kısmi siparişler
        bekleyen_raw = con.execute("""
            SELECT
                ss.id           AS siparis_id,
                ss.siparis_no,
                ss.siparis_miktari_kg    AS siparis_kg,
                ss.durum,
                ss.beklenen_teslim,
                ss.olusturma_tarihi,
                t.id            AS tedarikci_id,
                t.kod           AS ted_kod,
                t.ad            AS ted_ad,
                sk.id           AS stok_kart_id,
                sk.kod          AS stok_kod,
                sk.ad           AS stok_ad,
                sk.birim,
                COALESCE(
                    (SELECT SUM(mk.miktar_kg)
                     FROM nexgen_mal_kabul mk
                     WHERE mk.satin_siparis_id = ss.id), 0
                ) AS gelen_kg
            FROM nexgen_satin_siparis ss
            JOIN nexgen_tedarikci  t  ON t.id  = ss.tedarikci_id
            JOIN nexgen_stok_kart  sk ON sk.id = ss.stok_kart_id
            WHERE ss.durum IN ('BEKLIYOR', 'KISMI_TESLIM')
              AND ss.onay_durumu = 'ONAYLANDI'
            ORDER BY ss.beklenen_teslim ASC NULLS LAST, ss.id DESC
        """).fetchall()

        # Giriş/Çıkış geçmişi (son 100 hareket — hem giriş hem çıkış)
        gecmis_raw = con.execute("""
            SELECT
                h.id,
                h.hareket_tipi,
                h.miktar_kg,
                h.onceki_stok,
                h.sonraki_stok,
                h.aciklama,
                h.olusturma_tarihi AS hareket_tarihi,
                h.referans_tip,
                sk.kod  AS stok_kod,
                sk.ad   AS stok_ad,
                ku.KullaniciAdi AS islem_yapan_ad,
                -- Mal kabule ait detaylar (sadece GİRİŞ'lerde dolu)
                mk.irsaliye_no,
                mk.lot_no,
                mk.satin_siparis_id,
                ss.siparis_no,
                t.ad    AS ted_ad
            FROM nexgen_stok_hareket h
            JOIN nexgen_stok_kart  sk ON sk.id = h.stok_kart_id
            LEFT JOIN sistem_kullanici ku ON ku.Id = h.olusturan_id
            LEFT JOIN nexgen_mal_kabul  mk ON mk.stok_hareket_id = h.id
            LEFT JOIN nexgen_tedarikci  t  ON t.id  = mk.tedarikci_id
            LEFT JOIN nexgen_satin_siparis ss ON ss.id = mk.satin_siparis_id
            WHERE h.hareket_tipi IN ('GIRIS', 'URETIM_TUKETIM', 'ARGE_DENEME',
                                     'FIRE_ZAYI', 'SAYIM_DUZELTME', 'DIREKT_GIRIS')
               OR h.referans_tip = 'DEPO_CIKIS'
            ORDER BY h.olusturma_tarihi DESC, h.id DESC
            LIMIT 100
        """).fetchall()

        # Tedarikçi ve stok listesi (direkt giriş modalı için)
        tedarikciler_raw = con.execute(
            "SELECT id, kod, ad FROM nexgen_tedarikci WHERE aktif=1 ORDER BY ad"
        ).fetchall()
        kartlar_raw = con.execute(
            "SELECT id, kod, ad, birim FROM nexgen_stok_kart WHERE aktif=1 ORDER BY kod"
        ).fetchall()

        # KPI
        bugun_giren_kg = con.execute("""
            SELECT COALESCE(SUM(miktar_kg), 0)
            FROM nexgen_mal_kabul
            WHERE date(kabul_tarihi) = date('now')
        """).fetchone()[0]
        bu_hafta_kabul_say = con.execute("""
            SELECT COUNT(*) FROM nexgen_mal_kabul
            WHERE kabul_tarihi >= datetime('now', '-7 days')
        """).fetchone()[0]

    finally:
        con.close()

    bekleyen = []
    for r in bekleyen_raw:
        d = dict(r)
        d['kalan_kg'] = round(d['siparis_kg'] - d['gelen_kg'], 3)
        bekleyen.append(d)

    return render_template(
        'nexgen/depo.html',
        active='nexgen',
        can_giris=can_giris,
        bekleyen=bekleyen,
        gecmis=[dict(g) for g in gecmis_raw],
        tedarikciler=[dict(t) for t in tedarikciler_raw],
        kartlar=[dict(k) for k in kartlar_raw],
        bugun_giren_kg=round(bugun_giren_kg, 1),
        bu_hafta_kabul_say=bu_hafta_kabul_say,
        bekleyen_say=len(bekleyen),
    )


# ─────────────────────────────────────────────────────────────
# API — Mal Kabul
# POST /nexgen/api/depo/mal-kabul
# Yetki: nexgen.depo.giris can_create
#
# Kurallar:
#   1) nexgen_mal_kabul INSERT (belge)
#   2) nexgen_stok_hareket INSERT hareket_tipi=GIRIS
#   3) Siparişe bağlıysa sipariş durum güncelle
#   4) Fiyat tablolarına dokunulmaz
# ─────────────────────────────────────────────────────────────
@nexgen_bp.route('/api/depo/mal-kabul', methods=['POST'])
@yetki_gerekli('nexgen.depo.giris', 'can_create')
def api_depo_mal_kabul():
    d = request.get_json(silent=True) or {}

    tedarikci_id   = d.get('tedarikci_id')
    stok_kart_id   = d.get('stok_kart_id')
    miktar_kg      = d.get('miktar_kg')
    irsaliye_no    = d.get('irsaliye_no', '').strip() or None
    lot_no         = d.get('lot_no', '').strip() or None
    aciklama       = d.get('aciklama', '').strip() or None
    siparis_id     = d.get('satin_siparis_id') or None  # None → direkt giriş

    # Zorunlu alan kontrolü
    if not tedarikci_id or not stok_kart_id or not miktar_kg:
        return jsonify({"ok": False, "hata": "tedarikci_id, stok_kart_id ve miktar_kg zorunlu"}), 400
    try:
        miktar_kg = float(miktar_kg)
        if miktar_kg <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"ok": False, "hata": "miktar_kg sıfırdan büyük olmalı"}), 400

    kullanici_id = _kullanici_id()
    con = _db()
    try:
        # Tedarikçi ve stok kartı var mı?
        ted = con.execute("SELECT id FROM nexgen_tedarikci WHERE id=?", (tedarikci_id,)).fetchone()
        if not ted:
            return jsonify({"ok": False, "hata": "Tedarikçi bulunamadı"}), 404
        kart = con.execute("SELECT id, ad FROM nexgen_stok_kart WHERE id=?", (stok_kart_id,)).fetchone()
        if not kart:
            return jsonify({"ok": False, "hata": "Stok kartı bulunamadı"}), 404

        # Siparişe bağlıysa kontrol
        siparis = None
        if siparis_id:
            siparis = con.execute(
                "SELECT id, siparis_miktari_kg, durum, stok_kart_id, tedarikci_id FROM nexgen_satin_siparis WHERE id=?",
                (siparis_id,)
            ).fetchone()
            if not siparis:
                return jsonify({"ok": False, "hata": "Sipariş bulunamadı"}), 404
            if siparis['durum'] in ('TAMAMLANDI', 'IPTAL'):
                return jsonify({"ok": False, "hata": f"Sipariş durumu {siparis['durum']}, mal kabul yapılamaz"}), 400

        # ── Stok mevcut durumu ─────────────────────────────────
        onceki_stok = _mevcut_stok(con, stok_kart_id)
        sonraki_stok = round(onceki_stok + miktar_kg, 3)

        # ── referans_tip belirle ───────────────────────────────
        if siparis_id:
            referans_tip = 'SATIN_ALMA_SIPARIS'
            ref_id = siparis_id
        else:
            referans_tip = 'DIREKT_GIRIS'
            ref_id = None  # mal_kabul INSERT sonrası doldurulacak

        # ── nexgen_stok_hareket INSERT ─────────────────────────
        aciklama_hareket = aciklama or (
            f"Mal kabul — "
            + (f"Sipariş #{siparis_id}" if siparis_id else "Direkt giriş")
            + (f" İrsaliye:{irsaliye_no}" if irsaliye_no else "")
        )
        con.execute("""
            INSERT INTO nexgen_stok_hareket
              (stok_kart_id, hareket_tipi, miktar_kg,
               onceki_stok, sonraki_stok,
               aciklama, referans_tip, referans_id,
               olusturan_id, olusturma_tarihi)
            VALUES (?, 'GIRIS', ?, ?, ?, ?, ?, ?, ?, datetime('now'))
        """, (stok_kart_id, miktar_kg, onceki_stok, sonraki_stok,
              aciklama_hareket, referans_tip, ref_id, kullanici_id))
        hareket_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]

        # ── nexgen_mal_kabul INSERT ────────────────────────────
        con.execute("""
            INSERT INTO nexgen_mal_kabul
              (satin_siparis_id, tedarikci_id, stok_kart_id,
               miktar_kg, irsaliye_no, lot_no,
               kabul_eden_id, kabul_tarihi, aciklama, stok_hareket_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now'), ?, ?)
        """, (siparis_id, tedarikci_id, stok_kart_id,
              miktar_kg, irsaliye_no, lot_no,
              kullanici_id, aciklama, hareket_id))
        mal_kabul_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]

        # Direkt girişte referans_id = mal_kabul_id
        if not siparis_id:
            con.execute(
                "UPDATE nexgen_stok_hareket SET referans_id=? WHERE id=?",
                (mal_kabul_id, hareket_id)
            )

        # ── Sipariş durumunu güncelle ──────────────────────────
        yeni_durum = None
        toplam_gelen = None
        if siparis and siparis_id:
            toplam_gelen_row = con.execute(
                "SELECT COALESCE(SUM(miktar_kg), 0) FROM nexgen_mal_kabul WHERE satin_siparis_id=?",
                (siparis_id,)
            ).fetchone()[0]
            toplam_gelen = round(toplam_gelen_row, 3)

            if toplam_gelen >= siparis['siparis_miktari_kg']:
                yeni_durum = 'TAMAMLANDI'
            else:
                yeni_durum = 'KISMI_TESLIM'

            con.execute(
                "UPDATE nexgen_satin_siparis SET durum=? WHERE id=?",
                (yeni_durum, siparis_id)
            )

        con.commit()

    except Exception as e:
        con.close()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    sonuc = {
        "ok": True,
        "mal_kabul_id": mal_kabul_id,
        "hareket_id": hareket_id,
        "onceki_stok": onceki_stok,
        "sonraki_stok": sonraki_stok,
        "miktar_kg": miktar_kg,
    }
    if yeni_durum:
        sonuc["siparis_yeni_durum"] = yeni_durum
        sonuc["toplam_gelen_kg"] = toplam_gelen
    return jsonify(sonuc)


# ─────────────────────────────────────────────────────────────
# API — Depo Çıkış
# POST /nexgen/api/depo/cikis
# Yetki: nexgen.depo.giris can_create
#
# Kurallar:
#   1) Stok yeterliliği kontrol et
#   2) nexgen_stok_hareket INSERT hareket_tipi = seçilen tip, miktar_kg NEGATİF
#   3) nexgen_stok_kart'a dokunulmuyor
#   4) Fiyat tablolarına dokunulmuyor
# ─────────────────────────────────────────────────────────────
CIKIS_TIPLERI = {'URETIM_TUKETIM', 'ARGE_DENEME', 'FIRE_ZAYI', 'SAYIM_DUZELTME'}

@nexgen_bp.route('/api/depo/cikis', methods=['POST'])
@yetki_gerekli('nexgen.depo.giris', 'can_create')
def api_depo_cikis():
    d = request.get_json(silent=True) or {}

    stok_kart_id  = d.get('stok_kart_id')
    miktar_kg     = d.get('miktar_kg')
    hareket_tipi  = (d.get('hareket_tipi') or '').strip().upper()
    aciklama      = (d.get('aciklama') or '').strip() or None

    if not stok_kart_id or not miktar_kg or not hareket_tipi:
        return jsonify({"ok": False, "hata": "stok_kart_id, miktar_kg ve hareket_tipi zorunlu"}), 400

    if hareket_tipi not in CIKIS_TIPLERI:
        return jsonify({"ok": False,
                        "hata": f"Geçersiz hareket tipi. Geçerli: {', '.join(sorted(CIKIS_TIPLERI))}"}), 400

    try:
        miktar_kg = float(miktar_kg)
        if miktar_kg <= 0:
            raise ValueError
    except (TypeError, ValueError):
        return jsonify({"ok": False, "hata": "miktar_kg sıfırdan büyük olmalı"}), 400

    kullanici_id = _kullanici_id()
    con = _db()
    try:
        kart = con.execute(
            "SELECT id, ad FROM nexgen_stok_kart WHERE id=? AND aktif=1", (stok_kart_id,)
        ).fetchone()
        if not kart:
            return jsonify({"ok": False, "hata": "Stok kartı bulunamadı veya pasif"}), 404

        onceki_stok = _mevcut_stok(con, stok_kart_id)
        if onceki_stok < miktar_kg:
            return jsonify({
                "ok": False,
                "hata": f"Yetersiz stok. Mevcut: {onceki_stok:.3f} KG, İstenen: {miktar_kg:.3f} KG"
            }), 400

        miktar_negatif = -round(miktar_kg, 3)
        sonraki_stok   = round(onceki_stok + miktar_negatif, 3)

        aciklama_hareket = aciklama or f"Depo çıkış — {hareket_tipi}"

        con.execute("""
            INSERT INTO nexgen_stok_hareket
              (stok_kart_id, hareket_tipi, miktar_kg,
               onceki_stok, sonraki_stok,
               aciklama, referans_tip,
               olusturan_id, olusturma_tarihi)
            VALUES (?, ?, ?, ?, ?, ?, 'DEPO_CIKIS', ?, datetime('now'))
        """, (stok_kart_id, hareket_tipi, miktar_negatif,
              onceki_stok, sonraki_stok,
              aciklama_hareket, kullanici_id))
        hareket_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
        con.commit()

    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()

    return jsonify({
        "ok": True,
        "hareket_id":   hareket_id,
        "hareket_tipi": hareket_tipi,
        "onceki_stok":  onceki_stok,
        "sonraki_stok": sonraki_stok,
        "miktar_kg":    miktar_kg,
    })


# ═══════════════════════════════════════════════════════════════
# NEXGEN FAZ-5A — TABLET EKRANI
# Stok hareketi yapılmaz. Sadece batch kodu üretimi + AR-GE testi.
# ═══════════════════════════════════════════════════════════════

def _batch_kodu_uret(con):
    """NG-PRD-YYYY-NNNN formatında benzersiz üretim batch kodu üretir."""
    import datetime
    yil = datetime.datetime.now().year
    son = con.execute(
        "SELECT batch_kodu FROM nexgen_uretim_batch "
        "WHERE batch_kodu LIKE ? ORDER BY id DESC LIMIT 1",
        (f"NG-PRD-{yil}-%",)
    ).fetchone()
    if son:
        try:
            son_no = int(son['batch_kodu'].split('-')[-1])
        except Exception:
            son_no = 0
    else:
        son_no = 0
    yeni_no = son_no + 1
    return f"NG-PRD-{yil}-{yeni_no:05d}"


def _lot_kodu_uret(con):
    """NG-LOT-YYYY-NNNNN formatında benzersiz fiziksel LOT kodu üretir.
    batch_kodu = işlem kaydı; lot_kodu = fiziksel ürün/compound takip kodu.
    """
    import datetime
    yil = datetime.datetime.now().year
    son = con.execute(
        "SELECT lot_kodu FROM nexgen_uretim_batch "
        "WHERE lot_kodu LIKE ? ORDER BY id DESC LIMIT 1",
        (f"NG-LOT-{yil}-%",)
    ).fetchone()
    if son and son['lot_kodu']:
        try:
            son_no = int(son['lot_kodu'].split('-')[-1])
        except Exception:
            son_no = 0
    else:
        son_no = 0
    return f"NG-LOT-{yil}-{son_no + 1:05d}"


@nexgen_bp.route('/tablet')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_ana():
    con = _db()
    try:
        devam_eden, plan_isler = _tablet_ana_veri(con)
    finally:
        con.close()

    return render_template(
        'nexgen/tablet.html',
        active='nexgen',
        can_uretim=yetki_var('nexgen.tablet.uretim', 'can_uretim'),
        devam_eden_batches=devam_eden,
        plan_isler=plan_isler,
    )


@nexgen_bp.route('/tablet/devam-edenler')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_devam_edenler():
    """Devam eden üretimler — sipariş bazında gruplandırılmış, L/S kırılımlı."""
    con = _db()
    try:
        _cols = [c['name'] for c in con.execute(
            "PRAGMA table_info(nexgen_uretim_batch)"
        ).fetchall()]
        if 'plan_id' in _cols:
            _cari_join = ""
            _musteri_sel = "np.musteri_adi"
            if _plan_cari_kolonu_var(con):
                _cari_join = "LEFT JOIN nexgen_cari c ON c.id = np.cari_id"
                _musteri_sel = "COALESCE(c.unvan, np.musteri_adi) AS musteri_adi"
            batches_raw = con.execute(f"""
                SELECT nb.id, nb.batch_kodu, nb.lot_kodu, nb.plan_id,
                       nb.uretim_varyant_id,
                       nb.planlanan_kg, nb.durum,
                       nb.olusturma_tarihi, nb.notlar,
                       uv.ad AS uv_ad, uv.boyut,
                       rv.ad AS renk_ad,
                       f.ad AS formul_ad, f.kod AS formul_kod,
                       ku.KullaniciAdi AS olusturan_ad,
                       {_musteri_sel},
                       np.siparis_no,
                       np.plan_tarihi, np.created_at AS plan_created_at
                FROM nexgen_uretim_batch nb
                JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f          ON f.id  = rv.formul_id
                LEFT JOIN sistem_kullanici ku ON ku.Id  = nb.olusturan_id
                LEFT JOIN nexgen_uretim_plan np ON np.id = nb.plan_id
                {_cari_join}
                WHERE nb.durum IN ('TASLAK','HAZIR','DEVAM','BEKLEME')
                ORDER BY nb.id DESC
            """).fetchall()
        else:
            batches_raw = con.execute("""
                SELECT nb.id, nb.batch_kodu, nb.lot_kodu,
                       nb.uretim_varyant_id,
                       nb.planlanan_kg, nb.durum,
                       nb.olusturma_tarihi, nb.notlar,
                       uv.ad AS uv_ad, uv.boyut,
                       rv.ad AS renk_ad,
                       f.ad AS formul_ad, f.kod AS formul_kod,
                       ku.KullaniciAdi AS olusturan_ad,
                       NULL AS musteri_adi, NULL AS siparis_no,
                       NULL AS plan_tarihi, NULL AS plan_created_at
                FROM nexgen_uretim_batch nb
                JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f          ON f.id  = rv.formul_id
                LEFT JOIN sistem_kullanici ku ON ku.Id  = nb.olusturan_id
                WHERE nb.durum IN ('TASLAK','HAZIR','DEVAM','BEKLEME')
                ORDER BY nb.id DESC
            """).fetchall()
        batches_raw = [dict(b) for b in batches_raw]

        for b in batches_raw:
            rf = _batch_plan_rf_bilgi(con, batch_kodu=b.get('batch_kodu'),
                                      uretim_varyant_id=b.get('uretim_varyant_id'))
            if rf:
                b['rf_kod'] = rf['rf_kod']
                b['rf_renk_ad'] = rf.get('renk_ad')
            else:
                b['rf_kod'] = None
                b['rf_renk_ad'] = None

        # Alt emir sayaçları + KG
        parca_tablosu = _parca_tablosu_var(con)
        if parca_tablosu:
            for b in batches_raw:
                sayac = con.execute("""
                    SELECT COUNT(*) AS toplam,
                           SUM(CASE WHEN durum='HAZIR'   THEN 1 ELSE 0 END) AS hazir,
                           SUM(CASE WHEN durum='DEVAM'   THEN 1 ELSE 0 END) AS devam,
                           SUM(CASE WHEN durum='BEKLEME' THEN 1 ELSE 0 END) AS bekleme,
                           SUM(CASE WHEN durum='BITTI'   THEN 1 ELSE 0 END) AS biten,
                           SUM(CASE WHEN durum='BITTI'   THEN uretilen_kg ELSE 0 END) AS uretilen
                    FROM nexgen_uretim_parca WHERE batch_kodu=?
                """, (b['batch_kodu'],)).fetchone()
                b['toplam_alt_emir'] = sayac['toplam'] or 0
                b['biten_alt_emir']  = sayac['biten']  or 0
                b['hazir_alt_emir']  = sayac['hazir']  or 0
                b['devam_alt_emir']  = sayac['devam']  or 0
                b['bekleme_alt_emir']= sayac['bekleme']or 0
                _batch_kg_rf_enrich(con, b, liste_modu=True)
        else:
            for b in batches_raw:
                b['toplam_alt_emir'] = b['biten_alt_emir'] = 0
                b['hazir_alt_emir']  = b['devam_alt_emir'] = b['bekleme_alt_emir'] = 0
                b['uretilen_kg']     = 0.0
                b['kalan_kg']        = float(b['planlanan_kg'])
                _batch_kg_rf_enrich(con, b, liste_modu=True)

        # Sipariş bazında gruplama:
        # siparis_no varsa → gruplama anahtarı siparis_no
        # yoksa → her batch kendi grubu (batch_kodu anahtar)
        from collections import OrderedDict
        siparis_gruplari = OrderedDict()

        for b in batches_raw:
            sip = b.get('siparis_no') or ('__batch_' + b['batch_kodu'])
            if sip not in siparis_gruplari:
                siparis_gruplari[sip] = {
                    'siparis_no':   b.get('siparis_no'),
                    'musteri_adi':  b.get('musteri_adi'),
                    'plan_tarihi':  b.get('plan_tarihi'),
                    'formul_ad':    b.get('formul_ad'),
                    'renk_ad':      b.get('renk_ad'),
                    'boyutlar':     [],   # [{boyut, batch_kodu, durum, ...}]
                    'tek_batch':    None, # tek boyut varsa direkt batch_kodu
                }
            grup = siparis_gruplari[sip]
            grup['boyutlar'].append({
                'boyut':          b.get('boyut') or 'STD',
                'batch_kodu':     b['batch_kodu'],
                'lot_kodu':       b.get('lot_kodu'),
                'durum':          b['durum'],
                'planlanan_kg':   b['planlanan_kg'],
                'siparis_toplam_kg': b.get('siparis_toplam_kg'),
                'uretilen_kg':    b['uretilen_kg'],
                'kalan_kg':       b['kalan_kg'],
                'toplam_alt_emir':b['toplam_alt_emir'],
                'biten_alt_emir': b['biten_alt_emir'],
                'hazir_alt_emir': b['hazir_alt_emir'],
                'devam_alt_emir': b['devam_alt_emir'],
                'bekleme_alt_emir':b['bekleme_alt_emir'],
            })

        # Tek boyutlu gruplarda tek_batch doldur
        for sip, grup in siparis_gruplari.items():
            if len(grup['boyutlar']) == 1:
                grup['tek_batch'] = grup['boyutlar'][0]['batch_kodu']

        siparis_listesi = list(siparis_gruplari.values())

    finally:
        con.close()

    return render_template(
        'nexgen/tablet_devam_edenler.html',
        active='nexgen',
        siparis_listesi=siparis_listesi,
        parca_tablosu=parca_tablosu,
    )


@nexgen_bp.route('/tablet/uretim-islem/<batch_kodu>')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_uretim_islem(batch_kodu):
    """Üretim işlem takip ekranı — operatör karışım/dolum sürecini buradan görür."""
    con = _db()
    try:
        # plan_id kolonu migration 071 sonrası mevcut — graceful JOIN
        _cols = [c['name'] for c in con.execute(
            "PRAGMA table_info(nexgen_uretim_batch)"
        ).fetchall()]
        if 'plan_id' in _cols:
            _cari_join = ""
            _musteri_sel = "np.musteri_adi"
            if _plan_cari_kolonu_var(con):
                _cari_join = "LEFT JOIN nexgen_cari c ON c.id = np.cari_id"
                _musteri_sel = "COALESCE(c.unvan, np.musteri_adi) AS musteri_adi"
            batch = con.execute(f"""
                SELECT nb.id, nb.batch_kodu, nb.lot_kodu, nb.plan_id,
                       nb.uretim_varyant_id,
                       nb.planlanan_kg, nb.durum, nb.olusturma_tarihi, nb.notlar,
                       uv.boyut,
                       rv.ad AS renk_ad,
                       f.ad AS formul_ad,
                       {_musteri_sel},
                       np.siparis_no, np.plan_tarihi
                FROM nexgen_uretim_batch nb
                JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f          ON f.id  = rv.formul_id
                LEFT JOIN nexgen_uretim_plan np ON np.id = nb.plan_id
                {_cari_join}
                WHERE nb.batch_kodu = ?
            """, (batch_kodu,)).fetchone()
        else:
            batch = con.execute("""
                SELECT nb.id, nb.batch_kodu, nb.lot_kodu,
                       nb.uretim_varyant_id,
                       nb.planlanan_kg, nb.durum, nb.olusturma_tarihi, nb.notlar,
                       uv.boyut,
                       rv.ad AS renk_ad,
                       f.ad AS formul_ad,
                       NULL AS musteri_adi, NULL AS siparis_no, NULL AS plan_tarihi
                FROM nexgen_uretim_batch nb
                JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f          ON f.id  = rv.formul_id
                WHERE nb.batch_kodu = ?
            """, (batch_kodu,)).fetchone()
        if not batch:
            from flask import abort
            abort(404)
        batch = dict(batch)

        # Alt Emirler (nexgen_uretim_parca) — Migration 072+ sonrası aktif
        alt_emirler = []
        toplam_uretilen_kg = 0.0
        parca_tablosu = _parca_tablosu_var(con)
        formul_batch_kg = 0.0
        sayaclar = {'toplam': 0, 'hazir': 0, 'devam': 0, 'bekleme': 0, 'bitti': 0}

        if parca_tablosu:
            # formul_batch_kg kolonu var mı?
            _pk_cols = [c['name'] for c in con.execute(
                "PRAGMA table_info(nexgen_uretim_parca)"
            ).fetchall()]
            has_formul_kg_col = 'formul_batch_kg' in _pk_cols

            if has_formul_kg_col:
                _select = """
                    SELECT np.id, np.parca_no, np.hedef_kg, np.formul_batch_kg,
                           np.uretilen_kg, np.durum, np.baslama_zamani, np.bitis_zamani,
                           np.bekleme_sebebi, np.notlar, np.vardiya,
                           ku.KullaniciAdi AS operator_ad
                    FROM nexgen_uretim_parca np
                    LEFT JOIN sistem_kullanici ku ON ku.Id = np.operator_id
                    WHERE np.batch_kodu = ?
                    ORDER BY np.parca_no ASC
                """
            else:
                _select = """
                    SELECT np.id, np.parca_no, np.hedef_kg,
                           np.hedef_kg AS formul_batch_kg,
                           np.uretilen_kg, np.durum, np.baslama_zamani, np.bitis_zamani,
                           np.bekleme_sebebi, np.notlar, np.vardiya,
                           ku.KullaniciAdi AS operator_ad
                    FROM nexgen_uretim_parca np
                    LEFT JOIN sistem_kullanici ku ON ku.Id = np.operator_id
                    WHERE np.batch_kodu = ?
                    ORDER BY np.parca_no ASC
                """

            _kayitlar = con.execute(_select, (batch_kodu,)).fetchall()
            alt_emirler = [dict(r) for r in _kayitlar]

            for ae in alt_emirler:
                sayaclar['toplam'] += 1
                d_lower = ae['durum'].lower()
                if d_lower in sayaclar:
                    sayaclar[d_lower] += 1

            # formul_batch_kg: ilk kayittan (plan bilgisi, uretim hesabi degil)
            if alt_emirler:
                formul_batch_kg = float(alt_emirler[0].get('formul_batch_kg') or 0)
        else:
            formul_batch_kg = 0.0

        kg_kontrol = _batch_uretim_kontrol(con, batch_kodu) or {}
        siparis_toplam_kg = float(
            kg_kontrol.get('siparis_toplam_kg') or batch['planlanan_kg']
        )
        toplam_uretilen_kg = float(kg_kontrol.get('uretilen_kg') or 0.0)
        kalan_kg = float(kg_kontrol.get('kalan_kg') or max(0.0, siparis_toplam_kg - toplam_uretilen_kg))
        kontrol_durum = kg_kontrol.get('kontrol_durum')
        kontrol_uyari = kg_kontrol.get('uyari')
        kontrol_mesaj = kg_kontrol.get('mesaj')
        planlanan_kg = siparis_toplam_kg

        # Aynı siparis_no'daki kardeş batch'ler (L/S seçim ekranı için)
        kardes_boyutlar = []
        sip_no = batch.get('siparis_no')
        if sip_no and 'plan_id' in _cols:
            kardesler = con.execute("""
                SELECT nb.batch_kodu, uv.boyut
                FROM nexgen_uretim_batch nb
                JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
                LEFT JOIN nexgen_uretim_plan np ON np.id = nb.plan_id
                WHERE np.siparis_no = ?
                  AND nb.durum IN ('TASLAK','HAZIR','DEVAM','BEKLEME','BITTI')
                ORDER BY uv.boyut
            """, (sip_no,)).fetchall()
            # Tüm boyutlar (bu batch dahil)
            kardes_boyutlar = [
                {'boyut': r['boyut'] or 'STD', 'batch_kodu': r['batch_kodu'],
                 'aktif': r['batch_kodu'] == batch_kodu}
                for r in kardesler
            ]

        # Barkod/sevk özet: BITTI olan alt emirlerin toplamı
        # (ileride sevke_hazir flag gerekir; şimdilik BITTI = barkoda hazır)
        barkod_hazir_adet = sayaclar.get('bitti', 0)
        barkod_hazir_kg   = round(toplam_uretilen_kg, 1)

        rf_bilgi = _batch_plan_rf_bilgi(
            con, batch_kodu=batch_kodu,
            plan_id=batch.get('plan_id'),
            uretim_varyant_id=batch['uretim_varyant_id'],
        )
        rf_kod = rf_bilgi['rf_kod'] if rf_bilgi else None
        rf_renk_ad = rf_bilgi.get('renk_ad') if rf_bilgi else None
        rf_uyari = None if rf_bilgi else (
            'Bu formül/renk için ONAYLI RF eşleşmesi bulunamadı. '
            'RF kullanım logu oluşturulmayabilir.'
        )

    finally:
        con.close()

    return render_template(
        'nexgen/tablet_uretim_islem.html',
        active='nexgen',
        batch=batch,
        alt_emirler=alt_emirler,
        toplam_uretilen_kg=round(toplam_uretilen_kg, 3),
        siparis_toplam_kg=round(siparis_toplam_kg, 3),
        kalan_kg=round(kalan_kg, 3),
        kontrol_durum=kontrol_durum,
        kontrol_uyari=kontrol_uyari,
        kontrol_mesaj=kontrol_mesaj,
        parca_tablosu=parca_tablosu,
        formul_batch_kg=round(formul_batch_kg, 3),
        sayaclar=sayaclar,
        kardes_boyutlar=kardes_boyutlar,
        barkod_hazir_adet=barkod_hazir_adet,
        barkod_hazir_kg=barkod_hazir_kg,
        rf_kod=rf_kod,
        rf_renk_ad=rf_renk_ad,
        rf_uyari=rf_uyari,
    )


@nexgen_bp.route('/tablet/geri-donusum')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_geri_donusum():
    """Geri Dönüşüm Günlük Girişi — RECYCLE kategorisi kartlarını listele."""
    con = _db()
    try:
        recycle_kartlar = con.execute("""
            SELECT sk.id, sk.kod, sk.ad, sk.alt_kategori,
                   COALESCE(SUM(sh.miktar_kg), 0) AS toplam_kg
            FROM nexgen_stok_kart sk
            LEFT JOIN nexgen_stok_hareket sh ON sh.stok_kart_id = sk.id
            WHERE sk.kategori = 'RECYCLE' AND sk.aktif = 1
            GROUP BY sk.id
            ORDER BY sk.ad
        """).fetchall()
        recycle_kartlar = [dict(k) for k in recycle_kartlar]
    finally:
        con.close()

    return render_template(
        'nexgen/tablet_geri_donusum.html',
        active='nexgen',
        recycle_kartlar=recycle_kartlar,
    )


@nexgen_bp.route('/api/tablet/geri-donusum-kaydet', methods=['POST'])
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def api_tablet_geri_donusum_kaydet():
    """Geri dönüşüm günlük girişlerini stok hareketi olarak kaydet.

    POST JSON:
        tarih      : str (YYYY-MM-DD)
        vardiya    : str
        personel   : str
        kalemler   : [{stok_kart_id: int, miktar_kg: float}]

    Hareket tipi: GERI_DONUSUM_DEVIR
    Stok ARTAR (pozitif miktar).
    """
    d = request.get_json(silent=True) or {}
    tarih    = (d.get('tarih') or '').strip()
    vardiya  = (d.get('vardiya') or '').strip()
    personel = (d.get('personel') or '').strip()
    kalemler = d.get('kalemler') or []

    if not tarih:
        return jsonify({'ok': False, 'hata': 'Tarih zorunlu'}), 400
    if not kalemler:
        return jsonify({'ok': False, 'hata': 'En az bir kalem giriniz'}), 400

    # Sıfır veya negatif kalemleri filtrele
    kalemler_gecerli = [
        k for k in kalemler
        if k.get('stok_kart_id') and float(k.get('miktar_kg') or 0) > 0
    ]
    if not kalemler_gecerli:
        return jsonify({'ok': False, 'hata': 'Geçerli miktar bulunamadı'}), 400

    aciklama = f'Geri dönüşüm günlük girişi — {tarih}'
    if vardiya:
        aciklama += f' — Vardiya: {vardiya}'
    if personel:
        aciklama += f' — Personel: {personel}'

    olusturan = session.get('user_id')
    now_str   = __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    con = _db()
    try:
        eklenen = 0
        for k in kalemler_gecerli:
            kart_id  = int(k['stok_kart_id'])
            miktar   = round(float(k['miktar_kg']), 3)
            # Kart RECYCLE kategorisi kontrolü
            kart = con.execute(
                "SELECT id, kategori FROM nexgen_stok_kart WHERE id=? AND aktif=1",
                (kart_id,)
            ).fetchone()
            if not kart or kart['kategori'] != 'RECYCLE':
                continue
            con.execute("""
                INSERT INTO nexgen_stok_hareket
                    (stok_kart_id, hareket_tipi, miktar_kg, aciklama,
                     referans_no, olusturan_id, olusturma_tarihi)
                VALUES (?,?,?,?,?,?,?)
            """, (kart_id, 'GERI_DONUSUM_DEVIR', miktar, aciklama,
                  tarih, olusturan, now_str))
            eklenen += 1
        con.commit()
    finally:
        con.close()

    return jsonify({'ok': True, 'eklenen': eklenen})


# ─────────────────────────────────────────────────────────────
# NEXGEN FAZ-5C-1 — BARKOD OKUTMA + KAYIT AÇMA
# KURAL: Stok hareketi yok. Sadece kod tanıma + yönlendirme.
# ─────────────────────────────────────────────────────────────

@nexgen_bp.route('/tablet/barkod')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_barkod():
    """Barkod / LOT kodu okutma giriş ekranı."""
    return render_template('nexgen/tablet_barkod.html', active='nexgen')


@nexgen_bp.route('/tablet/barkod/sonuc')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_barkod_sonuc():
    """Barkod sorgu sonucu ekranı (GET ile, ?kod= parametresiyle)."""
    kod = (request.args.get('kod') or '').strip().upper()
    if not kod:
        return render_template('nexgen/tablet_barkod.html', active='nexgen', hata='Kod boş olamaz.')

    con = _db()
    try:
        # 1) NG-LOT-YYYY-NNNNN — üretim lot kodu
        if kod.startswith('NG-LOT-'):
            batch = con.execute("""
                SELECT nb.id, nb.batch_kodu, nb.lot_kodu, nb.planlanan_kg, nb.durum,
                       nb.olusturma_tarihi, nb.notlar,
                       uv.boyut,
                       rv.ad AS renk_ad,
                       f.ad AS formul_ad
                FROM nexgen_uretim_batch nb
                JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f          ON f.id  = rv.formul_id
                WHERE nb.lot_kodu = ?
            """, (kod,)).fetchone()
            if batch:
                return render_template(
                    'nexgen/tablet_barkod_sonuc.html', active='nexgen',
                    tip='URETIM_LOT', kayit=dict(batch), aranan_kod=kod,
                )

        # 2) NG-PRD-YYYY-NNNNN — üretim batch kodu
        if kod.startswith('NG-PRD-'):
            batch = con.execute("""
                SELECT nb.id, nb.batch_kodu, nb.lot_kodu, nb.planlanan_kg, nb.durum,
                       nb.olusturma_tarihi, nb.notlar,
                       uv.boyut,
                       rv.ad AS renk_ad,
                       f.ad AS formul_ad
                FROM nexgen_uretim_batch nb
                JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f          ON f.id  = rv.formul_id
                WHERE nb.batch_kodu = ?
            """, (kod,)).fetchone()
            if batch:
                return render_template(
                    'nexgen/tablet_barkod_sonuc.html', active='nexgen',
                    tip='URETIM_BATCH', kayit=dict(batch), aranan_kod=kod,
                )

        # 3) AT-YYYY-NNNNN — AR-GE test kodu
        if kod.startswith('AT-'):
            test = con.execute("""
                SELECT at.id, at.test_no, at.test_batch_kg, at.makina,
                       at.yeni_renk_adi, at.durum, at.olusturma_tarihi,
                       uv.boyut,
                       rv.ad AS renk_ad,
                       f.ad AS formul_ad
                FROM nexgen_arge_test at
                JOIN nexgen_uretim_varyant uv ON uv.id = at.kaynak_uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f          ON f.id  = rv.formul_id
                WHERE at.test_no = ?
            """, (kod,)).fetchone()
            if test:
                return render_template(
                    'nexgen/tablet_barkod_sonuc.html', active='nexgen',
                    tip='ARGE', kayit=dict(test), aranan_kod=kod,
                )

        # Hiçbiri bulunamadı
        return render_template(
            'nexgen/tablet_barkod.html', active='nexgen',
            hata=f'"{kod}" kodu bulunamadı. Kontrol edin.',
            son_kod=kod,
        )
    finally:
        con.close()


@nexgen_bp.route('/api/tablet/barkod-bul', methods=['POST'])
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def api_tablet_barkod_bul():
    """Barkod / LOT kodu arama API.

    POST JSON: { "kod": "NG-LOT-2026-00001" }
    Response:
        { ok: true, tip: "URETIM_LOT"|"URETIM_BATCH"|"ARGE",
          redirect_url: "...", ozet: {...} }
    """
    d = request.get_json(silent=True) or {}
    kod = (d.get('kod') or '').strip().upper()
    if not kod:
        return jsonify({'ok': False, 'hata': 'Kod boş olamaz'}), 400

    con = _db()
    try:
        # NG-LOT
        if kod.startswith('NG-LOT-'):
            row = con.execute(
                "SELECT nb.batch_kodu, nb.lot_kodu, nb.planlanan_kg, nb.durum, "
                "uv.boyut, rv.ad AS renk_ad, f.ad AS formul_ad "
                "FROM nexgen_uretim_batch nb "
                "JOIN nexgen_uretim_varyant uv ON uv.id=nb.uretim_varyant_id "
                "JOIN nexgen_renk_varyant rv ON rv.id=uv.renk_varyant_id "
                "JOIN nexgen_formul f ON f.id=rv.formul_id "
                "WHERE nb.lot_kodu=?", (kod,)
            ).fetchone()
            if row:
                r = dict(row)
                return jsonify({
                    'ok': True, 'tip': 'URETIM_LOT',
                    'redirect_url': f'/nexgen/tablet/barkod/sonuc?kod={kod}',
                    'ozet': r,
                })

        # NG-PRD
        if kod.startswith('NG-PRD-'):
            row = con.execute(
                "SELECT nb.batch_kodu, nb.lot_kodu, nb.planlanan_kg, nb.durum, "
                "uv.boyut, rv.ad AS renk_ad, f.ad AS formul_ad "
                "FROM nexgen_uretim_batch nb "
                "JOIN nexgen_uretim_varyant uv ON uv.id=nb.uretim_varyant_id "
                "JOIN nexgen_renk_varyant rv ON rv.id=uv.renk_varyant_id "
                "JOIN nexgen_formul f ON f.id=rv.formul_id "
                "WHERE nb.batch_kodu=?", (kod,)
            ).fetchone()
            if row:
                r = dict(row)
                return jsonify({
                    'ok': True, 'tip': 'URETIM_BATCH',
                    'redirect_url': f'/nexgen/tablet/barkod/sonuc?kod={kod}',
                    'ozet': r,
                })

        # AT-
        if kod.startswith('AT-'):
            row = con.execute(
                "SELECT at.test_no, at.test_batch_kg, at.makina, "
                "at.yeni_renk_adi, at.durum, "
                "rv.ad AS renk_ad, f.ad AS formul_ad "
                "FROM nexgen_arge_test at "
                "JOIN nexgen_uretim_varyant uv ON uv.id=at.kaynak_uretim_varyant_id "
                "JOIN nexgen_renk_varyant rv ON rv.id=uv.renk_varyant_id "
                "JOIN nexgen_formul f ON f.id=rv.formul_id "
                "WHERE at.test_no=?", (kod,)
            ).fetchone()
            if row:
                r = dict(row)
                return jsonify({
                    'ok': True, 'tip': 'ARGE',
                    'redirect_url': f'/nexgen/tablet/barkod/sonuc?kod={kod}',
                    'ozet': r,
                })

        return jsonify({'ok': False, 'hata': f'"{kod}" kodu bulunamadı'}), 404
    finally:
        con.close()


@nexgen_bp.route('/tablet/uretim')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_uretim():
    con = _db()
    try:
        varyantlar = _uretime_acik_receteler(con)
        liste = []
        for uv in varyantlar:
            uv_dict = dict(uv)
            kalemler = con.execute("""
                SELECT rk.miktar_kg, rk.stok_kart_id
                FROM nexgen_recete_kalem rk
                WHERE rk.uretim_varyant_id = ? AND rk.aktif = 1
            """, (uv_dict['id'],)).fetchall()
            toplam_kg = round(sum(k['miktar_kg'] for k in kalemler), 3) if kalemler else 0.0
            toplam_mal = 0.0
            for k in kalemler:
                fr = con.execute("""
                    SELECT fiyat, para_birimi, kur, fiyat_try
                    FROM nexgen_hammadde_fiyat
                    WHERE stok_kart_id=? AND aktif=1
                    ORDER BY fiyat_tarihi DESC, id DESC LIMIT 1
                """, (k['stok_kart_id'],)).fetchone()
                if fr:
                    if fr['fiyat_try'] and fr['fiyat_try'] > 0:
                        bp = float(fr['fiyat_try'])
                    elif fr['para_birimi'] == 'TRY':
                        bp = float(fr['fiyat'] or 0)
                    elif fr['kur'] and fr['kur'] > 0:
                        bp = float(fr['fiyat'] or 0) * float(fr['kur'])
                    else:
                        bp = 0.0
                    toplam_mal += float(k['miktar_kg']) * bp
            uv_dict['toplam_kg']   = toplam_kg
            uv_dict['kg_maliyet']  = round(toplam_mal / toplam_kg, 2) if toplam_kg > 0 else 0.0
            liste.append(uv_dict)

        # Açık TASLAK batch'lerden "bugünkü işler" — planlanan KG → kazan sayısı
        bugun = date.today().isoformat()
        acik_isler = con.execute("""
            SELECT nb.id, nb.batch_kodu, nb.planlanan_kg, nb.durum,
                   nb.olusturma_tarihi, nb.notlar,
                   uv.id AS uv_id, uv.boyut, uv.ad AS uv_ad,
                   rv.ad AS renk_ad,
                   f.ad AS formul_ad, f.kod AS formul_kod
            FROM nexgen_uretim_batch nb
            JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            WHERE nb.durum IN ('TASLAK','HAZIR')
              AND substr(nb.olusturma_tarihi,1,10) = ?
            ORDER BY nb.id ASC
        """, (bugun,)).fetchall()

        # Üretim planından PLANLANDI kayıtlarını da ekle (bugünkü + tarih uygun olanlar)
        plan_isler = con.execute("""
            SELECT np.id AS plan_id, np.plan_kodu, np.planlanan_kg,
                   np.durum AS plan_durum, np.plan_tarihi, np.notlar,
                   np.oncelik_sira,
                   uv.id AS uv_id, uv.boyut, uv.ad AS uv_ad,
                   rv.ad AS renk_ad,
                   f.ad AS formul_ad, f.kod AS formul_kod
            FROM nexgen_uretim_plan np
            JOIN nexgen_uretim_varyant uv ON uv.id = np.uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            WHERE np.durum = 'PLANLANDI'
              AND np.plan_tarihi <= ?
            ORDER BY np.oncelik_sira ASC, np.id ASC
        """, (bugun,)).fetchall()

        # Her iş için batch_kg (reçeteden) ve kazan sayısı hesapla
        acik_isler_liste = []
        for is_ in acik_isler:
            is_dict = dict(is_)
            is_dict['kaynak'] = 'BATCH'
            # Reçete batch KG
            kalemler2 = con.execute("""
                SELECT SUM(miktar_kg) as toplam
                FROM nexgen_recete_kalem
                WHERE uretim_varyant_id=? AND aktif=1
            """, (is_dict['uv_id'],)).fetchone()
            batch_kg = round(float(kalemler2['toplam'] or 0), 3) if kalemler2 else 0.0
            is_dict['batch_kg'] = batch_kg
            if batch_kg > 0 and is_dict['planlanan_kg'] > 0:
                q = is_dict['planlanan_kg'] / batch_kg
                is_dict['kazan_sayisi'] = int(q) if q == int(q) else int(q) + 1
            else:
                is_dict['kazan_sayisi'] = 0
            acik_isler_liste.append(is_dict)

        # Plan kayıtlarını da ekle (PLANLANDI + tarih uygun)
        for p_ in plan_isler:
            p_dict = dict(p_)
            p_dict['kaynak'] = 'PLAN'
            p_dict['batch_kodu'] = None
            p_dict['durum'] = p_dict.get('plan_durum', 'PLANLANDI')
            kalemler3 = con.execute("""
                SELECT SUM(miktar_kg) as toplam
                FROM nexgen_recete_kalem
                WHERE uretim_varyant_id=? AND aktif=1
            """, (p_dict['uv_id'],)).fetchone()
            batch_kg = round(float(kalemler3['toplam'] or 0), 3) if kalemler3 else 0.0
            p_dict['batch_kg'] = batch_kg
            if batch_kg > 0 and p_dict['planlanan_kg'] > 0:
                q = p_dict['planlanan_kg'] / batch_kg
                p_dict['kazan_sayisi'] = int(q) if q == int(q) else int(q) + 1
            else:
                p_dict['kazan_sayisi'] = 0
            acik_isler_liste.append(p_dict)
    finally:
        con.close()

    return render_template(
        'nexgen/tablet_uretim.html',
        active='nexgen',
        varyantlar=liste,
        acik_isler=acik_isler_liste,
    )


@nexgen_bp.route('/tablet/arge')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_arge():
    con = _db()
    try:
        # Kaynak reçetesi olan tüm aktif varyantlar (TASLAK dahil — AR-GE kaynak seçimi)
        varyantlar = con.execute("""
            SELECT uv.id, uv.boyut, uv.ad, uv.recete_durum,
                   rv.ad AS renk_ad,
                   f.id AS formul_id, f.kod AS formul_kod, f.ad AS formul_ad,
                   COUNT(rk.id) AS kalem_say,
                   SUM(rk.miktar_kg) AS toplam_kg
            FROM nexgen_uretim_varyant uv
            JOIN nexgen_renk_varyant rv ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f        ON f.id  = rv.formul_id
            JOIN nexgen_recete_kalem rk
                   ON rk.uretim_varyant_id = uv.id AND rk.aktif = 1
            WHERE uv.aktif = 1 AND rv.aktif = 1 AND f.aktif = 1
              AND uv.recete_durum != 'PASIF'
            GROUP BY uv.id
            HAVING kalem_say > 0
            ORDER BY f.kod, rv.ad, uv.boyut
        """).fetchall()
        varyantlar = [dict(v) for v in varyantlar]
        for v in varyantlar:
            v['toplam_kg'] = round(float(v['toplam_kg'] or 0), 3)

        son_numuneler = con.execute("""
            SELECT t.id, t.test_no, t.durum, t.yeni_renk_adi, t.test_batch_kg,
                   t.olusturma_tarihi, t.shore_degeri,
                   f.ad AS formul_ad,
                   rv.ad AS kaynak_renk_ad,
                   ku.KullaniciAdi AS olusturan_ad,
                   c.unvan AS cari_unvan
            FROM nexgen_arge_test t
            JOIN nexgen_uretim_varyant uv ON uv.id = t.kaynak_uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            LEFT JOIN sistem_kullanici ku ON ku.Id = t.olusturan_id
            LEFT JOIN nexgen_cari c       ON c.id  = t.cari_id
            WHERE t.aktif = 1
            ORDER BY t.id DESC
            LIMIT 5
        """).fetchall()
        son_numuneler = [dict(r) for r in son_numuneler]
        for n in son_numuneler:
            ozet, fazla = _arge_test_boya_ozet(con, n['id'])
            n['boya_ozet'] = ozet
            n['boya_fazla'] = fazla

        boya_stok = con.execute("""
            SELECT id, kod, ad, kategori
            FROM nexgen_stok_kart
            WHERE aktif = 1 AND kategori = 'BOYA'
            ORDER BY kod, ad
        """).fetchall()
        boya_stok = [dict(r) for r in boya_stok]
        for s in boya_stok:
            s['etiket'] = _stok_boya_dropdown_etiket(s.get('kod'), s.get('ad'))

        cariler = con.execute(
            "SELECT id, cari_kod, unvan FROM nexgen_cari WHERE aktif=1 ORDER BY unvan"
        ).fetchall()
        cariler = [dict(c) for c in cariler]

        bugun_numune = con.execute("""
            SELECT COUNT(*) AS c FROM nexgen_arge_test
            WHERE aktif = 1 AND date(olusturma_tarihi) = date('now')
        """).fetchone()['c']
    finally:
        con.close()

    return render_template(
        'nexgen/tablet_arge.html',
        active='nexgen',
        varyantlar=varyantlar,
        son_numuneler=son_numuneler,
        boya_stok=boya_stok,
        cariler=cariler,
        bugun_numune=bugun_numune,
        kullanici_ad=_kullanici_ad(),
    )


@nexgen_bp.route('/api/tablet/uretim-onizleme', methods=['POST'])
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def api_tablet_uretim_onizleme():
    """Planlanan KG için reçete ölçekler, stok yeterliliğini döner.
    Stok hareketi YAZMAZ — sadece hesap ve okuma.

    POST JSON:
        uretim_varyant_id : int
        planlanan_kg      : float
    """
    d = request.get_json(silent=True) or {}
    uv_id     = d.get('uretim_varyant_id')
    planlanan = d.get('planlanan_kg')

    if not uv_id or not planlanan:
        return jsonify({"ok": False, "hata": "uretim_varyant_id ve planlanan_kg zorunlu"}), 400
    try:
        planlanan = float(planlanan)
        if planlanan <= 0:
            return jsonify({"ok": False, "hata": "planlanan_kg sıfırdan büyük olmalı"}), 400
    except Exception:
        return jsonify({"ok": False, "hata": "Geçersiz planlanan_kg"}), 400

    con = _db()
    try:
        uv = con.execute(
            "SELECT uv.id, uv.boyut, uv.recete_durum, uv.ad AS uv_ad,"
            " rv.ad AS renk_ad, f.ad AS formul_ad, f.kod AS formul_kod"
            " FROM nexgen_uretim_varyant uv"
            " JOIN nexgen_renk_varyant rv ON rv.id=uv.renk_varyant_id"
            " JOIN nexgen_formul f ON f.id=rv.formul_id"
            " WHERE uv.id=? AND uv.aktif=1",
            (uv_id,)
        ).fetchone()
        if not uv:
            return jsonify({"ok": False, "hata": "Üretim varyantı bulunamadı"}), 404
        if uv['recete_durum'] != 'URETIME_ACIK':
            return jsonify({
                "ok": False,
                "hata": f"Sadece URETIME_ACIK reçeteler kullanılabilir. Mevcut: {uv['recete_durum']}"
            }), 400

        # Reçete kalemleri
        kalemler_db = con.execute("""
            SELECT rk.id, rk.stok_kart_id, rk.miktar_kg, rk.sira,
                   sk.ad AS stok_ad, sk.birim, sk.kod AS stok_kod
            FROM nexgen_recete_kalem rk
            JOIN nexgen_stok_kart sk ON sk.id = rk.stok_kart_id
            WHERE rk.uretim_varyant_id = ? AND rk.aktif = 1
            ORDER BY rk.sira
        """, (uv_id,)).fetchall()

        if not kalemler_db:
            return jsonify({"ok": False, "hata": "Bu varyanta ait aktif reçete kalemi bulunamadı"}), 400

        # Batch KG = kalem miktarları toplamı
        batch_kg = round(sum(float(k['miktar_kg']) for k in kalemler_db), 3)
        if batch_kg <= 0:
            return jsonify({"ok": False, "hata": "Reçete batch KG sıfır olamaz"}), 400

        carpan = round(planlanan / batch_kg, 6)

        # Her kalem için ölçekle ve stok kontrol et
        tum_yeterli = True
        kalemler_sonuc = []
        for k in kalemler_db:
            recete_kg  = round(float(k['miktar_kg']), 3)
            gerekli_kg = round(recete_kg * carpan, 3)
            mevcut_kg  = _mevcut_stok(con, k['stok_kart_id'])
            fark_kg    = round(mevcut_kg - gerekli_kg, 3)
            yeterli    = mevcut_kg >= gerekli_kg

            if not yeterli:
                tum_yeterli = False

            kalemler_sonuc.append({
                "stok_kart_id":    k['stok_kart_id'],
                "stok_kod":        k['stok_kod'] or '',
                "hammadde_adi":    k['stok_ad'],
                "birim":           k['birim'] or 'KG',
                "recete_miktar_kg": recete_kg,
                "gerekli_kg":      gerekli_kg,
                "mevcut_stok_kg":  mevcut_kg,
                "fark_kg":         fark_kg,
                "yeterli":         yeterli,
                "durum":           "YETERLİ" if yeterli else "EKSİK",
            })

        return jsonify({
            "ok":          True,
            "uv_id":       uv_id,
            "formul_ad":   uv['formul_ad'],
            "renk_ad":     uv['renk_ad'],
            "boyut":       uv['boyut'],
            "batch_kg":    batch_kg,
            "planlanan_kg": round(planlanan, 3),
            "carpan":      carpan,
            "toplam_kalem": len(kalemler_sonuc),
            "yeterli_mi":  tum_yeterli,
            "kalemler":    kalemler_sonuc,
        })
    except Exception as e:
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/tablet/uretim-kodu-olustur', methods=['POST'])
@yetki_gerekli('nexgen.tablet.uretim', 'can_uretim')
def api_tablet_uretim_kodu():
    """Üretim batch kodu oluşturur. Stok hareketi yapmaz."""
    d = request.get_json(silent=True) or {}
    uv_id      = d.get('uretim_varyant_id')
    planlanan  = d.get('planlanan_kg')
    notlar     = (d.get('notlar') or '').strip() or None

    if not uv_id or not planlanan:
        return jsonify({"ok": False, "hata": "uretim_varyant_id ve planlanan_kg zorunlu"}), 400
    try:
        planlanan = float(planlanan)
        if planlanan <= 0:
            return jsonify({"ok": False, "hata": "Planlanan KG sıfırdan büyük olmalı"}), 400
    except Exception:
        return jsonify({"ok": False, "hata": "Geçersiz planlanan_kg"}), 400

    import math as _math
    con = _db()
    try:
        uv = con.execute(
            "SELECT uv.id, uv.boyut, uv.recete_durum, rv.ad AS renk_ad, f.ad AS formul_ad "
            "FROM nexgen_uretim_varyant uv "
            "JOIN nexgen_renk_varyant rv ON rv.id=uv.renk_varyant_id "
            "JOIN nexgen_formul f ON f.id=rv.formul_id "
            "WHERE uv.id=? AND uv.aktif=1",
            (uv_id,)
        ).fetchone()
        if not uv:
            return jsonify({"ok": False, "hata": "Üretim varyantı bulunamadı"}), 404
        if uv['recete_durum'] != 'URETIME_ACIK':
            return jsonify({"ok": False, "hata": f"Sadece URETIME_ACIK reçeteler kullanılabilir. Mevcut: {uv['recete_durum']}"}), 400

        batch_kodu = _batch_kodu_uret(con)
        lot_kodu   = _lot_kodu_uret(con)
        uid = _kullanici_id()
        con.execute(
            "INSERT INTO nexgen_uretim_batch(batch_kodu, lot_kodu, uretim_varyant_id, planlanan_kg, durum, olusturan_id, notlar) "
            "VALUES(?,?,?,?,'HAZIR',?,?)",
            (batch_kodu, lot_kodu, uv_id, round(planlanan, 3), uid, notlar)
        )

        # Alt emirleri otomatik oluştur (wizard'dan açılan batch'ler için)
        alt_emir_sayisi = 0
        if _parca_tablosu_var(con):
            batch_row = con.execute(
                "SELECT id FROM nexgen_uretim_batch WHERE batch_kodu=?",
                (batch_kodu,)
            ).fetchone()
            batch_id = batch_row['id'] if batch_row else None
            formul_batch_kg = _formul_batch_kg_hesapla(con, uv_id)
            if batch_id and formul_batch_kg > 0:
                alt_emir_sayisi = _math.ceil(planlanan / formul_batch_kg)
                parca_cols = [c['name'] for c in con.execute(
                    "PRAGMA table_info(nexgen_uretim_parca)"
                ).fetchall()]
                has_formul_kg_col = 'formul_batch_kg' in parca_cols
                for no in range(1001, 1001 + alt_emir_sayisi):
                    if has_formul_kg_col:
                        con.execute("""
                            INSERT INTO nexgen_uretim_parca
                                (batch_id, batch_kodu, parca_no,
                                 hedef_kg, formul_batch_kg, uretilen_kg,
                                 durum, operator_id)
                            VALUES (?, ?, ?, ?, ?, 0, 'HAZIR', ?)
                        """, (batch_id, batch_kodu, no,
                              round(formul_batch_kg, 3),
                              round(formul_batch_kg, 3), uid))
                    else:
                        con.execute("""
                            INSERT INTO nexgen_uretim_parca
                                (batch_id, batch_kodu, parca_no,
                                 hedef_kg, uretilen_kg, durum, operator_id)
                            VALUES (?, ?, ?, ?, 0, 'HAZIR', ?)
                        """, (batch_id, batch_kodu, no,
                              round(formul_batch_kg, 3), uid))

        con.commit()
        return jsonify({
            "ok": True,
            "batch_kodu":      batch_kodu,
            "lot_kodu":        lot_kodu,
            "formul_ad":       uv['formul_ad'],
            "renk_ad":         uv['renk_ad'],
            "boyut":           uv['boyut'],
            "planlanan_kg":    round(planlanan, 3),
            "alt_emir_sayisi": alt_emir_sayisi,
        })
    except Exception as e:
        con.rollback()
        return jsonify({"ok": False, "hata": str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/tablet/kod/<batch_kodu>')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_kod_goster(batch_kodu):
    """Oluşturulan batch kodunu büyük ekranda göster."""
    con = _db()
    try:
        batch = con.execute("""
            SELECT nb.*, uv.boyut,
                   rv.ad AS renk_ad,
                   f.ad AS formul_ad, f.kod AS formul_kod
            FROM nexgen_uretim_batch nb
            JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            WHERE nb.batch_kodu = ?
        """, (batch_kodu,)).fetchone()
        if not batch:
            from flask import abort
            abort(404)
        batch = dict(batch)
    finally:
        con.close()

    return render_template(
        'nexgen/tablet_kod.html',
        active='nexgen',
        batch=batch,
        kod_tipi='URETIM',
    )


@nexgen_bp.route('/tablet/arge-kod/<test_no>')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_arge_kod_goster(test_no):
    """AR-GE test kodunu büyük ekranda göster."""
    con = _db()
    try:
        test = con.execute("""
            SELECT at.*, uv.boyut,
                   rv.ad AS renk_ad,
                   f.ad AS formul_ad, f.kod AS formul_kod,
                   ku.KullaniciAdi AS olusturan_ad
            FROM nexgen_arge_test at
            JOIN nexgen_uretim_varyant uv ON uv.id = at.kaynak_uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            LEFT JOIN sistem_kullanici ku ON ku.Id  = at.olusturan_id
            WHERE at.test_no = ?
        """, (test_no,)).fetchone()
        if not test:
            from flask import abort
            abort(404)
        test = dict(test)
    finally:
        con.close()

    return render_template(
        'nexgen/tablet_kod.html',
        active='nexgen',
        batch=None,
        test=test,
        kod_tipi='ARGE',
    )


# ─────────────────────────────────────────────────────────────
# NEXGEN FAZ-5C-0 — LOT / BARKOD ETİKET SAYFALARI
# KURAL: Stok hareketi yok. Sadece etiket görüntüleme + yazdırma.
# ─────────────────────────────────────────────────────────────

@nexgen_bp.route('/tablet/etiket/uretim/<batch_kodu>')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_etiket_uretim(batch_kodu):
    """Üretim batch için LOT / barkod etiket önizleme sayfası."""
    con = _db()
    try:
        _ecols = [c['name'] for c in con.execute(
            "PRAGMA table_info(nexgen_uretim_batch)"
        ).fetchall()]
        if 'plan_id' in _ecols:
            batch = con.execute("""
                SELECT nb.id, nb.batch_kodu, nb.lot_kodu, nb.planlanan_kg,
                       nb.durum, nb.olusturma_tarihi, nb.notlar,
                       uv.boyut,
                       rv.ad AS renk_ad,
                       f.ad AS formul_ad,
                       ku.KullaniciAdi AS olusturan_ad,
                       np.musteri_adi, np.siparis_no
                FROM nexgen_uretim_batch nb
                JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f          ON f.id  = rv.formul_id
                LEFT JOIN sistem_kullanici ku ON ku.Id  = nb.olusturan_id
                LEFT JOIN nexgen_uretim_plan np ON np.id = nb.plan_id
                WHERE nb.batch_kodu = ?
            """, (batch_kodu,)).fetchone()
        else:
            batch = con.execute("""
                SELECT nb.id, nb.batch_kodu, nb.lot_kodu, nb.planlanan_kg,
                       nb.durum, nb.olusturma_tarihi, nb.notlar,
                       uv.boyut,
                       rv.ad AS renk_ad,
                       f.ad AS formul_ad,
                       ku.KullaniciAdi AS olusturan_ad,
                       NULL AS musteri_adi, NULL AS siparis_no
                FROM nexgen_uretim_batch nb
                JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f          ON f.id  = rv.formul_id
                LEFT JOIN sistem_kullanici ku ON ku.Id  = nb.olusturan_id
                WHERE nb.batch_kodu = ?
            """, (batch_kodu,)).fetchone()
        if not batch:
            from flask import abort
            abort(404)
        batch = dict(batch)

        # lot_kodu yoksa (eski kayıt) — anında oluştur ve kaydet
        if not batch.get('lot_kodu'):
            yeni_lot = _lot_kodu_uret(con)
            con.execute(
                "UPDATE nexgen_uretim_batch SET lot_kodu=? WHERE batch_kodu=?",
                (yeni_lot, batch_kodu)
            )
            con.commit()
            batch['lot_kodu'] = yeni_lot

        # Kazan hesabı
        kalemler = con.execute(
            "SELECT SUM(miktar_kg) AS toplam FROM nexgen_recete_kalem "
            "WHERE uretim_varyant_id=("
            "  SELECT uretim_varyant_id FROM nexgen_uretim_batch WHERE batch_kodu=?"
            ") AND aktif=1",
            (batch_kodu,)
        ).fetchone()
        batch_kg = float((kalemler['toplam'] or 0) if kalemler else 0)
        if batch_kg > 0 and batch['planlanan_kg'] > 0:
            q = batch['planlanan_kg'] / batch_kg
            batch['kazan_sayisi'] = int(q) if q == int(q) else int(q) + 1
            batch['batch_kg']     = round(batch_kg, 3)
        else:
            batch['kazan_sayisi'] = 0
            batch['batch_kg']     = batch_kg

    finally:
        con.close()

    from flask import session as _session
    operator_ad = _session.get('kullanici_ad') or _session.get('ad') or '—'
    from datetime import date as _date
    return render_template(
        'nexgen/tablet_etiket_uretim.html',
        active='nexgen',
        batch=batch,
        operator_ad=operator_ad,
        bugun=_date.today().isoformat(),
    )


@nexgen_bp.route('/tablet/etiket/arge/<test_no>')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def tablet_etiket_arge(test_no):
    """AR-GE testi için barkod etiket önizleme sayfası."""
    con = _db()
    try:
        test = con.execute("""
            SELECT at.id, at.test_no, at.test_batch_kg, at.makina,
                   at.yeni_renk_adi, at.durum, at.olusturma_tarihi,
                   uv.boyut,
                   rv.ad AS renk_ad,
                   f.ad AS formul_ad,
                   ku.KullaniciAdi AS olusturan_ad
            FROM nexgen_arge_test at
            JOIN nexgen_uretim_varyant uv ON uv.id = at.kaynak_uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            LEFT JOIN sistem_kullanici ku ON ku.Id  = at.olusturan_id
            WHERE at.test_no = ?
        """, (test_no,)).fetchone()
        if not test:
            from flask import abort
            abort(404)
        test = dict(test)
    finally:
        con.close()

    from flask import session as _session
    operator_ad = _session.get('kullanici_ad') or _session.get('ad') or '—'
    from datetime import date as _date
    return render_template(
        'nexgen/tablet_etiket_arge.html',
        active='nexgen',
        test=test,
        operator_ad=operator_ad,
        bugun=_date.today().isoformat(),
    )


# ─────────────────────────────────────────────────────────────
# NEXGEN FAZ-5C-2 — TSPL BARKOD ETİKET KOMUT ÇIKTISI
# KURAL: Stok hareketi yok. Sadece TSPL metin çıktısı.
# Desteklenen cihazlar: TSC, Xprinter, Godex (TSPL/TSPL2)
# Etiket boyutu: 100×60 mm
# ─────────────────────────────────────────────────────────────

def _tspl_uretim(b, operator_ad):
    """100×60 mm TSPL komutu üretir — üretim LOT etiketi.

    b: dict ile batch + join alanları (formul_ad, renk_ad, boyut,
       lot_kodu, batch_kodu, planlanan_kg, olusturma_tarihi)
    """
    from datetime import date as _date
    bugun = _date.today().strftime('%d.%m.%Y')
    lot    = b.get('lot_kodu') or b.get('batch_kodu') or '—'
    batch  = b.get('batch_kodu') or '—'
    formul = (b.get('formul_ad') or '').upper()[:24]
    renk   = (b.get('renk_ad') or '').upper()[:24]
    boyut_raw = b.get('boyut') or ''
    boyut  = 'BUYUK BOY' if boyut_raw == 'LARGE' else ('KUCUK BOY' if boyut_raw == 'SMALL' else boyut_raw.upper())
    kg     = str(int(b.get('planlanan_kg') or 0)) + ' KG'
    op     = (operator_ad or '').upper()[:20]

    lines = [
        "SIZE 100 mm,60 mm",
        "GAP 3 mm,0 mm",
        "DIRECTION 0",
        "REFERENCE 0,0",
        "OFFSET 0 mm",
        "SET PEEL OFF",
        "SET TEAR ON",
        "CLS",
        # ── Başlık bandı (siyah dolgu) ──
        "BAR 0,0,800,40",
        f'REVERSE 10,6,780,30,"ARIAL.TTF",0,8,8,"SOLARIZ NEXGEN - URETIM ETIKETI"',
        # ── LOT kodu (büyük font) ──
        f'TEXT 10,48,"ARIAL.TTF",0,16,16,"{lot}"',
        # ── Barkod (Code128) ──
        f'BARCODE 10,100,"128",70,1,0,3,3,"{lot}"',
        # ── Bilgi satırları ──
        f'TEXT 10,182,"ARIAL.TTF",0,7,7,"FORMUL: {formul}"',
        f'TEXT 10,196,"ARIAL.TTF",0,7,7,"RENK  : {renk}"',
        f'TEXT 10,210,"ARIAL.TTF",0,7,7,"BOYUT : {boyut}    KG: {kg}"',
        f'TEXT 10,224,"ARIAL.TTF",0,7,7,"BATCH : {batch}"',
        # ── Alt şerit ──
        "BAR 0,238,800,2",
        f'TEXT 10,244,"ARIAL.TTF",0,6,6,"Tarih: {bugun}    Op: {op}"',
        "PRINT 1,1",
    ]
    return "\r\n".join(lines)


def _tspl_arge(t, operator_ad):
    """100×60 mm TSPL komutu üretir — AR-GE test etiketi."""
    from datetime import date as _date
    bugun    = _date.today().strftime('%d.%m.%Y')
    test_no  = t.get('test_no') or '—'
    formul   = (t.get('formul_ad') or '').upper()[:24]
    renk     = (t.get('renk_ad') or '').upper()[:20]
    yeni_renk = (t.get('yeni_renk_adi') or '—').upper()[:20]
    kg       = str(t.get('test_batch_kg') or '—') + ' KG'
    makina   = (t.get('makina') or '—').upper()[:20]
    op       = (operator_ad or '').upper()[:20]

    lines = [
        "SIZE 100 mm,60 mm",
        "GAP 3 mm,0 mm",
        "DIRECTION 0",
        "REFERENCE 0,0",
        "OFFSET 0 mm",
        "SET PEEL OFF",
        "SET TEAR ON",
        "CLS",
        # ── Başlık bandı (mor/lacivert) ──
        "BAR 0,0,800,40",
        f'REVERSE 10,6,780,30,"ARIAL.TTF",0,8,8,"SOLARIZ NEXGEN - AR-GE TEST"',
        # ── Test kodu ──
        f'TEXT 10,48,"ARIAL.TTF",0,14,14,"{test_no}"',
        # ── Barkod ──
        f'BARCODE 10,96,"128",65,1,0,3,3,"{test_no}"',
        # ── Bilgi satırları ──
        f'TEXT 10,174,"ARIAL.TTF",0,7,7,"FORMUL    : {formul}"',
        f'TEXT 10,188,"ARIAL.TTF",0,7,7,"YENİ RENK : {yeni_renk}   KG: {kg}"',
        f'TEXT 10,202,"ARIAL.TTF",0,7,7,"MAKİNA    : {makina}"',
        # ── Alt şerit ──
        "BAR 0,218,800,2",
        f'TEXT 10,224,"ARIAL.TTF",0,6,6,"Tarih: {bugun}    Hazirlayan: {op}"',
        "PRINT 1,1",
    ]
    return "\r\n".join(lines)


@nexgen_bp.route('/api/etiket/uretim/<batch_kodu>/tspl')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def api_etiket_uretim_tspl(batch_kodu):
    """100×60 mm TSPL komut dosyası üretir — TSC/Xprinter/Godex için.

    Stok hareketi yapılmaz. Sadece metin çıktısı.
    Content-Type: text/plain; charset=utf-8
    Content-Disposition: attachment; filename=etiket_<batch_kodu>.tspl
    """
    con = _db()
    try:
        batch = con.execute("""
            SELECT nb.batch_kodu, nb.lot_kodu, nb.planlanan_kg, nb.olusturma_tarihi,
                   uv.boyut,
                   rv.ad AS renk_ad,
                   f.ad AS formul_ad,
                   ku.KullaniciAdi AS olusturan_ad
            FROM nexgen_uretim_batch nb
            JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            LEFT JOIN sistem_kullanici ku ON ku.Id  = nb.olusturan_id
            WHERE nb.batch_kodu = ?
        """, (batch_kodu,)).fetchone()
        if not batch:
            abort(404)
        b = dict(batch)
    finally:
        con.close()

    operator_ad = session.get('kullanici_ad') or session.get('ad') or b.get('olusturan_ad') or '—'
    tspl = _tspl_uretim(b, operator_ad)

    return Response(
        tspl,
        mimetype='text/plain',
        headers={
            'Content-Disposition': f'attachment; filename="etiket_{batch_kodu}.tspl"',
            'Content-Type': 'text/plain; charset=utf-8',
        }
    )


@nexgen_bp.route('/api/etiket/arge/<test_no>/tspl')
@yetki_gerekli('nexgen.tablet.view', 'can_view')
def api_etiket_arge_tspl(test_no):
    """100×60 mm TSPL komut dosyası üretir — AR-GE etiketi.

    Stok hareketi yapılmaz. Sadece metin çıktısı.
    """
    con = _db()
    try:
        test = con.execute("""
            SELECT at.test_no, at.test_batch_kg, at.makina,
                   at.yeni_renk_adi, at.durum, at.olusturma_tarihi,
                   rv.ad AS renk_ad,
                   f.ad AS formul_ad,
                   ku.KullaniciAdi AS olusturan_ad
            FROM nexgen_arge_test at
            JOIN nexgen_uretim_varyant uv ON uv.id = at.kaynak_uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            LEFT JOIN sistem_kullanici ku ON ku.Id  = at.olusturan_id
            WHERE at.test_no = ?
        """, (test_no,)).fetchone()
        if not test:
            abort(404)
        t = dict(test)
    finally:
        con.close()

    operator_ad = session.get('kullanici_ad') or session.get('ad') or t.get('olusturan_ad') or '—'
    tspl = _tspl_arge(t, operator_ad)

    return Response(
        tspl,
        mimetype='text/plain',
        headers={
            'Content-Disposition': f'attachment; filename="etiket_{test_no}.tspl"',
            'Content-Type': 'text/plain; charset=utf-8',
        }
    )


# ─────────────────────────────────────────────────────────────
# NEXGEN FAZ-5D — ÜRETİM PLAN / İŞ KUYRUĞU
# KURAL: Ana Planlama modülüne dokunma. Korgun bağlantısı yok.
#        Stok hareketi yok. Sadece plan CRUD + tablet bağlantısı.
# ─────────────────────────────────────────────────────────────

def _plan_kodu_uret(con):
    """NP-YYYY-NNNNN formatında benzersiz plan kodu üretir."""
    import datetime
    yil = datetime.datetime.now().year
    son = con.execute(
        "SELECT plan_kodu FROM nexgen_uretim_plan "
        "WHERE plan_kodu LIKE ? ORDER BY id DESC LIMIT 1",
        (f"NP-{yil}-%",)
    ).fetchone()
    if son:
        try:
            son_no = int(son['plan_kodu'].split('-')[-1])
        except Exception:
            son_no = 0
    else:
        son_no = 0
    return f"NP-{yil}-{son_no + 1:05d}"


def _nexgen_siparis_no_uret(con):
    """NSP-YYYY-NNNNN formatında benzersiz sipariş no üretir.

    Kullanıcı sipariş no girmezse otomatik çağrılır.
    LARGE+SMALL aynı request'ten geliyorsa JS aynı no'yu tekrar gönderir.
    """
    import datetime
    yil = datetime.datetime.now().year
    prefix = f"NSP-{yil}-%"
    son_no = 0
    for tbl in ('nexgen_planlama_siparis', 'nexgen_uretim_plan'):
        try:
            row = con.execute(
                f"SELECT siparis_no FROM {tbl} "
                "WHERE siparis_no LIKE ? ORDER BY id DESC LIMIT 1",
                (prefix,),
            ).fetchone()
            if row and row['siparis_no']:
                try:
                    n = int(row['siparis_no'].split('-')[-1])
                    son_no = max(son_no, n)
                except Exception:
                    pass
        except Exception:
            pass
    return f"NSP-{yil}-{son_no + 1:05d}"


def _tablet_ana_veri(con):
    """Tablet ana ekran — üretim plan ile aynı aktif plan/batch listesi."""
    from datetime import date as _date
    bugun = _date.today().isoformat()
    _AKTIF_BATCH = ('TASLAK', 'HAZIR', 'DEVAM', 'BEKLEME')
    planlar = [dict(p) for p in _plan_liste_sorgu(con, sadece_aktif=True)]

    devam_eden = []
    plan_isler = []
    seen_batches = set()

    for pl in planlar:
        bk = pl.get('batch_kodu')
        bd = pl.get('batch_durum')
        if bk and bd in _AKTIF_BATCH and bk not in seen_batches:
            seen_batches.add(bk)
            b = con.execute("""
                SELECT nb.batch_kodu, nb.lot_kodu, nb.planlanan_kg, nb.durum,
                       nb.olusturma_tarihi, nb.uretim_varyant_id,
                       uv.ad AS uv_ad, uv.boyut,
                       rv.ad AS renk_ad, f.ad AS formul_ad
                FROM nexgen_uretim_batch nb
                JOIN nexgen_uretim_varyant uv ON uv.id = nb.uretim_varyant_id
                JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
                JOIN nexgen_formul f          ON f.id  = rv.formul_id
                WHERE nb.batch_kodu = ?
            """, (bk,)).fetchone()
            if b:
                row = dict(b)
                row['musteri_adi'] = pl.get('musteri_adi')
                row['siparis_no'] = pl.get('siparis_no') or pl.get('plan_kodu')
                rf = _batch_plan_rf_bilgi(
                    con, batch_kodu=bk,
                    uretim_varyant_id=row.get('uretim_varyant_id') or pl.get('uv_id'),
                )
                if rf:
                    row['rf_kod'] = rf['rf_kod']
                    row['rf_renk_ad'] = rf.get('renk_ad')
                else:
                    row['rf_kod'] = None
                    row['rf_renk_ad'] = None
                devam_eden.append(row)
        elif pl['durum'] == 'PLANLANDI' and (pl.get('plan_tarihi') or '') <= bugun and not bk:
            pi = {
                'plan_id': pl['id'],
                'plan_kodu': pl['plan_kodu'],
                'planlanan_kg': pl['planlanan_kg'],
                'plan_durum': pl['durum'],
                'oncelik_sira': pl.get('oncelik_sira'),
                'notlar': pl.get('notlar'),
                'musteri_adi': pl.get('musteri_adi'),
                'uv_id': pl['uv_id'],
                'boyut': pl['boyut'],
                'renk_ad': pl['renk_ad'],
                'formul_ad': pl['formul_ad'],
            }
            rf = _plan_rf_bilgi(con, rf_renk_id=pl.get('rf_renk_id'), uretim_varyant_id=pl['uv_id'])
            if rf:
                pi['rf_kod'] = rf['rf_kod']
                pi['rf_renk_ad'] = rf.get('renk_ad')
            plan_isler.append(pi)

    plan_isler.sort(key=lambda p: (p.get('oncelik_sira') or 999, p.get('plan_id') or 0))
    devam_eden.sort(key=lambda b: b.get('olusturma_tarihi') or '', reverse=True)
    devam_eden = devam_eden[:20]
    plan_isler = plan_isler[:20]

    if _parca_tablosu_var(con):
        for b in devam_eden:
            bk = b['batch_kodu']
            sayac = con.execute("""
                SELECT COUNT(*) AS toplam,
                       SUM(CASE WHEN durum='BITTI' THEN 1 ELSE 0 END) AS biten,
                       SUM(CASE WHEN durum='BITTI' THEN uretilen_kg ELSE 0 END) AS uretilen
                FROM nexgen_uretim_parca WHERE batch_kodu=?
            """, (bk,)).fetchone()
            b['toplam_alt_emir'] = sayac['toplam'] or 0
            b['biten_alt_emir']  = sayac['biten']  or 0
            _batch_kg_rf_enrich(con, b, liste_modu=True)
    else:
        for b in devam_eden:
            b['toplam_alt_emir'] = None
            b['biten_alt_emir']  = None
            b['uretilen_kg']     = 0.0
            b['kalan_kg']        = float(b['planlanan_kg'])
            _batch_kg_rf_enrich(con, b, liste_modu=True)

    return devam_eden, plan_isler


def _plan_cari_kolonu_var(con):
    return 'cari_id' in [
        c['name'] for c in con.execute(
            "PRAGMA table_info(nexgen_uretim_plan)"
        ).fetchall()
    ]


def _plan_rf_renk_kolonu_var(con):
    return 'rf_renk_id' in [
        c['name'] for c in con.execute(
            "PRAGMA table_info(nexgen_uretim_plan)"
        ).fetchall()
    ]


def _plan_termin_kolonu_var(con):
    return 'termin_tarihi' in [
        c['name'] for c in con.execute(
            "PRAGMA table_info(nexgen_uretim_plan)"
        ).fetchall()
    ]


def _plan_planlama_siparis_kolonu_var(con):
    return 'planlama_siparis_id' in [
        c['name'] for c in con.execute(
            "PRAGMA table_info(nexgen_uretim_plan)"
        ).fetchall()
    ]


def _planlama_siparis_tablosu_var(con):
    return con.execute(
        "SELECT name FROM sqlite_master WHERE type='table' "
        "AND name='nexgen_planlama_siparis'"
    ).fetchone() is not None


def _planlama_siparis_olustur(con, siparis_no, cari_id, cari_unvan,
                                termin_tarihi=None, notlar=None,
                                talep_referansi=None, olusturan_id=None,
                                durum='TALEP'):
    """Yeni planlama sipariş header kaydı oluşturur."""
    cur = con.execute("""
        INSERT INTO nexgen_planlama_siparis
            (siparis_no, cari_id, cari_unvan, termin_tarihi, talep_referansi,
             durum, notlar, olusturan_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        siparis_no, cari_id, cari_unvan, termin_tarihi, talep_referansi,
        durum, notlar, olusturan_id,
    ))
    return cur.lastrowid


def _planlama_siparis_bul_veya_olustur(con, siparis_no, cari_id, cari_unvan,
                                         termin_tarihi=None, notlar=None,
                                         olusturan_id=None):
    """siparis_no ile header bulur; yoksa oluşturur."""
    if not _planlama_siparis_tablosu_var(con):
        return None
    row = con.execute(
        "SELECT id FROM nexgen_planlama_siparis WHERE siparis_no=?",
        (siparis_no,),
    ).fetchone()
    if row:
        return row['id']
    return _planlama_siparis_olustur(
        con, siparis_no, cari_id, cari_unvan,
        termin_tarihi=termin_tarihi, notlar=notlar,
        olusturan_id=olusturan_id,
    )


def _plan_cari_select_sql(con):
    """Plan sorgularında cari gösterim: FK unvan, yoksa musteri_adi."""
    if _plan_cari_kolonu_var(con):
        return (
            "np.cari_id, COALESCE(c.unvan, np.musteri_adi) AS musteri_adi, "
            "np.musteri_adi AS musteri_adi_legacy",
            "LEFT JOIN nexgen_cari c ON c.id = np.cari_id",
        )
    return "NULL AS cari_id, np.musteri_adi, np.musteri_adi AS musteri_adi_legacy", ""


def _plan_liste_sorgu(con, sadece_aktif=False):
    """Plan listesini join'li şekilde döner.
    plan_id üzerinden bağlı batch_kodu da getirilir (BASLADI → Devam linki için).
    """
    where = "WHERE np.durum NOT IN ('BITTI','IPTAL')" if sadece_aktif else ""
    # plan_id kolonu varlığına göre batch JOIN kararı
    _bcols = [c['name'] for c in con.execute(
        "PRAGMA table_info(nexgen_uretim_batch)"
    ).fetchall()]
    batch_join = ""
    batch_select = "NULL AS batch_kodu, NULL AS batch_durum"
    if 'plan_id' in _bcols:
        batch_join = (
            "LEFT JOIN nexgen_uretim_batch nb "
            "ON nb.plan_id = np.id AND nb.durum NOT IN ('BITTI')"
        )
        batch_select = "nb.batch_kodu, nb.durum AS batch_durum"
    cari_select, cari_join = _plan_cari_select_sql(con)
    if _plan_rf_renk_kolonu_var(con):
        mpr_select = (
            "np.rf_renk_id, np.termin_tarihi, "
            "rf.rf_kod AS rf_kod_fk, rf.ad AS rf_ad_fk, "
        )
        mpr_join = "LEFT JOIN nexgen_rf_renk rf ON rf.id = np.rf_renk_id "
    else:
        mpr_select = (
            "NULL AS rf_renk_id, NULL AS termin_tarihi, "
            "NULL AS rf_kod_fk, NULL AS rf_ad_fk, "
        )
        mpr_join = ""
    hdr_select = ""
    hdr_join = ""
    if _plan_planlama_siparis_kolonu_var(con) and _planlama_siparis_tablosu_var(con):
        hdr_select = (
            "np.planlama_siparis_id, "
            "ps.siparis_no AS hdr_siparis_no, ps.cari_unvan AS hdr_cari_unvan, "
            "ps.termin_tarihi AS hdr_termin_tarihi, ps.durum AS hdr_durum, "
            "ps.talep_referansi AS hdr_talep_referansi, "
        )
        hdr_join = (
            "LEFT JOIN nexgen_planlama_siparis ps "
            "ON ps.id = np.planlama_siparis_id "
        )
    else:
        hdr_select = (
            "NULL AS planlama_siparis_id, NULL AS hdr_siparis_no, "
            "NULL AS hdr_cari_unvan, NULL AS hdr_termin_tarihi, "
            "NULL AS hdr_durum, NULL AS hdr_talep_referansi, "
        )
    return con.execute(f"""
        SELECT np.id, np.plan_kodu, np.kaynak, np.siparis_no,
               {cari_select},
               {hdr_select}
               {mpr_select}
               np.planlanan_kg, np.oncelik_sira, np.plan_tarihi,
               np.durum, np.notlar, np.created_at,
               uv.id AS uv_id, uv.boyut,
               rv.ad AS renk_ad,
               f.ad AS formul_ad, f.kod AS formul_kod, f.id AS formul_id,
               ku.KullaniciAdi AS olusturan_ad,
               {batch_select}
        FROM nexgen_uretim_plan np
        {cari_join}
        {hdr_join}
        {mpr_join}
        JOIN nexgen_uretim_varyant uv ON uv.id = np.uretim_varyant_id
        JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
        JOIN nexgen_formul f          ON f.id  = rv.formul_id
        LEFT JOIN sistem_kullanici ku ON ku.Id  = np.created_by
        {batch_join}
        {where}
        ORDER BY np.id DESC
    """).fetchall()


@nexgen_bp.route('/uretim-plan')
@yetki_gerekli('nexgen.plan.view', 'can_view')
def uretim_plan_liste():
    """Üretim plan listesi."""
    con = _db()
    try:
        planlar = [dict(p) for p in _plan_liste_sorgu(con)]

        for pl in planlar:
            rf = _plan_rf_bilgi(
                con,
                rf_renk_id=pl.get('rf_renk_id'),
                uretim_varyant_id=pl['uv_id'],
            )
            if rf:
                pl['rf_kod'] = rf['rf_kod']
                pl['rf_renk_ad'] = rf.get('renk_ad')
                pl['rf_uyari'] = None
            else:
                pl['rf_kod'] = pl.get('rf_kod_fk')
                pl['rf_renk_ad'] = pl.get('rf_ad_fk')
                pl['rf_uyari'] = (
                    'Bu formül/renk için ONAYLI RF eşleşmesi bulunamadı.'
                )
            if pl.get('rf_renk_id') and pl.get('cari_id') and rf and rf.get('cari_id'):
                pl['rf_cari_uyari'] = _rf_cari_uyari(con, rf['cari_id'], pl['cari_id'])
            else:
                pl['rf_cari_uyari'] = None

            sk = _siparis_kontrol_get(con, pl['id'])
            if sk:
                pl['siparis_toplam_kg'] = sk['siparis_toplam_kg']
                pl['uretilen_kg'] = sk['uretilen_kg']
                pl['kalan_kg'] = max(0.0, float(sk['fark_kg']))
                pl['kontrol_durum'] = sk.get('kontrol_durum')
                pl['kontrol_uyari'] = sk.get('uyari')
            else:
                pl['siparis_toplam_kg'] = pl.get('planlanan_kg', 0)
                pl['uretilen_kg'] = 0.0
                pl['kalan_kg'] = pl['siparis_toplam_kg']

        # Alt emir sayaçlarını planlara ekle (Migration 072+ gerekli)
        if _parca_tablosu_var(con):
            for pl in planlar:
                bk = pl.get('batch_kodu')
                if bk:
                    sayac = con.execute("""
                        SELECT COUNT(*) AS toplam,
                               SUM(CASE WHEN durum='BITTI' THEN 1 ELSE 0 END) AS biten
                        FROM nexgen_uretim_parca WHERE batch_kodu=?
                    """, (bk,)).fetchone()
                    pl['toplam_alt_emir'] = sayac['toplam'] or 0
                    pl['biten_alt_emir']  = sayac['biten']  or 0
                else:
                    pl['toplam_alt_emir'] = 0
                    pl['biten_alt_emir']  = 0

        # Aktif varyantları formül > renk > boyut hiyerarşisinde grupla
        # Yapı: [{formul_id, formul_ad, renkler: [{renk_id, renk_ad,
        #          large_uv_id, small_uv_id}]}]
        varyant_rows = con.execute("""
            SELECT uv.id AS uv_id, uv.boyut, uv.recete_durum,
                   rv.id AS renk_id, rv.ad AS renk_ad,
                   f.id  AS formul_id, f.ad AS formul_ad
            FROM nexgen_uretim_varyant uv
            JOIN nexgen_renk_varyant rv ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f        ON f.id  = rv.formul_id
            WHERE uv.aktif = 1
              AND uv.recete_durum = 'URETIME_ACIK'
            ORDER BY f.ad, rv.ad, uv.boyut
        """).fetchall()

        # Hiyerarşik dict oluştur
        formul_map = {}
        for r in varyant_rows:
            fid = r['formul_id']
            if fid not in formul_map:
                formul_map[fid] = {'formul_id': fid, 'formul_ad': r['formul_ad'], 'renkler': {}}
            rid = r['renk_id']
            if rid not in formul_map[fid]['renkler']:
                formul_map[fid]['renkler'][rid] = {
                    'renk_id': rid, 'renk_ad': r['renk_ad'],
                    'large_uv_id': None, 'small_uv_id': None,
                }
            if r['boyut'] == 'LARGE':
                formul_map[fid]['renkler'][rid]['large_uv_id'] = r['uv_id']
            elif r['boyut'] == 'SMALL':
                formul_map[fid]['renkler'][rid]['small_uv_id'] = r['uv_id']

        formuller = []
        for f in formul_map.values():
            formuller.append({
                'formul_id': f['formul_id'],
                'formul_ad': f['formul_ad'],
                'renkler': list(f['renkler'].values()),
            })

        # Aktif cariler — dropdown için
        try:
            cari_rows = con.execute(
                "SELECT id, cari_kod, unvan FROM nexgen_cari WHERE aktif=1 ORDER BY cari_kod"
            ).fetchall()
            cariler = [dict(c) for c in cari_rows]
        except Exception:
            cariler = []

    finally:
        con.close()
    return render_template(
        'nexgen/uretim_plan.html',
        active='nexgen',
        planlar=planlar,
        formuller=formuller,
        cariler=cariler,
        can_manage=yetki_var('nexgen.plan.manage', 'can_manage'),
    )


@nexgen_bp.route('/api/plan/rf-secenekleri')
@login_gerekli
def api_plan_rf_secenekleri():
    """Formül için ONAYLI RF renk listesi (plan oluşturma)."""
    formul_id = request.args.get('formul_id')
    if not formul_id:
        return jsonify({'ok': False, 'hata': 'formul_id zorunlu'}), 400
    try:
        formul_id = int(formul_id)
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'hata': 'formul_id geçersiz'}), 400

    con = _db()
    try:
        rows = con.execute("""
            SELECT rf.id, rf.rf_kod, rf.ad, rf.cari_id, rf.ilk_talep_cari_id
            FROM nexgen_rf_renk rf
            JOIN nexgen_rf_formul_uygunluk u ON u.rf_renk_id = rf.id
            WHERE u.formul_id = ?
              AND u.durum = 'ONAYLI' AND u.aktif = 1
              AND rf.aktif = 1 AND rf.durum = 'ONAYLI'
            ORDER BY rf.rf_kod
        """, (formul_id,)).fetchall()
        return jsonify({'ok': True, 'liste': [dict(r) for r in rows]})
    finally:
        con.close()


@nexgen_bp.route('/api/plan/ekle', methods=['POST'])
@yetki_gerekli('nexgen.plan.manage', 'can_manage')
def api_plan_ekle():
    """Yeni üretim planı oluştur. Stok hareketi yapılmaz."""
    d = request.get_json(silent=True) or {}
    uv_id      = d.get('uretim_varyant_id')
    kg         = d.get('planlanan_kg')
    oncelik    = d.get('oncelik_sira', 10)
    plan_tarihi = (d.get('plan_tarihi') or '').strip()
    notlar     = (d.get('notlar') or '').strip() or None
    siparis_no  = (d.get('siparis_no') or '').strip() or None
    musteri_adi = (d.get('musteri_adi') or '').strip() or None
    cari_id_raw = d.get('cari_id')
    rf_renk_id_raw = d.get('rf_renk_id')
    termin_tarihi = (d.get('termin_tarihi') or '').strip()

    if not uv_id or not kg:
        return jsonify({'ok': False, 'hata': 'uretim_varyant_id ve planlanan_kg zorunlu'}), 400
    try:
        kg = float(kg)
        if kg <= 0:
            return jsonify({'ok': False, 'hata': 'KG sıfırdan büyük olmalı'}), 400
    except Exception:
        return jsonify({'ok': False, 'hata': 'Geçersiz planlanan_kg'}), 400

    if not plan_tarihi:
        from datetime import date as _date
        plan_tarihi = _date.today().isoformat()

    con = _db()
    try:
        uv = con.execute(
            "SELECT uv.id, uv.boyut, rv.ad AS renk_ad, f.ad AS formul_ad "
            "FROM nexgen_uretim_varyant uv "
            "JOIN nexgen_renk_varyant rv ON rv.id=uv.renk_varyant_id "
            "JOIN nexgen_formul f ON f.id=rv.formul_id "
            "WHERE uv.id=? AND uv.aktif=1",
            (uv_id,)
        ).fetchone()
        if not uv:
            return jsonify({'ok': False, 'hata': 'Üretim varyantı bulunamadı'}), 404

        cari_id = None
        if cari_id_raw is not None and cari_id_raw != '':
            try:
                cari_id = int(cari_id_raw)
            except (TypeError, ValueError):
                return jsonify({'ok': False, 'hata': 'cari_id geçersiz'}), 400
            cari = con.execute(
                "SELECT id, unvan FROM nexgen_cari WHERE id=? AND aktif=1",
                (cari_id,)
            ).fetchone()
            if not cari:
                return jsonify({'ok': False, 'hata': f'Cari bulunamadı (id={cari_id})'}), 400
            musteri_adi = cari['unvan']
        elif not musteri_adi:
            return jsonify({'ok': False, 'hata': 'Cari seçimi zorunludur.'}), 400

        rf_renk_id = None
        rf_cari_uyari = None
        if _plan_rf_renk_kolonu_var(con):
            if not rf_renk_id_raw:
                return jsonify({'ok': False, 'hata': 'rf_renk_id zorunlu'}), 400
            try:
                rf_renk_id = int(rf_renk_id_raw)
            except (TypeError, ValueError):
                return jsonify({'ok': False, 'hata': 'rf_renk_id geçersiz'}), 400
            rf = con.execute("""
                SELECT rf.id, rf.rf_kod, rf.ad, rf.cari_id
                FROM nexgen_rf_renk rf
                JOIN nexgen_rf_formul_uygunluk u ON u.rf_renk_id = rf.id
                JOIN nexgen_uretim_varyant uv ON uv.id = ?
                JOIN nexgen_renk_varyant rv ON rv.id = uv.renk_varyant_id
                    AND rv.formul_id = u.formul_id
                WHERE rf.id = ?
                  AND rf.durum = 'ONAYLI' AND rf.aktif = 1
                  AND u.durum = 'ONAYLI' AND u.aktif = 1
            """, (uv_id, rf_renk_id)).fetchone()
            if not rf:
                return jsonify({
                    'ok': False,
                    'hata': 'Seçilen RF bu formül için ONAYLI değil veya bulunamadı.',
                }), 400
            if cari_id and rf['cari_id']:
                rf_cari_uyari = _rf_cari_uyari(con, rf['cari_id'], cari_id)

        if _plan_termin_kolonu_var(con) and not termin_tarihi:
            return jsonify({'ok': False, 'hata': 'termin_tarihi zorunlu'}), 400

        # Sipariş no boşsa otomatik üret
        if not siparis_no:
            siparis_no = _nexgen_siparis_no_uret(con)

        planlama_siparis_id = None
        if (_plan_planlama_siparis_kolonu_var(con)
                and _planlama_siparis_tablosu_var(con)):
            planlama_siparis_id = _planlama_siparis_bul_veya_olustur(
                con, siparis_no, cari_id, musteri_adi,
                termin_tarihi=termin_tarihi, notlar=notlar,
                olusturan_id=_kullanici_id(),
            )

        plan_kodu = _plan_kodu_uret(con)
        uid = _kullanici_id()

        cols = [
            'plan_kodu', 'kaynak', 'siparis_no', 'musteri_adi',
            'uretim_varyant_id', 'planlanan_kg', 'oncelik_sira', 'plan_tarihi',
            'durum', 'notlar', 'created_by',
        ]
        vals = [
            plan_kodu, 'MANUEL', siparis_no, musteri_adi, uv_id,
            round(kg, 3), int(oncelik), plan_tarihi, 'PLANLANDI', notlar, uid,
        ]
        if _plan_cari_kolonu_var(con):
            cols.insert(4, 'cari_id')
            vals.insert(4, cari_id)
        if _plan_planlama_siparis_kolonu_var(con) and planlama_siparis_id:
            cols.insert(-3, 'planlama_siparis_id')
            vals.insert(-3, planlama_siparis_id)
        if _plan_rf_renk_kolonu_var(con):
            cols.insert(-3, 'rf_renk_id')
            vals.insert(-3, rf_renk_id)
        if _plan_termin_kolonu_var(con):
            cols.insert(-3, 'termin_tarihi')
            vals.insert(-3, termin_tarihi)

        placeholders = ','.join(['?'] * len(cols))
        con.execute(
            f"INSERT INTO nexgen_uretim_plan ({','.join(cols)}) VALUES ({placeholders})",
            vals,
        )
        con.commit()
        resp = {
            'ok': True,
            'plan_kodu': plan_kodu,
            'siparis_no': siparis_no,
            'formul_ad': uv['formul_ad'],
            'renk_ad': uv['renk_ad'],
            'boyut': uv['boyut'],
            'planlanan_kg': round(kg, 3),
            'musteri_adi': musteri_adi,
        }
        if cari_id is not None:
            resp['cari_id'] = cari_id
        if planlama_siparis_id is not None:
            resp['planlama_siparis_id'] = planlama_siparis_id
        if rf_renk_id is not None:
            resp['rf_renk_id'] = rf_renk_id
            resp['rf_kod'] = rf['rf_kod']
            resp['rf_renk_ad'] = rf['ad']
        if termin_tarihi:
            resp['termin_tarihi'] = termin_tarihi
        if rf_cari_uyari:
            resp['rf_cari_uyari'] = rf_cari_uyari
        return jsonify(resp)
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/planlama-siparis/<int:ps_id>/mpr-ekle', methods=['POST'])
@yetki_gerekli('nexgen.plan.manage', 'can_manage')
def api_planlama_siparis_mpr_ekle(ps_id):
    """Mevcut planlama sipariş header'ına yeni MPR satırı ekler."""
    d = request.get_json(silent=True) or {}
    uv_id = d.get('uretim_varyant_id')
    kg = d.get('planlanan_kg')
    oncelik = d.get('oncelik_sira', 10)
    plan_tarihi = (d.get('plan_tarihi') or '').strip()
    notlar = (d.get('notlar') or '').strip() or None
    rf_renk_id_raw = d.get('rf_renk_id')

    if not uv_id or not kg:
        return jsonify({'ok': False, 'hata': 'uretim_varyant_id ve planlanan_kg zorunlu'}), 400
    try:
        kg = float(kg)
        if kg <= 0:
            return jsonify({'ok': False, 'hata': 'KG sıfırdan büyük olmalı'}), 400
    except Exception:
        return jsonify({'ok': False, 'hata': 'Geçersiz planlanan_kg'}), 400

    if not plan_tarihi:
        from datetime import date as _date
        plan_tarihi = _date.today().isoformat()

    con = _db()
    try:
        if not _planlama_siparis_tablosu_var(con):
            return jsonify({'ok': False, 'hata': 'Planlama sipariş tablosu yok'}), 400

        hdr = con.execute(
            "SELECT id, siparis_no, cari_id, cari_unvan, termin_tarihi, durum "
            "FROM nexgen_planlama_siparis WHERE id=?",
            (ps_id,),
        ).fetchone()
        if not hdr:
            return jsonify({'ok': False, 'hata': 'Planlama siparişi bulunamadı'}), 404
        if hdr['durum'] == 'IPTAL':
            return jsonify({'ok': False, 'hata': 'İptal edilmiş siparişe satır eklenemez'}), 400

        uv = con.execute(
            "SELECT uv.id, uv.boyut, rv.ad AS renk_ad, f.ad AS formul_ad "
            "FROM nexgen_uretim_varyant uv "
            "JOIN nexgen_renk_varyant rv ON rv.id=uv.renk_varyant_id "
            "JOIN nexgen_formul f ON f.id=rv.formul_id "
            "WHERE uv.id=? AND uv.aktif=1",
            (uv_id,),
        ).fetchone()
        if not uv:
            return jsonify({'ok': False, 'hata': 'Üretim varyantı bulunamadı'}), 404

        cari_id = hdr['cari_id']
        musteri_adi = hdr['cari_unvan']
        siparis_no = hdr['siparis_no']
        termin_tarihi = hdr['termin_tarihi']

        rf_renk_id = None
        rf_cari_uyari = None
        rf = None
        if _plan_rf_renk_kolonu_var(con):
            if not rf_renk_id_raw:
                return jsonify({'ok': False, 'hata': 'rf_renk_id zorunlu'}), 400
            try:
                rf_renk_id = int(rf_renk_id_raw)
            except (TypeError, ValueError):
                return jsonify({'ok': False, 'hata': 'rf_renk_id geçersiz'}), 400
            rf = con.execute("""
                SELECT rf.id, rf.rf_kod, rf.ad, rf.cari_id
                FROM nexgen_rf_renk rf
                JOIN nexgen_rf_formul_uygunluk u ON u.rf_renk_id = rf.id
                JOIN nexgen_uretim_varyant uv2 ON uv2.id = ?
                JOIN nexgen_renk_varyant rv ON rv.id = uv2.renk_varyant_id
                    AND rv.formul_id = u.formul_id
                WHERE rf.id = ?
                  AND rf.durum = 'ONAYLI' AND rf.aktif = 1
                  AND u.durum = 'ONAYLI' AND u.aktif = 1
            """, (uv_id, rf_renk_id)).fetchone()
            if not rf:
                return jsonify({
                    'ok': False,
                    'hata': 'Seçilen RF bu formül için ONAYLI değil veya bulunamadı.',
                }), 400
            if cari_id and rf['cari_id']:
                rf_cari_uyari = _rf_cari_uyari(con, rf['cari_id'], cari_id)

        if _plan_termin_kolonu_var(con) and not termin_tarihi:
            return jsonify({'ok': False, 'hata': 'Header termin tarihi eksik'}), 400

        planlama_siparis_id = hdr['id']
        plan_kodu = _plan_kodu_uret(con)
        uid = _kullanici_id()

        cols = [
            'plan_kodu', 'kaynak', 'siparis_no', 'musteri_adi',
            'uretim_varyant_id', 'planlanan_kg', 'oncelik_sira', 'plan_tarihi',
            'durum', 'notlar', 'created_by',
        ]
        vals = [
            plan_kodu, 'MANUEL', siparis_no, musteri_adi, uv_id,
            round(kg, 3), int(oncelik), plan_tarihi, 'PLANLANDI', notlar, uid,
        ]
        if _plan_cari_kolonu_var(con):
            cols.insert(4, 'cari_id')
            vals.insert(4, cari_id)
        if _plan_planlama_siparis_kolonu_var(con) and planlama_siparis_id:
            cols.insert(-3, 'planlama_siparis_id')
            vals.insert(-3, planlama_siparis_id)
        if _plan_rf_renk_kolonu_var(con):
            cols.insert(-3, 'rf_renk_id')
            vals.insert(-3, rf_renk_id)
        if _plan_termin_kolonu_var(con):
            cols.insert(-3, 'termin_tarihi')
            vals.insert(-3, termin_tarihi)

        placeholders = ','.join(['?'] * len(cols))
        con.execute(
            f"INSERT INTO nexgen_uretim_plan ({','.join(cols)}) VALUES ({placeholders})",
            vals,
        )
        con.commit()
        resp = {
            'ok': True,
            'plan_kodu': plan_kodu,
            'siparis_no': siparis_no,
            'planlama_siparis_id': planlama_siparis_id,
            'formul_ad': uv['formul_ad'],
            'renk_ad': uv['renk_ad'],
            'boyut': uv['boyut'],
            'planlanan_kg': round(kg, 3),
            'musteri_adi': musteri_adi,
        }
        if cari_id is not None:
            resp['cari_id'] = cari_id
        if rf_renk_id is not None and rf:
            resp['rf_renk_id'] = rf_renk_id
            resp['rf_kod'] = rf['rf_kod']
            resp['rf_renk_ad'] = rf['ad']
        if termin_tarihi:
            resp['termin_tarihi'] = termin_tarihi
        if rf_cari_uyari:
            resp['rf_cari_uyari'] = rf_cari_uyari
        return jsonify(resp)
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/plan/<int:plan_id>/iptal', methods=['POST'])
@yetki_gerekli('nexgen.plan.manage', 'can_manage')
def api_plan_iptal(plan_id):
    """Planı IPTAL durumuna çeker."""
    con = _db()
    try:
        p = con.execute(
            "SELECT id, durum FROM nexgen_uretim_plan WHERE id=?", (plan_id,)
        ).fetchone()
        if not p:
            return jsonify({'ok': False, 'hata': 'Plan bulunamadı'}), 404
        if p['durum'] == 'BITTI':
            return jsonify({'ok': False, 'hata': 'Tamamlanan plan iptal edilemez'}), 400
        con.execute(
            "UPDATE nexgen_uretim_plan SET durum='IPTAL' WHERE id=?", (plan_id,)
        )
        con.commit()
        return jsonify({'ok': True})
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/batch/<batch_kodu>/durum', methods=['POST'])
@login_gerekli
def api_batch_durum_guncelle(batch_kodu):
    """Batch durumunu güncelle.

    Kabul edilen geçişler:
      HAZIR      → DEVAM      (operatör başlattı)
      DEVAM      → BEKLEME    (beklemede)
      BEKLEME    → DEVAM      (devam ettirdi)
      DEVAM      → BITTI      (üretim bitti)
      HAZIR      → BITTI      (direkt bitti)

    Yetki: nexgen.tablet.uretim VEYA nexgen.plan.manage
    """
    if not (yetki_var('nexgen.tablet.uretim', 'can_uretim') or
            yetki_var('nexgen.plan.manage', 'can_manage')):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    d = request.get_json(silent=True) or {}
    yeni_durum = (d.get('durum') or '').strip().upper()
    notlar = (d.get('notlar') or '').strip() or None

    GECERLI_DURUMLAR = {'HAZIR', 'DEVAM', 'BEKLEME', 'BITTI'}
    GECISLER = {
        'HAZIR':   {'DEVAM', 'BITTI'},
        'DEVAM':   {'BEKLEME', 'BITTI'},
        'BEKLEME': {'DEVAM'},
    }

    if yeni_durum not in GECERLI_DURUMLAR:
        return jsonify({'ok': False, 'hata': f'Geçersiz durum: {yeni_durum}'}), 400

    con = _db()
    try:
        batch = con.execute(
            "SELECT batch_kodu, durum, plan_id FROM nexgen_uretim_batch WHERE batch_kodu=?",
            (batch_kodu,)
        ).fetchone()
        if not batch:
            return jsonify({'ok': False, 'hata': 'Batch bulunamadı'}), 404

        mevcut = batch['durum']
        izin = GECISLER.get(mevcut, set())
        if yeni_durum not in izin:
            return jsonify({
                'ok': False,
                'hata': f'{mevcut} → {yeni_durum} geçişi geçersiz'
            }), 400

        if notlar:
            con.execute(
                "UPDATE nexgen_uretim_batch SET durum=?, notlar=? WHERE batch_kodu=?",
                (yeni_durum, notlar, batch_kodu)
            )
        else:
            con.execute(
                "UPDATE nexgen_uretim_batch SET durum=? WHERE batch_kodu=?",
                (yeni_durum, batch_kodu)
            )

        # Üretim bitti → bağlı planı BITTI yap
        if yeni_durum == 'BITTI' and batch['plan_id']:
            con.execute(
                "UPDATE nexgen_uretim_plan SET durum='BITTI' WHERE id=? AND durum='BASLADI'",
                (batch['plan_id'],)
            )

        if yeni_durum == 'DEVAM':
            _rf_kullanim_tablet_sync(con, batch_kodu)
        elif yeni_durum == 'BITTI':
            _rf_kullanim_tablet_sync(con, batch_kodu, tamamlandi=True)

        con.commit()
        return jsonify({'ok': True, 'durum': yeni_durum, 'batch_kodu': batch_kodu})
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


def _formul_batch_kg_hesapla(con, uretim_varyant_id):
    """Bir üretim varyantının aktif reçete kalemlerinden standart batch KG'sini döner.

    BOYA kalemleri hariç — renk formülü RF tarafında yönetilir (FAZ-1A/2A).
    Returns: float (0.0 reçete bulunamazsa)
    """
    row = con.execute("""
        SELECT COALESCE(SUM(rk.miktar_kg), 0) AS toplam
        FROM nexgen_recete_kalem rk
        JOIN nexgen_stok_kart sk ON sk.id = rk.stok_kart_id
        WHERE rk.uretim_varyant_id = ? AND rk.aktif = 1
          AND UPPER(COALESCE(sk.kategori, '')) != 'BOYA'
    """, (uretim_varyant_id,)).fetchone()
    return round(float(row['toplam']), 3) if row else 0.0


def _mpr_stok_ihtiyac_hesapla(con, uretim_varyant_id, rf_renk_id, planlanan_kg):
    """MPR stok ihtiyacı önizleme — taban hammadde (BOYA hariç) + RF boya.

    Stok hareketi YAZMAZ. Mevcut stok: nexgen_stok_hareket SUM (_mevcut_stok).
    """
    try:
        planlanan_kg = float(planlanan_kg)
    except (TypeError, ValueError):
        return {'ok': False, 'hata': 'Geçersiz planlanan_kg'}
    if planlanan_kg <= 0:
        return {'ok': False, 'hata': 'planlanan_kg sıfırdan büyük olmalı'}

    taban_batch_kg = _formul_batch_kg_hesapla(con, uretim_varyant_id)
    if taban_batch_kg <= 0:
        return {
            'ok': False,
            'hata': 'Taban reçete batch KG bulunamadı (BOYA hariç kalem yok).',
            'taban_batch_kg': 0.0,
        }

    carpan = planlanan_kg / taban_batch_kg
    satirlar = []

    taban_rows = con.execute("""
        SELECT rk.stok_kart_id, rk.miktar_kg,
               sk.kod AS stok_kod, sk.ad AS stok_ad
        FROM nexgen_recete_kalem rk
        JOIN nexgen_stok_kart sk ON sk.id = rk.stok_kart_id
        WHERE rk.uretim_varyant_id = ? AND rk.aktif = 1
          AND UPPER(COALESCE(sk.kategori, '')) != 'BOYA'
        ORDER BY rk.sira, rk.id
    """, (uretim_varyant_id,)).fetchall()

    for row in taban_rows:
        gerekli = round(float(row['miktar_kg']) * carpan, 3)
        if gerekli <= 0:
            continue
        satirlar.append({
            'kaynak': 'TABAN',
            'stok_kart_id': row['stok_kart_id'],
            'stok_kod': row['stok_kod'] or '',
            'stok_ad': row['stok_ad'] or '',
            'gerekli_kg': gerekli,
        })

    rf_id = None
    if rf_renk_id not in (None, ''):
        try:
            rf_id = int(rf_renk_id)
        except (TypeError, ValueError):
            pass
    if rf_id:
        rf_rows = con.execute("""
            SELECT rk.stok_kart_id, rk.miktar_kg,
                   sk.kod AS stok_kod, sk.ad AS stok_ad
            FROM nexgen_rf_kalem rk
            JOIN nexgen_stok_kart sk ON sk.id = rk.stok_kart_id
            WHERE rk.rf_renk_id = ? AND rk.aktif = 1
            ORDER BY rk.sira, rk.id
        """, (rf_id,)).fetchall()
        for row in rf_rows:
            gerekli = round(float(row['miktar_kg']) * carpan, 3)
            if gerekli <= 0:
                continue
            satirlar.append({
                'kaynak': 'RF',
                'stok_kart_id': row['stok_kart_id'],
                'stok_kod': row['stok_kod'] or '',
                'stok_ad': row['stok_ad'] or '',
                'gerekli_kg': gerekli,
            })

    talep = {}
    for s in satirlar:
        sid = s['stok_kart_id']
        talep[sid] = round(talep.get(sid, 0.0) + s['gerekli_kg'], 3)

    mevcut_cache = {sid: _mevcut_stok(con, sid) for sid in talep}
    eksik_stok_ids = {
        sid for sid, gerekli in talep.items()
        if mevcut_cache[sid] < gerekli
    }

    kalemler = []
    for s in satirlar:
        sid = s['stok_kart_id']
        toplam_gerekli = talep[sid]
        mevcut = mevcut_cache[sid]
        kalemler.append({
            'kaynak': s['kaynak'],
            'stok_kart_id': sid,
            'stok_kod': s['stok_kod'],
            'stok_ad': s['stok_ad'],
            'gerekli_kg': s['gerekli_kg'],
            'mevcut_kg': mevcut,
            'fark_kg': round(mevcut - toplam_gerekli, 3),
            'yeterli': sid not in eksik_stok_ids,
        })

    return {
        'ok': True,
        'yeterli_mi': len(eksik_stok_ids) == 0,
        'eksik_sayisi': len(eksik_stok_ids),
        'taban_batch_kg': taban_batch_kg,
        'planlanan_kg': round(planlanan_kg, 3),
        'kalemler': kalemler,
    }


@nexgen_bp.route('/api/plan/stok-onizle', methods=['POST'])
@login_gerekli
def api_plan_stok_onizle():
    """MPR stok ihtiyacı önizleme — salt okuma, hareket yazmaz."""
    if not (yetki_var('nexgen.plan.manage', 'can_manage') or
            yetki_var('nexgen.plan.view', 'can_view')):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    d = request.get_json(silent=True) or {}
    uv_id = d.get('uretim_varyant_id')
    rf_renk_id = d.get('rf_renk_id')
    planlanan_kg = d.get('planlanan_kg')

    if not uv_id:
        return jsonify({'ok': False, 'hata': 'uretim_varyant_id zorunlu'}), 400

    con = _db()
    try:
        uv = con.execute(
            "SELECT id FROM nexgen_uretim_varyant WHERE id=? AND aktif=1",
            (uv_id,),
        ).fetchone()
        if not uv:
            return jsonify({'ok': False, 'hata': 'Üretim varyantı bulunamadı'}), 404

        sonuc = _mpr_stok_ihtiyac_hesapla(con, uv_id, rf_renk_id, planlanan_kg)
        if not sonuc.get('ok'):
            return jsonify(sonuc), 400
        return jsonify(sonuc)
    finally:
        con.close()


@nexgen_bp.route('/api/plan/<int:plan_id>/alt-emir-onizle')
@login_gerekli
def api_alt_emir_onizle(plan_id):
    """Alt Emir önizleme — DB'ye yazmaz, sadece hesap döner.

    Planlama ekranı "Üretime Gönder" öncesi bu endpoint'i çağırarak
    özet bilgiyi modal'da gösterir.

    Yanıt:
      planlanan_kg      : float
      formul_batch_kg   : float  (reçeteden)
      alt_emir_sayisi   : int    (ceil)
      toplam_uretim_kg  : float  (alt_emir_sayisi × formul_batch_kg)
      fazla_kg          : float  (toplam_uretim_kg - planlanan_kg)
      uyari             : str | None (reçete yoksa)

    Yetki: nexgen.plan.manage
    """
    if not yetki_var('nexgen.plan.manage', 'can_manage'):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    con = _db()
    try:
        p = con.execute("""
            SELECT np.id, np.planlanan_kg, np.uretim_varyant_id, np.durum,
                   f.ad AS formul_ad, rv.ad AS renk_ad, uv.boyut
            FROM nexgen_uretim_plan np
            JOIN nexgen_uretim_varyant uv ON uv.id = np.uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            WHERE np.id = ?
        """, (plan_id,)).fetchone()
        if not p:
            return jsonify({'ok': False, 'hata': 'Plan bulunamadı'}), 404

        planlanan = float(p['planlanan_kg'])
        formul_batch_kg = _formul_batch_kg_hesapla(con, p['uretim_varyant_id'])

        uyari = None
        if formul_batch_kg <= 0:
            uyari = 'Bu formül için aktif reçete kalemi bulunamadı.'
            alt_emir_sayisi = 0
            toplam_uretim_kg = 0.0
            fazla_kg = 0.0
        else:
            import math
            alt_emir_sayisi  = math.ceil(planlanan / formul_batch_kg)
            toplam_uretim_kg = round(alt_emir_sayisi * formul_batch_kg, 3)
            fazla_kg         = round(toplam_uretim_kg - planlanan, 3)

        return jsonify({
            'ok': True,
            'plan_id': plan_id,
            'formul_ad': p['formul_ad'],
            'renk_ad': p['renk_ad'],
            'boyut': p['boyut'],
            'planlanan_kg': round(planlanan, 3),
            'formul_batch_kg': formul_batch_kg,
            'alt_emir_sayisi': alt_emir_sayisi,
            'toplam_uretim_kg': toplam_uretim_kg,
            'fazla_kg': fazla_kg,
            'uyari': uyari,
        })
    finally:
        con.close()


@nexgen_bp.route('/api/plan/<int:plan_id>/basla', methods=['POST'])
@login_gerekli
def api_plan_basla(plan_id):
    """Plandan üretim başlat: Batch + LOT oluştur, Alt Emirleri otomatik oluştur.

    FAZ-5G: api_plan_basla artık nexgen_uretim_parca tablosuna
    alt emirleri toplu olarak HAZIR durumuyla INSERT eder.
    Alt emir sayısı = ceil(planlanan_kg / formul_batch_kg).
    Her alt emirin hedef_kg = formul_batch_kg (formül KG).
    Operatör KG girmez.

    Yetki: nexgen.tablet.uretim VEYA nexgen.plan.manage
    Stok hareketi yapılmaz. MRP yok.
    """
    if not (yetki_var('nexgen.tablet.uretim', 'can_uretim') or
            yetki_var('nexgen.plan.manage', 'can_manage')):
        return jsonify({'ok': False, 'hata': 'Yetki yok (tablet.uretim veya plan.manage gerekli)'}), 403

    import math

    d = request.get_json(silent=True) or {}
    notlar = (d.get('notlar') or '').strip() or None

    con = _db()
    try:
        rf_col = ", np.rf_renk_id" if _plan_rf_renk_kolonu_var(con) else ", NULL AS rf_renk_id"
        p = con.execute(f"""
            SELECT np.id, np.plan_kodu, np.uretim_varyant_id,
                   np.planlanan_kg, np.durum, np.notlar{rf_col},
                   uv.recete_durum,
                   uv.boyut, rv.ad AS renk_ad, f.ad AS formul_ad
            FROM nexgen_uretim_plan np
            JOIN nexgen_uretim_varyant uv ON uv.id = np.uretim_varyant_id
            JOIN nexgen_renk_varyant rv   ON rv.id = uv.renk_varyant_id
            JOIN nexgen_formul f          ON f.id  = rv.formul_id
            WHERE np.id = ?
        """, (plan_id,)).fetchone()
        if not p:
            return jsonify({'ok': False, 'hata': 'Plan bulunamadı'}), 404
        if p['durum'] not in ('PLANLANDI',):
            return jsonify({'ok': False, 'hata': f'Plan durumu uygun değil: {p["durum"]}'}), 400

        rf_renk_id = p['rf_renk_id'] if 'rf_renk_id' in p.keys() else None
        stok = _mpr_stok_ihtiyac_hesapla(
            con, p['uretim_varyant_id'], rf_renk_id, float(p['planlanan_kg'])
        )
        if not stok.get('ok'):
            return jsonify({'ok': False, 'hata': stok.get('hata', 'Stok hesaplanamadı')}), 400
        if not stok.get('yeterli_mi'):
            eksik = [k for k in stok.get('kalemler', []) if not k.get('yeterli')]
            return jsonify({
                'ok': False,
                'hata': 'Stok yetersiz',
                'eksik_sayisi': stok.get('eksik_sayisi', len(eksik)),
                'eksik_kalemler': eksik,
                'kalemler': stok.get('kalemler', []),
            }), 400

        # Batch + LOT oluştur
        batch_kodu   = _batch_kodu_uret(con)
        lot_kodu     = _lot_kodu_uret(con)
        uid          = _kullanici_id()
        batch_notlar = notlar or p['notlar']

        # plan_id kolonu migration 071 sonrası mevcut — graceful INSERT
        batch_cols = [c['name'] for c in con.execute(
            "PRAGMA table_info(nexgen_uretim_batch)"
        ).fetchall()]
        if 'plan_id' in batch_cols:
            con.execute(
                "INSERT INTO nexgen_uretim_batch"
                "(batch_kodu, lot_kodu, uretim_varyant_id, planlanan_kg,"
                " durum, olusturan_id, notlar, plan_id)"
                " VALUES(?,?,?,?,'HAZIR',?,?,?)",
                (batch_kodu, lot_kodu, p['uretim_varyant_id'],
                 round(p['planlanan_kg'], 3), uid, batch_notlar, plan_id)
            )
        else:
            con.execute(
                "INSERT INTO nexgen_uretim_batch"
                "(batch_kodu, lot_kodu, uretim_varyant_id, planlanan_kg,"
                " durum, olusturan_id, notlar)"
                " VALUES(?,?,?,?,'HAZIR',?,?)",
                (batch_kodu, lot_kodu, p['uretim_varyant_id'],
                 round(p['planlanan_kg'], 3), uid, batch_notlar)
            )

        # Batch id'yi al
        batch_row = con.execute(
            "SELECT id FROM nexgen_uretim_batch WHERE batch_kodu=?",
            (batch_kodu,)
        ).fetchone()
        batch_id = batch_row['id'] if batch_row else None

        # Alt Emirleri toplu oluştur (FAZ-5G)
        alt_emir_sayisi = 0
        formul_batch_kg = 0.0

        if _parca_tablosu_var(con) and batch_id:
            formul_batch_kg = _formul_batch_kg_hesapla(con, p['uretim_varyant_id'])
            if formul_batch_kg > 0:
                planlanan = float(p['planlanan_kg'])
                alt_emir_sayisi = math.ceil(planlanan / formul_batch_kg)

                # formul_batch_kg kolonu var mı? (Migration 073)
                parca_cols = [c['name'] for c in con.execute(
                    "PRAGMA table_info(nexgen_uretim_parca)"
                ).fetchall()]
                has_formul_kg_col = 'formul_batch_kg' in parca_cols

                # Alt emir numaraları 1001'den başlar (1001, 1002, 1003, ...)
                for no in range(1001, 1001 + alt_emir_sayisi):
                    if has_formul_kg_col:
                        con.execute("""
                            INSERT INTO nexgen_uretim_parca
                                (batch_id, batch_kodu, plan_id, parca_no,
                                 hedef_kg, formul_batch_kg, uretilen_kg,
                                 durum, operator_id)
                            VALUES (?, ?, ?, ?, ?, ?, 0, 'HAZIR', ?)
                        """, (batch_id, batch_kodu, plan_id, no,
                              round(formul_batch_kg, 3),
                              round(formul_batch_kg, 3),
                              uid))
                    else:
                        con.execute("""
                            INSERT INTO nexgen_uretim_parca
                                (batch_id, batch_kodu, plan_id, parca_no,
                                 hedef_kg, uretilen_kg, durum, operator_id)
                            VALUES (?, ?, ?, ?, ?, 0, 'HAZIR', ?)
                        """, (batch_id, batch_kodu, plan_id, no,
                              round(formul_batch_kg, 3), uid))

        con.execute(
            "UPDATE nexgen_uretim_plan SET durum='BASLADI' WHERE id=?", (plan_id,)
        )

        _rf_kullanim_tablet_sync(con, batch_kodu)

        rf_bilgi = _plan_rf_bilgi(
            con,
            rf_renk_id=p['rf_renk_id'],
            uretim_varyant_id=p['uretim_varyant_id'],
        )
        rf_uyari = None
        if not rf_bilgi:
            rf_uyari = (
                'Bu formül/renk için ONAYLI RF eşleşmesi bulunamadı. '
                'Üretim başladı; RF kullanım logu oluşturulamadı.'
            )

        con.commit()

        resp = {
            'ok': True,
            'batch_kodu': batch_kodu,
            'lot_kodu': lot_kodu,
            'formul_ad': p['formul_ad'],
            'renk_ad': p['renk_ad'],
            'boyut': p['boyut'],
            'planlanan_kg': round(p['planlanan_kg'], 3),
            'formul_batch_kg': round(formul_batch_kg, 3),
            'alt_emir_sayisi': alt_emir_sayisi,
        }
        if rf_bilgi:
            resp['rf_kod'] = rf_bilgi['rf_kod']
            resp['rf_renk_ad'] = rf_bilgi.get('renk_ad')
        if rf_uyari:
            resp['rf_uyari'] = rf_uyari
        return jsonify(resp)
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


# ─────────────────────────────────────────────────────────────
# NEXGEN FAZ-4F — RECYCLE İZİN ALTYAPISI

def _uretim_varyant_recycle_izinleri(con, uretim_varyant_id):
    """Bir üretim varyantına tanımlı recycle izinlerini + mevcut stok KG ile döner.

    Returns: list of dict:
        id, recycle_kart_id, recycle_kart_ad, recycle_kart_kod,
        mevcut_stok_kg, yerine_kart_id, yerine_kart_ad,
        max_oran_pct, aktif, notlar, created_at
    """
    izinler_raw = con.execute("""
        SELECT ri.id, ri.recycle_stok_kart_id, ri.yerine_stok_kart_id,
               ri.max_oran_pct, ri.aktif, ri.notlar, ri.created_at,
               sk_rc.ad  AS recycle_kart_ad,
               sk_rc.kod AS recycle_kart_kod,
               sk_yc.ad  AS yerine_kart_ad
        FROM nexgen_recete_recycle_izin ri
        JOIN nexgen_stok_kart sk_rc ON sk_rc.id = ri.recycle_stok_kart_id
        LEFT JOIN nexgen_stok_kart sk_yc ON sk_yc.id = ri.yerine_stok_kart_id
        WHERE ri.uretim_varyant_id = ?
        ORDER BY ri.id
    """, (uretim_varyant_id,)).fetchall()

    sonuc = []
    for iz in izinler_raw:
        mevcut = float(con.execute(
            "SELECT COALESCE(SUM(miktar_kg),0) FROM nexgen_stok_hareket WHERE stok_kart_id=?",
            (iz['recycle_stok_kart_id'],)
        ).fetchone()[0])
        sonuc.append({
            'id':                iz['id'],
            'recycle_kart_id':   iz['recycle_stok_kart_id'],
            'recycle_kart_ad':   iz['recycle_kart_ad'],
            'recycle_kart_kod':  iz['recycle_kart_kod'],
            'mevcut_stok_kg':    mevcut,
            'yerine_kart_id':    iz['yerine_stok_kart_id'],
            'yerine_kart_ad':    iz['yerine_kart_ad'],
            'max_oran_pct':      iz['max_oran_pct'],
            'aktif':             iz['aktif'],
            'notlar':            iz['notlar'],
            'created_at':        iz['created_at'],
        })
    return sonuc


@nexgen_bp.route('/api/recycle-izin/ekle', methods=['POST'])
@yetki_gerekli('nexgen.recycle.manage', 'can_manage')
def api_recycle_izin_ekle():
    """Bir üretim varyantına recycle izni ekle.
    POST JSON:
        uretim_varyant_id    : int
        recycle_stok_kart_id : int
        yerine_stok_kart_id  : int | null
        max_oran_pct         : float (default 10)
        notlar               : str
    Stok hareketi YAZMAZ.
    """
    d = request.get_json(silent=True) or {}
    uv_id    = d.get('uretim_varyant_id')
    rc_id    = d.get('recycle_stok_kart_id')
    yc_id    = d.get('yerine_stok_kart_id') or None
    oran     = float(d.get('max_oran_pct', 10))
    notlar   = (d.get('notlar') or '').strip()

    if not uv_id or not rc_id:
        return jsonify({'ok': False, 'hata': 'uretim_varyant_id ve recycle_stok_kart_id zorunlu'}), 400

    con = _db()
    try:
        # Recycle kategorisi kontrolü
        rc_kart = con.execute(
            "SELECT id, ad, kategori FROM nexgen_stok_kart WHERE id=? AND aktif=1",
            (rc_id,)
        ).fetchone()
        if not rc_kart:
            return jsonify({'ok': False, 'hata': 'Recycle stok kartı bulunamadı'}), 404
        if rc_kart['kategori'] != 'RECYCLE':
            return jsonify({'ok': False, 'hata': 'Sadece RECYCLE kategorisi seçilebilir'}), 400

        # Varyant kontrolü
        uv = con.execute(
            "SELECT id, ad FROM nexgen_uretim_varyant WHERE id=? AND aktif=1", (uv_id,)
        ).fetchone()
        if not uv:
            return jsonify({'ok': False, 'hata': 'Üretim varyantı bulunamadı'}), 404

        # Duplicate kontrolü
        mev = con.execute(
            "SELECT id FROM nexgen_recete_recycle_izin WHERE uretim_varyant_id=? AND recycle_stok_kart_id=?",
            (uv_id, rc_id)
        ).fetchone()
        if mev:
            return jsonify({'ok': False, 'hata': 'Bu recycle izni zaten tanımlı'}), 409

        con.execute("""
            INSERT INTO nexgen_recete_recycle_izin
                (uretim_varyant_id, recycle_stok_kart_id, yerine_stok_kart_id,
                 max_oran_pct, aktif, notlar, created_at, created_by)
            VALUES (?,?,?,?,1,?,datetime('now','localtime'),?)
        """, (uv_id, rc_id, yc_id, oran, notlar or None,
              session.get('user_id')))
        con.commit()
        new_id = con.execute("SELECT last_insert_rowid()").fetchone()[0]
    finally:
        con.close()

    return jsonify({'ok': True, 'id': new_id})


@nexgen_bp.route('/api/recycle-izin/<int:izin_id>/toggle', methods=['POST'])
@yetki_gerekli('nexgen.recycle.manage', 'can_manage')
def api_recycle_izin_toggle(izin_id):
    """Recycle iznini aktif/pasif yap. Stok hareketi YAZMAZ."""
    con = _db()
    try:
        izin = con.execute(
            "SELECT id, aktif FROM nexgen_recete_recycle_izin WHERE id=?", (izin_id,)
        ).fetchone()
        if not izin:
            return jsonify({'ok': False, 'hata': 'İzin bulunamadı'}), 404
        yeni = 0 if izin['aktif'] else 1
        con.execute(
            "UPDATE nexgen_recete_recycle_izin SET aktif=? WHERE id=?", (yeni, izin_id)
        )
        con.commit()
    finally:
        con.close()
    return jsonify({'ok': True, 'aktif': yeni})


@nexgen_bp.route('/api/recycle-izin/stok-kartlari')
@yetki_gerekli('nexgen.recycle.manage', 'can_manage')
def api_recycle_stok_kartlari():
    """Sadece RECYCLE kategorisindeki aktif stok kartlarını döner (modal dropdown için)."""
    con = _db()
    try:
        kartlar = con.execute("""
            SELECT sk.id, sk.kod, sk.ad, sk.alt_kategori,
                   COALESCE(SUM(sh.miktar_kg), 0) AS mevcut_kg
            FROM nexgen_stok_kart sk
            LEFT JOIN nexgen_stok_hareket sh ON sh.stok_kart_id = sk.id
            WHERE sk.kategori = 'RECYCLE' AND sk.aktif = 1
            GROUP BY sk.id
            ORDER BY sk.ad
        """).fetchall()
    finally:
        con.close()
    return jsonify([dict(k) for k in kartlar])


# ─────────────────────────────────────────────────────────────
# NEXGEN FAZ-5F — Parçalı Üretim (nexgen_uretim_parca)
# Migration 072 sonrası aktif.
# MRP/Stok hareketi YOK. Barkod batch bazlı korunur.
# ─────────────────────────────────────────────────────────────

def _parca_tablosu_var(con):
    """nexgen_uretim_parca tablosunun var olup olmadığını kontrol eder."""
    tablolar = [r[0] for r in con.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='nexgen_uretim_parca'"
    ).fetchall()]
    return bool(tablolar)


def _batch_toplam_uretilen(con, batch_kodu):
    """Bir batch'in BITTI parçalarından toplam üretilen KG'yi döner."""
    row = con.execute(
        "SELECT COALESCE(SUM(uretilen_kg), 0) AS toplam "
        "FROM nexgen_uretim_parca "
        "WHERE batch_kodu=? AND durum='BITTI'",
        (batch_kodu,)
    ).fetchone()
    return float(row['toplam']) if row else 0.0


def _rf_kullanim_kolonlari(con):
    """nexgen_rf_kullanim tablo kolonlari; tablo yoksa bos set."""
    if not con.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='nexgen_rf_kullanim'"
    ).fetchone():
        return set()
    return {c['name'] for c in con.execute(
        "PRAGMA table_info(nexgen_rf_kullanim)"
    ).fetchall()}


def _rf_kullanim_tablet_aktif(con):
    cols = _rf_kullanim_kolonlari(con)
    return {'tablet_session_id', 'durum', 'miktar_kg'}.issubset(cols)


def _rf_renk_uv_icin(con, uretim_varyant_id):
    """UV + formul uygunluk uzerinden en uygun RF renk kaydini bulur."""
    row = con.execute("""
        SELECT rf.id AS rf_renk_id, rf.rf_kod, rf.cari_id, u.formul_id, rv.ad AS renk_ad
        FROM nexgen_uretim_varyant uv
        JOIN nexgen_renk_varyant rv ON rv.id = uv.renk_varyant_id
        JOIN nexgen_rf_formul_uygunluk u ON u.formul_id = rv.formul_id
            AND u.durum = 'ONAYLI' AND u.aktif = 1
        JOIN nexgen_rf_renk rf ON rf.id = u.rf_renk_id
            AND rf.aktif = 1 AND rf.durum = 'ONAYLI'
        WHERE uv.id = ?
        ORDER BY
            CASE
                WHEN UPPER(rf.ad) = UPPER(rv.ad) THEN 0
                WHEN UPPER(rf.ad) LIKE '%' || UPPER(rv.ad) || '%' THEN 1
                WHEN UPPER(rv.ad) LIKE '%' || UPPER(rf.ad) || '%' THEN 2
                ELSE 3
            END,
            rf.id DESC
        LIMIT 1
    """, (uretim_varyant_id,)).fetchone()
    return dict(row) if row else None


def _rf_cari_uyari(con, rf_cari_id, plan_cari_id):
    """RF cari ile plan cari farkliysa uyari metni (engelleme yok)."""
    if not rf_cari_id or not plan_cari_id or rf_cari_id == plan_cari_id:
        return None
    return (
        'Seçilen RF farklı bir cari kaydına bağlı; '
        'genel kullanılabilir renk olabilir.'
    )


def _plan_rf_bilgi(con, rf_renk_id=None, uretim_varyant_id=None):
    """Plan/batch RF bilgisi — önce rf_renk_id FK, yoksa UV otomatik eşleşme."""
    if rf_renk_id:
        row = con.execute("""
            SELECT rf.id AS rf_renk_id, rf.rf_kod, rf.cari_id, rf.ad AS renk_ad
            FROM nexgen_rf_renk rf
            WHERE rf.id = ? AND rf.aktif = 1 AND rf.durum = 'ONAYLI'
        """, (rf_renk_id,)).fetchone()
        if row:
            return dict(row)
    if uretim_varyant_id:
        return _rf_renk_uv_icin(con, uretim_varyant_id)
    return None


def _batch_plan_rf_bilgi(con, batch_kodu=None, plan_id=None, uretim_varyant_id=None):
    """Batch/plan bağlamında RF — plan.rf_renk_id öncelikli."""
    rf_renk_id = None
    if batch_kodu:
        row = con.execute("""
            SELECT nb.plan_id, nb.uretim_varyant_id, np.rf_renk_id
            FROM nexgen_uretim_batch nb
            LEFT JOIN nexgen_uretim_plan np ON np.id = nb.plan_id
            WHERE nb.batch_kodu = ?
        """, (batch_kodu,)).fetchone()
        if row:
            plan_id = plan_id or row['plan_id']
            uretim_varyant_id = uretim_varyant_id or row['uretim_varyant_id']
            rf_renk_id = row['rf_renk_id']
    elif plan_id and _plan_rf_renk_kolonu_var(con):
        row = con.execute(
            "SELECT rf_renk_id, uretim_varyant_id FROM nexgen_uretim_plan WHERE id=?",
            (plan_id,),
        ).fetchone()
        if row:
            rf_renk_id = row['rf_renk_id']
            uretim_varyant_id = uretim_varyant_id or row['uretim_varyant_id']
    return _plan_rf_bilgi(
        con, rf_renk_id=rf_renk_id, uretim_varyant_id=uretim_varyant_id,
    )


def _batch_uretim_miktar_kg(con, batch_kodu):
    """BITTI uretilen + DEVAM alt emir hedeflerinin toplami."""
    miktar = _batch_toplam_uretilen(con, batch_kodu)
    if _parca_tablosu_var(con):
        row = con.execute("""
            SELECT COALESCE(SUM(hedef_kg), 0) AS s
            FROM nexgen_uretim_parca
            WHERE batch_kodu = ? AND durum = 'DEVAM'
        """, (batch_kodu,)).fetchone()
        miktar += float(row['s'] or 0) if row else 0.0
    return round(miktar, 3)


def _rf_kullanim_tablet_sync(con, batch_kodu, uretim_emir_id=None, tamamlandi=False):
    """Tablet uretim akisindan RF kullanim kaydini olusturur veya gunceller."""
    if not _rf_kullanim_tablet_aktif(con):
        return None

    baglam = con.execute("""
        SELECT nb.batch_kodu, nb.plan_id, nb.uretim_varyant_id
        FROM nexgen_uretim_batch nb
        WHERE nb.batch_kodu = ?
    """, (batch_kodu,)).fetchone()
    if not baglam:
        return None

    rf = _batch_plan_rf_bilgi(
        con,
        batch_kodu=batch_kodu,
        plan_id=baglam['plan_id'],
        uretim_varyant_id=baglam['uretim_varyant_id'],
    )
    if not rf:
        return None

    cols = _rf_kullanim_kolonlari(con)
    miktar_kg = _batch_uretim_miktar_kg(con, batch_kodu)
    durum = 'TAMAMLANDI' if tamamlandi else 'URETIM'
    tablet_session_id = batch_kodu
    uid = _kullanici_id()

    mevcut = con.execute("""
        SELECT id FROM nexgen_rf_kullanim
        WHERE tablet_session_id = ? AND aktif = 1
        ORDER BY id DESC LIMIT 1
    """, (tablet_session_id,)).fetchone()

    if mevcut:
        sets = ["durum = ?", "miktar_kg = ?", "guncelleme_tarihi = datetime('now')"]
        params = [durum, miktar_kg]
        if uretim_emir_id is not None and 'uretim_emir_id' in cols:
            sets.append("uretim_emir_id = ?")
            params.append(uretim_emir_id)
        params.append(mevcut['id'])
        con.execute(
            f"UPDATE nexgen_rf_kullanim SET {', '.join(sets)} WHERE id = ?",
            params
        )
        return mevcut['id']

    col_names = [
        'rf_renk_id', 'formul_id', 'cari_id', 'siparis_id', 'aciklama',
        'olusturan_id', 'aktif', 'durum', 'miktar_kg', 'tablet_session_id',
    ]
    values = [
        rf['rf_renk_id'], rf.get('formul_id'), rf.get('cari_id'),
        baglam['plan_id'], 'Tablet uretim', uid, 1, durum, miktar_kg,
        tablet_session_id,
    ]
    if uretim_emir_id is not None and 'uretim_emir_id' in cols:
        col_names.append('uretim_emir_id')
        values.append(uretim_emir_id)

    ph = ','.join(['?'] * len(values))
    con.execute(
        f"INSERT INTO nexgen_rf_kullanim ({','.join(col_names)}) VALUES ({ph})",
        values
    )
    return con.execute("SELECT last_insert_rowid()").fetchone()[0]


def _batch_kalan_kg(con, batch_kodu):
    """Bir batch için kalan KG = planlanan_kg - toplam_uretilen (BITTI parçalar)."""
    batch = con.execute(
        "SELECT planlanan_kg FROM nexgen_uretim_batch WHERE batch_kodu=?",
        (batch_kodu,)
    ).fetchone()
    if not batch:
        return 0.0
    toplam = _batch_toplam_uretilen(con, batch_kodu)
    return max(0.0, float(batch['planlanan_kg']) - toplam)


def _parca_no_uret(con, batch_kodu):
    """Batch içinde bir sonraki parca_no'yu döner.
    Yeni numara sistemi: 1001'den başlar (1001, 1002, 1003, ...).
    Eski kayıtlar (1, 2, 3) varsa mevcut max'ın üstünden devam eder.
    """
    row = con.execute(
        "SELECT COALESCE(MAX(parca_no), 1000) AS maks "
        "FROM nexgen_uretim_parca WHERE batch_kodu=?",
        (batch_kodu,)
    ).fetchone()
    maks = row['maks'] if row else 1000
    # Eski kayıt (1-999) varsa 1001'den başlat
    return max(maks, 1000) + 1


@nexgen_bp.route('/api/batch/<batch_kodu>/parca/ekle', methods=['POST'])
@login_gerekli
def api_parca_ekle(batch_kodu):
    """Batch altına manuel üretim parçası ekle (admin override).

    FAZ-5G'den itibaren normal akışta alt emirler api_plan_basla ile
    otomatik oluşturulur. Bu endpoint yalnızca plan.manage yetkisiyle
    kullanılabilir (operatör erişemez).

    Body (JSON):
      hedef_kg  : float  — zorunlu, kalan KG'yi geçemez
      notlar    : str    — opsiyonel
      vardiya   : str    — opsiyonel (SABAH / OGLEN / AKSAM / GECE)

    Yetki: SADECE nexgen.plan.manage (operatöre kapalı)
    """
    if not yetki_var('nexgen.plan.manage', 'can_manage'):
        return jsonify({'ok': False, 'hata': 'Bu işlem için plan.manage yetkisi gerekli'}), 403

    d = request.get_json(silent=True) or {}
    hedef_kg = d.get('hedef_kg')
    notlar   = (d.get('notlar') or '').strip() or None
    vardiya  = (d.get('vardiya') or '').strip() or None

    if hedef_kg is None:
        return jsonify({'ok': False, 'hata': 'hedef_kg zorunlu'}), 400
    try:
        hedef_kg = float(hedef_kg)
        if hedef_kg <= 0:
            return jsonify({'ok': False, 'hata': 'hedef_kg sıfırdan büyük olmalı'}), 400
    except (TypeError, ValueError):
        return jsonify({'ok': False, 'hata': 'Geçersiz hedef_kg'}), 400

    con = _db()
    try:
        if not _parca_tablosu_var(con):
            return jsonify({'ok': False, 'hata': 'Migration 072 çalıştırılmadı'}), 500

        batch = con.execute(
            "SELECT id, batch_kodu, plan_id, planlanan_kg, durum "
            "FROM nexgen_uretim_batch WHERE batch_kodu=?",
            (batch_kodu,)
        ).fetchone()
        if not batch:
            return jsonify({'ok': False, 'hata': 'Batch bulunamadı'}), 404
        if batch['durum'] == 'BITTI':
            return jsonify({'ok': False, 'hata': 'Tamamlanmış batch\'e parça eklenemez'}), 400

        kalan = _batch_kalan_kg(con, batch_kodu)
        if hedef_kg > round(kalan, 3):
            return jsonify({
                'ok': False,
                'hata': f'Girilen miktar ({hedef_kg} KG) kalan KG\'yi ({round(kalan,3)} KG) aşıyor'
            }), 400

        parca_no = _parca_no_uret(con, batch_kodu)
        uid      = _kullanici_id()

        con.execute("""
            INSERT INTO nexgen_uretim_parca
                (batch_id, batch_kodu, plan_id, parca_no,
                 hedef_kg, uretilen_kg, durum,
                 operator_id, vardiya, notlar)
            VALUES (?, ?, ?, ?, ?, 0, 'HAZIR', ?, ?, ?)
        """, (
            batch['id'], batch_kodu, batch['plan_id'], parca_no,
            round(hedef_kg, 3), uid, vardiya, notlar
        ))
        con.commit()

        parca = con.execute(
            "SELECT * FROM nexgen_uretim_parca WHERE batch_kodu=? AND parca_no=?",
            (batch_kodu, parca_no)
        ).fetchone()

        return jsonify({
            'ok': True,
            'parca_no': parca_no,
            'parca_id': parca['id'],
            'hedef_kg': hedef_kg,
            'kalan_kg': round(kalan - hedef_kg, 3),
        })
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/batch/<batch_kodu>/alt-emir-olustur', methods=['POST'])
@login_gerekli
def api_batch_alt_emir_olustur(batch_kodu):
    """Mevcut bir batch için alt emirleri otomatik oluşturur.

    Alt emir yoksa formül KG'sine göre hesaplayıp toplu HAZIR olarak ekler.
    Alt emir zaten varsa dokunmaz (idempotent).
    Yetki: nexgen.tablet.uretim VEYA nexgen.plan.manage
    """
    if not (yetki_var('nexgen.tablet.uretim', 'can_uretim') or
            yetki_var('nexgen.plan.manage', 'can_manage')):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    import math as _math

    con = _db()
    try:
        # Batch'i getir
        batch = con.execute("""
            SELECT nb.id, nb.batch_kodu, nb.planlanan_kg, nb.durum,
                   nb.uretim_varyant_id, nb.plan_id
            FROM nexgen_uretim_batch nb
            WHERE nb.batch_kodu = ?
        """, (batch_kodu,)).fetchone()
        if not batch:
            return jsonify({'ok': False, 'hata': 'Batch bulunamadı'}), 404

        if not _parca_tablosu_var(con):
            return jsonify({'ok': False, 'hata': 'Alt emir tablosu (nexgen_uretim_parca) mevcut değil'}), 400

        # Zaten alt emir var mı?
        mevcut = con.execute(
            "SELECT COUNT(*) AS adet FROM nexgen_uretim_parca WHERE batch_kodu=?",
            (batch_kodu,)
        ).fetchone()
        if mevcut and mevcut['adet'] > 0:
            return jsonify({
                'ok': True,
                'mesaj': f'Bu batch için zaten {mevcut["adet"]} alt emir mevcut. Dokunulmadı.',
                'alt_emir_sayisi': mevcut['adet'],
                'olusturuldu': False,
            })

        # Formül batch KG hesapla
        formul_batch_kg = _formul_batch_kg_hesapla(con, batch['uretim_varyant_id'])
        if formul_batch_kg <= 0:
            return jsonify({'ok': False, 'hata': 'Formül reçete KG hesaplanamadı (aktif reçete kalemi yok)'}), 400

        planlanan = float(batch['planlanan_kg'])
        alt_emir_sayisi = _math.ceil(planlanan / formul_batch_kg)
        uid = _kullanici_id()
        plan_id = batch['plan_id'] if 'plan_id' in batch.keys() else None

        parca_cols = [c['name'] for c in con.execute(
            "PRAGMA table_info(nexgen_uretim_parca)"
        ).fetchall()]
        has_formul_kg_col = 'formul_batch_kg' in parca_cols

        for no in range(1001, 1001 + alt_emir_sayisi):
            if has_formul_kg_col:
                con.execute("""
                    INSERT INTO nexgen_uretim_parca
                        (batch_id, batch_kodu, plan_id, parca_no,
                         hedef_kg, formul_batch_kg, uretilen_kg,
                         durum, operator_id)
                    VALUES (?, ?, ?, ?, ?, ?, 0, 'HAZIR', ?)
                """, (batch['id'], batch_kodu, plan_id, no,
                      round(formul_batch_kg, 3),
                      round(formul_batch_kg, 3), uid))
            else:
                con.execute("""
                    INSERT INTO nexgen_uretim_parca
                        (batch_id, batch_kodu, plan_id, parca_no,
                         hedef_kg, uretilen_kg, durum, operator_id)
                    VALUES (?, ?, ?, ?, ?, 0, 'HAZIR', ?)
                """, (batch['id'], batch_kodu, plan_id, no,
                      round(formul_batch_kg, 3), uid))

        con.commit()
        return jsonify({
            'ok': True,
            'batch_kodu': batch_kodu,
            'alt_emir_sayisi': alt_emir_sayisi,
            'formul_batch_kg': formul_batch_kg,
            'olusturuldu': True,
            'mesaj': f'{alt_emir_sayisi} alt emir oluşturuldu (1001–{1000 + alt_emir_sayisi})',
        })
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/batch/<batch_kodu>/parca/<int:parca_id>/baslat', methods=['POST'])
@login_gerekli
def api_parca_baslat(batch_kodu, parca_id):
    """Parçayı HAZIR → DEVAM'a al, baslama_zamani atar.

    Yetki: nexgen.tablet.uretim VEYA nexgen.plan.manage
    """
    if not (yetki_var('nexgen.tablet.uretim', 'can_uretim') or
            yetki_var('nexgen.plan.manage', 'can_manage')):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    con = _db()
    try:
        if not _parca_tablosu_var(con):
            return jsonify({'ok': False, 'hata': 'Migration 072 çalıştırılmadı'}), 500

        parca = con.execute(
            "SELECT id, batch_kodu, durum FROM nexgen_uretim_parca "
            "WHERE id=? AND batch_kodu=?",
            (parca_id, batch_kodu)
        ).fetchone()
        if not parca:
            return jsonify({'ok': False, 'hata': 'Parça bulunamadı'}), 404
        if parca['durum'] not in ('HAZIR', 'BEKLEME'):
            return jsonify({
                'ok': False,
                'hata': f'{parca["durum"]} durumundaki parça başlatılamaz'
            }), 400

        con.execute("""
            UPDATE nexgen_uretim_parca
            SET durum='DEVAM',
                baslama_zamani=datetime('now','localtime'),
                updated_at=datetime('now','localtime')
            WHERE id=?
        """, (parca_id,))

        # Ana batch'i de DEVAM'a çek (henüz HAZIR ise)
        con.execute("""
            UPDATE nexgen_uretim_batch SET durum='DEVAM'
            WHERE batch_kodu=? AND durum='HAZIR'
        """, (batch_kodu,))

        _rf_kullanim_tablet_sync(con, batch_kodu, uretim_emir_id=parca_id)

        con.commit()
        return jsonify({'ok': True, 'parca_id': parca_id, 'durum': 'DEVAM'})
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/batch/<batch_kodu>/parca/<int:parca_id>/bitir', methods=['POST'])
@login_gerekli
def api_parca_bitir(batch_kodu, parca_id):
    """Alt Emiri DEVAM → BITTI yap.

    FAZ-5G: Operatör KG girmez.
    uretilen_kg = hedef_kg (formül KG) otomatik atanır.

    Body (JSON — opsiyonel):
      notlar : str — opsiyonel not

    Yetki: nexgen.tablet.uretim VEYA nexgen.plan.manage
    """
    if not (yetki_var('nexgen.tablet.uretim', 'can_uretim') or
            yetki_var('nexgen.plan.manage', 'can_manage')):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    d = request.get_json(silent=True) or {}
    notlar = (d.get('notlar') or '').strip() or None

    con = _db()
    try:
        if not _parca_tablosu_var(con):
            return jsonify({'ok': False, 'hata': 'Migration 072 çalıştırılmadı'}), 500

        parca = con.execute(
            "SELECT id, batch_kodu, durum, hedef_kg FROM nexgen_uretim_parca "
            "WHERE id=? AND batch_kodu=?",
            (parca_id, batch_kodu)
        ).fetchone()
        if not parca:
            return jsonify({'ok': False, 'hata': 'Alt Emir bulunamadı'}), 404
        if parca['durum'] not in ('DEVAM', 'HAZIR'):
            return jsonify({
                'ok': False,
                'hata': f'{parca["durum"]} durumundaki Alt Emir bitirilemez'
            }), 400

        # uretilen_kg = hedef_kg (formül KG), operatör girmez
        uretilen_kg = round(float(parca['hedef_kg']), 3)

        update_fields = [
            "durum='BITTI'",
            "uretilen_kg=?",
            "bitis_zamani=datetime('now','localtime')",
            "updated_at=datetime('now','localtime')",
        ]
        params = [uretilen_kg]
        if notlar:
            update_fields.append("notlar=?")
            params.append(notlar)
        params.append(parca_id)

        con.execute(
            f"UPDATE nexgen_uretim_parca SET {', '.join(update_fields)} WHERE id=?",
            params
        )

        _rf_kullanim_tablet_sync(con, batch_kodu, uretim_emir_id=parca_id)

        con.commit()
        kg = _batch_uretim_kontrol(con, batch_kodu) or {}

        # Toplam alt emir ve biten sayısı
        sayac = con.execute("""
            SELECT COUNT(*) AS toplam,
                   SUM(CASE WHEN durum='BITTI' THEN 1 ELSE 0 END) AS biten
            FROM nexgen_uretim_parca WHERE batch_kodu=?
        """, (batch_kodu,)).fetchone()
        toplam_emir = sayac['toplam'] if sayac else 0
        biten_emir  = sayac['biten'] or 0

        return jsonify({
            'ok': True,
            'parca_id': parca_id,
            'durum': 'BITTI',
            'uretilen_kg': uretilen_kg,
            'siparis_toplam_kg': kg.get('siparis_toplam_kg'),
            'toplam_uretilen_kg': round(float(kg.get('uretilen_kg') or 0), 3),
            'kalan_kg': round(float(kg.get('kalan_kg') or 0), 3),
            'kontrol_durum': kg.get('kontrol_durum'),
            'kontrol_uyari': kg.get('uyari'),
            'toplam_emir': toplam_emir,
            'biten_emir': biten_emir,
        })
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/batch/<batch_kodu>/parca/<int:parca_id>/beklet', methods=['POST'])
@login_gerekli
def api_parca_beklet(batch_kodu, parca_id):
    """Parçayı DEVAM → BEKLEME'ye al, bekleme sebebi kaydet.

    Body (JSON):
      sebep  : str — opsiyonel
      notlar : str — opsiyonel

    Yetki: nexgen.tablet.uretim VEYA nexgen.plan.manage
    """
    if not (yetki_var('nexgen.tablet.uretim', 'can_uretim') or
            yetki_var('nexgen.plan.manage', 'can_manage')):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    d = request.get_json(silent=True) or {}
    sebep  = (d.get('sebep') or '').strip() or None
    notlar = (d.get('notlar') or '').strip() or None

    con = _db()
    try:
        if not _parca_tablosu_var(con):
            return jsonify({'ok': False, 'hata': 'Migration 072 çalıştırılmadı'}), 500

        parca = con.execute(
            "SELECT id, batch_kodu, durum FROM nexgen_uretim_parca "
            "WHERE id=? AND batch_kodu=?",
            (parca_id, batch_kodu)
        ).fetchone()
        if not parca:
            return jsonify({'ok': False, 'hata': 'Parça bulunamadı'}), 404
        if parca['durum'] != 'DEVAM':
            return jsonify({
                'ok': False,
                'hata': f'Sadece DEVAM durumundaki parça bekletilebilir (şu an: {parca["durum"]})'
            }), 400

        con.execute("""
            UPDATE nexgen_uretim_parca
            SET durum='BEKLEME',
                bekleme_sebebi=?,
                notlar=COALESCE(?, notlar),
                updated_at=datetime('now','localtime')
            WHERE id=?
        """, (sebep, notlar, parca_id))
        con.commit()
        return jsonify({'ok': True, 'parca_id': parca_id, 'durum': 'BEKLEME'})
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/batch/<batch_kodu>/parca/liste')
@login_gerekli
def api_parca_liste(batch_kodu):
    """Bir batch'in tüm parçalarını listeler.

    Yetki: nexgen.tablet.view
    """
    if not yetki_var('nexgen.tablet.view', 'can_view'):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    con = _db()
    try:
        if not _parca_tablosu_var(con):
            return jsonify({'ok': True, 'parcalar': [], 'migration_eksik': True})

        parcalar = con.execute("""
            SELECT np.id, np.parca_no, np.hedef_kg, np.uretilen_kg,
                   np.durum, np.baslama_zamani, np.bitis_zamani,
                   np.bekleme_sebebi, np.notlar, np.vardiya,
                   np.created_at,
                   ku.KullaniciAdi AS operator_ad
            FROM nexgen_uretim_parca np
            LEFT JOIN sistem_kullanici ku ON ku.Id = np.operator_id
            WHERE np.batch_kodu = ?
            ORDER BY np.parca_no ASC
        """, (batch_kodu,)).fetchall()

        return jsonify({
            'ok': True,
            'batch_kodu': batch_kodu,
            'parcalar': [dict(p) for p in parcalar],
        })
    finally:
        con.close()


@nexgen_bp.route('/api/batch/<batch_kodu>/ilerleme')
@login_gerekli
def api_batch_ilerleme(batch_kodu):
    """Batch KG ilerleme özeti: planlanan / toplam_uretilen / kalan.

    Tablet ana ekranı ve uretim_islem ekranı için.
    Migration 072 yoksa placeholder sıfırlar döner.
    """
    if not yetki_var('nexgen.tablet.view', 'can_view'):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    con = _db()
    try:
        kg = _batch_uretim_kontrol(con, batch_kodu)
        if not kg:
            return jsonify({'ok': False, 'hata': 'Batch bulunamadı'}), 404

        parca_sayisi = 0
        if _parca_tablosu_var(con):
            parca_sayisi = con.execute(
                "SELECT COUNT(*) AS cnt FROM nexgen_uretim_parca WHERE batch_kodu=?",
                (batch_kodu,)
            ).fetchone()['cnt']

        return jsonify({
            'ok': True,
            'batch_kodu': batch_kodu,
            'siparis_toplam_kg': kg['siparis_toplam_kg'],
            'planlanan_kg': kg['siparis_toplam_kg'],
            'toplam_uretilen_kg': kg['uretilen_kg'],
            'kalan_kg': kg['kalan_kg'],
            'kontrol_durum': kg.get('kontrol_durum'),
            'kontrol_uyari': kg.get('uyari'),
            'kontrol_mesaj': kg.get('mesaj'),
            'parca_sayisi': parca_sayisi,
            'migration_eksik': not _parca_tablosu_var(con),
        })
    finally:
        con.close()


@nexgen_bp.route('/api/batch/<batch_kodu>/parca/toplu-baslat', methods=['POST'])
@login_gerekli
def api_parca_toplu_baslat(batch_kodu):
    """İlk N adet HAZIR alt emiri toplu DEVAM'a al.

    Body (JSON):
      adet : int — kaç tane başlatılacak (1-500)

    Mantık:
      parca_no sırasıyla ilk <adet> HAZIR kaydı DEVAM yapılır.
      BITTI / DEVAM emirler etkilenmez.
      Her emire baslama_zamani atanır.
      Ana batch HAZIR ise DEVAM'a çekilir.
    """
    if not (yetki_var('nexgen.tablet.uretim', 'can_uretim') or
            yetki_var('nexgen.plan.manage', 'can_manage')):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    d    = request.get_json(silent=True) or {}
    adet = d.get('adet', 1)
    try:
        adet = max(1, min(500, int(adet)))
    except Exception:
        return jsonify({'ok': False, 'hata': 'Geçersiz adet'}), 400

    con = _db()
    try:
        if not _parca_tablosu_var(con):
            return jsonify({'ok': False, 'hata': 'Migration 072 çalıştırılmadı'}), 500

        hedefler = con.execute("""
            SELECT id FROM nexgen_uretim_parca
            WHERE batch_kodu=? AND durum='HAZIR'
            ORDER BY parca_no ASC
            LIMIT ?
        """, (batch_kodu, adet)).fetchall()

        if not hedefler:
            return jsonify({'ok': False, 'hata': 'Başlatılacak HAZIR alt emir yok'}), 400

        ids = [r['id'] for r in hedefler]
        placeholders = ','.join(['?'] * len(ids))
        con.execute(f"""
            UPDATE nexgen_uretim_parca
            SET durum='DEVAM',
                baslama_zamani=datetime('now','localtime'),
                updated_at=datetime('now','localtime')
            WHERE id IN ({placeholders})
        """, ids)

        con.execute("""
            UPDATE nexgen_uretim_batch SET durum='DEVAM'
            WHERE batch_kodu=? AND durum='HAZIR'
        """, (batch_kodu,))

        ilk_id = ids[0] if ids else None
        _rf_kullanim_tablet_sync(con, batch_kodu, uretim_emir_id=ilk_id)

        con.commit()

        sayac = con.execute("""
            SELECT
              COUNT(*) AS toplam,
              SUM(CASE WHEN durum='HAZIR'   THEN 1 ELSE 0 END) AS hazir,
              SUM(CASE WHEN durum='DEVAM'   THEN 1 ELSE 0 END) AS devam,
              SUM(CASE WHEN durum='BEKLEME' THEN 1 ELSE 0 END) AS bekleme,
              SUM(CASE WHEN durum='BITTI'   THEN 1 ELSE 0 END) AS bitti
            FROM nexgen_uretim_parca WHERE batch_kodu=?
        """, (batch_kodu,)).fetchone()

        return jsonify({
            'ok': True,
            'baslanan': len(ids),
            'sayac': {
                'toplam':  sayac['toplam']  or 0,
                'hazir':   sayac['hazir']   or 0,
                'devam':   sayac['devam']   or 0,
                'bekleme': sayac['bekleme'] or 0,
                'bitti':   sayac['bitti']   or 0,
            },
        })
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/batch/<batch_kodu>/parca/toplu-bitir', methods=['POST'])
@login_gerekli
def api_parca_toplu_bitir(batch_kodu):
    """İlk N adet DEVAM alt emiri toplu BITTI yap.

    Body (JSON):
      adet : int — kaç tane bitirilecek (1-500)

    Mantık:
      parca_no sırasıyla ilk <adet> DEVAM kaydı BITTI yapılır.
      uretilen_kg = hedef_kg (formül KG) otomatik atanır.
      Tekrar basılırsa zaten BITTI olanlar atlanır, sonraki DEVAM emirlerden devam eder.
    """
    if not (yetki_var('nexgen.tablet.uretim', 'can_uretim') or
            yetki_var('nexgen.plan.manage', 'can_manage')):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    d    = request.get_json(silent=True) or {}
    adet = d.get('adet', 1)
    try:
        adet = max(1, min(500, int(adet)))
    except Exception:
        return jsonify({'ok': False, 'hata': 'Geçersiz adet'}), 400

    con = _db()
    try:
        if not _parca_tablosu_var(con):
            return jsonify({'ok': False, 'hata': 'Migration 072 çalıştırılmadı'}), 500

        hedefler = con.execute("""
            SELECT id, hedef_kg FROM nexgen_uretim_parca
            WHERE batch_kodu=? AND durum='DEVAM'
            ORDER BY parca_no ASC
            LIMIT ?
        """, (batch_kodu, adet)).fetchall()

        if not hedefler:
            return jsonify({'ok': False, 'hata': 'Bitirilecek DEVAM alt emir yok'}), 400

        biten_kg = 0.0
        for r in hedefler:
            uretilen = round(float(r['hedef_kg']), 3)
            con.execute("""
                UPDATE nexgen_uretim_parca
                SET durum='BITTI',
                    uretilen_kg=?,
                    bitis_zamani=datetime('now','localtime'),
                    updated_at=datetime('now','localtime')
                WHERE id=?
            """, (uretilen, r['id']))
            biten_kg += uretilen

        _rf_kullanim_tablet_sync(
            con, batch_kodu,
            uretim_emir_id=hedefler[-1]['id'] if hedefler else None
        )

        con.commit()

        kg = _batch_uretim_kontrol(con, batch_kodu) or {}

        sayac = con.execute("""
            SELECT
              COUNT(*) AS toplam,
              SUM(CASE WHEN durum='HAZIR'   THEN 1 ELSE 0 END) AS hazir,
              SUM(CASE WHEN durum='DEVAM'   THEN 1 ELSE 0 END) AS devam,
              SUM(CASE WHEN durum='BEKLEME' THEN 1 ELSE 0 END) AS bekleme,
              SUM(CASE WHEN durum='BITTI'   THEN 1 ELSE 0 END) AS bitti
            FROM nexgen_uretim_parca WHERE batch_kodu=?
        """, (batch_kodu,)).fetchone()

        return jsonify({
            'ok': True,
            'biten': len(hedefler),
            'biten_kg': round(biten_kg, 3),
            'siparis_toplam_kg': kg.get('siparis_toplam_kg'),
            'toplam_uretilen_kg': round(float(kg.get('uretilen_kg') or 0), 3),
            'kalan_kg': round(float(kg.get('kalan_kg') or 0), 3),
            'kontrol_durum': kg.get('kontrol_durum'),
            'kontrol_uyari': kg.get('uyari'),
            'sayac': {
                'toplam':  sayac['toplam']  or 0,
                'hazir':   sayac['hazir']   or 0,
                'devam':   sayac['devam']   or 0,
                'bekleme': sayac['bekleme'] or 0,
                'bitti':   sayac['bitti']   or 0,
            },
        })
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()


@nexgen_bp.route('/api/batch/<batch_kodu>/parca/secili-isle', methods=['POST'])
@login_gerekli
def api_parca_secili_isle(batch_kodu):
    """Seçili alt emirleri toplu işle: baslat / bitir / beklet.

    Body (JSON):
      islem   : 'baslat' | 'bitir' | 'beklet'
      ids     : [int, ...]  — parca id listesi
      sebep   : str (opsiyonel, beklet için)
      notlar  : str (opsiyonel)
    """
    if not (yetki_var('nexgen.tablet.uretim', 'can_uretim') or
            yetki_var('nexgen.plan.manage', 'can_manage')):
        return jsonify({'ok': False, 'hata': 'Yetki yok'}), 403

    d      = request.get_json(silent=True) or {}
    islem  = (d.get('islem') or '').strip()
    ids    = d.get('ids', [])
    sebep  = (d.get('sebep')  or '').strip() or None
    notlar = (d.get('notlar') or '').strip() or None

    if islem not in ('baslat', 'bitir', 'beklet', 'not'):
        return jsonify({'ok': False, 'hata': 'Geçersiz işlem'}), 400
    if not ids or not isinstance(ids, list):
        return jsonify({'ok': False, 'hata': 'id listesi boş'}), 400

    ids = [int(i) for i in ids[:200]]  # max 200

    con = _db()
    try:
        if not _parca_tablosu_var(con):
            return jsonify({'ok': False, 'hata': 'Migration 072 çalıştırılmadı'}), 500

        placeholders = ','.join(['?'] * len(ids))
        durum_filtresi = "" if islem == 'not' else ""
        parcalar = con.execute(
            f"SELECT id, durum, hedef_kg FROM nexgen_uretim_parca "
            f"WHERE id IN ({placeholders}) AND batch_kodu=?",
            ids + [batch_kodu]
        ).fetchall()

        islenen = 0
        atlanan = 0
        for p in parcalar:
            pid   = p['id']
            durum = p['durum']

            if islem == 'baslat':
                if durum not in ('HAZIR', 'BEKLEME'):
                    atlanan += 1; continue
                con.execute("""
                    UPDATE nexgen_uretim_parca
                    SET durum='DEVAM', baslama_zamani=datetime('now','localtime'),
                        updated_at=datetime('now','localtime')
                    WHERE id=?
                """, (pid,))

            elif islem == 'bitir':
                if durum not in ('DEVAM', 'HAZIR'):
                    atlanan += 1; continue
                uretilen = round(float(p['hedef_kg']), 3)
                fields = ["durum='BITTI'", "uretilen_kg=?",
                          "bitis_zamani=datetime('now','localtime')",
                          "updated_at=datetime('now','localtime')"]
                params = [uretilen]
                if notlar:
                    fields.append("notlar=?"); params.append(notlar)
                params.append(pid)
                con.execute(f"UPDATE nexgen_uretim_parca SET {', '.join(fields)} WHERE id=?", params)

            elif islem == 'beklet':
                if durum != 'DEVAM':
                    atlanan += 1; continue
                con.execute("""
                    UPDATE nexgen_uretim_parca
                    SET durum='BEKLEME', bekleme_sebebi=?,
                        notlar=COALESCE(?, notlar),
                        updated_at=datetime('now','localtime')
                    WHERE id=?
                """, (sebep, notlar, pid))

            elif islem == 'not':
                con.execute("""
                    UPDATE nexgen_uretim_parca
                    SET notlar=?,
                        updated_at=datetime('now','localtime')
                    WHERE id=?
                """, (notlar, pid))

            islenen += 1

        if islem == 'baslat' and islenen:
            con.execute("""
                UPDATE nexgen_uretim_batch SET durum='DEVAM'
                WHERE batch_kodu=? AND durum='HAZIR'
            """, (batch_kodu,))

        if islem in ('baslat', 'bitir') and islenen:
            son_id = ids[0] if ids else None
            _rf_kullanim_tablet_sync(con, batch_kodu, uretim_emir_id=son_id)

        con.commit()

        kg = _batch_uretim_kontrol(con, batch_kodu) or {}

        sayac = con.execute("""
            SELECT COUNT(*) AS toplam,
              SUM(CASE WHEN durum='HAZIR'   THEN 1 ELSE 0 END) AS hazir,
              SUM(CASE WHEN durum='DEVAM'   THEN 1 ELSE 0 END) AS devam,
              SUM(CASE WHEN durum='BEKLEME' THEN 1 ELSE 0 END) AS bekleme,
              SUM(CASE WHEN durum='BITTI'   THEN 1 ELSE 0 END) AS bitti
            FROM nexgen_uretim_parca WHERE batch_kodu=?
        """, (batch_kodu,)).fetchone()

        return jsonify({
            'ok': True,
            'islenen': islenen,
            'atlanan': atlanan,
            'siparis_toplam_kg': kg.get('siparis_toplam_kg'),
            'toplam_uretilen_kg': round(float(kg.get('uretilen_kg') or 0), 3),
            'kalan_kg': round(float(kg.get('kalan_kg') or 0), 3),
            'kontrol_durum': kg.get('kontrol_durum'),
            'kontrol_uyari': kg.get('uyari'),
            'sayac': {
                'toplam':  sayac['toplam']  or 0,
                'hazir':   sayac['hazir']   or 0,
                'devam':   sayac['devam']   or 0,
                'bekleme': sayac['bekleme'] or 0,
                'bitti':   sayac['bitti']   or 0,
            },
        })
    except Exception as e:
        con.rollback()
        return jsonify({'ok': False, 'hata': str(e)}), 500
    finally:
        con.close()

