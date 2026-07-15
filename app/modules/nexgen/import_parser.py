# -*- coding: utf-8 -*-
"""P4A — Excel parser (read-only, DB bağlantısı yok)."""
from __future__ import annotations

import hashlib
import os
import re
import unicodedata
from datetime import datetime
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from modules.nexgen.import_models import (
    GECERLI_AILELER,
    GECERLI_BOYUTLAR,
    HamAramaKaydi,
    HamCariKaydi,
    HamExcelVerisi,
    HamFormulKalemi,
    HamFormulSutunu,
    HamStokKaydi,
    KaynakIzi,
)

BEKLENEN_SAYFALAR = (
    "STOK_KARTLARI",
    "ARAMA_LISTESI",
    "TUM_FORMULLER",
    "KULLANIM",
    "CARI_LISTESI",
)

# TUM_FORMULLER sabit satır eşlemesi (kolon A etiketleri)
TUM_META_SATIRLAR = {
    3: "mamul_uretim_kodu",
    4: "musteri_formul_kodu",
    5: "formul_ad",
    6: "urun_ailesi",
    7: "formul_baslik",
    8: "varyant",
    9: "boyut",
    10: "kalip_carpani",
    11: "renk_kodu",
    12: "renk_adi",
    13: "formul_tarihi",
    14: "durum",
    15: "cari_kodu",
    16: "cari_unvan",
    17: "not_alan",
}
TUM_KONTROL_SATIR = 125
TUM_MATERIAL_BASLIK = 18
TUM_MATERIAL_BASLANGIC = 19
TUM_MATERIAL_BITIS = 119
TUM_ILK_FORMUL_SUTUN = 7  # G


def sha256_dosya(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _hucre_metin(cell) -> str:
    """Baştaki sıfırları koruyarak metin döner."""
    if cell is None or cell.value is None:
        return ""
    val = cell.value
    if isinstance(val, str):
        s = val.strip()
        if s.startswith("="):
            return s
        return s
    if isinstance(val, (int, float)):
        nf = (cell.number_format or "").strip()
        if nf == "@" or "@" in nf:
            return str(val)
        if isinstance(val, float) and val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _hucre_sayi(cell) -> float | None:
    if cell is None or cell.value is None:
        return None
    val = cell.value
    if isinstance(val, (int, float)):
        if val != val or val in (float("inf"), float("-inf")):
            return None
        return float(val)
    if isinstance(val, str):
        s = val.strip().replace(",", ".")
        if not s or s.startswith("="):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _kaynak(sayfa: str, satir: int, col: int, baslik: str = "") -> KaynakIzi:
    return KaynakIzi(
        sayfa=sayfa,
        satir=satir,
        sutun_baslik=baslik,
        hucre=f"{get_column_letter(col)}{satir}",
    )


def _normalize_boyut(raw: str) -> str:
    s = (raw or "").strip().upper()
    if s in GECERLI_BOYUTLAR:
        return s
    aliases = {"L": "LARGE", "S": "SMALL", "M": "MEDIUM", "STD": "STANDART"}
    return aliases.get(s, s)


def _normalize_aile(raw: str) -> str:
    s = (raw or "").strip().upper()
    s = s.replace("Ö", "O").replace("Ü", "U").replace("İ", "I")
    if "TERLIK" in s or "TERLİK" in (raw or "").upper():
        return "TERLIK"
    if "TABAN" in s:
        return "TABAN"
    return s if s in GECERLI_AILELER else s


def _infer_uretim_tipi(formul_ad: str, urun_ailesi: str) -> str:
    ad = (formul_ad or "").upper()
    if "DOKME" in ad or "DÖKME" in (formul_ad or "").upper():
        return "DOKME"
    if "ENJEK" in ad:
        return "ENJEKSIYON"
    return "ENJEKSIYON"


def parse_stok_kartlari(ws) -> list[HamStokKaydi]:
    out = []
    for r in range(2, (ws.max_row or 1) + 1):
        kod = _hucre_metin(ws.cell(r, 1))
        if not kod:
            continue
        out.append(HamStokKaydi(
            stok_kodu=kod,
            resmi_ad=_hucre_metin(ws.cell(r, 2)),
            kategori=_hucre_metin(ws.cell(r, 3)).upper(),
            birim=_hucre_metin(ws.cell(r, 4)).upper(),
            kaynak=_kaynak("STOK_KARTLARI", r, 1, "Stok Kodu"),
        ))
    return out


def parse_arama_listesi(ws) -> list[HamAramaKaydi]:
    out = []
    for r in range(2, (ws.max_row or 1) + 1):
        ara = _hucre_metin(ws.cell(r, 1))
        kod = _hucre_metin(ws.cell(r, 2))
        if not ara and not kod:
            continue
        out.append(HamAramaKaydi(
            kolay_arama=ara,
            stok_kodu=kod,
            kaynak=_kaynak("ARAMA_LISTESI", r, 1, "Kolay Arama Adı"),
        ))
    return out


def parse_cari_listesi(ws) -> list[HamCariKaydi]:
    out = []
    for r in range(2, (ws.max_row or 1) + 1):
        kod = _hucre_metin(ws.cell(r, 1))
        if not kod:
            continue
        out.append(HamCariKaydi(
            cari_kod=kod,
            unvan=_hucre_metin(ws.cell(r, 2)),
            durum=_hucre_metin(ws.cell(r, 3)),
            kaynak=_kaynak("CARI_LISTESI", r, 1, "Cari Kod"),
        ))
    return out


def parse_kullanim(ws) -> list[str]:
    lines = []
    for r in range(1, min(20, (ws.max_row or 1) + 1)):
        parts = []
        for c in range(1, min(6, (ws.max_column or 1) + 1)):
            t = _hucre_metin(ws.cell(r, c))
            if t:
                parts.append(t)
        if parts:
            lines.append(" | ".join(parts))
    return lines


def _formul_sutunlari_bul(ws) -> list[int]:
    """Dolu formül sütunlarını bul (satır 4'te müşteri formül kodu olan)."""
    cols = []
    max_col = ws.max_column or TUM_ILK_FORMUL_SUTUN
    for c in range(TUM_ILK_FORMUL_SUTUN, max_col + 1):
        kod = _hucre_metin(ws.cell(4, c))
        if kod and not kod.startswith("="):
            cols.append(c)
    return cols


def parse_tum_formuller(ws) -> tuple[list[HamFormulSutunu], list[str]]:
    uyarilar = []
    sutunlar = []

    for col_idx in _formul_sutunlari_bul(ws):
        harf = get_column_letter(col_idx)
        meta: dict[str, Any] = {}
        meta_kaynak: dict[str, KaynakIzi] = {}

        for satir, alan in TUM_META_SATIRLAR.items():
            cell = ws.cell(satir, col_idx)
            if alan == "kalip_carpani":
                meta[alan] = _hucre_sayi(cell)
            else:
                meta[alan] = _hucre_metin(cell)
            meta_kaynak[alan] = _kaynak("TUM_FORMULLER", satir, col_idx, alan)

        kontrol = _hucre_metin(ws.cell(TUM_KONTROL_SATIR, col_idx))

        kalemler: list[HamFormulKalemi] = []
        for r in range(TUM_MATERIAL_BASLANGIC, TUM_MATERIAL_BITIS + 1):
            malzeme = _hucre_metin(ws.cell(r, 2))
            stok_kod = _hucre_metin(ws.cell(r, 3))
            if not malzeme and not stok_kod:
                qty = _hucre_sayi(ws.cell(r, col_idx))
                if qty is None or qty == 0:
                    continue
            miktar = _hucre_sayi(ws.cell(r, col_idx))
            if miktar is None or miktar <= 0:
                continue
            if not malzeme and not stok_kod:
                continue

            kategori = _hucre_metin(ws.cell(r, 5)).upper()
            birim = _hucre_metin(ws.cell(r, 6)).upper()
            sira_raw = ws.cell(r, 1).value
            try:
                sira = int(sira_raw) if sira_raw is not None else r - TUM_MATERIAL_BASLANGIC + 1
            except (TypeError, ValueError):
                sira = r - TUM_MATERIAL_BASLANGIC + 1

            kalemler.append(HamFormulKalemi(
                sira=sira,
                malzeme_ara=malzeme,
                stok_kodu=stok_kod,
                resmi_ad=_hucre_metin(ws.cell(r, 4)),
                kategori=kategori,
                birim=birim,
                miktar_ham=miktar,
                formul_sutun=harf,
                formul_sutun_idx=col_idx,
                kaynak=KaynakIzi(
                    sayfa="TUM_FORMULLER",
                    satir=r,
                    sutun_baslik=harf,
                    hucre=f"{harf}{r}",
                    ham_deger=miktar,
                ),
            ))

        boyut = _normalize_boyut(meta.get("boyut", ""))
        if boyut and boyut not in GECERLI_BOYUTLAR:
            uyarilar.append(
                f"{harf}: Geçersiz boyut '{meta.get('boyut')}' → '{boyut}'"
            )

        aile = _normalize_aile(meta.get("urun_ailesi", ""))
        renk_kodu = meta.get("renk_kodu") or meta.get("musteri_formul_kodu", "")

        fs = HamFormulSutunu(
            sutun_harf=harf,
            sutun_idx=col_idx,
            mamul_uretim_kodu=meta.get("mamul_uretim_kodu", ""),
            musteri_formul_kodu=meta.get("musteri_formul_kodu", ""),
            formul_ad=meta.get("formul_ad", ""),
            urun_ailesi=aile,
            formul_baslik=meta.get("formul_baslik", ""),
            varyant=meta.get("varyant", ""),
            boyut=boyut,
            kalip_carpani=meta.get("kalip_carpani"),
            renk_kodu=renk_kodu,
            renk_adi=meta.get("renk_adi", ""),
            formul_tarihi=meta.get("formul_tarihi", ""),
            durum=meta.get("durum", ""),
            cari_kodu=meta.get("cari_kodu", ""),
            cari_unvan=meta.get("cari_unvan", ""),
            not_alan=meta.get("not_alan", ""),
            kontrol=kontrol,
            kalemler=kalemler,
            meta_kaynak=meta_kaynak,
        )
        sutunlar.append(fs)

    return sutunlar, uyarilar


def parse_excel(dosya_yolu: str) -> HamExcelVerisi:
    """Excel dosyasını okur; DB'ye yazmaz."""
    if not os.path.isfile(dosya_yolu):
        raise FileNotFoundError(f"Excel bulunamadı: {dosya_yolu}")

    st = os.stat(dosya_yolu)
    sha = sha256_dosya(dosya_yolu)
    modified = datetime.fromtimestamp(st.st_mtime).isoformat(sep=" ")

    wb = load_workbook(dosya_yolu, read_only=False, data_only=True)
    sayfalar = wb.sheetnames
    eksik = [s for s in BEKLENEN_SAYFALAR if s not in sayfalar]
    uyarilar = []
    if eksik:
        uyarilar.append(f"Eksik sayfalar: {eksik}")

    stok = []
    arama = []
    cari = []
    formul_cols = []
    kullanim = []

    if "STOK_KARTLARI" in wb.sheetnames:
        stok = parse_stok_kartlari(wb["STOK_KARTLARI"])
    if "ARAMA_LISTESI" in wb.sheetnames:
        arama = parse_arama_listesi(wb["ARAMA_LISTESI"])
    if "CARI_LISTESI" in wb.sheetnames:
        cari = parse_cari_listesi(wb["CARI_LISTESI"])
    if "KULLANIM" in wb.sheetnames:
        kullanim = parse_kullanim(wb["KULLANIM"])
    if "TUM_FORMULLER" in wb.sheetnames:
        formul_cols, tf_uyari = parse_tum_formuller(wb["TUM_FORMULLER"])
        uyarilar.extend(tf_uyari)

    wb.close()

    return HamExcelVerisi(
        dosya_yolu=os.path.abspath(dosya_yolu),
        dosya_sha256=sha,
        dosya_boyut=st.st_size,
        dosya_modified=modified,
        sayfa_adlari=sayfalar,
        stok_kartlari=stok,
        arama_listesi=arama,
        cari_listesi=cari,
        formul_sutunlari=formul_cols,
        kullanim_metni=kullanim,
        parser_uyarilari=uyarilar,
    )
