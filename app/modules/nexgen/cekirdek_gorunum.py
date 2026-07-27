"""Çekirdek formül/renk görünürlük ve Reçete Merkezi ağaç gruplama yardımcıları."""
from __future__ import annotations

import hashlib
import os
import re
import unicodedata
from typing import Any

# Gelecekte 4BA, 5BA genişletilebilir — şimdilik 1/2/3BA aktif çekirdek
CEKIRDEK_KOD_ONEKLERI = ('1BA-', '2BA-', '3BA-')
_CEKIRDEK_RE = re.compile(r'^(1|2|3)BA-', re.IGNORECASE)

_AILE_SIRA = {'TERLIK': 1, 'TABAN': 2}
_VARYANT_SIRA = {
    'TERLIK': {'18-28': 1, '18-22': 2, '18-POE': 3},
    'TABAN': {'TABAN 18-28': 1, 'DÖKME 18-28': 2},
}
_BOYUT_SIRA = {'LARGE': 1, 'SMALL': 2, 'MEDIUM': 3}


def cekirdek_formul_mu(kod: str | None) -> bool:
    """Formül kodu yeni çekirdek ailesinden mi (1BA / 2BA / 3BA)."""
    if not kod:
        return False
    return bool(_CEKIRDEK_RE.match(str(kod).strip()))


def yeni_secimde_gosterilebilir_mi(
    kod: str | None,
    *,
    formul: dict | None = None,
) -> bool:
    """Yeni kullanıcı seçimlerinde gösterilebilir mi — legacy ve test hariç."""
    del formul  # gelecekte formul.durum vb. için genişletme noktası
    return cekirdek_formul_mu(kod)


def cekirdek_kod_sql_filter(alias: str = 'f') -> str:
    """SQL WHERE parçası: yalnız aktif çekirdek kodları."""
    a = alias
    return (
        f"({a}.kod LIKE '1BA-%' OR {a}.kod LIKE '2BA-%' OR {a}.kod LIKE '3BA-%')"
    )


def _cekirdek_boyut(kod: str) -> str:
    k = kod.upper()
    if '-FM' in k:
        return 'MEDIUM'
    if '-FL' in k:
        return 'LARGE'
    if '-FS' in k:
        return 'SMALL'
    return ''


def _cekirdek_varyant_baslik(kod: str) -> str:
    k = kod.upper()
    if k.startswith('1BA-'):
        if k.endswith('01'):
            return '18-28'
        if k.endswith('02'):
            return '18-22'
        if k.endswith('03'):
            return '18-POE'
    if k.startswith('2BA-'):
        return 'TABAN 18-28'
    if k.startswith('3BA-'):
        return 'DÖKME 18-28'
    return k


def _cekirdek_aile(kod: str) -> tuple[str, str]:
    k = kod.upper()
    if k.startswith('1BA-'):
        return 'TERLIK', 'TERLİK'
    if k.startswith('2BA-'):
        return 'TABAN', 'TABAN'
    if k.startswith('3BA-'):
        return 'DOKME', 'DÖKME'
    return 'DIGER', 'DİĞER'


def _varyant_key(baslik: str) -> str:
    return baslik.upper().replace('İ', 'I').replace('Ö', 'O').replace('Ü', 'U')


def rc_cekirdek_agac_hazirla(formuller_guncel: list[dict]) -> list[dict]:
    """Reçete Merkezi sol ağaç: aile → varyant → boyut hiyerarşisi."""
    aile_map: dict[str, dict[str, Any]] = {}

    for f in formuller_guncel:
        kod = (f.get('kod') or '').strip()
        if not yeni_secimde_gosterilebilir_mi(kod):
            continue

        aile_key, aile_etiket = _cekirdek_aile(kod)
        varyant_baslik = _cekirdek_varyant_baslik(kod)
        var_key = _varyant_key(varyant_baslik)

        if aile_key not in aile_map:
            aile_map[aile_key] = {
                'aile_key': aile_key,
                'aile': aile_etiket,
                'varyantlar': {},
            }

        var_bucket = aile_map[aile_key]['varyantlar']
        if var_key not in var_bucket:
            var_bucket[var_key] = {
                'varyant_key': var_key,
                'baslik': varyant_baslik,
                'boyutlar': [],
            }

        for rv in f.get('rv_ozet') or []:
            for uv in rv.get('varyantlar') or []:
                if not uv.get('agac_goster', True):
                    continue
                boyut = (uv.get('boyut') or _cekirdek_boyut(kod)).upper()
                var_bucket[var_key]['boyutlar'].append({
                    'boyut': boyut,
                    'formul_id': f['id'],
                    'kod': kod,
                    'rv_id': rv['rv_id'],
                    'uv_id': uv['uv_id'],
                    'recete_durum': uv.get('recete_durum') or 'TASLAK',
                    'aile_key': aile_key,
                    'varyant_key': var_key,
                })

    agac: list[dict] = []
    for aile_key in sorted(aile_map.keys(), key=lambda k: _AILE_SIRA.get(k, 99)):
        node = aile_map[aile_key]
        varyant_list: list[dict] = []
        sira_map = _VARYANT_SIRA.get(aile_key, {})
        for var_key, var_node in sorted(
            node['varyantlar'].items(),
            key=lambda x: (sira_map.get(x[1]['baslik'], 99), x[1]['baslik']),
        ):
            boyutlar = sorted(
                var_node['boyutlar'],
                key=lambda b: (_BOYUT_SIRA.get(b['boyut'], 99), b['boyut']),
            )
            if boyutlar:
                varyant_list.append({
                    'varyant_key': var_node['varyant_key'],
                    'baslik': var_node['baslik'],
                    'boyutlar': boyutlar,
                })
        if varyant_list:
            agac.append({
                'aile_key': node['aile_key'],
                'aile': node['aile'],
                'varyantlar': varyant_list,
            })
    return agac


def _modul02_aile_norm(v: str | None) -> str:
    a = (v or '').strip().upper().replace('İ', 'I').replace('Ö', 'O')
    if a in ('TERLIK', 'TERLİK'):
        return 'TERLIK'
    if a == 'TABAN':
        return 'TABAN'
    if a in ('DOKME', 'DÖKME'):
        return 'DOKME'
    return a


def modul02_rc_formul_gruplari_hazirla(
    formuller_guncel: list[dict],
    *,
    tip_n: str | None = None,
) -> tuple[list[dict], list[dict], list[dict]]:
    """Reçete Merkezi ağacından MOD-02 Ana Formül grupları (birebir görünürlük)."""
    agac = rc_cekirdek_agac_hazirla(formuller_guncel)
    kod_map = {int(f['id']): f for f in formuller_guncel if f.get('id') is not None}

    formul_gruplar: list[dict] = []
    varyantlar: list[dict] = []
    formul_by_id: dict[int, dict] = {}

    for aile_node in agac:
        aile_key = aile_node.get('aile_key') or ''
        aile_etiket = aile_node.get('aile') or ''
        if tip_n and _modul02_aile_norm(aile_key) != tip_n:
            continue
        for var in aile_node.get('varyantlar') or []:
            baslik = (var.get('baslik') or '').strip()
            if not baslik:
                continue
            secenekler_map: dict[str, dict] = {}
            rv_ad = ''
            for b in var.get('boyutlar') or []:
                boyut = (b.get('boyut') or '').upper()
                try:
                    uv_id = int(b.get('uv_id'))
                    formul_id = int(b.get('formul_id'))
                    rv_id = int(b.get('rv_id')) if b.get('rv_id') is not None else None
                except (TypeError, ValueError):
                    continue
                if not boyut or boyut in secenekler_map:
                    continue
                kod = (b.get('kod') or '').strip()
                f_rec = kod_map.get(formul_id) or {}
                renk_ad = ''
                master_kg = 0.0
                for rv in f_rec.get('rv_ozet') or []:
                    if rv.get('rv_id') == rv_id:
                        renk_ad = (rv.get('ad') or '').strip()
                        for uv in rv.get('varyantlar') or []:
                            if uv.get('uv_id') == uv_id:
                                master_kg = float(uv.get('toplam_kg') or 0)
                                break
                        break
                if not rv_ad:
                    rv_ad = baslik
                secenekler_map[boyut] = {
                    'boyut': boyut,
                    'boyut_harf': boyut_kisaltma(boyut),
                    'formul_id': formul_id,
                    'formul_kod': kod,
                    'uv_id': uv_id,
                    'rv_id': rv_id,
                }
                varyantlar.append({
                    'id': uv_id,
                    'renk_varyant_id': rv_id,
                    'boyut': boyut,
                    'recete_durum': b.get('recete_durum') or 'TASLAK',
                    'rv_id': rv_id,
                    'renk_ad': renk_ad,
                    'formul_id': formul_id,
                    'formul_kod': kod,
                    'formul_ad': f_rec.get('ad') or kod,
                    'urun_ailesi': f_rec.get('urun_ailesi') or aile_etiket,
                    'master_kg': round(master_kg, 3),
                })
                if formul_id not in formul_by_id:
                    formul_by_id[formul_id] = {
                        'id': formul_id,
                        'kod': kod,
                        'ad': baslik,
                        'urun_ailesi': f_rec.get('urun_ailesi') or aile_etiket,
                    }
            secenekler = [
                secenekler_map[b]
                for b in _BOYUT_SIRA_SECIM
                if b in secenekler_map
            ]
            if not secenekler:
                continue
            formul_gruplar.append({
                'grup_key': f'{aile_etiket}|{baslik}',
                'baslik': baslik,
                'aile': aile_etiket,
                'rv_ad': rv_ad,
                'secenekler': secenekler,
            })

    formuller = sorted(
        formul_by_id.values(),
        key=lambda f: (f.get('kod') or '', f.get('ad') or ''),
    )
    return formul_gruplar, varyantlar, formuller


# ── Renk çekirdek görünürlük ─────────────────────────────────────────────

_RENK_SAYISAL_RE = re.compile(r'^(\d{3,4})', re.IGNORECASE)
_RF_LEGACY_PREFIX_RE = re.compile(r'^RF-', re.IGNORECASE)
_NX_PREFIX_RE = re.compile(r'^NX-', re.IGNORECASE)
_EXCEL_RENK_KODLARI_CACHE: frozenset[str] | None = None
_EXCEL_RENK_ONEK_CACHE: frozenset[str] | None = None
_EXCEL_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), '..', '..', '..', 'import_files', 'nexgen_renk.xlsx')
)


def _normalize_renk_kod_metin(s: str | None) -> str:
    if not s:
        return ''
    t = unicodedata.normalize('NFKC', str(s))
    t = t.replace('–', '-').replace('—', '-')
    t = t.strip().upper()
    return re.sub(r'\s+', ' ', t)


def renk_sayisal_onek(rf_kod: str | None) -> str | None:
    """Renk kodunun başındaki 3–4 haneli sayısal önek (ör. 0677, 099)."""
    if not rf_kod:
        return None
    m = _RENK_SAYISAL_RE.match(str(rf_kod).strip())
    return m.group(1) if m else None


def excel_renk_kodlari_yukle() -> frozenset[str]:
    """nexgen_renk.xlsx içindeki kanonik renk kodları (normalize edilmiş)."""
    global _EXCEL_RENK_KODLARI_CACHE, _EXCEL_RENK_ONEK_CACHE
    if _EXCEL_RENK_KODLARI_CACHE is not None:
        return _EXCEL_RENK_KODLARI_CACHE

    kodlar: set[str] = set()
    if os.path.isfile(_EXCEL_PATH):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(_EXCEL_PATH, read_only=True, data_only=True)
            if 'Renkler' in wb.sheetnames:
                ws = wb['Renkler']
                for row in ws.iter_rows(values_only=True):
                    for c in range(0, len(row), 3):
                        if c + 1 < len(row) and row[c] and str(row[c + 1]).strip().lower() == 'miktar (gr)':
                            kodlar.add(_normalize_renk_kod_metin(str(row[c])))
            wb.close()
        except Exception:
            pass

    _EXCEL_RENK_KODLARI_CACHE = frozenset(k for k in kodlar if k)
    onekler = {o for o in (renk_sayisal_onek(k) for k in _EXCEL_RENK_KODLARI_CACHE) if o}
    _EXCEL_RENK_ONEK_CACHE = frozenset(onekler)
    return _EXCEL_RENK_KODLARI_CACHE


def excel_renk_onekleri_yukle() -> frozenset[str]:
    """Excel kaynaklı sayısal renk önekleri (0001, 0677, 099 …)."""
    excel_renk_kodlari_yukle()
    return _EXCEL_RENK_ONEK_CACHE or frozenset()


def _rf_alan(rf: dict | Any, *anahtarlar: str, varsayilan=None):
    get = rf.get if hasattr(rf, 'get') else (lambda k, d=None: rf[k] if k in rf.keys() else d)
    for anahtar in anahtarlar:
        val = get(anahtar)
        if val is not None:
            return val
    return varsayilan


def cekirdek_renk_mi(rf: dict | Any) -> bool:
    """Sayısal kodlu, legacy/test olmayan çekirdek renk mi."""
    if rf is None:
        return False
    kod = (_rf_alan(rf, 'rf_kod', 'kod', 'rf_kodu') or '').strip()
    if not kod:
        return False
    if _RF_LEGACY_PREFIX_RE.match(kod) or _NX_PREFIX_RE.match(kod):
        return False
    if _rf_alan(rf, 'kaynak_arge_test_id'):
        return False

    kod_n = _normalize_renk_kod_metin(kod)
    if kod_n in ('BORDO',):
        return False
    for tok in ('TEST', 'RECON', 'DENEME', 'DEBUG'):
        if tok in kod_n:
            return False
    if not renk_sayisal_onek(kod):
        return False

    onek = renk_sayisal_onek(kod)
    excel_onekler = excel_renk_onekleri_yukle()
    if excel_onekler:
        return onek in excel_onekler
    excel_set = excel_renk_kodlari_yukle()
    if excel_set:
        return kod_n in excel_set
    return True


def yeni_secimde_renk_gosterilebilir_mi(rf: dict | Any) -> bool:
    """Yeni kullanıcı seçimlerinde gösterilebilir çekirdek renk."""
    if not cekirdek_renk_mi(rf):
        return False
    if _rf_alan(rf, 'aktif', 'rf_aktif', varsayilan=1) != 1:
        return False
    durum = (_rf_alan(rf, 'durum', 'rf_durum') or '').upper()
    return durum in ('ONAYLI', 'AKTIF', 'URETIME_ACIK')


def renk_merkezi_listede_gosterilebilir_mi(rf: dict | Any) -> bool:
    """Renk Merkezi sol liste — aktif ve pasif çekirdek renkler."""
    if not cekirdek_renk_mi(rf):
        return False
    durum = (_rf_alan(rf, 'durum', 'rf_durum') or '').upper()
    return durum in ('ONAYLI', 'AKTIF', 'URETIME_ACIK')


_BOYUT_HARF = {'LARGE': 'L', 'SMALL': 'S', 'MEDIUM': 'M'}


def boyut_kisaltma(boyut: str | None) -> str:
    """UV.boyut (LARGE/SMALL/MEDIUM) → seçim harfi (L/S/M)."""
    if not boyut:
        return ''
    return _BOYUT_HARF.get(str(boyut).strip().upper(), '')


def cekirdek_secim_adi(kod: str | None) -> str:
    """Seçim ekranı adı — DB ad kullanılmaz (ENJEKSİYON vb. engellenir).

    1BA → TERLİK 18-28 / 18-22 / 18-POE
    2BA → TABAN 18-28
    3BA → DÖKME 18-28
    """
    k = (kod or '').strip()
    if not cekirdek_formul_mu(k):
        return ''
    ku = k.upper()
    _, aile = _cekirdek_aile(k)
    varyant = _cekirdek_varyant_baslik(k)
    if ku.startswith('1BA-'):
        return f'{aile} {varyant}'.strip()
    if ku.startswith('3BA-'):
        return varyant  # DÖKME 18-28
    if ku.startswith('2BA-'):
        return varyant  # TABAN 18-28
    return varyant or aile


def cekirdek_formul_gosterim(
    kod: str | None,
    ad: str | None = None,
    *,
    uv_boyut: str | None = None,
) -> dict:
    """Çekirdek formül seçim listesi etiketi.

    uv_boyut: gerçek nexgen_uretim_varyant.boyut — harf (L/S/M) yalnız bundan üretilir.
    """
    k = (kod or '').strip()
    aile_key, aile = _cekirdek_aile(k)
    if k.upper().startswith('3BA-'):
        aile = 'DÖKME'
    varyant = _cekirdek_varyant_baslik(k)
    boyut_kod = _cekirdek_boyut(k)
    secim_adi = cekirdek_secim_adi(k) or (ad or '').strip()
    harf = boyut_kisaltma(uv_boyut)
    etiket_ad = f'{secim_adi} ({harf})' if secim_adi and harf else secim_adi
    boyut_goster = (str(uv_boyut).strip().upper() if uv_boyut else '') or boyut_kod
    parcalar = [secim_adi] if secim_adi else [p for p in (aile, varyant) if p]
    if boyut_goster:
        parcalar.append(boyut_goster)
    return {
        'kod': k,
        'ad': (ad or '').strip(),
        'aile': aile,
        'varyant': varyant,
        'boyut': boyut_goster,
        'secim_adi': secim_adi,
        'boyut_harf': harf,
        'etiket_ad': etiket_ad,
        'alt_metin': ' • '.join(p for p in parcalar if p),
    }


def formul_secim_gosterim_uygula(
    satirlar: list[dict],
    *,
    kod_alan: str = 'kod',
    ad_alan: str = 'ad',
    id_alan: str = 'id',
    boyut_alan: str | None = None,
    uv_boyut_map: dict[int, str] | None = None,
    harf_ekle: bool = True,
) -> list[dict]:
    """Çekirdek satırlarda ad alanını seçim adı (+ isteğe bağlı L/S/M) ile değiştirir."""
    out: list[dict] = []
    for s in satirlar:
        d = dict(s)
        kod = d.get(kod_alan)
        if not cekirdek_formul_mu(kod):
            out.append(d)
            continue
        uv_b = None
        if boyut_alan and d.get(boyut_alan):
            uv_b = d.get(boyut_alan)
        elif uv_boyut_map is not None:
            try:
                uv_b = uv_boyut_map.get(int(d.get(id_alan)))
            except (TypeError, ValueError):
                uv_b = uv_boyut_map.get(d.get(id_alan))
        g = cekirdek_formul_gosterim(kod, d.get(ad_alan), uv_boyut=uv_b)
        if harf_ekle:
            d[ad_alan] = g['etiket_ad'] or g['secim_adi'] or d.get(ad_alan)
        else:
            d[ad_alan] = g['secim_adi'] or d.get(ad_alan)
        d['secim_adi'] = g['secim_adi']
        d['boyut_harf'] = g['boyut_harf']
        out.append(d)
    return out


_BOYUT_SIRA_SECIM = ('LARGE', 'SMALL', 'MEDIUM', 'STANDART')
_AILE_SIRA_SECIM = {'TERLIK': 1, 'TABAN': 2, 'DOKME': 3, 'DÖKME': 3}
# Model sırası — rv.ad / başlıktaki model parçası (kod FL/FS parse edilmez)
_MODEL_SIRA_SECIM = {
    '18-28': 1,
    '18-22': 2,
    '18-POE': 3,
}


def _secim_model_anahtari(rv_ad: str | None, baslik: str | None) -> str:
    """Grup sıralaması için model anahtarı — önce rv.ad, yoksa başlıktan son parça."""
    rv = (rv_ad or '').strip().upper().replace('İ', 'I')
    if rv:
        return rv
    b = (baslik or '').strip().upper().replace('İ', 'I')
    if not b:
        return ''
    parcalar = b.split()
    return parcalar[-1] if parcalar else b


def formul_secim_gruplari_hazirla(
    uv_satirlari: list[dict],
    *,
    include_non_cekirdek: bool = False,
) -> list[dict]:
    """Aynı ana formülün LARGE/SMALL UV kayıtlarını tek seçim grubunda birleştirir.

    Grup anahtarı: urun_ailesi + secim_adi (DB formul.ad kardeşleri aynı adı paylaşır;
    gösterim adı ENJEKSİYON yerine TABAN/DÖKME olur).
    Boyut seçenekleri nexgen_uretim_varyant.boyut üzerinden gelir.

    include_non_cekirdek=True: aktif urun_ailesi dolu legacy formülleri de gruplar
    (FAZ-3 hydrate TERLİK tam liste — yalnız 1BA üçlüsü değil).
    """
    buckets: dict[str, dict[str, Any]] = {}
    for uv in uv_satirlari:
        kod = (uv.get('formul_kod') or uv.get('kod') or '').strip()
        boyut = (uv.get('boyut') or '').strip().upper()
        if not boyut:
            continue
        aile_db = (uv.get('urun_ailesi') or '').strip().upper()
        if cekirdek_formul_mu(kod):
            g = cekirdek_formul_gosterim(kod, uv.get('formul_ad'), uv_boyut=boyut)
            baslik = g.get('secim_adi') or ''
            if not baslik:
                continue
            aile = aile_db or (g.get('aile') or '').upper().replace('İ', 'I')
            boyut_harf = g.get('boyut_harf') or boyut_kisaltma(boyut)
        else:
            if not include_non_cekirdek:
                continue
            # Legacy / NX aktif ana formül — aile yoksa ad/koddan çıkar
            baslik = (uv.get('formul_ad') or uv.get('ad') or kod).strip()
            if not baslik:
                continue
            if not aile_db:
                ad_u = baslik.upper().replace('İ', 'I').replace('Ö', 'O')
                if 'TERLIK' in ad_u:
                    aile_db = 'TERLIK'
                elif 'DOKME' in ad_u:
                    aile_db = 'DOKME'
                elif 'TABAN' in ad_u:
                    aile_db = 'TABAN'
            if not aile_db:
                continue
            aile = aile_db
            boyut_harf = boyut_kisaltma(boyut)
            g = {'secim_adi': baslik, 'aile': aile, 'boyut_harf': boyut_harf}
        if baslik.upper().startswith('DÖKME') or baslik.upper().startswith('DOKME'):
            aile_grup = 'DÖKME'
        elif aile in ('TERLIK', 'TERLİK'):
            aile_grup = 'TERLİK'
        elif aile == 'TABAN':
            aile_grup = 'TABAN'
        else:
            aile_grup = g.get('aile') or aile or 'DİĞER'
        grup_key = f'{aile_grup}|{baslik}'
        bucket = buckets.get(grup_key)
        if not bucket:
            bucket = {
                'grup_key': grup_key,
                'baslik': baslik,
                'aile': aile_grup,
                'rv_ad': (uv.get('renk_ad') or uv.get('rv_ad') or '').strip(),
                'secenekler_map': {},
            }
            buckets[grup_key] = bucket
        if boyut in bucket['secenekler_map']:
            continue
        try:
            uv_id = int(uv.get('id') or uv.get('uv_id'))
            formul_id = int(uv.get('formul_id'))
            rv_id = int(uv.get('rv_id')) if uv.get('rv_id') is not None else None
        except (TypeError, ValueError):
            continue
        bucket['secenekler_map'][boyut] = {
            'boyut': boyut,
            'boyut_harf': boyut_harf,
            'formul_id': formul_id,
            'formul_kod': kod,
            'uv_id': uv_id,
            'rv_id': rv_id,
        }

    sonuc: list[dict] = []
    for bucket in buckets.values():
        secenekler = [
            bucket['secenekler_map'][b]
            for b in _BOYUT_SIRA_SECIM
            if b in bucket['secenekler_map']
        ]
        if not secenekler:
            continue
        sonuc.append({
            'grup_key': bucket['grup_key'],
            'baslik': bucket['baslik'],
            'aile': bucket['aile'],
            'rv_ad': bucket['rv_ad'],
            'secenekler': secenekler,
        })

    def _sira(g: dict) -> tuple:
        aile = (g.get('aile') or '').upper().replace('İ', 'I').replace('Ö', 'O')
        aile_s = _AILE_SIRA_SECIM.get(aile, 50)
        model = _secim_model_anahtari(g.get('rv_ad'), g.get('baslik'))
        model_s = _MODEL_SIRA_SECIM.get(model, 50)
        return (aile_s, model_s, g.get('baslik') or '')

    sonuc.sort(key=_sira)
    return sonuc


def renk_liste_grubu(rf: dict) -> str:
    """Renk Merkezi liste grubu: BEKLEYEN / AKTİF / REVİZE / PASİF."""
    if _rf_alan(rf, 'aktif', 'rf_aktif', varsayilan=1) == 0:
        return 'PASİF'
    rev_no = int(_rf_alan(rf, 'aktif_rev_no', varsayilan=1) or 1)
    if rev_no > 1:
        return 'REVİZE'
    durum = (_rf_alan(rf, 'durum', 'rf_durum') or '').upper()
    if durum == 'ONAYLI':
        return 'AKTİF'
    return 'PASİF'


def renk_kart_tekillestir(rf_satirlari: list[dict]) -> list[dict]:
    """Aynı normalize renk kodunda tek kart — en düşük rf_id kalır."""
    by_kod: dict[str, dict] = {}
    for r in rf_satirlari:
        nk = _normalize_renk_kod_metin(r.get('rf_kod') or r.get('rf_kodu') or '')
        if not nk:
            continue
        rid = int(r.get('rf_id') or r.get('id') or 0)
        prev = by_kod.get(nk)
        if not prev or rid < int(prev.get('rf_id') or prev.get('id') or 0):
            by_kod[nk] = r
    return sorted(by_kod.values(), key=lambda x: _normalize_renk_kod_metin(x.get('rf_kod') or ''))


def renk_icerik_fp_stok_bazli(kalemler: list[dict]) -> str:
    """İçerik fingerprint: stok_kart_id + miktar_gr (isim yazımı yok sayılır)."""
    parts: list[str] = []
    for k in sorted(
        kalemler,
        key=lambda x: (int(x.get('stok_kart_id') or x.get('stok_id') or 0), int(x.get('sira') or 0)),
    ):
        sk = int(k.get('stok_kart_id') or k.get('stok_id') or 0)
        if 'miktar_gr' in k and k['miktar_gr'] is not None:
            gr = round(float(k['miktar_gr']), 6)
        else:
            gr = round(float(k.get('miktar_kg') or 0) * 1000, 6)
        parts.append(f'{sk}|{gr:.6f}')
    raw = ';'.join(parts)
    return hashlib.sha256(raw.encode('utf-8')).hexdigest()
