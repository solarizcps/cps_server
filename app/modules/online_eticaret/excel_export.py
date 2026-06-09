"""Sipariş paketlerini Excel (.xlsx) dosyasına çevirir.

İki sayfa üretir:
  1) "Kargo Listesi": her ürün satırı ayrı satır. Kargo son tarihine göre sıralı,
     gecikenler kırmızı. Ayrıca 'Model+Renk+Beden' birleşik kolonu içerir.
  2) "Özet": aynı Model+Renk+Beden'den toplam kaç adet gerektiği (depo toplama).
     Aciliyete göre sıralı: en acil (en çok geciken / en az süre kalan) en üstte.
     Her grup, içindeki en acil ürünün son tarihine göre sıralanır.
"""

import time
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    "Mağaza",
    "Gecikme Durumu",
    "Sipariş Tarihi",
    "Teslimat Son Tarihi",
    "Sipariş No",
    "Paket/Teslimat No",
    "Ürün Adı",
    "Stok Kodu",
    "Model Kodu",
    "Renk",
    "Beden",
    "Model+Renk+Beden",
    "Barkod",
    "Adet",
    "Birim Fiyat (indirimsiz)",
    "Net Tutar (faturalanacak)",
    "Alıcı Adı",
    "Kargo Takip No",
    "Kargo Firması",
]

SUMMARY_HEADERS = [
    "Model+Renk+Beden",
    "Ürün Adı",
    "Model Kodu",
    "Renk",
    "Beden",
    "Toplam Adet",
    "Sipariş Sayısı",
    "Gecikmeli Sipariş",
    "En Acil Durum",
]

# Boyama için kolon indeksleri
_DELAY_COL = HEADERS.index("Gecikme Durumu")
_SUM_STATUS_COL = SUMMARY_HEADERS.index("En Acil Durum")
_SUM_DELAYED_COL = SUMMARY_HEADERS.index("Gecikmeli Sipariş")
_DELAYED_PREFIX = "GECİKTİ"
_NO_DEADLINE = float("inf")

# Stiller (modül düzeyinde tek sefer)
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="F27A1A")  # Trendyol turuncu
_ROW_FILL = PatternFill("solid", fgColor="FDE0E0")  # açık kırmızı (tüm satır)
_DELAY_FILL = PatternFill("solid", fgColor="D32F2F")  # koyu kırmızı (hücre)
_DELAY_FONT = Font(bold=True, color="FFFFFF")


def _num(value):
    """Sayıya çevirir, çevrilemezse None döner."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _qty(line):
    """Satırın adedini tam sayı olarak döndürür (yoksa 0)."""
    v = _num(line.get("quantity"))
    return int(v) if v is not None else 0


def _unit_price(line):
    """İndirim uygulanmamış birim fiyat."""
    for key in ("price", "amount"):
        val = _num(line.get(key))
        if val is not None:
            return val
    return None


def _net_amount(line):
    """İndirimler düşülmüş net satır tutarı (faturalanacak)."""
    amount = _num(line.get("amount"))
    if amount is None:
        amount = _num(line.get("price"))
    if amount is None:
        return None
    discount = _num(line.get("discount")) or 0
    ty_discount = _num(line.get("tyDiscount")) or 0
    return round(amount - discount - ty_discount, 2)


def _model_code_from_name(name, size):
    """Ürün adından model kodunu çıkarır (yedek yöntem).

    Model kodu, ürün adında bedenden hemen önceki son kelimedir:
    'Kaptan Çocuk ... Terlik CRF-8240, 35' -> 'CRF-8240'
    """
    name = (name or "").strip()
    if not name:
        return ""
    parts = [p.strip() for p in name.split(",")]
    size = str(size or "").strip()
    if len(parts) > 1 and size and parts[-1] == size:
        parts = parts[:-1]
    head = parts[-1] if parts else ""
    tokens = head.split()
    return tokens[-1] if tokens else ""


def _model_code(line, model_map):
    """Önce ürün API'sinden (barkod->productMainId), bulunamazsa ad'dan çıkarır."""
    barcode = str(line.get("barcode") or "")
    if model_map and barcode in model_map:
        return model_map[barcode]
    return _model_code_from_name(line.get("productName"), line.get("productSize"))


def _variant_key(model_code, color, size):
    """Model+Renk+Beden birleşik anahtarı: 'V-COMFRT-Z6188_Beyaz_38'."""
    return f"{model_code}_{(color or '').strip()}_{(size or '').strip()}"


# Dışarıdan (preview ucu) kullanım için açık adlar
def variant_model_code(line, model_map=None):
    return _model_code(line, model_map)


def variant_key(model_code, color, size):
    return _variant_key(model_code, color, size)


def _customer_name(order):
    first = (order.get("customerFirstName") or "").strip()
    last = (order.get("customerLastName") or "").strip()
    return (first + " " + last).strip()


def _fmt_dt(ms):
    """Milisaniye epoch -> 'GG.AA.YYYY SS:DD' (yoksa boş)."""
    try:
        ms = int(ms)
    except (TypeError, ValueError):
        return ""
    return datetime.fromtimestamp(ms / 1000).strftime("%d.%m.%Y %H:%M")


def _human_dur(ms):
    """Milisaniyeyi 'X gün Y saat' biçiminde okunur süreye çevirir."""
    total_min = int(ms // 60000)
    days, rem = divmod(total_min, 1440)
    hours, minutes = divmod(rem, 60)
    if days:
        return f"{days} gün {hours} saat"
    if hours:
        return f"{hours} saat {minutes} dk"
    return f"{minutes} dk"


def _deadline_ms(order):
    """Siparişin kargo son tarihi (agreedDeliveryDate) ms; yoksa +sonsuz."""
    try:
        return int(order.get("agreedDeliveryDate"))
    except (TypeError, ValueError):
        return _NO_DEADLINE


def _delay_text(deadline_ms, now_ms):
    """(metin, gecikti_mi) — son tarihe göre."""
    if deadline_ms == _NO_DEADLINE:
        return ("", False)
    if deadline_ms < now_ms:
        return (f"{_DELAYED_PREFIX} ({_human_dur(now_ms - deadline_ms)})", True)
    return (f"Kalan: {_human_dur(deadline_ms - now_ms)}", False)


def orders_to_rows(orders, store_name="", model_map=None, now_ms=None):
    """Sipariş paketlerini düz satır listesine çevirir (her ürün ayrı satır).

    Siparişler kargo son tarihine göre artan sıralanır (en acil en üstte).
    """
    if now_ms is None:
        now_ms = int(time.time() * 1000)

    rows = []
    for order in sorted(orders, key=_deadline_ms):
        order_number = order.get("orderNumber", "")
        package_no = order.get("id", "")
        customer = _customer_name(order)
        cargo_no = order.get("cargoTrackingNumber", "")
        cargo_provider = order.get("cargoProviderName", "")
        order_date = _fmt_dt(order.get("orderDate"))
        deadline = _fmt_dt(order.get("agreedDeliveryDate"))
        delay_str, _ = _delay_text(_deadline_ms(order), now_ms)

        for line in order.get("lines") or []:
            model = _model_code(line, model_map)
            color = line.get("productColor", "")
            size = line.get("productSize", "")
            rows.append(
                [
                    store_name,
                    delay_str,
                    order_date,
                    deadline,
                    order_number,
                    package_no,
                    line.get("productName", ""),
                    line.get("merchantSku", ""),
                    model,
                    color,
                    size,
                    _variant_key(model, color, size),
                    line.get("barcode", ""),
                    _qty(line),
                    _unit_price(line),
                    _net_amount(line),
                    customer,
                    cargo_no,
                    cargo_provider,
                ]
            )
    return rows


def summary_rows(orders, model_map=None, now_ms=None):
    """Model+Renk+Beden bazında toplama özeti döndürür (depo toplama için).

    Her grup: toplam adet, sipariş (satır) sayısı, gecikmeli sipariş sayısı ve
    grubun en acil ürününün durumu. Sıralama: en acil (en erken son tarih) en üstte.
    """
    if now_ms is None:
        now_ms = int(time.time() * 1000)

    groups = {}  # key -> dict
    for order in orders:
        d_ms = _deadline_ms(order)
        is_delayed = d_ms != _NO_DEADLINE and d_ms < now_ms
        for line in order.get("lines") or []:
            model = _model_code(line, model_map)
            color = line.get("productColor", "")
            size = line.get("productSize", "")
            key = _variant_key(model, color, size)
            g = groups.get(key)
            if g is None:
                g = {
                    "key": key, "name": "", "model": model, "color": color,
                    "size": size, "qty": 0, "orders": 0, "delayed": 0,
                    "min_deadline": _NO_DEADLINE,
                }
                groups[key] = g
            if not g["name"]:
                g["name"] = line.get("productName", "")
            g["qty"] += _qty(line)
            g["orders"] += 1
            if is_delayed:
                g["delayed"] += 1
            if d_ms < g["min_deadline"]:
                g["min_deadline"] = d_ms

    # En acil (en erken son tarih) en üstte; eşitlikte anahtara göre
    ordered = sorted(groups.values(), key=lambda g: (g["min_deadline"], g["key"]))

    rows = []
    for g in ordered:
        status, _ = _delay_text(g["min_deadline"], now_ms)
        rows.append([
            g["key"], g["name"], g["model"], g["color"], g["size"],
            g["qty"], g["orders"], g["delayed"], status,
        ])
    return rows


def dashboard_stats(orders, model_map=None, now_ms=None):
    """Panel için özet istatistikleri hesaplar."""
    if now_ms is None:
        now_ms = int(time.time() * 1000)
    day = 24 * 60 * 60 * 1000

    stats = {
        "total_orders": len(orders),
        "total_lines": 0,
        "total_qty": 0,
        "delayed_orders": 0,
        "urgent": 0,       # 24 saatten az kalan
        "soon": 0,         # 1-3 gün kalan
        "later": 0,        # 3+ gün kalan
        "no_deadline": 0,  # tarih bilgisi yok
        "cargo": {},       # firma -> {orders, qty, delayed}
    }
    variants = set()

    for order in orders:
        d_ms = _deadline_ms(order)
        is_delayed = d_ms != _NO_DEADLINE and d_ms < now_ms
        if is_delayed:
            stats["delayed_orders"] += 1
        elif d_ms == _NO_DEADLINE:
            stats["no_deadline"] += 1
        else:
            rem = d_ms - now_ms
            if rem < day:
                stats["urgent"] += 1
            elif rem < 3 * day:
                stats["soon"] += 1
            else:
                stats["later"] += 1

        cname = (order.get("cargoProviderName") or "—").strip() or "—"
        c = stats["cargo"].setdefault(cname, {"orders": 0, "qty": 0, "delayed": 0})
        c["orders"] += 1
        if is_delayed:
            c["delayed"] += 1

        for line in order.get("lines") or []:
            stats["total_lines"] += 1
            q = _qty(line)
            stats["total_qty"] += q
            c["qty"] += q
            variants.add(
                _variant_key(
                    _model_code(line, model_map),
                    line.get("productColor", ""),
                    line.get("productSize", ""),
                )
            )

    stats["variants"] = len(variants)
    return stats


def _write_dashboard(ws, stats, store_name, info, now_ms):
    """Panel sayfasını yazar: genel toplamlar, gecikme özeti, kargo dağılımı."""
    title_font = Font(bold=True, size=14, color="F27A1A")
    section_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="2A2F36")
    label_font = Font(bold=True)

    def section(row, text, span=2):
        ws.cell(row=row, column=1, value=text).font = section_font
        for c in range(1, span + 1):
            ws.cell(row=row, column=c).fill = section_fill
        return row + 1

    def kv(row, label, value, value_font=None):
        ws.cell(row=row, column=1, value=label).font = label_font
        cell = ws.cell(row=row, column=2, value=value)
        if value_font:
            cell.font = value_font
        return row + 1

    r = 1
    ws.cell(row=r, column=1, value="TRENDYOL KARGO ÖZET PANELİ").font = title_font
    r += 2

    info = info or {}
    r = kv(r, "Mağaza", store_name or "—")
    if info.get("status"):
        r = kv(r, "Sipariş Durumu", info["status"])
    if info.get("start") and info.get("end"):
        r = kv(r, "Tarih Aralığı", f"{info['start']} – {info['end']}")
    r = kv(r, "Rapor Tarihi", _fmt_dt(now_ms))
    r += 1

    # --- Genel ---
    r = section(r, "GENEL")
    r = kv(r, "Toplam Sipariş (paket)", stats["total_orders"])
    r = kv(r, "Toplam Ürün Satırı", stats["total_lines"])
    r = kv(r, "Toplam Adet (ürün)", stats["total_qty"])
    r = kv(r, "Benzersiz Ürün (Model+Renk+Beden)", stats["variants"])
    r += 1

    # --- Gecikme durumu ---
    r = section(r, "GECİKME DURUMU")
    total = stats["total_orders"] or 1
    pct = round(stats["delayed_orders"] * 100 / total, 1)
    r = kv(r, "Geciken Sipariş", f"{stats['delayed_orders']}  (%{pct})",
           Font(bold=True, color="D32F2F"))
    r = kv(r, "Acil — 24 saatten az kalan", stats["urgent"])
    r = kv(r, "1–3 gün kalan", stats["soon"])
    r = kv(r, "3+ gün kalan", stats["later"])
    if stats["no_deadline"]:
        r = kv(r, "Tarih bilgisi yok", stats["no_deadline"])
    r += 1

    # --- Kargo firması dağılımı ---
    r = section(r, "KARGO FİRMASINA GÖRE", span=4)
    headers = ["Kargo Firması", "Sipariş", "Adet", "Gecikmeli"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    r += 1
    for name, c in sorted(stats["cargo"].items(), key=lambda kv: -kv[1]["orders"]):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=c["orders"])
        ws.cell(row=r, column=3, value=c["qty"])
        dcell = ws.cell(row=r, column=4, value=c["delayed"])
        if c["delayed"]:
            dcell.font = Font(bold=True, color="D32F2F")
        r += 1

    # Genişlikler
    ws.column_dimensions["A"].width = 38
    for col in ("B", "C", "D"):
        ws.column_dimensions[col].width = 14


def _write_header(ws, headers):
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws, headers):
    widths = [len(h) for h in headers]
    for row in ws.iter_rows(min_row=2, values_only=True):
        for i, value in enumerate(row):
            widths[i] = max(widths[i], len(str(value)) if value is not None else 0)
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 55)
    ws.freeze_panes = "A2"


def _paint_delayed_row(ws, ncols, status_col_index):
    """Son eklenen satırı kırmızıya boyar; durum hücresini koyu kırmızı yapar."""
    r = ws.max_row
    for c in range(1, ncols + 1):
        ws.cell(row=r, column=c).fill = _ROW_FILL
    cell = ws.cell(row=r, column=status_col_index + 1)
    cell.fill = _DELAY_FILL
    cell.font = _DELAY_FONT


def build_workbook(orders, store_name="", model_map=None, now_ms=None, info=None):
    """Üç sayfalı .xlsx üretir (Panel + Kargo Listesi + Özet), BytesIO döndürür."""
    if now_ms is None:
        now_ms = int(time.time() * 1000)

    wb = Workbook()

    # --- Sayfa 1: Panel (dashboard) ---
    ws_dash = wb.active
    ws_dash.title = "Panel"
    stats = dashboard_stats(orders, model_map, now_ms)
    _write_dashboard(ws_dash, stats, store_name, info, now_ms)

    # --- Sayfa 2: Kargo Listesi (detay) ---
    ws = wb.create_sheet("Kargo Listesi")
    _write_header(ws, HEADERS)
    ncols = len(HEADERS)
    for row in orders_to_rows(orders, store_name, model_map, now_ms):
        ws.append(row)
        if str(row[_DELAY_COL]).startswith(_DELAYED_PREFIX):
            _paint_delayed_row(ws, ncols, _DELAY_COL)
    _autofit(ws, HEADERS)

    # --- Sayfa 3: Özet (Model+Renk+Beden toplama) ---
    ws2 = wb.create_sheet("Özet")
    _write_header(ws2, SUMMARY_HEADERS)
    ncols2 = len(SUMMARY_HEADERS)
    for row in summary_rows(orders, model_map, now_ms):
        ws2.append(row)
        if (row[_SUM_DELAYED_COL] or 0) > 0:
            _paint_delayed_row(ws2, ncols2, _SUM_STATUS_COL)
    _autofit(ws2, SUMMARY_HEADERS)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
