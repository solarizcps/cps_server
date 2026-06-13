# -*- coding: utf-8 -*-
# GARDA 2026 Urun Excel Import Scripti
# Sheetler: "2026 Garda", "2026 Garda (2)"
# Kurallar:
#   - Bos model_no satirlari atla
#   - Ayni sheet_adi + excel_satir_no tekrar import edilmesin
#   - Ayni modelden birden fazla satir olabilir (silme)
#   - Turkce karakter bozulmaz (openpyxl UTF-8)
import sys, os, sqlite3, glob
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# Glob ile dosyayi bul (Turkce karakter os.path.exists sorununu atlatir)
_EXCEL_GLOB = os.path.join(
    os.path.expanduser('~'), 'Desktop', 'Solariz Fuar', '*GARDA*rv3.xlsx'
)
_matches = [x for x in glob.glob(_EXCEL_GLOB) if not os.path.basename(x).startswith('~$')]
EXCEL_PATH = _matches[0] if _matches else ''
DB_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                        '..', '..', 'mock_data.db')
FUAR_ADI = 'GARDA_2026'
SHEETS   = ['2026 Garda', '2026 Garda (2)']

# Excel kolon indeksleri (0-tabanli, A=0) — gercek yapiya gore
# Satir 1 baslik: A=Urun Foto, B=Taban, C=Model No, D=Kategori, E=Tip, F=urun cinsi,
#                 G=Asorti, H=Asorti Dagilimi, ...(merge)..., T=Birim Fiyat,
#                 U=Malzeme Bilgisi, AC=Maliyet, AF=Kur, AI=Marj
COL_MODEL_NO        = 2   # C
COL_KATEGORI        = 3   # D
COL_TIP             = 4   # E
COL_URUN_CINSI      = 5   # F
COL_ASORTI          = 6   # G
COL_ASORTI_DAGILIMI = 7   # H (merged — ilk hucre)
COL_BIRIM_FIYAT     = 19  # T
COL_MALZEME_BILGISI = 20  # U
COL_SARFIYAT        = 21  # V (Saya maliyeti)
COL_MALIYET         = 28  # AD (Toplam maliyet) — 0-indexed
COL_KUR             = 31  # AF
COL_MARJ            = 34  # AI


def _cell(row, idx):
    cells = list(row)
    if idx < len(cells):
        v = cells[idx].value
        if v is None:
            return None
        if isinstance(v, str):
            return v.strip()
        return v
    return None


def _cell_str_multiline(row, idx):
    """Multiline (newline) iceren hucreyi tek satira indirger."""
    v = _cell(row, idx)
    if v is None:
        return None
    s = str(v).replace('\n', ' / ').strip()
    return s if s else None


def _float_safe(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return None


def _str_safe(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def run():
    try:
        import openpyxl
    except ImportError:
        print('[HATA] openpyxl yuklu degil. pip install openpyxl')
        sys.exit(1)

    if not os.path.exists(EXCEL_PATH):
        print('[HATA] Excel dosyasi bulunamadi:')
        print('       ' + EXCEL_PATH)
        sys.exit(1)

    print('[INFO] Excel aciliyor...')
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()

    toplam_eklenen   = 0
    toplam_atlanan   = 0
    toplam_bos_model = 0

    for sheet_adi in SHEETS:
        if sheet_adi not in wb.sheetnames:
            print(f'[UYARI] Sheet bulunamadi: "{sheet_adi}" — atlandi')
            continue

        ws = wb[sheet_adi]
        print(f'[INFO] Sheet isleniyor: "{sheet_adi}"')

        sheet_eklenen = 0
        sheet_atlanan = 0

        for excel_row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            model_no = _str_safe(_cell(row, COL_MODEL_NO))

            if not model_no:
                toplam_bos_model += 1
                continue

            mevcut = cur.execute(
                "SELECT id FROM crm_urun WHERE fuar_adi=? AND sheet_adi=? AND excel_satir_no=?",
                (FUAR_ADI, sheet_adi, excel_row_idx)
            ).fetchone()

            if mevcut:
                sheet_atlanan  += 1
                toplam_atlanan += 1
                continue

            cur.execute("""
                INSERT INTO crm_urun
                    (fuar_adi, sheet_adi, excel_satir_no, model_no, kategori, tip,
                     urun_cinsi, asorti, asorti_dagilimi, birim_fiyat, malzeme_bilgisi,
                     sarfiyat, maliyet, kur, marj, aktif)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1)
            """, (
                FUAR_ADI, sheet_adi, excel_row_idx,
                model_no,
                _str_safe(_cell(row, COL_KATEGORI)),
                _str_safe(_cell(row, COL_TIP)),
                _str_safe(_cell(row, COL_URUN_CINSI)),
                _str_safe(_cell(row, COL_ASORTI)),
                _cell_str_multiline(row, COL_ASORTI_DAGILIMI),
                _float_safe(_cell(row, COL_BIRIM_FIYAT)),
                _str_safe(_cell(row, COL_MALZEME_BILGISI)),
                _float_safe(_cell(row, COL_SARFIYAT)),
                _float_safe(_cell(row, COL_MALIYET)),
                _str_safe(_cell(row, COL_KUR)),
                _str_safe(_cell(row, COL_MARJ)),
            ))

            sheet_eklenen  += 1
            toplam_eklenen += 1

        print(f'         Eklenen: {sheet_eklenen}  |  Atlanan (duplicate): {sheet_atlanan}')

    conn.commit()
    conn.close()

    print()
    print(f'[SONUC] Toplam eklenen : {toplam_eklenen}')
    print(f'        Toplam atlanan : {toplam_atlanan} (duplicate)')
    print(f'        Bos model_no   : {toplam_bos_model} (atildi)')
    print('[DONE]  Import tamamlandi.')


if __name__ == '__main__':
    run()
