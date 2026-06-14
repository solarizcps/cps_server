# -*- coding: utf-8 -*-
"""
CPS — Fuar CRM Modulu (Faz 2)
================================
Ekranlar:
  /fuar-crm/                           -> Dashboard (8 istatistik + tablolar)
  /fuar-crm/firma                      -> Firma listesi (zengin kolonlar)
  /fuar-crm/firma/ekle                 -> Yeni firma
  /fuar-crm/firma/<id>                 -> Detay + hizli aksiyonlar
  /fuar-crm/firma/<id>/gorusme-ekle    -> Gorusme notu POST
  /fuar-crm/firma/<id>/dosya-yukle     -> Kartvizit/gorsel yukle POST
  /fuar-crm/firma/<id>/dosya/<did>/sil -> Dosya sil POST
"""
from flask import (Blueprint, render_template, request, redirect,
                   url_for, session, abort, flash, send_from_directory,
                   jsonify, make_response)
from db import q as _q, qone as _qone, get_conn as _get_conn
from modules.auth import login_gerekli
import os, datetime, uuid, mimetypes

fuar_crm_bp = Blueprint(
    'fuar_crm', __name__,
    url_prefix='/fuar-crm',
    template_folder='../../templates/fuar_crm'
)

# Yuklemeler: app/static/uploads/fuar_crm/
UPLOAD_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    'static', 'uploads', 'fuar_crm'
)
ALLOWED_EXT = {'jpg', 'jpeg', 'png', 'webp', 'gif', 'pdf'}
MAX_MB = 10


# ── Yardimcilar ---------------------------------------------------------------

def _u():
    k = session.get('kullanici')
    return k['KullaniciAdi'] if k else 'sistem'


def _aktif_katalog_id():
    """Aktif katalog ID'sini döndürür. Yoksa None (tüm ürünler listelenir)."""
    row = _qone("SELECT id FROM crm_katalog WHERE aktif = 1 ORDER BY id DESC LIMIT 1")
    return row['id'] if row else None


def _crm_erisim():
    u = session.get('kullanici')
    if not u:
        return False
    kadi = (u.get('KullaniciAdi') or '').strip().lower()
    rol  = (u.get('RolAd') or u.get('Rol') or '').strip()
    if kadi == 'admin':
        return True
    if rol in ('Yonetim', 'Muhasebe', 'Finans'):
        return True
    # Turkce rol adi da kabul et
    if 'yonetim' in rol.lower():
        return True
    try:
        from modules.auth import yetki_var
        return yetki_var('fuar_crm', 'can_view')
    except Exception:
        return False


@fuar_crm_bp.before_request
def _erisim_kontrol():
    if not session.get('kullanici'):
        return redirect(url_for('auth.login', next=request.path))
    if not _crm_erisim():
        abort(403)


def _ext_ok(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


def _is_image(filename):
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    return ext in ('jpg', 'jpeg', 'png', 'webp', 'gif')


def _safe_name(original):
    ext = original.rsplit('.', 1)[-1].lower() if '.' in original else 'bin'
    return f"{uuid.uuid4().hex}.{ext}"


# ── DASHBOARD -----------------------------------------------------------------

@fuar_crm_bp.route('/')
@login_gerekli
def dashboard():
    today = datetime.date.today().isoformat()

    toplam_firma    = (_qone("SELECT COUNT(*) AS c FROM crm_firma WHERE aktif=1") or {}).get('c', 0)
    toplam_gorusme  = (_qone("SELECT COUNT(*) AS c FROM crm_gorusme") or {}).get('c', 0)
    takip_bekleyen  = (_qone("SELECT COUNT(*) AS c FROM crm_gorusme WHERE durum='takip_bekliyor'") or {}).get('c', 0)
    numune_isteyen  = (_qone("SELECT COUNT(*) AS c FROM crm_gorusme WHERE numune=1") or {}).get('c', 0)
    fiyat_verilecek = (_qone("SELECT COUNT(*) AS c FROM crm_gorusme WHERE fiyat_verildi=0 AND durum NOT IN ('kapandi','olumsuz')") or {}).get('c', 0)
    bugun_gorusme   = (_qone(
        "SELECT COUNT(*) AS c FROM crm_gorusme WHERE substr(created_at,1,10)=?", (today,)
    ) or {}).get('c', 0)
    toplam_urun     = (_qone("SELECT COUNT(*) AS c FROM crm_urun WHERE aktif=1") or {}).get('c', 0)

    # Ulke dagilimi
    ulke_dagilim = _q("""
        SELECT ulke, COUNT(*) AS cnt
        FROM crm_firma
        WHERE aktif=1 AND ulke IS NOT NULL AND ulke != ''
        GROUP BY ulke ORDER BY cnt DESC LIMIT 8
    """) or []

    # Son eklenen firmalar
    son_firmalar = _q("""
        SELECT f.id, f.firma_adi, f.yetkili, f.ulke, f.firma_tipi,
               f.telefon, f.whatsapp, f.created_at,
               COUNT(g.id) AS gorusme_sayisi,
               MAX(g.durum) AS son_durum
        FROM crm_firma f
        LEFT JOIN crm_gorusme g ON g.firma_id = f.id
        WHERE f.aktif = 1
        GROUP BY f.id
        ORDER BY f.created_at DESC
        LIMIT 8
    """) or []

    # Yaklasan takipler (7 gun icinde)
    from datetime import timedelta
    bitis = (datetime.date.today() + timedelta(days=7)).isoformat()
    yaklasan_takip = _q("""
        SELECT g.takip_tarihi, g.durum, g.urun_ilgisi,
               f.id AS firma_id, f.firma_adi, f.yetkili
        FROM crm_gorusme g
        JOIN crm_firma f ON f.id = g.firma_id
        WHERE g.takip_tarihi IS NOT NULL
          AND g.takip_tarihi >= ?
          AND g.takip_tarihi <= ?
          AND g.durum NOT IN ('kapandi','olumsuz')
        ORDER BY g.takip_tarihi ASC
        LIMIT 10
    """, (today, bitis)) or []

    # En cok ilgi goren urunler (gorusme_urun tablosundan)
    en_cok_urun = _q("""
        SELECT u.model_no, u.kategori, u.tip, u.urun_cinsi,
               COUNT(gu.id) AS ilgi_sayisi
        FROM crm_gorusme_urun gu
        JOIN crm_urun u ON u.id = gu.urun_id
        GROUP BY u.id
        ORDER BY ilgi_sayisi DESC
        LIMIT 5
    """) or []

    return render_template(
        'fuar_crm/dashboard.html',
        toplam_firma=toplam_firma,
        toplam_gorusme=toplam_gorusme,
        takip_bekleyen=takip_bekleyen,
        numune_isteyen=numune_isteyen,
        fiyat_verilecek=fiyat_verilecek,
        bugun_gorusme=bugun_gorusme,
        toplam_urun=toplam_urun,
        en_cok_urun=en_cok_urun,
        ulke_dagilim=ulke_dagilim,
        son_firmalar=son_firmalar,
        yaklasan_takip=yaklasan_takip,
        aktif_katalog=_qone("SELECT id, ad FROM crm_katalog WHERE aktif=1 LIMIT 1"),
    )


# ── FIRMA LISTESI -------------------------------------------------------------

@fuar_crm_bp.route('/firma')
@login_gerekli
def firma_liste():
    arama       = request.args.get('q', '').strip()
    ulke_filtre = request.args.get('ulke', '').strip()
    durum_filtre= request.args.get('durum', '').strip()

    params = []
    where  = ["f.aktif = 1"]

    if arama:
        where.append("(f.firma_adi LIKE ? OR f.yetkili LIKE ? OR f.email LIKE ? OR f.marka_ilgisi LIKE ?)")
        like = f"%{arama}%"
        params += [like, like, like, like]

    if ulke_filtre:
        where.append("f.ulke = ?")
        params.append(ulke_filtre)

    where_sql = " AND ".join(where)

    firmalar = _q(f"""
        SELECT f.id, f.firma_adi, f.yetkili, f.telefon, f.whatsapp,
               f.email, f.ulke, f.sehir, f.firma_tipi, f.marka_ilgisi,
               f.kaynak, f.created_at,
               COUNT(DISTINCT g.id)          AS gorusme_sayisi,
               MAX(g.durum)                  AS son_durum,
               MAX(g.created_at)             AS son_gorusme_tarihi,
               MAX(g.takip_tarihi)           AS son_takip_tarihi,
               SUM(CASE WHEN substr(g.created_at,1,10)=date('now') THEN 1 ELSE 0 END) AS bugun_gorusme,
               COUNT(DISTINCT gu.urun_id)    AS urun_sayisi
        FROM crm_firma f
        LEFT JOIN crm_gorusme g  ON g.firma_id = f.id
        LEFT JOIN crm_gorusme_urun gu ON gu.gorusme_id = g.id
        WHERE {where_sql}
        GROUP BY f.id
        ORDER BY f.firma_adi
    """, params) or []

    ulkeler = _q(
        "SELECT DISTINCT ulke FROM crm_firma WHERE aktif=1 AND ulke IS NOT NULL AND ulke != '' ORDER BY ulke"
    ) or []
    ulke_listesi = [r['ulke'] for r in ulkeler]

    return render_template(
        'fuar_crm/firma_liste.html',
        firmalar=firmalar,
        arama=arama,
        ulke_filtre=ulke_filtre,
        durum_filtre=durum_filtre,
        ulke_listesi=ulke_listesi,
    )


# ── FIRMA EKLE ----------------------------------------------------------------

@fuar_crm_bp.route('/firma/ekle', methods=['GET', 'POST'])
@login_gerekli
def firma_ekle():
    if request.method == 'POST':
        firma_adi = request.form.get('firma_adi', '').strip()
        if not firma_adi:
            flash('Firma adi zorunludur.', 'hata')
            return redirect(url_for('fuar_crm.firma_ekle'))

        fields = {
            'yetkili':      request.form.get('yetkili', '').strip() or None,
            'telefon':      request.form.get('telefon', '').strip() or None,
            'whatsapp':     request.form.get('whatsapp', '').strip() or None,
            'email':        request.form.get('email', '').strip() or None,
            'ulke':         request.form.get('ulke', '').strip() or None,
            'sehir':        request.form.get('sehir', '').strip() or None,
            'firma_tipi':   request.form.get('firma_tipi', '').strip() or None,
            'marka_ilgisi': request.form.get('marka_ilgisi', '').strip() or None,
            'erp_cari_kodu':request.form.get('erp_cari_kodu', '').strip() or None,
            'kaynak':       request.form.get('kaynak', '').strip() or None,
        }

        conn = _get_conn()
        try:
            cur = conn.cursor()

            # Duplicate uyarisi: ayni firma_adi + ulke/sehir kombine
            ulke_check = fields.get('ulke') or ''
            sehir_check = fields.get('sehir') or ''
            force_ekle = request.form.get('force_ekle') == '1'
            dup = cur.execute(
                "SELECT id FROM crm_firma WHERE LOWER(firma_adi)=LOWER(?) LIMIT 1",
                (firma_adi,)
            ).fetchone()
            if dup and not force_ekle:
                flash(f'UYARI: "{firma_adi}" adinda benzer bir firma zaten var (id={dup[0]}). '
                      'Yine de eklemek istiyorsaniz formu tekrar gonderin.', 'uyari')
                return render_template('fuar_crm/firma_ekle.html',
                                       dup_uyari=True, form_data=request.form)

            cur.execute("""
                INSERT INTO crm_firma
                    (firma_adi, yetkili, telefon, whatsapp, email,
                     ulke, sehir, firma_tipi, marka_ilgisi,
                     erp_cari_kodu, kaynak, aktif, created_by)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,1,?)
            """, (firma_adi,
                  fields['yetkili'], fields['telefon'], fields['whatsapp'],
                  fields['email'], fields['ulke'], fields['sehir'],
                  fields['firma_tipi'], fields['marka_ilgisi'],
                  fields['erp_cari_kodu'], fields['kaynak'], _u()))
            conn.commit()
            new_id = cur.lastrowid
            flash(f'"{firma_adi}" basariyla eklendi.', 'basari')
            return redirect(url_for('fuar_crm.firma_detay', firma_id=new_id))
        except Exception as exc:
            conn.rollback()
            flash(f'Kayit hatasi: {exc}', 'hata')
        finally:
            conn.close()

    return render_template('fuar_crm/firma_ekle.html')


# ── FIRMA DUZENLE -------------------------------------------------------------

@fuar_crm_bp.route('/firma/<int:firma_id>/duzenle', methods=['GET', 'POST'])
@login_gerekli
def firma_duzenle(firma_id):
    firma = _qone("SELECT * FROM crm_firma WHERE id=?", (firma_id,))
    if not firma:
        abort(404)

    if request.method == 'POST':
        firma_adi = request.form.get('firma_adi', '').strip()
        if not firma_adi:
            flash('Firma adi zorunludur.', 'hata')
            return redirect(url_for('fuar_crm.firma_duzenle', firma_id=firma_id))

        fields = {
            'yetkili':       request.form.get('yetkili', '').strip() or None,
            'telefon':       request.form.get('telefon', '').strip() or None,
            'whatsapp':      request.form.get('whatsapp', '').strip() or None,
            'email':         request.form.get('email', '').strip() or None,
            'ulke':          request.form.get('ulke', '').strip() or None,
            'sehir':         request.form.get('sehir', '').strip() or None,
            'firma_tipi':    request.form.get('firma_tipi', '').strip() or None,
            'marka_ilgisi':  request.form.get('marka_ilgisi', '').strip() or None,
            'erp_cari_kodu': request.form.get('erp_cari_kodu', '').strip() or None,
            'kaynak':        request.form.get('kaynak', '').strip() or None,
            'notlar':        request.form.get('notlar', '').strip() or None,
        }

        conn = _get_conn()
        try:
            cur = conn.cursor()
            cur.execute("""
                UPDATE crm_firma SET
                    firma_adi=?, yetkili=?, telefon=?, whatsapp=?, email=?,
                    ulke=?, sehir=?, firma_tipi=?, marka_ilgisi=?,
                    erp_cari_kodu=?, kaynak=?, notlar=?
                WHERE id=?
            """, (firma_adi,
                  fields['yetkili'], fields['telefon'], fields['whatsapp'],
                  fields['email'], fields['ulke'], fields['sehir'],
                  fields['firma_tipi'], fields['marka_ilgisi'],
                  fields['erp_cari_kodu'], fields['kaynak'], fields['notlar'],
                  firma_id))
            conn.commit()
            flash(f'"{firma_adi}" guncellendi.', 'basari')
            return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id))
        except Exception as exc:
            conn.rollback()
            flash(f'Guncelleme hatasi: {exc}', 'hata')
        finally:
            conn.close()

    return render_template('fuar_crm/firma_duzenle.html', firma=firma)


# ── FIRMA SIL -----------------------------------------------------------------

@fuar_crm_bp.route('/firma/<int:firma_id>/sil', methods=['POST'])
@login_gerekli
def firma_sil(firma_id):
    firma = _qone("SELECT * FROM crm_firma WHERE id=?", (firma_id,))
    if not firma:
        abort(404)

    # Gorusme sayisi kontrolu
    gorusme_sayisi = _qone(
        "SELECT COUNT(*) as cnt FROM crm_gorusme WHERE firma_id=?", (firma_id,)
    )
    cnt = gorusme_sayisi['cnt'] if gorusme_sayisi else 0

    if cnt > 0:
        flash(f'"{firma["firma_adi"]}" firmasina bagli {cnt} gorusme kaydi var. '
              'Silme engellendi. Once gorusmeleri kontrol edin.', 'hata')
        return redirect(url_for('fuar_crm.firma_liste'))

    conn = _get_conn()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM crm_firma WHERE id=?", (firma_id,))
        conn.commit()
        flash(f'"{firma["firma_adi"]}" silindi.', 'basari')
    except Exception as exc:
        conn.rollback()
        flash(f'Silme hatasi: {exc}', 'hata')
    finally:
        conn.close()

    return redirect(url_for('fuar_crm.firma_liste'))


# ── FIRMA DETAY ---------------------------------------------------------------

@fuar_crm_bp.route('/firma/<int:firma_id>')
@login_gerekli
def firma_detay(firma_id):
    firma = _qone("SELECT * FROM crm_firma WHERE id = ? AND aktif = 1", (firma_id,))
    if not firma:
        abort(404)

    gorusmeler = _q("""
        SELECT * FROM crm_gorusme
        WHERE firma_id = ?
        ORDER BY created_at DESC
    """, (firma_id,)) or []

    # Her gorusme icin urunleri de cek
    gorusme_listesi = []
    today = datetime.date.today().isoformat()
    for g in gorusmeler:
        g = dict(g)
        g['urunler'] = _q("""
            SELECT gu.id AS gu_id, gu.fiyat_konusuldu, gu.numune_istendi, gu.not_text,
                   gu.verilen_fiyat, gu.para_birimi, gu.eski_fiyat, gu.indirim_notu,
                   gu.numune_adet, gu.numune_beden, gu.istenen_renk,
                   gu.renk_basi_adet, gu.toplam_adet, gu.teslim_notu, gu.urun_notu,
                   u.model_no, u.kategori, u.tip, u.urun_cinsi, u.asorti,
                   u.birim_fiyat, u.maliyet,
                   gorsel.dosya_yolu AS gorsel_yolu
            FROM crm_gorusme_urun gu
            JOIN crm_urun u ON u.id = gu.urun_id
            LEFT JOIN crm_urun_gorsel gorsel ON gorsel.urun_id = u.id
            WHERE gu.gorusme_id = ?
        """, (g['id'],)) or []
        g['bugun'] = (g.get('created_at') or '')[:10] == today
        gorusme_listesi.append(g)

    dosyalar = _q(
        "SELECT * FROM crm_dosya WHERE firma_id = ? ORDER BY created_at DESC",
        (firma_id,)
    ) or []

    dosya_listesi = []
    for d in dosyalar:
        d = dict(d)
        fn = os.path.basename(d.get('dosya_yolu', ''))
        d['is_image'] = _is_image(fn)
        d['filename'] = fn
        dosya_listesi.append(d)

    # Urun istatistigi
    toplam_urun = (_qone("SELECT COUNT(*) AS c FROM crm_urun WHERE aktif=1") or {}).get('c', 0)

    return render_template(
        'fuar_crm/firma_detay.html',
        firma=firma,
        gorusmeler=gorusme_listesi,
        dosyalar=dosya_listesi,
        toplam_urun=toplam_urun,
        today=today,
    )


# ── GORUSME EKLE -------------------------------------------------------------

@fuar_crm_bp.route('/firma/<int:firma_id>/gorusme-ekle', methods=['POST'])
@login_gerekli
def gorusme_ekle(firma_id):
    firma = _qone("SELECT id FROM crm_firma WHERE id = ? AND aktif = 1", (firma_id,))
    if not firma:
        abort(404)

    fuar_adi      = request.form.get('fuar_adi', '').strip() or None
    gorusen       = request.form.get('gorusen', '').strip() or None
    not_text      = request.form.get('not_text', '').strip() or None
    urun_ilgisi   = request.form.get('urun_ilgisi', '').strip() or None
    numune        = 1 if request.form.get('numune') == '1' else 0
    fiyat_verildi = 1 if request.form.get('fiyat_verildi') == '1' else 0
    takip_tarihi  = request.form.get('takip_tarihi', '').strip() or None
    durum         = request.form.get('durum', 'beklemede').strip()

    # Secilen urunler: urun_ids[] listesi
    urun_ids = request.form.getlist('urun_ids[]')

    conn = _get_conn()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO crm_gorusme
                (firma_id, fuar_adi, gorusen, not_text, urun_ilgisi,
                 numune, fiyat_verildi, takip_tarihi, durum, created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (firma_id, fuar_adi, gorusen, not_text, urun_ilgisi,
              numune, fiyat_verildi, takip_tarihi, durum, _u()))
        gorusme_id = cur.lastrowid

        # Urun iliskileri kaydet
        for uid in urun_ids:
            try:
                uid_int = int(uid)
                fiyat_k      = 1 if request.form.get(f'fiyat_konusuldu_{uid}') == '1' else 0
                numune_i     = 1 if request.form.get(f'numune_{uid}') == '1' else 0
                urun_not     = request.form.get(f'urun_notu_{uid}', '').strip() or None

                def _float_f(key):
                    v = request.form.get(key, '').strip()
                    try: return float(v) if v else None
                    except ValueError: return None

                def _int_f(key):
                    v = request.form.get(key, '').strip()
                    try: return int(v) if v else None
                    except ValueError: return None

                verilen_fiyat  = _float_f(f'verilen_fiyat_{uid}')
                para_birimi    = request.form.get(f'para_birimi_{uid}', 'USD').strip() or 'USD'
                eski_fiyat     = _float_f(f'eski_fiyat_{uid}')
                indirim_notu   = request.form.get(f'indirim_notu_{uid}', '').strip() or None
                numune_adet    = _int_f(f'numune_adet_{uid}')
                numune_beden   = request.form.get(f'numune_beden_{uid}', '').strip() or None
                istenen_renk   = request.form.get(f'istenen_renk_{uid}', '').strip() or None
                renk_basi_adet = _int_f(f'renk_basi_adet_{uid}')
                toplam_adet    = _int_f(f'toplam_adet_{uid}')
                teslim_notu    = request.form.get(f'teslim_notu_{uid}', '').strip() or None
                urun_notu      = request.form.get(f'urun_notu_{uid}', '').strip() or None

                cur.execute("""
                    INSERT INTO crm_gorusme_urun
                        (gorusme_id, urun_id, not_text, fiyat_konusuldu, numune_istendi,
                         verilen_fiyat, para_birimi, eski_fiyat, indirim_notu,
                         numune_adet, numune_beden, istenen_renk,
                         renk_basi_adet, toplam_adet, teslim_notu, urun_notu)
                    VALUES (?,?,?,?,?, ?,?,?,?, ?,?,?, ?,?,?,?)
                """, (gorusme_id, uid_int, urun_not, fiyat_k, numune_i,
                      verilen_fiyat, para_birimi, eski_fiyat, indirim_notu,
                      numune_adet, numune_beden, istenen_renk,
                      renk_basi_adet, toplam_adet, teslim_notu, urun_notu))
            except (ValueError, TypeError):
                continue

        conn.commit()
        flash('Gorusme notu eklendi.', 'basari')
    except Exception as exc:
        conn.rollback()
        flash(f'Hata: {exc}', 'hata')
    finally:
        conn.close()

    return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id) + '#gecmis')


# ── GORUSME DUZENLE -----------------------------------------------------------

@fuar_crm_bp.route('/firma/<int:firma_id>/gorusme/<int:gorusme_id>/duzenle', methods=['POST'])
@login_gerekli
def gorusme_duzenle(firma_id, gorusme_id):
    g = _qone("SELECT id, firma_id FROM crm_gorusme WHERE id=? AND firma_id=?",
              (gorusme_id, firma_id))
    if not g:
        abort(404)

    not_text      = request.form.get('not_text', '').strip() or None
    durum         = request.form.get('durum', 'beklemede').strip()
    takip_tarihi  = request.form.get('takip_tarihi', '').strip() or None
    gorusen       = request.form.get('gorusen', '').strip() or None
    urun_ilgisi   = request.form.get('urun_ilgisi', '').strip() or None
    numune        = 1 if request.form.get('numune') == '1' else 0
    fiyat_verildi = 1 if request.form.get('fiyat_verildi') == '1' else 0

    conn = _get_conn()
    try:
        conn.execute("""
            UPDATE crm_gorusme
            SET not_text=?, durum=?, takip_tarihi=?,
                gorusen=?, urun_ilgisi=?, numune=?, fiyat_verildi=?
            WHERE id=?
        """, (not_text, durum, takip_tarihi,
              gorusen, urun_ilgisi, numune, fiyat_verildi,
              gorusme_id))
        conn.commit()
        flash('Görüşme güncellendi.', 'basari')
    except Exception as exc:
        conn.rollback()
        flash(f'Hata: {exc}', 'hata')
    finally:
        conn.close()

    return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id) + '#gecmis')


# ── GORUSME SIL ---------------------------------------------------------------

@fuar_crm_bp.route('/firma/<int:firma_id>/gorusme/<int:gorusme_id>/sil', methods=['POST'])
@login_gerekli
def gorusme_sil(firma_id, gorusme_id):
    g = _qone("SELECT id, firma_id FROM crm_gorusme WHERE id=? AND firma_id=?",
              (gorusme_id, firma_id))
    if not g:
        abort(404)

    conn = _get_conn()
    try:
        # Once urun iliskileri sil, sonra gorusmeyi sil
        conn.execute("DELETE FROM crm_gorusme_urun WHERE gorusme_id=?", (gorusme_id,))
        conn.execute("DELETE FROM crm_gorusme WHERE id=?", (gorusme_id,))
        conn.commit()
        flash('Görüşme silindi.', 'basari')
    except Exception as exc:
        conn.rollback()
        flash(f'Hata: {exc}', 'hata')
    finally:
        conn.close()

    return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id) + '#gecmis')


# ── URUN KATALOGU SAYFASI -----------------------------------------------------

@fuar_crm_bp.route('/urunler')
@login_gerekli
def urun_katalogu():
    q_str    = request.args.get('q', '').strip()
    kat_filt = request.args.get('kategori', '').strip()
    tip_filt = request.args.get('tip', '').strip()

    params = ["u.aktif = 1"]
    vals   = []

    # Aktif katalog filtresi
    aktif_katalog = _aktif_katalog_id()
    if aktif_katalog is not None:
        params.append("u.katalog_id = ?")
        vals.append(aktif_katalog)

    if q_str:
        like = f"%{q_str}%"
        params.append("(u.model_no LIKE ? OR u.kategori LIKE ? OR u.tip LIKE ? OR u.urun_cinsi LIKE ? OR u.malzeme_bilgisi LIKE ?)")
        vals += [like, like, like, like, like]
    if kat_filt:
        params.append("u.kategori = ?")
        vals.append(kat_filt)
    if tip_filt:
        params.append("u.tip = ?")
        vals.append(tip_filt)

    where_sql = " AND ".join(params)
    urunler = _q(f"""
        SELECT u.id, u.model_no, u.kategori, u.tip, u.urun_cinsi, u.asorti,
               u.birim_fiyat, u.maliyet, u.malzeme_bilgisi, u.sheet_adi, u.excel_satir_no,
               g.dosya_yolu AS gorsel_yolu
        FROM crm_urun u
        LEFT JOIN crm_urun_gorsel g ON g.urun_id = u.id
        WHERE {where_sql}
        ORDER BY u.model_no, u.sheet_adi
    """, vals) or []

    kategoriler = [r['kategori'] for r in (_q(
        "SELECT DISTINCT kategori FROM crm_urun WHERE aktif=1 AND kategori IS NOT NULL ORDER BY kategori"
    ) or []) if r.get('kategori')]
    tipler = [r['tip'] for r in (_q(
        "SELECT DISTINCT tip FROM crm_urun WHERE aktif=1 AND tip IS NOT NULL ORDER BY tip"
    ) or []) if r.get('tip')]

    toplam_urun = (_qone("SELECT COUNT(*) AS c FROM crm_urun WHERE aktif=1") or {}).get('c', 0)

    # Gruplu görünüm için: model bazlı gruplayarak template'e gönder
    from collections import OrderedDict
    gruplar = OrderedDict()
    for u in urunler:
        mn = u['model_no'] or '?'
        if mn not in gruplar:
            gruplar[mn] = {
                'model_no':    mn,
                'kategori':    u['kategori'],
                'tip':         u['tip'],
                'urun_cinsi':  u['urun_cinsi'],
                'ilk_gorsel':  u['gorsel_yolu'],
                'min_fiyat':   u['birim_fiyat'],
                'max_fiyat':   u['birim_fiyat'],
                'asortiler':   [],
                'variants':    [],
            }
        g = gruplar[mn]
        # min/max fiyat güncelle
        if u['birim_fiyat'] is not None:
            if g['min_fiyat'] is None or u['birim_fiyat'] < g['min_fiyat']:
                g['min_fiyat'] = u['birim_fiyat']
            if g['max_fiyat'] is None or u['birim_fiyat'] > g['max_fiyat']:
                g['max_fiyat'] = u['birim_fiyat']
        if g['ilk_gorsel'] is None and u['gorsel_yolu']:
            g['ilk_gorsel'] = u['gorsel_yolu']
        # benzersiz asorti listesi
        if u['asorti'] and u['asorti'] not in g['asortiler']:
            g['asortiler'].append(u['asorti'])
        g['variants'].append(u)

    return render_template(
        'fuar_crm/urun_katalogu.html',
        urunler=urunler,
        gruplar=list(gruplar.values()),
        q_str=q_str,
        kat_filt=kat_filt,
        tip_filt=tip_filt,
        kategoriler=kategoriler,
        tipler=tipler,
        toplam_urun=toplam_urun,
    )


# ── URUN ARAMA API (JSON) -----------------------------------------------------

@fuar_crm_bp.route('/api/urun-ara')
@login_gerekli
def urun_ara():
    q_str   = request.args.get('q', '').strip()
    limit   = min(int(request.args.get('limit', 30)), 200)
    gruplu  = request.args.get('gruplu', '0') == '1'

    params = []
    where  = ["u.aktif = 1"]

    # Aktif katalog filtresi — sadece aktif katalog ürünleri listelenir
    aktif_katalog = _aktif_katalog_id()
    if aktif_katalog is not None:
        where.append("u.katalog_id = ?")
        params.append(aktif_katalog)

    if q_str:
        like = f"%{q_str}%"
        where.append("""(
            u.model_no   LIKE ? OR
            u.kategori   LIKE ? OR
            u.tip        LIKE ? OR
            u.urun_cinsi LIKE ?
        )""")
        params += [like, like, like, like]

    where_sql = " AND ".join(where)

    if gruplu:
        # Model bazlı grup: her modelin tüm varyantlarını döndür
        # Önce grupları bul (min/max fiyat, sayı, ilk görsel)
        grup_rows = _q(f"""
            SELECT u.model_no,
                   u.kategori, u.tip, u.urun_cinsi,
                   COUNT(*) AS varyant_sayisi,
                   MIN(u.birim_fiyat) AS min_fiyat,
                   MAX(u.birim_fiyat) AS max_fiyat,
                   (SELECT g2.dosya_yolu FROM crm_urun u2
                    LEFT JOIN crm_urun_gorsel g2 ON g2.urun_id = u2.id
                    WHERE u2.model_no = u.model_no AND g2.dosya_yolu IS NOT NULL
                    LIMIT 1) AS ilk_gorsel
            FROM crm_urun u
            WHERE {where_sql}
            GROUP BY u.model_no, u.kategori, u.tip, u.urun_cinsi
            ORDER BY u.model_no
            LIMIT {limit}
        """, params) or []

        # Her grup için varyantları al
        result = []
        for g in grup_rows:
            varyant_params = [g['model_no']]
            varyant_where  = "u.aktif = 1 AND u.model_no = ?"
            if aktif_katalog is not None:
                varyant_where += " AND u.katalog_id = ?"
                varyant_params.append(aktif_katalog)
            varyants = _q(f"""
                SELECT u.id, u.model_no, u.kategori, u.tip, u.urun_cinsi,
                       u.asorti, u.birim_fiyat, u.maliyet, u.malzeme_bilgisi,
                       u.sheet_adi, u.excel_satir_no,
                       gr.dosya_yolu AS gorsel_yolu
                FROM crm_urun u
                LEFT JOIN crm_urun_gorsel gr ON gr.urun_id = u.id
                WHERE {varyant_where}
                ORDER BY u.birim_fiyat, u.id
            """, varyant_params) or []

            result.append({
                'model_no':      g['model_no'],
                'kategori':      g['kategori'],
                'tip':           g['tip'],
                'urun_cinsi':    g['urun_cinsi'],
                'varyant_sayisi': g['varyant_sayisi'],
                'min_fiyat':     g['min_fiyat'],
                'max_fiyat':     g['max_fiyat'],
                'ilk_gorsel':    g['ilk_gorsel'],
                'variants':      [dict(v) for v in varyants],
            })
        return jsonify(result)

    # Eski düz liste (gruplu=0, geriye dönük uyumluluk)
    rows = _q(f"""
        SELECT u.id, u.model_no, u.kategori, u.tip, u.urun_cinsi,
               u.asorti, u.birim_fiyat, u.maliyet, u.malzeme_bilgisi,
               u.sheet_adi, u.excel_satir_no,
               g.dosya_yolu AS gorsel_yolu
        FROM crm_urun u
        LEFT JOIN crm_urun_gorsel g ON g.urun_id = u.id
        WHERE {where_sql}
        ORDER BY u.model_no
        LIMIT {limit}
    """, params) or []

    return jsonify(rows)


# ── GORUSME URUN SIL ----------------------------------------------------------

# ── GORUSME URUN GUNCELLE -----------------------------------------------------

@fuar_crm_bp.route('/gorusme-urun/<int:gu_id>/guncelle', methods=['POST'])
@login_gerekli
def gorusme_urun_guncelle(gu_id):
    gu = _qone("SELECT gorusme_id FROM crm_gorusme_urun WHERE id=?", (gu_id,))
    if not gu:
        abort(404)
    gorusme_id = gu['gorusme_id']
    firma = _qone("SELECT g.firma_id FROM crm_gorusme g WHERE g.id=?", (gorusme_id,))
    firma_id = firma['firma_id'] if firma else None

    def _f(key):
        v = request.form.get(key, '').strip()
        try: return float(v) if v else None
        except ValueError: return None

    def _i(key):
        v = request.form.get(key, '').strip()
        try: return int(v) if v else None
        except ValueError: return None

    verilen_fiyat  = _f('verilen_fiyat')
    para_birimi    = request.form.get('para_birimi', 'USD').strip() or 'USD'
    fiyat_konusuldu= 1 if request.form.get('fiyat_konusuldu') == '1' else 0
    istenen_renk   = request.form.get('istenen_renk', '').strip() or None
    renk_basi_adet = _i('renk_basi_adet')
    toplam_adet    = _i('toplam_adet')
    numune_istendi = 1 if request.form.get('numune_istendi') == '1' else 0
    numune_adet    = _i('numune_adet')
    numune_beden   = request.form.get('numune_beden', '').strip() or None
    urun_notu      = request.form.get('urun_notu', '').strip() or None
    teslim_notu    = request.form.get('teslim_notu', '').strip() or None
    indirim_notu   = request.form.get('indirim_notu', '').strip() or None

    conn = _get_conn()
    try:
        conn.execute("""
            UPDATE crm_gorusme_urun
            SET verilen_fiyat=?, para_birimi=?, fiyat_konusuldu=?,
                istenen_renk=?, renk_basi_adet=?, toplam_adet=?,
                numune_istendi=?, numune_adet=?, numune_beden=?,
                urun_notu=?, teslim_notu=?, indirim_notu=?
            WHERE id=?
        """, (verilen_fiyat, para_birimi, fiyat_konusuldu,
              istenen_renk, renk_basi_adet, toplam_adet,
              numune_istendi, numune_adet, numune_beden,
              urun_notu, teslim_notu, indirim_notu,
              gu_id))
        conn.commit()
        flash('Ürün bilgisi güncellendi.', 'basari')
    except Exception as exc:
        conn.rollback()
        flash(f'Hata: {exc}', 'hata')
    finally:
        conn.close()

    if firma_id:
        return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id) + '#gecmis')
    return redirect(url_for('fuar_crm.firma_liste'))


# ── GORUSME URUN SIL ----------------------------------------------------------

@fuar_crm_bp.route('/gorusme-urun/<int:gu_id>/sil', methods=['POST'])
@login_gerekli
def gorusme_urun_sil(gu_id):
    gu = _qone("SELECT gorusme_id FROM crm_gorusme_urun WHERE id=?", (gu_id,))
    if not gu:
        abort(404)
    gorusme_id = gu['gorusme_id']
    firma = _qone("""
        SELECT g.firma_id FROM crm_gorusme g WHERE g.id = ?
    """, (gorusme_id,))
    firma_id = firma['firma_id'] if firma else None

    conn = _get_conn()
    try:
        conn.execute("DELETE FROM crm_gorusme_urun WHERE id=?", (gu_id,))
        conn.commit()
        flash('Urun gorunden kaldirildi.', 'basari')
    except Exception as exc:
        conn.rollback()
        flash(f'Hata: {exc}', 'hata')
    finally:
        conn.close()

    if firma_id:
        return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id) + '#gecmis')
    return redirect(url_for('fuar_crm.firma_liste'))


# ── DOSYA YUKLE --------------------------------------------------------------

@fuar_crm_bp.route('/firma/<int:firma_id>/dosya-yukle', methods=['POST'])
@login_gerekli
def dosya_yukle(firma_id):
    firma = _qone("SELECT id FROM crm_firma WHERE id = ? AND aktif = 1", (firma_id,))
    if not firma:
        abort(404)

    f = request.files.get('dosya')
    if not f or not f.filename:
        flash('Dosya secilmedi.', 'hata')
        return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id))

    if not _ext_ok(f.filename):
        flash('Desteklenmeyen dosya tipi. Izin verilenler: jpg, jpeg, png, pdf', 'hata')
        return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id))

    aciklama = request.form.get('aciklama', '').strip() or None
    tip      = request.form.get('tip', 'kartvizit').strip()

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    safe_fn  = _safe_name(f.filename)
    abs_path = os.path.join(UPLOAD_DIR, safe_fn)

    try:
        f.save(abs_path)
    except Exception as exc:
        flash(f'Dosya kaydedilemedi: {exc}', 'hata')
        return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id))

    # DB'ye kaydet — dosya_yolu = relative web path
    web_path = f"uploads/fuar_crm/{safe_fn}"

    conn = _get_conn()
    try:
        cur = conn.cursor()
        # crm_dosya tablosuna aciklama kolonu olmayabilir, ALTER ile ekle
        try:
            cur.execute("ALTER TABLE crm_dosya ADD COLUMN aciklama TEXT")
            conn.commit()
        except Exception:
            pass  # zaten varsa hata verir, sorun degil

        cur.execute("""
            INSERT INTO crm_dosya (firma_id, dosya_yolu, tip, aciklama, created_by)
            VALUES (?,?,?,?,?)
        """, (firma_id, web_path, tip, aciklama, _u()))
        conn.commit()
        flash('Dosya basariyla yuklendi.', 'basari')
    except Exception as exc:
        conn.rollback()
        # Dosyayi geri sil
        try:
            os.remove(abs_path)
        except Exception:
            pass
        flash(f'DB hatasi: {exc}', 'hata')
    finally:
        conn.close()

    return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id))


# ── DOSYA SIL ----------------------------------------------------------------

@fuar_crm_bp.route('/firma/<int:firma_id>/dosya/<int:dosya_id>/sil', methods=['POST'])
@login_gerekli
def dosya_sil(firma_id, dosya_id):
    d = _qone("SELECT * FROM crm_dosya WHERE id=? AND firma_id=?", (dosya_id, firma_id))
    if not d:
        abort(404)

    web_path = d['dosya_yolu']
    # Disk'ten sil
    static_root = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        'static'
    )
    abs_path = os.path.join(static_root, web_path)
    try:
        if os.path.exists(abs_path):
            os.remove(abs_path)
    except Exception:
        pass

    conn = _get_conn()
    try:
        cur = conn.cursor()
        cur.execute("DELETE FROM crm_dosya WHERE id=?", (dosya_id,))
        conn.commit()
        flash('Dosya silindi.', 'basari')
    except Exception as exc:
        conn.rollback()
        flash(f'Silme hatasi: {exc}', 'hata')
    finally:
        conn.close()

    return redirect(url_for('fuar_crm.firma_detay', firma_id=firma_id))


# ── STATIK DOSYA SERVISI (opsiyonel fallback) ---------------------------------

@fuar_crm_bp.route('/uploads/<path:filename>')
@login_gerekli
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)


# ── PDF EXPORT ----------------------------------------------------------------

@fuar_crm_bp.route('/firma/<int:firma_id>/gorusme/<int:gorusme_id>/pdf')
@login_gerekli
def gorusme_pdf(firma_id, gorusme_id):
    """Bir görüşmenin PDF çıktısını üretip indirir. DB'ye dokunmaz."""
    from flask import make_response
    import io

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, Image as RLImage,
                                        HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        abort(500, 'reportlab kurulu değil')

    # ── Veri çek ──────────────────────────────────────────────────────────────
    firma = _qone(
        "SELECT * FROM crm_firma WHERE id=?", (firma_id,)
    )
    if not firma:
        abort(404)

    gorusme = _qone(
        "SELECT * FROM crm_gorusme WHERE id=? AND firma_id=?",
        (gorusme_id, firma_id)
    )
    if not gorusme:
        abort(404)

    urunler = _q("""
        SELECT gu.*, u.model_no, u.kategori, u.tip, u.urun_cinsi, u.asorti,
               u.birim_fiyat as liste_fiyat, u.malzeme_bilgisi,
               g.dosya_yolu as gorsel_yolu
        FROM crm_gorusme_urun gu
        JOIN crm_urun u ON gu.urun_id = u.id
        LEFT JOIN crm_urun_gorsel g ON g.urun_id = u.id
        WHERE gu.gorusme_id = ?
        ORDER BY gu.id
    """, (gorusme_id,))

    # ── Türkçe font desteği için DejaVu ───────────────────────────────────────
    # Sisteminizde DejaVu yoksa reportlab varsayılan Helvetica kullanır
    APP_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    font_registered = False
    for font_path in [
        os.path.join(APP_DIR, 'static', 'fonts', 'DejaVuSans.ttf'),
        'C:/Windows/Fonts/arial.ttf',
        'C:/Windows/Fonts/tahoma.ttf',
    ]:
        if os.path.exists(font_path):
            try:
                font_name = os.path.splitext(os.path.basename(font_path))[0]
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                font_registered = True
                break
            except Exception:
                continue

    BASE_FONT = font_name if font_registered else 'Helvetica'
    SOL_COLOR = colors.HexColor('#1E6B45')

    # ── Stiller ───────────────────────────────────────────────────────────────
    styles = getSampleStyleSheet()
    def _style(name, parent='Normal', **kw):
        s = ParagraphStyle(name, parent=styles[parent], fontName=BASE_FONT, **kw)
        return s

    s_title   = _style('PDFTitle',   fontSize=16, textColor=SOL_COLOR, spaceAfter=2, leading=20)
    s_sub     = _style('PDFSub',     fontSize=10, textColor=colors.grey, spaceAfter=6)
    s_heading = _style('PDFHead',    fontSize=11, textColor=SOL_COLOR, spaceBefore=8, spaceAfter=3)
    s_normal  = _style('PDFNormal',  fontSize=9,  leading=13)
    s_small   = _style('PDFSmall',   fontSize=8,  textColor=colors.grey)
    s_bold    = _style('PDFBold',    fontSize=9,  leading=13)
    s_price   = _style('PDFPrice',   fontSize=11, textColor=colors.HexColor('#059669'), leading=14)
    s_strike  = _style('PDFStrike',  fontSize=9,  textColor=colors.grey)
    s_center  = _style('PDFCenter',  fontSize=9,  alignment=TA_CENTER)

    def P(text, style=None):
        return Paragraph(str(text or '—'), style or s_normal)

    # ── PDF belgesi ───────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    story = []

    # Başlık
    story.append(P('SOLARIZ — GARDA 2026 FUARI', s_title))
    story.append(P('Müşteri Görüşme Özeti', s_sub))
    story.append(HRFlowable(width='100%', thickness=1.5, color=SOL_COLOR, spaceAfter=8))

    # Firma bilgisi
    story.append(P('FİRMA BİLGİSİ', s_heading))
    firma_data = [
        ['Firma Adı:', firma.get('firma_adi','') or '—'],
        ['Yetkili:', firma.get('yetkili','') or '—'],
        ['Telefon:', firma.get('telefon','') or '—'],
        ['WhatsApp:', firma.get('whatsapp','') or '—'],
        ['E-Posta:', firma.get('email','') or '—'],
        ['Ülke / Şehir:', (firma.get('ulke','') or '') + (' / ' + firma.get('sehir','') if firma.get('sehir') else '')],
    ]
    firma_tbl = Table(
        [[P(r[0], s_bold), P(r[1], s_normal)] for r in firma_data],
        colWidths=[3.5*cm, None]
    )
    firma_tbl.setStyle(TableStyle([
        ('VALIGN',    (0,0), (-1,-1), 'TOP'),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.whitesmoke, colors.white]),
        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
    ]))
    story.append(firma_tbl)
    story.append(Spacer(1, 8))

    # Görüşme bilgisi
    story.append(P('GÖRÜŞME BİLGİSİ', s_heading))
    tarih_str = ''
    if gorusme.get('tarih'):
        try:
            tarih_str = str(gorusme['tarih'])[:10]
        except Exception:
            tarih_str = str(gorusme.get('tarih',''))

    gor_data = [
        ['Tarih:', tarih_str or '—'],
        ['Görüşen:', gorusme.get('gorusen','') or '—'],
        ['Durum:', gorusme.get('durum','') or '—'],
        ['Takip Tarihi:', str(gorusme.get('takip_tarihi','') or '—')],
        ['Not:', gorusme.get('not_text','') or '—'],
    ]
    gor_tbl = Table(
        [[P(r[0], s_bold), P(r[1], s_normal)] for r in gor_data],
        colWidths=[3.5*cm, None]
    )
    gor_tbl.setStyle(TableStyle([
        ('VALIGN',    (0,0), (-1,-1), 'TOP'),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.whitesmoke, colors.white]),
        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),
    ]))
    story.append(gor_tbl)
    story.append(Spacer(1, 10))

    # Ürünler
    if urunler:
        story.append(P(f'GÖRÜŞÜLEN ÜRÜNLER ({len(urunler)} adet)', s_heading))
        story.append(HRFlowable(width='100%', thickness=0.5, color=colors.lightgrey, spaceAfter=6))

        STATIC_DIR = os.path.join(APP_DIR, 'static')

        for idx, u in enumerate(urunler, 1):
            # Görsel
            gorsel_cell = Spacer(1.8*cm, 1.8*cm)
            if u.get('gorsel_yolu'):
                img_path = os.path.join(STATIC_DIR, u['gorsel_yolu'].lstrip('/').replace('/', os.sep))
                if os.path.exists(img_path):
                    try:
                        gorsel_cell = RLImage(img_path, width=1.8*cm, height=1.8*cm)
                    except Exception:
                        pass

            # Ürün bilgi sütunu
            urun_info = []
            urun_info.append(P(f"#{idx}  {u.get('model_no','')}", s_bold))
            cat_parts = [x for x in [u.get('kategori'), u.get('tip'), u.get('urun_cinsi')] if x]
            if cat_parts:
                urun_info.append(P(' / '.join(cat_parts), s_small))
            if u.get('asorti'):
                urun_info.append(P(f"Asorti: {u['asorti']}", s_small))
            if u.get('malzeme_bilgisi'):
                urun_info.append(P(f"Malzeme: {u['malzeme_bilgisi']}", s_small))

            # Fiyat sütunu
            fiyat_info = []
            if u.get('liste_fiyat'):
                fiyat_info.append(P(f"Liste: {u['liste_fiyat']:.2f} USD", s_strike))
            if u.get('verilen_fiyat'):
                pb = u.get('para_birimi') or 'USD'
                fiyat_info.append(P(f"Verilen: {u['verilen_fiyat']:.2f} {pb}", s_price))
            if u.get('fiyat_konusuldu'):
                fiyat_info.append(P('✓ Fiyat konuşuldu', s_small))
            if u.get('indirim_notu'):
                fiyat_info.append(P(f"İndirim: {u['indirim_notu']}", s_small))
            if not fiyat_info:
                fiyat_info.append(P('—', s_small))

            # Sipariş/talep sütunu
            siparis_info = []
            if u.get('istenen_renk'):
                siparis_info.append(P(f"Renk: {u['istenen_renk']}", s_normal))
            if u.get('renk_basi_adet'):
                siparis_info.append(P(f"Renk başı: {u['renk_basi_adet']}", s_small))
            if u.get('toplam_adet'):
                siparis_info.append(P(f"Toplam: {u['toplam_adet']}", s_small))
            if u.get('teslim_notu'):
                siparis_info.append(P(f"Teslim: {u['teslim_notu']}", s_small))
            if not siparis_info:
                siparis_info.append(P('—', s_small))

            # Numune sütunu
            numune_info = []
            if u.get('numune_istendi'):
                numune_info.append(P('✓ Numune istendi', s_bold))
                if u.get('numune_adet'):
                    numune_info.append(P(f"Adet: {u['numune_adet']}", s_small))
                if u.get('numune_beden'):
                    numune_info.append(P(f"Beden: {u['numune_beden']}", s_small))
            else:
                numune_info.append(P('Numune yok', s_small))

            if u.get('urun_notu'):
                numune_info.append(Spacer(1, 3))
                numune_info.append(P(f"Not: {u['urun_notu']}", s_small))

            urun_row = Table(
                [[gorsel_cell, urun_info, fiyat_info, siparis_info, numune_info]],
                colWidths=[2*cm, 5.5*cm, 4*cm, 3.5*cm, 3*cm]
            )
            urun_row.setStyle(TableStyle([
                ('VALIGN',    (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING',  (0,0), (-1,-1), 4),
                ('RIGHTPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING',   (0,0), (-1,-1), 4),
                ('BOTTOMPADDING',(0,0), (-1,-1), 4),
                ('BOX', (0,0), (-1,-1), 0.5, colors.lightgrey),
                ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white]),
            ]))
            story.append(urun_row)
            story.append(Spacer(1, 4))
    else:
        story.append(P('Bu görüşmeye ürün eklenmemiş.', s_small))

    # Alt bilgi
    story.append(Spacer(1, 12))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.lightgrey))
    story.append(Spacer(1, 4))
    now_str = datetime.datetime.now().strftime('%d.%m.%Y %H:%M')
    story.append(P(f'Oluşturulma: {now_str}  |  Solariz CPS — Garda 2026 Fuar CRM', s_small))

    # ── PDF oluştur ────────────────────────────────────────────────────────────
    doc.build(story)
    buffer.seek(0)

    # Dosya adı
    firma_adi_safe = (firma.get('firma_adi') or 'Firma').replace(' ', '_')[:30]
    tarih_safe     = (tarih_str or datetime.date.today().isoformat()).replace('-','')
    filename       = f"{firma_adi_safe}_Garda2026_{tarih_safe}.pdf"

    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── KATALOG YÖNETİMİ ----------------------------------------------------------

@fuar_crm_bp.route('/kataloglar')
@login_gerekli
def katalog_listesi():
    if not _crm_erisim():
        abort(403)
    kataloglar = _q("""
        SELECT k.id, k.ad, k.fuar_adi, k.aciklama, k.aktif, k.created_at,
               COUNT(u.id) AS urun_sayisi
        FROM crm_katalog k
        LEFT JOIN crm_urun u ON u.katalog_id = k.id AND u.aktif = 1
        GROUP BY k.id
        ORDER BY k.id DESC
    """) or []
    return render_template('fuar_crm/katalog_listesi.html', kataloglar=kataloglar)


@fuar_crm_bp.route('/katalog/olustur', methods=['GET', 'POST'])
@login_gerekli
def katalog_olustur():
    if not _crm_erisim():
        abort(403)
    if request.method == 'POST':
        ad       = (request.form.get('ad') or '').strip()
        fuar_adi = (request.form.get('fuar_adi') or '').strip()
        aciklama = (request.form.get('aciklama') or '').strip()
        if not ad:
            flash('Katalog adi zorunlu.', 'warning')
            return redirect(url_for('fuar_crm.katalog_olustur'))
        conn = _get_conn()
        conn.execute(
            "INSERT INTO crm_katalog (ad, fuar_adi, aciklama, aktif) VALUES (?,?,?,0)",
            (ad, fuar_adi or ad, aciklama or None)
        )
        conn.commit()
        flash(f'Katalog oluşturuldu: {ad}', 'success')
        return redirect(url_for('fuar_crm.katalog_listesi'))
    return render_template('fuar_crm/katalog_olustur.html')


@fuar_crm_bp.route('/katalog/<int:katalog_id>/aktif-yap', methods=['POST'])
@login_gerekli
def katalog_aktif_yap(katalog_id):
    if not _crm_erisim():
        abort(403)
    katalog = _qone("SELECT id, ad FROM crm_katalog WHERE id = ?", (katalog_id,))
    if not katalog:
        abort(404)
    conn = _get_conn()
    conn.execute("UPDATE crm_katalog SET aktif = 0")
    conn.execute("UPDATE crm_katalog SET aktif = 1 WHERE id = ?", (katalog_id,))
    conn.commit()
    flash(f'Kullanımdaki katalog değiştirildi: {katalog["ad"]}', 'success')
    return redirect(url_for('fuar_crm.katalog_listesi'))


@fuar_crm_bp.route('/katalog/ornek-excel')
@login_gerekli
def ornek_excel_indir():
    if not _crm_erisim():
        abort(403)
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io as _io
    except ImportError:
        flash('openpyxl yuklu degil.', 'danger')
        return redirect(url_for('fuar_crm.katalog_listesi'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2026 Katalog'

    basliklar = {
        1:  'Urun Fotograf',
        2:  'Taban',
        3:  'Model No',
        4:  'Kategori',
        5:  'Tip',
        6:  'urun cinsi',
        7:  'Asorti',
        8:  'Asorti Dagilimi',
        20: 'Birim Fiyat',
        21: 'Malzeme Bilgisi',
        22: 'Sarfiyat',
        29: 'Maliyet',
        32: 'Kur',
        35: 'Marj',
    }
    hdr_fill = PatternFill('solid', fgColor='7C3AED')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    for col, title in basliklar.items():
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')

    ornekler = [
        ('BRZ-9000', 'Terlik',   'Karbot', 'Kadin',  '36/40', '36:2/37:2/38:2/39:2/40:2', 4.95, 'EVA + Tekstil', 2.5, 3.96, 'USD', '1.25'),
        ('CRP-8100', 'Sandalet', 'Poli',   'COCUK',  '22/33', '22:2/24:2/26:2', 3.95, 'PVC + Eva', 1.8, 3.16, 'USD', '1.25'),
        ('Z107141',  'Terlik',   'Poli',   'ERKEK',  '36/41', '36:2/37:2/38:2/39:2/40:2/41:2', 6.50, 'Poli + Eva', 3.0, 5.20, 'USD', '1.25'),
    ]
    for row_idx, o in enumerate(ornekler, start=2):
        model_no, kat, tip, cinsi, asorti, asorti_dag, fiyat, malzeme, sarfiyat, maliyet, kur, marj = o
        ws.cell(row=row_idx, column=3,  value=model_no)
        ws.cell(row=row_idx, column=4,  value=kat)
        ws.cell(row=row_idx, column=5,  value=tip)
        ws.cell(row=row_idx, column=6,  value=cinsi)
        ws.cell(row=row_idx, column=7,  value=asorti)
        ws.cell(row=row_idx, column=8,  value=asorti_dag)
        ws.cell(row=row_idx, column=20, value=fiyat)
        ws.cell(row=row_idx, column=21, value=malzeme)
        ws.cell(row=row_idx, column=22, value=sarfiyat)
        ws.cell(row=row_idx, column=29, value=maliyet)
        ws.cell(row=row_idx, column=32, value=kur)
        ws.cell(row=row_idx, column=35, value=marj)

    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['U'].width = 22

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename="katalog_ornek.xlsx"'
    return resp


@fuar_crm_bp.route('/katalog/<int:katalog_id>/excel-yukle', methods=['GET', 'POST'])
@login_gerekli
def katalog_excel_yukle(katalog_id):
    if not _crm_erisim():
        abort(403)
    katalog = _qone("SELECT * FROM crm_katalog WHERE id = ?", (katalog_id,))
    if not katalog:
        abort(404)

    if request.method == 'GET':
        return render_template('fuar_crm/katalog_excel_yukle.html', katalog=katalog)

    dosya = request.files.get('excel_dosya')
    if not dosya or not dosya.filename:
        flash('Excel dosyasi secilmedi.', 'warning')
        return redirect(url_for('fuar_crm.katalog_excel_yukle', katalog_id=katalog_id))
    if not dosya.filename.lower().endswith('.xlsx'):
        flash('Sadece .xlsx dosyasi kabul edilir.', 'warning')
        return redirect(url_for('fuar_crm.katalog_excel_yukle', katalog_id=katalog_id))

    try:
        import openpyxl
        import io as _io
    except ImportError:
        flash('openpyxl yuklu degil.', 'danger')
        return redirect(url_for('fuar_crm.katalog_excel_yukle', katalog_id=katalog_id))

    try:
        wb = openpyxl.load_workbook(_io.BytesIO(dosya.read()), data_only=True)
    except Exception as e:
        flash(f'Excel acilamadi: {e}', 'danger')
        return redirect(url_for('fuar_crm.katalog_excel_yukle', katalog_id=katalog_id))

    COL_MODEL_NO        = 2
    COL_KATEGORI        = 3
    COL_TIP             = 4
    COL_URUN_CINSI      = 5
    COL_ASORTI          = 6
    COL_ASORTI_DAGILIMI = 7
    COL_BIRIM_FIYAT     = 19
    COL_MALZEME_BILGISI = 20
    COL_SARFIYAT        = 21
    COL_MALIYET         = 28
    COL_KUR             = 31
    COL_MARJ            = 34

    def _c(row, idx):
        cells = list(row)
        if idx < len(cells):
            v = cells[idx].value
            if v is None:
                return None
            return v.strip() if isinstance(v, str) else v
        return None

    def _cs(row, idx):
        v = _c(row, idx)
        if v is None:
            return None
        s = str(v).replace('\n', ' / ').strip()
        return s if s else None

    def _fl(v):
        if v is None:
            return None
        try:
            return float(str(v).replace(',', '.').strip())
        except (ValueError, TypeError):
            return None

    def _st(v):
        if v is None:
            return None
        s = str(v).strip()
        return s if s else None

    fuar_adi = katalog['fuar_adi'] or katalog['ad']
    eklenen = atlanan = bos_model = 0
    hatalar = []
    conn = _get_conn()

    for sheet_adi in wb.sheetnames:
        ws = wb[sheet_adi]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            model_no = _st(_c(row, COL_MODEL_NO))
            if not model_no:
                bos_model += 1
                continue
            mevcut = conn.execute(
                "SELECT id FROM crm_urun WHERE katalog_id=? AND sheet_adi=? AND excel_satir_no=?",
                (katalog_id, sheet_adi, row_idx)
            ).fetchone()
            if mevcut:
                atlanan += 1
                continue
            try:
                conn.execute("""
                    INSERT INTO crm_urun
                        (fuar_adi, sheet_adi, excel_satir_no, model_no, kategori, tip,
                         urun_cinsi, asorti, asorti_dagilimi, birim_fiyat, malzeme_bilgisi,
                         sarfiyat, maliyet, kur, marj, aktif, katalog_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,1,?)
                """, (
                    fuar_adi, sheet_adi, row_idx, model_no,
                    _st(_c(row, COL_KATEGORI)),
                    _st(_c(row, COL_TIP)),
                    _st(_c(row, COL_URUN_CINSI)),
                    _st(_c(row, COL_ASORTI)),
                    _cs(row, COL_ASORTI_DAGILIMI),
                    _fl(_c(row, COL_BIRIM_FIYAT)),
                    _st(_c(row, COL_MALZEME_BILGISI)),
                    _fl(_c(row, COL_SARFIYAT)),
                    _fl(_c(row, COL_MALIYET)),
                    _st(_c(row, COL_KUR)),
                    _st(_c(row, COL_MARJ)),
                    katalog_id,
                ))
                eklenen += 1
            except Exception as e:
                hatalar.append(f'Satir {row_idx}: {e}')

    conn.commit()
    ozet = f'{eklenen} urun eklendi, {atlanan} atlanda (duplicate), {bos_model} bos model atlanda.'
    if hatalar:
        ozet += f' {len(hatalar)} hata olustu.'
    flash(ozet, 'success' if not hatalar else 'warning')
    return redirect(url_for('fuar_crm.katalog_listesi'))
