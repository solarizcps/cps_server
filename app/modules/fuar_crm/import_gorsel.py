"""
import_gorsel.py — Excel'den urun gorsellerini cikartip DB'ye kaydet.

Kullanim:
    python app/modules/fuar_crm/import_gorsel.py [--dry-run]

Kurallar:
- Mevcut crm_urun_gorsel kayitlari guncellenmez (UNIQUE sheet+satir).
- crm_urun silme/degistirme yapilmaz.
- Gorsel dosyalari app/static/uploads/fuar_crm/urunler/ klasorune kaydedilir.
- Eslesme: sheet_adi + excel_satir_no -> crm_urun.id
"""
import sys, os, glob, sqlite3, re

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR   = os.path.join(os.path.dirname(__file__), '..', '..', '..')  # -> C:\Solariz_CPS_SERVER
APP_DIR    = os.path.join(BASE_DIR, 'app')
DB_PATH    = os.path.join(APP_DIR, 'mock_data.db')
OUT_DIR    = os.path.join(APP_DIR, 'static', 'uploads', 'fuar_crm', 'urunler')
DRY_RUN    = '--dry-run' in sys.argv

HEDEF_SHEETLER = ['2026 Garda', '2026 Garda (2)']

# Excel dosyasini bul
_matches = [p for p in glob.glob(
    r'C:\Users\LENOVO\Desktop\Solariz Fuar\*.xlsx'
) if not os.path.basename(p).startswith('~')]
EXCEL_PATH = _matches[0] if _matches else ''

if not EXCEL_PATH:
    print("[HATA] Excel dosyasi bulunamadi")
    sys.exit(1)

print(f"Excel : {EXCEL_PATH}")
print(f"Out   : {OUT_DIR}")
print(f"DB    : {DB_PATH}")
print(f"DryRun: {DRY_RUN}")
print()

os.makedirs(OUT_DIR, exist_ok=True)

# ── 1) crm_urun'dan sheet+satir -> id haritasi ──────────────────────────────
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
cur  = conn.cursor()

rows = cur.execute(
    "SELECT id, sheet_adi, excel_satir_no, model_no FROM crm_urun WHERE aktif=1"
).fetchall()

urun_map = {}   # (sheet_adi_normalized, excel_satir_no) -> (urun_id, model_no)
for r in rows:
    key = (r['sheet_adi'], r['excel_satir_no'])
    urun_map[key] = (r['id'], r['model_no'])

print(f"crm_urun haritalandi: {len(urun_map)} kayit")

# ── 2) Mevcut gorselleri oku (duplicate atlamak icin) ───────────────────────
mevcut = set()
for mr in cur.execute("SELECT sheet_adi, excel_satir_no FROM crm_urun_gorsel").fetchall():
    mevcut.add((mr['sheet_adi'], mr['excel_satir_no']))
print(f"Mevcut gorsel kayitlari: {len(mevcut)}")

# ── 3) Workbook'u yukle ─────────────────────────────────────────────────────
from openpyxl import load_workbook
print("Workbook yukleniyor...")
wb = load_workbook(EXCEL_PATH, data_only=True)
print("Workbook yuklendi.\n")

# ── 4) Her sheet'teki gorselleri isle ───────────────────────────────────────
sayac = {
    'toplam_gorsel'  : 0,
    'dosyaya_cikti'  : 0,
    'es_bulundu'     : 0,
    'es_bulunamadi'  : 0,
    'zaten_var'      : 0,
    'hata'           : 0,
}
es_bulunamadi_ornekler = []

for sheet_adi in HEDEF_SHEETLER:
    if sheet_adi not in wb.sheetnames:
        print(f"[ATLA] Sheet yok: {sheet_adi}")
        continue

    ws     = wb[sheet_adi]
    images = getattr(ws, '_images', [])
    print(f"Sheet: '{sheet_adi}'  =>  {len(images)} gorsel")

    for img in images:
        sayac['toplam_gorsel'] += 1

        # Anchor -> satir no
        anchor = getattr(img, 'anchor', None)
        if anchor is None:
            sayac['hata'] += 1
            continue

        if hasattr(anchor, '_from'):
            excel_satir = anchor._from.row + 1  # 0-indexed -> 1-indexed
        elif hasattr(anchor, 'row'):
            excel_satir = anchor.row
        else:
            sayac['hata'] += 1
            continue

        key = (sheet_adi, excel_satir)

        # Zaten var mi?
        if key in mevcut:
            sayac['zaten_var'] += 1
            continue

        # crm_urun eslesme
        if key not in urun_map:
            sayac['es_bulunamadi'] += 1
            if len(es_bulunamadi_ornekler) < 5:
                es_bulunamadi_ornekler.append(f"{sheet_adi} satir={excel_satir}")
            continue

        urun_id, model_no = urun_map[key]
        sayac['es_bulundu'] += 1

        # Dosya adi: guvenli karakter
        safe_sheet = re.sub(r'[^\w]', '_', sheet_adi)
        fname = f"GARDA_2026_{safe_sheet}_{excel_satir}_{urun_id}.png"
        fpath = os.path.join(OUT_DIR, fname)
        rel_path = f"uploads/fuar_crm/urunler/{fname}"

        if DRY_RUN:
            sayac['dosyaya_cikti'] += 1
            continue

        # Gorsel verisini yaz
        try:
            img_data = img._data()
            with open(fpath, 'wb') as f:
                f.write(img_data)
            sayac['dosyaya_cikti'] += 1
        except Exception as e:
            print(f"  [HATA] Dosya yazma: {fname} — {e}")
            sayac['hata'] += 1
            continue

        # DB'ye kaydet
        try:
            cur.execute("""
                INSERT OR IGNORE INTO crm_urun_gorsel
                    (urun_id, sheet_adi, excel_satir_no, dosya_yolu)
                VALUES (?, ?, ?, ?)
            """, (urun_id, sheet_adi, excel_satir, rel_path))
        except Exception as e:
            print(f"  [DB HATA] urun_id={urun_id}: {e}")
            sayac['hata'] += 1

if not DRY_RUN:
    conn.commit()

conn.close()

# ── 5) Rapor ────────────────────────────────────────────────────────────────
print()
print("=" * 50)
print("GORSEL IMPORT RAPORU")
print("=" * 50)
print(f"Excel'de bulunan gorsel     : {sayac['toplam_gorsel']}")
print(f"Dosyaya cikartilan          : {sayac['dosyaya_cikti']}")
print(f"crm_urun ile eslesti        : {sayac['es_bulundu']}")
print(f"Eslesme bulunamadi          : {sayac['es_bulunamadi']}")
print(f"Zaten var (atlandi)         : {sayac['zaten_var']}")
print(f"Hata                        : {sayac['hata']}")
if es_bulunamadi_ornekler:
    print(f"Eslesmeyenler (ornek)       : {es_bulunamadi_ornekler}")
print("=" * 50)
if DRY_RUN:
    print("DRY-RUN modu: hic bir sey yazilmadi.")
