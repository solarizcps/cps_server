"""
update_maliyet.py — crm_urun.maliyet alanini Excel AD sutunundan gunceller.

Kural:
- crm_urun kayitlarini SILMEZ.
- Sadece maliyet alanini UPDATE eder.
- Eslestirme: sheet_adi + excel_satir_no
- Bos maliyet → NULL (degistirmez)
- Diger alanlara (fiyat, gorsel, model_no, vb.) DOKUNMAZ.
- Tekrar calistirilabilir (idempotent).

Kullanim:
    python app/modules/fuar_crm/update_maliyet.py [--dry-run]
"""
import sys, os, glob, sqlite3

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', '..', '..'))
DB_PATH  = os.path.join(BASE_DIR, 'app', 'mock_data.db')
DRY_RUN  = '--dry-run' in sys.argv

_matches = [p for p in glob.glob(r'C:\Users\LENOVO\Desktop\Solariz Fuar\*.xlsx')
            if not os.path.basename(p).startswith('~')]
EXCEL_PATH = _matches[0] if _matches else ''

if not EXCEL_PATH:
    print("[HATA] Excel bulunamadi")
    sys.exit(1)

HEDEF_SHEETLER = ['2026 Garda', '2026 Garda (2)']
COL_MALIYET    = 28   # AD sutunu, 0-indexed

print(f"Excel  : {EXCEL_PATH}")
print(f"DB     : {DB_PATH}")
print(f"DryRun : {DRY_RUN}")
print()

def _float_safe(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError):
        return None

# ── Excel'den (sheet, satir_no, maliyet) ciftlerini topla ───────────────────
from openpyxl import load_workbook
print("Workbook yukleniyor...")
wb = load_workbook(EXCEL_PATH, data_only=True)
print("Yuklu.\n")

excel_maliyet = {}   # (sheet_adi, satir_no) -> float | None
excel_dolu    = 0

for sheet_adi in HEDEF_SHEETLER:
    if sheet_adi not in wb.sheetnames:
        print(f"  [ATLA] Sheet yok: {sheet_adi}")
        continue
    ws = wb[sheet_adi]
    for row in ws.iter_rows(min_row=2):
        satir_no = row[0].row
        val      = row[COL_MALIYET].value if COL_MALIYET < len(row) else None
        maliyet  = _float_safe(val)
        excel_maliyet[(sheet_adi, satir_no)] = maliyet
        if maliyet is not None:
            excel_dolu += 1

print(f"Excel'de maliyet dolu satir  : {excel_dolu}")
print(f"Excel'de toplam satir izlendi: {len(excel_maliyet)}")
print()

# ── DB guncelle ──────────────────────────────────────────────────────────────
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
cur  = conn.cursor()

rows_db = cur.execute(
    "SELECT id, sheet_adi, excel_satir_no, model_no FROM crm_urun WHERE aktif=1"
).fetchall()
print(f"crm_urun kayit sayisi: {len(rows_db)}")

guncellendi    = 0
maliyet_yazildi= 0
maliyet_bos    = 0
eslesmedi      = 0

for r in rows_db:
    key = (r['sheet_adi'], r['excel_satir_no'])
    if key not in excel_maliyet:
        eslesmedi += 1
        continue
    maliyet = excel_maliyet[key]
    guncellendi += 1
    if maliyet is not None:
        maliyet_yazildi += 1
        if not DRY_RUN:
            cur.execute(
                "UPDATE crm_urun SET maliyet=? WHERE id=?",
                (maliyet, r['id'])
            )
    else:
        maliyet_bos += 1
        # Zaten NULL, dokunma

if not DRY_RUN:
    conn.commit()

conn.close()

# ── Rapor ────────────────────────────────────────────────────────────────────
print()
print("=" * 55)
print("MALIYET UPDATE RAPORU")
print("=" * 55)
print(f"Excel'de maliyet dolu satir    : {excel_dolu}")
print(f"crm_urun ile eslesen kayit     : {guncellendi}")
print(f"  Maliyet yazilan              : {maliyet_yazildi}")
print(f"  Maliyet bos birakilan (NULL) : {maliyet_bos}")
print(f"  Eslesmedi (sheet/satir yok)  : {eslesmedi}")
print("=" * 55)
if DRY_RUN:
    print("DRY-RUN: Hicbir sey yazilmadi.")
